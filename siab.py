### 📌 Bloque 1 (imports, DB, inicio de clase App hasta mitad de `init_legajo`)

import os, sys, sqlite3, shutil, re
import tkinter as tk
import unicodedata
import threading
from pdf_manager import PDFManager
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from datetime import date, datetime
from tkinter import *
from tkinter.messagebox import showwarning
from ui_helpers import UIHelpers
import string
import secrets
import random
import time
import inspect
import pandas as pd
import types
import traceback
from tkinter import font
import tkinter.font as tkFont
from tkinter import END
from PIL import Image, ImageTk
from tkcalendar import DateEntry
from tkinter import ttk, filedialog, simpledialog, messagebox, Toplevel, Label, Button, Frame
from reportlab.platypus import Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import smtplib
from email.message import EmailMessage
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from tkinter import BOTH, RIGHT, Y, LEFT, X
from tkinter import Entry
try:
    RESAMPLE = Image.Resampling.LANCZOS  # Pillow >= 9.1
except AttributeError:
    RESAMPLE = getattr(Image, "LANCZOS", Image.BICUBIC)  # Pillow < 9.1

# ==========================
# Colores de texto por botón
# ==========================
COLORES_TEXTO_BOTON = {
    "Firmar actividad": "#2E8B57",     # verde
    "Firmar Supervisor": "#1E5AA8",    # azul
    "No firmar ahora": "#6C757D",      # gris
    "default": "#000000"
}
# ==========================
# DICCIONARIO PERMISOS ACTIVIDADES
# ==========================
PERMISOS_ACTIVIDADES = {
    "ADMIN": {
        "inicial": ("Nuevo", "Buscar", "Impr.Listado Actividad", "Pendientes de Firma"),
        "nuevo": ("Guardar", "Limpiar/Cancelar"),
        "cargado": (
            "Nuevo", "Modificar", "Anular", "Limpiar/Cancelar", "Buscar",
            "Exportar PDF", "Imprimir Actividad", "Impr.Listado Actividad",
            "No firmar ahora", "Firmar actividad", "Ver historial", "Firmar Supervisor", "Pendientes de Firma"
        ),
        "editando": ("Guardar", "Limpiar/Cancelar"),
        "anulada": ("Nuevo", "Buscar", "Limpiar/Cancelar", "Ver historial", "Pendientes de Firma"),
        "procesando": ()
    },

    "SUPERVISOR": {
        "inicial": ("Nuevo", "Buscar", "Impr.Listado Actividad", "Pendientes de Firma"),
        "nuevo": ("Guardar", "Limpiar/Cancelar"),
        "cargado": (
            "Nuevo", "Modificar", "Limpiar/Cancelar", "Buscar",
            "Exportar PDF", "Imprimir Actividad", "Impr.Listado Actividad",
            "No firmar ahora", "Firmar actividad", "Ver historial", "Firmar Supervisor", "Pendientes de Firma"
        ),
        "editando": ("Guardar", "Limpiar/Cancelar"),
        "anulada": ("Nuevo", "Buscar", "Limpiar/Cancelar", "Ver historial", "Pendientes de Firma"),
        "procesando": ()
    },

    "BOMBERO": {
        "inicial": ("Nuevo", "Buscar", "Pendientes de Firma"),
        "nuevo": ("Guardar", "Limpiar/Cancelar"),
        "cargado": (
            "Nuevo", "Modificar", "Limpiar/Cancelar", "Buscar",
            "Exportar PDF", "Imprimir Actividad", "No firmar ahora", 
            "Firmar actividad", "Pendientes de Firma"
        ),
        "firmado_bombero": (
            "Nuevo", "Buscar", "Limpiar/Cancelar", "Exportar PDF", 
            "Ver historial", "Pendientes de Firma" # 👈 Agregado
        ),
        "firmado_supervisor": (
            "Nuevo", "Buscar", "Limpiar/Cancelar", "Exportar PDF", 
            "Ver historial", "Pendientes de Firma" # 👈 Agregado
        ),        
        "editando": ("Guardar", "Limpiar/Cancelar"),
        "anulada": ("Nuevo", "Buscar", "Limpiar/Cancelar", "Pendientes de Firma"), # 👈 Agregado
        "procesando": ()
    },
}

# -------------------- rutas --------------------
import os
import sys
import sqlite3

def app_base_path():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.abspath(".")

def resource_path(relative_path):
    try:
        # Esto busca dentro del EXE (Carpeta temporal _MEIPASS)
        base_path = sys._MEIPASS
    except Exception:
        base_path = app_base_path()
    return os.path.join(base_path, relative_path)

# --- CONFIGURACIÓN DE LA BASE DE DATOS EN C: ---
ruta_disco_c = r"C:\SIAB_Sistema"
if not os.path.exists(ruta_disco_c):
    try:
        os.makedirs(ruta_disco_c)
    except Exception as e:
        print(f"Error al crear carpeta en C: {e}")
        # Si falla (por permisos), cae a la carpeta del usuario como respaldo
        ruta_disco_c = os.path.join(os.path.expanduser("~"), "SIAB_Datos")
        if not os.path.exists(ruta_disco_c): os.makedirs(ruta_disco_c)

db_path = os.path.join(ruta_disco_c, "siab.db")
conexion = sqlite3.connect(db_path)

# --- CARGA DEL ICONO (Usando resource_path) ---
def cargar_icono_global():
    from PIL import Image, ImageTk
    try:
        # Aquí usamos resource_path porque el PNG sí va dentro del EXE
        ruta = resource_path("Bomberos.png") 

        if os.path.exists(ruta):
            img = Image.open(ruta).resize((32, 32), Image.Resampling.LANCZOS)
            return ImageTk.PhotoImage(img)
        return None
    except Exception as e:
        print("⚠ Error cargando icono:", e)
        return None
    
def parse_ddmmyyyy(fecha_str):
    """
    Convierte una fecha en formato dd/mm/yyyy (string) a objeto datetime.date.
    Retorna None si está vacía o inválida.
    """
    if not fecha_str:
        return None
    try:
        return datetime.strptime(fecha_str.strip(), "%d/%m/%Y").date()
    except Exception:
        return None

APP_DIR = app_base_path()
DB_PATH = os.path.join(APP_DIR, "siab.db")
FOTOS_DIR = os.path.join(APP_DIR, "fotos")
LOGO_PATH = resource_path("bomberos.png")
os.makedirs(FOTOS_DIR, exist_ok=True)
TIEMPO_INACTIVIDAD = 60 * 1000  # 1 minuto
TIEMPO_AVISO = 30 * 1000              # 30 segundos antes

    # -------------------- DB --------------------
import mysql.connector # Asegúrate de añadir este import al inicio del archivo

# Configuración de conexión (Ajusta con tus datos de la base vacía)
DB_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': 'siab1234',
    'database': 'siab_vacia'
}

def get_db_connection():
    """Establece conexión con MySQL centralizado"""
    return mysql.connector.connect(**DB_CONFIG)

def init_db():
    try:
        # 1. Establecer conexión
        conn = get_db_connection()
        c = conn.cursor()

        # -------------------------------------------------
        # 2. TABLAS ORIGINALES (Migradas a MySQL)
        # -------------------------------------------------
        
        # LEGAJOS
        c.execute("""
        CREATE TABLE IF NOT EXISTS legajos (
            legajo VARCHAR(20) PRIMARY KEY,
            apellido VARCHAR(100), nombre VARCHAR(100), dni VARCHAR(20),
            grado VARCHAR(50), cargo VARCHAR(50), email VARCHAR(100),
            nro_cel VARCHAR(50), foto VARCHAR(255), situacion VARCHAR(50),
            autoriza VARCHAR(10), fecha_baja DATE
        ) ENGINE=InnoDB;
        """)

        # USUARIOS (Con el nuevo rol ENCARGADO)
        c.execute("""
        CREATE TABLE IF NOT EXISTS usuarios (
            id INT AUTO_INCREMENT PRIMARY KEY,
            username VARCHAR(50) UNIQUE NOT NULL,
            password_hash VARCHAR(255) NOT NULL,
            rol ENUM('BOMBERO', 'ENCARGADO', 'SUPERVISOR', 'ADMIN') NOT NULL,
            legajo VARCHAR(20),
            activo TINYINT DEFAULT 1,
            fecha_baja DATETIME,
            usuario_baja VARCHAR(50),
            debe_cambiar_password TINYINT DEFAULT 1,
            FOREIGN KEY (legajo) REFERENCES legajos(legajo) ON DELETE SET NULL
        ) ENGINE=InnoDB;
        """)

        # CONCEPTOS
        c.execute("""
        CREATE TABLE IF NOT EXISTS conceptos (
            id INT AUTO_INCREMENT PRIMARY KEY,
            concepto VARCHAR(100),
            puntos INT,
            detalle TEXT
        ) ENGINE=InnoDB;
        """)

        # ACTIVIDADES (La estructura completa que ya usas)
        c.execute("""
        CREATE TABLE IF NOT EXISTS actividades (
            id INT AUTO_INCREMENT PRIMARY KEY,
            legajo VARCHAR(20),
            asignado VARCHAR(100),
            actividad TEXT,
            concepto_id INT,
            area VARCHAR(100),
            fecha_inicio DATE,
            fecha_fin DATE,
            hora_inicio VARCHAR(10),
            hora_fin VARCHAR(10),
            descripcion TEXT,
            fecha_carga DATETIME,
            fecha_creacion DATETIME,
            fecha_modificacion DATETIME,
            usuario_id INT,
            creado_por INT,
            modificado_por INT,
            horas DECIMAL(10,2),
            firma_bombero_usuario VARCHAR(50),
            firma_bombero_fecha DATETIME,
            firma_supervisor_usuario VARCHAR(50),
            firma_supervisor_fecha DATETIME,
            anulada TINYINT DEFAULT 0,
            usuario_anula VARCHAR(50),
            fecha_anulacion DATETIME,
            motivo_anulacion TEXT,
            FOREIGN KEY (legajo) REFERENCES legajos(legajo)
        ) ENGINE=InnoDB;
        """)

        # HISTORIAL Y NOTIFICACIONES
        c.execute("CREATE TABLE IF NOT EXISTS actividades_historial (id INT AUTO_INCREMENT PRIMARY KEY, actividad_id INT, campo VARCHAR(50), valor_anterior TEXT, valor_nuevo TEXT, usuario_id INT, fecha DATETIME) ENGINE=InnoDB;")
        c.execute("CREATE TABLE IF NOT EXISTS notificaciones (id INT AUTO_INCREMENT PRIMARY KEY, actividad_id INT, tipo VARCHAR(50), destinatario VARCHAR(100), asunto VARCHAR(255), fecha_envio DATETIME, estado VARCHAR(20)) ENGINE=InnoDB;")

        # -------------------------------------------------
        # 3. TABLAS DE LA LÓGICA WEB (Puntajes Playón)
        # -------------------------------------------------
        
        # EVENTOS (Capacitaciones, reuniones, etc.)
        c.execute("""
        CREATE TABLE IF NOT EXISTS eventos (
            id INT AUTO_INCREMENT PRIMARY KEY,
            fecha DATE NOT NULL,
            tipo VARCHAR(50),
            nombre_departamento VARCHAR(100),
            descripcion TEXT,
            estado ENUM('BORRADOR', 'FINALIZADO', 'ANULADO') DEFAULT 'BORRADOR',
            usuario_creador VARCHAR(50)
        ) ENGINE=InnoDB;
        """)

        # TEMAS / POSTAS (Asignación de ENCARGADO)
        c.execute("""
        CREATE TABLE IF NOT EXISTS evento_temas (
            id INT AUTO_INCREMENT PRIMARY KEY,
            evento_id INT,
            nombre VARCHAR(100),
            calificador_legajo VARCHAR(20),
            orden INT,
            FOREIGN KEY (evento_id) REFERENCES eventos(id) ON DELETE CASCADE
        ) ENGINE=InnoDB;
        """)

        # NOTAS (La unión final)
        c.execute("""
        CREATE TABLE IF NOT EXISTS asistencia_notas (
            id INT AUTO_INCREMENT PRIMARY KEY,
            evento_id INT,
            legajo VARCHAR(20),
            tema_id INT,
            nota DECIMAL(3,2),
            fecha_carga TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE KEY (evento_id, legajo, tema_id),
            FOREIGN KEY (evento_id) REFERENCES eventos(id),
            FOREIGN KEY (tema_id) REFERENCES evento_temas(id)
        ) ENGINE=InnoDB;
        """)

        conn.commit()
        c.close()
        conn.close()
        
        # 4. Asegurar el Admin por defecto
        ensure_default_admin()
        print(">>> Base de datos MySQL inicializada y sincronizada.")

    except Exception as e:
        print(f"Error inicializando base de datos MySQL: {e}")

def fmt_date(fecha):
    return fecha.strftime("%Y-%m-%d")  # o "%d/%m/%Y" si preferís

# -------------------- seguridad: hashing de contraseñas --------------------
import hashlib, hmac, os

def hash_password(password: str) -> str:
    """Genera un hash seguro (salt + pbkdf2_hmac sha256)."""
    salt = os.urandom(16)
    dk = hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), salt, 100_000)
    return salt.hex() + ':' + dk.hex()

def verify_password(password: str, stored_hash: str) -> bool:
    """Verifica la contraseña contra el hash almacenado."""
    try:
        salt_hex, dk_hex = stored_hash.split(':')
        salt = bytes.fromhex(salt_hex)
        expected = bytes.fromhex(dk_hex)
        derived = hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), salt, 100_000)
        return hmac.compare_digest(derived, expected)
    except Exception:
        return False

def ensure_default_admin():
    """Si la tabla usuarios está vacía, crea un admin por defecto en MySQL."""
    try:
        conn = get_db_connection()
        c = conn.cursor()
        
        c.execute("SELECT COUNT(*) FROM usuarios")
        if c.fetchone()[0] == 0:
            # Aquí generamos el hash de la clave "admin"
            pwd_hash = hash_password("admin")
            
            query = "INSERT INTO usuarios (username, password_hash, rol) VALUES (%s, %s, %s)"
            c.execute(query, ("admin", pwd_hash, "ADMIN"))
            conn.commit()
            print(">>> Usuario admin por defecto creado con éxito en MySQL.")
            
        c.close()
        conn.close()
    except Exception as e:
        print(f"Error en ensure_default_admin: {e}")

class ToolTip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tipwindow = None

        widget.bind("<Enter>", self.show)
        widget.bind("<Leave>", self.hide)

    def show(self, event=None):

        try:
            if self.tipwindow:
                return

            if not self.widget.winfo_exists():
                return

            x = self.widget.winfo_rootx() + 20
            y = self.widget.winfo_rooty() + 20

            self.tipwindow = tw = tk.Toplevel(self.widget)

            # 🔴 CLAVE: hacer dependiente del widget
            tw.transient(self.widget)

            tw.wm_overrideredirect(True)
            tw.wm_geometry(f"+{x}+{y}")

            label = tk.Label(
                tw,
                text=self.text,
                bg="#ffffe0",
                relief="solid",
                borderwidth=1,
                font=("Arial", 8)
            )

            label.pack()

        except:
            self.tipwindow = None  # 🔴 limpiar referencia

    def hide(self, event=None):
        try:
            if self.tipwindow:
                if self.tipwindow.winfo_exists():
                    self.tipwindow.destroy()
        except:
            pass
        finally:
            self.tipwindow = None  # 🔴 SIEMPRE limpiar
            
class LoginWindow:
    def __init__(self, root, ui_helper):
        self.master = root
        self.ui = ui_helper
        self.logged_user = None
        
        # Una sola ventana secundaria
        self.top = tk.Toplevel(self.master)
        self.top.withdraw() # Ocultar mientras se configura

        # Esto pone el escudo de Bomberos Almafuerte en la esquina
        if hasattr(self.master, "_icono_global") and self.master._icono_global:
            self.top.iconphoto(True, self.master._icono_global)

        self.top.title("Login SIAB")
        self.top.configure(bg="#f0f0f0")
        
        # Diseño de la interfaz
        tk.Label(self.top, text="Usuario:", bg="#f0f0f0").grid(row=0, column=0, padx=10, pady=10)
        tk.Label(self.top, text="Contraseña:", bg="#f0f0f0").grid(row=1, column=0, padx=10, pady=10)

        self.user = tk.Entry(self.top)
        self.pwd = tk.Entry(self.top, show="*")
        self.user.grid(row=0, column=1, padx=10, pady=10)
        self.pwd.grid(row=1, column=1, padx=10, pady=10)

        self.btn_login = tk.Button(self.top, text="Entrar", command=self.login)
        self.btn_login.grid(row=2, column=0, columnspan=2, pady=10)

        # Centrado
        self.top.update_idletasks()
        ancho, alto = 300, 150
        x = (self.top.winfo_screenwidth() // 2) - (ancho // 2)
        y = (self.top.winfo_screenheight() // 2) - (alto // 2)
        self.top.geometry(f"{ancho}x{alto}+{x}+{y}")
        
        # Binds de teclado
        self.user.bind("<Return>", lambda e: self.pwd.focus_set())
        self.pwd.bind("<Return>", lambda e: self.btn_login.focus_set())
        self.btn_login.bind("<Return>", lambda e: self.login())

        # Mostrar ventana con foco
        self.top.after(150, lambda: (self.top.deiconify(), self.top.focus_force(), self.user.focus_force()))

    def login(self):
        user = self.user.get().strip().lower()
        password = self.pwd.get().strip()

        if not password:
            return

        try:
            # Conexión a MySQL usando la configuración centralizada
            conn = get_db_connection()
            c = conn.cursor()

            # En MySQL usamos %s en lugar de ?
            # Agregamos el LEFT JOIN para traer los datos del legajo asociados
            query = """
                SELECT 
                    u.id,
                    u.username,
                    u.password_hash,
                    u.rol,
                    u.debe_cambiar_password,
                    l.legajo,
                    l.apellido,
                    l.nombre
                FROM usuarios u
                LEFT JOIN legajos l ON u.legajo = l.legajo
                WHERE LOWER(u.username) = %s AND u.activo = 1
            """
            c.execute(query, (user,))
            row = c.fetchone()
            c.close()
            conn.close()

        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Error de Conexión", f"No se pudo conectar con el servidor: {e}")
            return

        # Validación de usuario existente
        if not row:
            from tkinter import messagebox
            messagebox.showerror("Error", "Usuario o contraseña incorrectos")
            return

        # Validación de contraseña (row[2] es el hash)
        if not verify_password(password, row[2]):
            from tkinter import messagebox
            messagebox.showerror("Error", "Usuario o contraseña incorrectos")
            return

        # 🔐 Usuario válido - Guardamos la sesión
        # El ROL ahora puede ser: 'ADMIN', 'SUPERVISOR', 'ENCARGADO' o 'BOMBERO'
        self.logged_user = {
            "id": row[0],
            "username": row[1],
            "rol": row[3].upper(), # Lo forzamos a mayúsculas para evitar errores
            "legajo": row[5],
            "apellido": row[6],
            "nombre": row[7],
        }

        # 🔴 Verificación de cambio obligatorio de password (campo debe_cambiar_password)
        if row[4] == 1:
            self.abrir_ventana_cambio_password(self.logged_user)
            return

        # ✅ Login exitoso
        print(f">>> Sesión iniciada: {self.logged_user['username']} ({self.logged_user['rol']})")
        self.top.destroy()

    def abrir_ventana_cambio_password(self, user_data):
        win = tk.Toplevel(self.top)
        win.title("Cambiar Contraseña Obligatorio")
        win.configure(bg="#f0f0f0")
        
        # Centrar la ventana usando el método que agregamos antes
        self.centrar_ventana(win, 400, 250)
        win.grab_set() # Bloquea la ventana de atrás

        tk.Label(win, text="Debe cambiar su contraseña inicial", 
                 font=("Arial", 10, "bold"), bg="#f0f0f0", fg="red").pack(pady=10)

        # Campos de texto
        tk.Label(win, text="Nueva Contraseña:", bg="#f0f0f0").pack()
        self.new_pwd = tk.Entry(win, show="*")
        self.new_pwd.pack(pady=5)

        tk.Label(win, text="Confirmar Contraseña:", bg="#f0f0f0").pack()
        self.conf_pwd = tk.Entry(win, show="*")
        self.conf_pwd.pack(pady=5)

        # FOCO INICIAL: Poner el cursor en la nueva contraseña
        self.new_pwd.focus_set()

        # Función interna para procesar el cambio
        def procesar_cambio():
            p1 = self.new_pwd.get()
            p2 = self.conf_pwd.get()
            if p1 and p1 == p2:
                # Aquí llamarías a tu función de base de datos para actualizar
                # update_password(user_data['id'], p1)
                from tkinter import messagebox
                messagebox.showinfo("SIAB", "Contraseña actualizada con éxito")
                win.destroy()
                self.top.destroy() # Cierra el login para entrar a la App
            else:
                from tkinter import messagebox
                messagebox.showerror("Error", "Las contraseñas no coinciden o están vacías")

        # BOTÓN ACEPTAR
        btn = tk.Button(win, text="Aceptar y Entrar", command=procesar_cambio, 
                         bg="#27ae60", fg="white", font=("Arial", 10, "bold"))
        btn.pack(pady=20)

        # EVENTOS DE TECLADO (Navegación con Enter)
        self.new_pwd.bind("<Return>", lambda e: self.conf_pwd.focus_set())
        self.conf_pwd.bind("<Return>", lambda e: btn.focus_set())
        btn.bind("<Return>", lambda e: procesar_cambio())

    def centrar_ventana(self, ventana, ancho, alto):

        ventana.update_idletasks()

        # referencia: ventana padre (login)
        parent = self.top

        x = parent.winfo_rootx()
        y = parent.winfo_rooty()
        w = parent.winfo_width()
        h = parent.winfo_height()

        pos_x = x + (w // 2) - (ancho // 2)
        pos_y = y + (h // 2) - (alto // 2)

        ventana.geometry(f"{ancho}x{alto}+{pos_x}+{pos_y}")

# ============================ APP ============================
class App:
    def __init__(self, master, usuario_actual, ui_helper): # 'root' es el nombre del parámetro
        self.master = master # Guardamos la referencia primero
        self.usuario_actual = usuario_actual
        self.ui = ui_helper
        
        # 🔴 CORRECCIÓN DE ICONO: Usa self.master
        if hasattr(self.master, "_icono_global") and self.master._icono_global:
            self.master.iconphoto(True, self.master._icono_global)

        self.master.title("SOCIEDAD BOMBEROS VOLUNTARIOS DE ALMAFUERTE")
        self.master.geometry("1100x550")
        self.master.configure(bg="red")

        # 2. CARGAR LOGO Y TÍTULO (Mover aquí arriba)
        if os.path.exists(LOGO_PATH):
            try:
                img = Image.open(LOGO_PATH).resize((40, 40), Image.Resampling.BILINEAR)
                self.logo_icon = ImageTk.PhotoImage(img)
                Label(self.master, image=self.logo_icon, bg="red").place(x=10, y=10)
                self.master.iconphoto(False, self.logo_icon)
            except:
                pass
        
        Label(self.master, text="SOCIEDAD BOMBEROS VOLUNTARIOS DE ALMAFUERTE",
              font=("Arial", 22, "bold"), fg="white", bg="red").place(x=60, y=15)

        # 3. FORZAR DIBUJO INMEDIATO
        self.master.update_idletasks()

        # --- Configurar Estilo 
        self.style = ttk.Style()
        self.style.theme_use("clam")
        # --- Estilo global para Treeview (mejor visibilidad) ---
        style = ttk.Style()
        style.theme_use("default")

        style.configure(
            "Custom.Horizontal.TProgressbar",
            troughcolor="#2c3e50",
            background="#27ae60",   # verde bombero
            thickness=8
        )
        style.configure("Treeview.Heading",
                        font=("Arial", 10, "bold"),
                        background="#e6e6e6",
                        relief="raised")

        style.configure("Treeview",
                        font=("Arial", 10),
                        rowheight=24,
                        borderwidth=1,
                        relief="solid")

        style.map("Treeview",
                background=[("selected", "#cce5ff")],
                foreground=[("selected", "black")])

        self.ui = UIHelpers(self.master)        
        # -----------------------------
        # CONTROL DE SESIÓN
        # -----------------------------

        self.tiempo_inactividad = 600000      # 10 minutos
        self.aviso_inactividad = 540000       # aviso 9 minutos

        self.timer_sesion = None
        self.timer_aviso = None
        self.ventana_aviso = None
        self.sesion_cerrada = False

        # detectar actividad del usuario
        self.master.bind_all("<Key>", self.reset_timer_sesion)
        self.master.bind_all("<Motion>", self.reset_timer_sesion)
        self.master.bind_all("<Button>", self.reset_timer_sesion)

        self.reset_timer_sesion()

        # Configuración SMTP
        self.smtp_user = "jefaturabva435@gmail.com"
        self.smtp_pass = "exat lbfr qpaq qcbl"
        self.smtp_server = "smtp.gmail.com"
        self.smtp_port = 465

        # --------------------------------------------------
        # ESTADO ACTUAL DE ACTIVIDAD (CLAVE)
        # --------------------------------------------------
        self.id_actividad_actual = None
        self.firma_bombero_fecha = None
        self.firma_supervisor_fecha = None
        self.modo_actividad = None
        self.actividad_legajo = None
        self.actividad_asignado = None
        self.actividad_anulada = False

        # --- Usuario actual, arriba a la derecha ---
        tk.Label(
            self.master,
            text=f"👤 {self.usuario_actual['username'].upper()}",
            font=("Segoe UI", 11, "bold"),
            fg="yellow",
            bg="red",
            anchor="e"
        ).place(x=920, y=50) # Ajusta x e y según veas necesario

        init_db()

        # --- Botón SALIR ---
        Button(
            self.master,
            text="Salir",
            bg="white", fg="red",
            font=("Arial", 10, "bold"),
            command=self.on_exit
        ).place(x=1000, y=15, width=60, height=30)

        # --- Botón CERRAR SESIÓN ---
        Button(
            self.master,
            text="Cerrar sesión",
            bg="white", fg="red",
            font=("Arial", 9, "bold"),
            command=self.cerrar_sesion
        ).place(x=900, y=15, width=90, height=30)

        # --- Capturar cierre con la X ---
        self.master.protocol("WM_DELETE_WINDOW", self.on_exit)

       # ---- variables
        self.var_id = StringVar()
        self.var_apellido = StringVar()
        self.var_nombre = StringVar()
        self.var_dni = StringVar()
        self.var_grado = StringVar()
        self.var_cargo = StringVar()
        self.var_apn_actividades = StringVar()
        self.var_id_actividad = StringVar()
        self.foto_path = None
        self.var_nro_cel = StringVar()
        self.var_email = StringVar()
        self.enforce_lowercase_var(self.var_email)
        for v in [self.var_id, self.var_apellido, self.var_nombre, self.var_grado, self.var_cargo, self.var_apn_actividades]:
            self.enforce_uppercase_var(v)

        # Notebook
        self.tab_control = ttk.Notebook(self.master)
        self.legajo_frame = Frame(self.tab_control, bg="red")
        self.actividades_frame = Frame(self.tab_control, bg="red")
        self.informes_frame = Frame(self.tab_control, bg="red")
        self.tab_control.add(self.legajo_frame, text="Legajo")
        self.tab_control.add(self.actividades_frame, text="Actividades")
        self.conceptos_frame = Frame(self.tab_control, bg="red")
        self.tab_control.add(self.conceptos_frame, text="Conceptos")
        self.tab_control.add(self.informes_frame, text="Informes")
        self.tab_control.place(x=20, y=80, width=1050, height=430)
        # seleccionar pestaña actividades al inicio
        self.tab_control.select(self.actividades_frame)

        if self.usuario_actual["rol"] == "ADMIN":
            self.usuarios_frame = Frame(self.tab_control, bg="red")
            self.tab_control.add(self.usuarios_frame, text="Usuarios")
            self.init_usuarios()

        self.GRADOS = ["ASPIRANTE", "DRAGONEANTE", "BOMBERO", "OFICIAL", "SUBOFICIAL"]
        self.AREAS = ["ADMINISTRACIÓN","JEFATURA","COMISIÓN DIRECTIVA","COMISIÓN DE EVENTOS","CAPACITACIÓN","BN1","PROTOCOLO Y CEREMONIAL","INCENDIO FORESTAL","INCENDIO ESTRUCTURAL", "RESCATE ACUATICO", "RESCATE CON CUERDAS", "RESCATE VEHICULAR","EMERGENTOLOGIA","PSICOLOGÍA","MATERIALES PELIGROSOS","VANT","COMUNICACIÓN","K9","DEPORTE Y RECREACIÓN","OTRA"]

        # --- Compatibilidad: asegurar que exista _on_situacion_enter ---
        # (parche robusto en caso de que la definición real no quede ligada a la clase)
        if not hasattr(self, "_on_situacion_enter"):
            def _stub_on_situacion_enter(event=None):
                # si existe _on_situacion_change lo delegamos, sino no hacemos nada
                handler = getattr(self, "_on_situacion_change", None)
                if callable(handler):
                    return handler(event)
                return None
            # asignar como método de instancia
            self._on_situacion_enter = _stub_on_situacion_enter

        # --- Inicializar todas las pestañas ---
        self.init_legajo()
        self.cargar_usuarios_dict()
        self.init_actividades()
        self.init_conceptos()
        self.init_informes()
        self.cargar_grilla_legajos()

        # 🔹 Reaplicar permisos al cambiar de pestaña
        self.tab_control.bind(
            "<<NotebookTabChanged>>",
            self._on_tab_changed
        )

        # 🔹 Proteger acciones (métodos) y luego aplicar permisos visibles
        # primero protegemos, luego aplicamos permisos visuales
        self.master.after(300, self._protect_ui_actions)
        self.master.after(800, self._aplicar_permisos)

        # 🔹 Refuerzo final: asegurar permisos después de que todo se dibuje
        self.master.after(1500, self._aplicar_permisos)

        # 🔹 Forzar estado inicial de botones de Usuarios
        self.master.after(1600, lambda: self._estado_botones_usuario("inicial"))
        self.master.after(1800, self._foco_inicial)
        self.cargar_legajos_combobox()

        def _focus_legajo():
            if hasattr(self, "e_legajo") and self.e_legajo.winfo_exists():
                self.e_legajo.focus_set()

    def reset_timer_sesion(self, event=None):
        # Lógica para reiniciar el tiempo de espera
        pass

    def on_exit(self):
        self.master.destroy()

    def _preparar_ventana(self):
            """Aplica el icono y recién ahí muestra la ventana"""
            if hasattr(self.master, "_icono_global") and self.master._icono_global:
                # Forzamos el icono (False y luego True ayuda a refrescar en Win10/11)
                self.top.iconphoto(False, self.master._icono_global)
                self.top.iconphoto(True, self.master._icono_global)
            
            # Centrar la ventana antes de mostrarla
            self.top.update_idletasks()
            ancho = 300 # Ajusta según tu diseño
            alto = 150
            x = (self.top.winfo_screenwidth() // 2) - (ancho // 2)
            y = (self.top.winfo_screenheight() // 2) - (alto // 2)
            self.top.geometry(f"{ancho}x{alto}+{x}+{y}")
            
            # Finalmente mostrar
            self.top.deiconify()
            self.top.lift()
            self.user.focus_force()

    def _convertir_horas_sql(self, horas_texto):
        """
        Función para SQLite que convierte cualquier formato de horas a decimal.
        Se registra con conn.create_function("CONVERTIR_HORAS", 1, self._convertir_horas_sql)
        """
        try:
            if horas_texto is None:
                return 0.0
            texto = str(horas_texto).strip()
            if not texto:
                return 0.0
            
            # Reemplazar coma por punto
            texto = texto.replace(',', '.')
            
            # Si tiene formato HH:MM
            if ':' in texto:
                partes = texto.split(':')
                if len(partes) >= 2:
                    h = float(partes[0]) if partes[0] else 0
                    m = float(partes[1]) if partes[1] else 0
                    return h + (m / 60.0)
            
            # Intentar como decimal
            return float(texto)
        except:
            return 0.0

        def guardar(event=None):
            nueva = var_new.get().strip()
            confirmar = var_conf.get().strip()

            # Validación de longitud (Ajustada a 4 para pruebas o 8 para producción)
            if len(nueva) < 8:
                self.ui.show_error("Error", "La contraseña debe tener mínimo 8 caracteres")
                e_new.focus_set()
                e_new.selection_range(0, tk.END)
                return

            if nueva != confirmar:
                self.ui.show_error("Error", "Las contraseñas no coinciden")
                e_conf.focus_set()
                e_conf.selection_range(0, tk.END)
                return

            nuevo_hash = hash_password(nueva)

            try:
                conn = get_db_connection()
                c = conn.cursor()

                # UPDATE en MySQL
                c.execute("""
                    UPDATE usuarios
                    SET password_hash=%s,
                        debe_cambiar_password=0
                    WHERE id=%s
                    """, (nuevo_hash, user["id"]))

                # 🔔 Registrar notificación de cambio de contraseña
                # MySQL usa NOW() para la fecha actual
                c.execute("""
                    INSERT INTO notificaciones
                    (actividad_id, tipo, destinatario, asunto, fecha_envio, estado)
                    VALUES (%s, %s, %s, %s, NOW(), %s)
                    """, (
                        None,
                        "CAMBIO_PASSWORD",
                        user["username"],
                        "Cambio de contraseña",
                        "OK"
                    ))

                conn.commit()
                c.close()
                conn.close()

                self.ui.show_info("Éxito", "Contraseña actualizada correctamente")
                win.destroy()
                
                # Limpieza de widgets para reiniciar el estado de la app
                for widget in self.master.winfo_children():
                    try:
                        widget.destroy()
                    except:
                        pass

            except Exception as e:
                print(f"Error en cambio de contraseña: {e}")
                self.ui.show_error("Error de base de datos", f"No se pudo actualizar: {e}")

        btn_guardar = Button(win, text="Guardar", command=guardar, width=15)
        btn_guardar.pack(pady=15)

        # 🔹 Atajos de teclado
        e_new.bind("<Return>", lambda e: e_conf.focus_set())
        e_conf.bind("<Return>", guardar)
        e_new.focus_set()

    def obtener_calificaciones_bombero(self, legajo):
        try:
            conn = get_db_connection()
            cursor = conn.cursor(dictionary=True)
            
            # Unimos notas con el nombre del evento y del tema
            query = """
                SELECT 
                    e.fecha, 
                    e.tipo, 
                    et.nombre AS tema, 
                    an.nota
                FROM asistencia_notas an
                JOIN eventos e ON an.evento_id = e.id
                JOIN evento_temas et ON an.tema_id = et.id
                WHERE an.legajo = %s
                ORDER BY e.fecha DESC
                LIMIT 20
            """
            cursor.execute(query, (legajo,))
            notas = cursor.fetchall()
            
            cursor.close()
            conn.close()
            return notas
        except Exception as e:
            print(f"Error al obtener notas: {e}")
            return []

    def _foco_inicial(self):
            try:
                # Asegurar que estamos en la pestaña de Actividades
                self.tab_control.select(self.actividades_frame)

                rol = self.usuario_actual.get("rol", "").upper()

                # 🔥 UNIFICADO: BOMBERO, SUPERVISOR y ADMIN hacen foco en "Nuevo"
                if rol in ["BOMBERO", "SUPERVISOR", "ADMIN"]:
                    btn = self.act_btns.get("Nuevo")
                else:
                    # Cualquier otro rol futuro (ej: CONSULTA) va a "Buscar"
                    btn = self.act_btns.get("Buscar")

                if btn and btn.winfo_exists():
                    btn.focus_set()

            except Exception as e:
                print("⚠ Error foco inicial:", e)

    def _focus_nuevo():
        btn = self.act_btns.get("Nuevo")
        if btn and btn.winfo_exists():
            btn.focus_set()

    def iniciar_timer_sesion(self):

        # cancelar timer anterior
        if self.timer_sesion:
            self.master.after_cancel(self.timer_sesion)

        # iniciar nuevo timer
        self.timer_sesion = self.master.after(60000, self.verificar_sesion)

    def monitor_sesion(self):

        if self.sesion_cerrada:
            return

        ahora = time.time()
        inactivo = ahora - self.ultimo_evento

        # mostrar aviso
        if inactivo >= self.aviso_inactividad and not self.ventana_aviso:
            self.mostrar_aviso_inactividad()

        # cerrar sesión
        if inactivo >= self.tiempo_inactividad:
            self.logout_silencioso()
            return

        self.timer_control = self.master.after(1000, self.monitor_sesion)

    def cargar_usuarios_dict(self):
        try:
            conn = sqlite3.connect(DB_PATH)
            c = conn.cursor()

            c.execute("""
                SELECT 
                    u.legajo,
                    u.username,
                    u.rol,
                    IFNULL(l.apellido, '') as apellido,
                    IFNULL(l.nombre, '') as nombre
                FROM usuarios u
                LEFT JOIN legajos l ON u.legajo = l.legajo
                WHERE u.activo = 1
                AND u.legajo IS NOT NULL
            """)

            self.usuarios_dict = {}

            for legajo, username, rol, apellido, nombre in c.fetchall():

                nombre_real = f"{apellido} {nombre}".strip()

                # fallback por si no existe en legajos
                if not nombre_real:
                    nombre_real = username

                self.usuarios_dict[str(legajo)] = {
                    "nombre": nombre_real,   # ✅ ahora es nombre real
                    "rol": rol.upper()
                }

            conn.close()

            print("USUARIOS_DICT cargado:", self.usuarios_dict)

        except Exception as e:
            print("Error cargando usuarios_dict:", e)
            self.usuarios_dict = {}

    def verificar_sesion(self):

        respuesta = messagebox.askyesno(
            "Sesión",
            "La sesión se cerrará por inactividad.\n¿Desea continuar?"
        )

        if respuesta:
            print(">>> Sesión continuada")
            self.iniciar_timer_sesion()

        else:
            self.cerrar_sesion()

    def reset_timer(self, event=None):

        if hasattr(self, "timer_aviso") and self.timer_aviso:
            try:
                self.master.after_cancel(self.timer_aviso)
            except:
                pass

        if hasattr(self, "timer_sesion") and self.timer_sesion:
            try:
                self.master.after_cancel(self.timer_sesion)
            except:
                pass

        # aviso antes del cierre
        self.timer_aviso = self.master.after(
            TIEMPO_INACTIVIDAD - TIEMPO_AVISO,
            self.mostrar_aviso_inactividad
        )

        # cierre final
        self.timer_sesion = self.master.after(
            TIEMPO_INACTIVIDAD,
            self.cerrar_sesion_por_inactividad
        )

    def registrar_actividad(self, event=None):
        self.ultimo_evento = time.time()

    def mostrar_aviso_inactividad(self):

        if self.sesion_cerrada:
            return
        
        if not self.master.winfo_exists():
            return

        if self.ventana_aviso and self.ventana_aviso.winfo_exists():
            return

        self.ventana_aviso = tk.Toplevel(self.master)
        self.ventana_aviso.title("Sesión por expirar")
        self.ventana_aviso.geometry("320x150")
        self.ventana_aviso.transient(self.master)
        self.ventana_aviso.grab_set()

        tk.Label(
            self.ventana_aviso,
            text="La sesión se cerrará por inactividad",
            font=("Arial", 11)
        ).pack(pady=15)

        # 🔹 LABEL DEL CONTADOR
        self.lbl_contador = tk.Label(
            self.ventana_aviso,
            text="Tiempo restante: 60 s",
            font=("Arial", 12, "bold"),
            fg="red"
        )
        self.lbl_contador.pack()

        frame = tk.Frame(self.ventana_aviso)
        frame.pack(pady=10)

        tk.Button(
            frame,
            text="Continuar sesión",
            command=self.continuar_sesion
        ).pack(side="left", padx=5)

        tk.Button(
            frame,
            text="Cerrar sesión",
            command=self.logout_silencioso
        ).pack(side="left", padx=5)

        # 🔹 CENTRAR LA VENTANA
        self.ventana_aviso.update_idletasks()
        ancho = self.ventana_aviso.winfo_width()
        alto = self.ventana_aviso.winfo_height()

        x = (self.ventana_aviso.winfo_screenwidth() // 2) - (ancho // 2)
        y = (self.ventana_aviso.winfo_screenheight() // 2) - (alto // 2)

        self.ventana_aviso.geometry(f"{ancho}x{alto}+{x}+{y}")

        # 🔹 MOMENTO EXACTO DE CIERRE
        self.tiempo_cierre = time.time() + 60

        # 🔹 INICIAR CONTADOR
        self.actualizar_contador()

    def actualizar_contador(self):

        if self.sesion_cerrada:
            return

        if not hasattr(self, "master") or not self.master.winfo_exists():
            return

        if not getattr(self, "ventana_aviso", None):
            return

        if not self.ventana_aviso.winfo_exists():
            return

        restante = int(self.tiempo_cierre - time.time())

        if restante <= 0:
            self.logout_silencioso()
            return

        self.lbl_contador.config(
            text=f"Tiempo restante: {restante} s"
        )

        self.ventana_aviso.after(1000, self.actualizar_contador)

    def reset_timer_sesion(self, event=None):

        if self.sesion_cerrada:
            return

        if not self.master.winfo_exists():
            return

        if self.timer_sesion:
            self.master.after_cancel(self.timer_sesion)

        if self.timer_aviso:
            self.master.after_cancel(self.timer_aviso)

        self.timer_aviso = self.master.after(
            self.aviso_inactividad,
            self.mostrar_aviso_inactividad
        )

        self.timer_sesion = self.master.after(
            self.tiempo_inactividad,
            self.logout_silencioso
        )

    def reiniciar_app(self, usuario):

        print(">>> Reiniciando aplicación con usuario:", usuario)

        self.usuario_actual = usuario
        self.sesion_cerrada = False

        # reset actividad actual
        self.id_actividad_actual = None
        self.modo_actividad = None

        # limpiar formularios
        self.limpiar_formulario_actividad()

        # refrescar datos
        self._refrescar_actividades()

        # aplicar permisos
        self.aplicar_permisos_por_rol()

        # reiniciar control de sesión
        self.reset_timer_sesion()

    def continuar_sesion(self):

        if self.ventana_aviso:
            try:
                self.ventana_aviso.destroy()
            except:
                pass

        self.ventana_aviso = None
        self.reset_timer_sesion()

    def cerrar_sesion_por_inactividad(self):

        if self.sesion_cerrada:
            return

        if not self.master.winfo_exists():
            return

        self.ui.show_info(
            "Sesión finalizada",
            "La sesión se cerró por inactividad."
        )

        self.logout_silencioso()

    def logout_silencioso(self):
        if self.sesion_cerrada:
            return

        self.sesion_cerrada = True
        self.master.withdraw() # 👈 OCULTA la ventana roja inmediatamente

        # Cancelar timers
        for t in [self.timer_sesion, self.timer_aviso]:
            if t: self.master.after_cancel(t)

        # Limpiar UI
        for widget in self.master.winfo_children():
            widget.destroy()

        # Mostrar login
        login = LoginWindow(self.master)
        self.master.wait_window(login.top)

        if not login.logged_user:
            self._cerrar_aplicacion()
            return

        self.sesion_cerrada = False
        self.master.deiconify() # 👈 VUELVE A MOSTRAR la ventana cuando el usuario entra
        App(self.master, login.logged_user)
    
    def cerrar_sesion(self):
        from tkinter.messagebox import askyesno

        if not messagebox.askyesno("Cerrar sesión", "¿Desea cerrar la sesión actual?"):
            return

        self.sesion_cerrada = True

        # cancelar timers
        if self.timer_sesion:
            self.master.after_cancel(self.timer_sesion)
            self.timer_sesion = None

        if self.timer_aviso:
            self.master.after_cancel(self.timer_aviso)
            self.timer_aviso = None

        # desactivar control de actividad
        self.master.unbind_all("<Key>")
        self.master.unbind_all("<Button>")
        self.master.unbind_all("<Motion>")

        # cerrar ventanas secundarias
        for widget in list(self.master.winfo_children()):
            if isinstance(widget, tk.Toplevel):
                try:
                    widget.destroy()
                except:
                    pass

        # ocultar la app
        self.master.withdraw()

        # mostrar login otra vez
        login = LoginWindow(self.master)
        self.master.wait_window(login.top)

        if not hasattr(login, "logged_user") or not login.logged_user:
            self._cerrar_aplicacion()
            return

        print(">>> Usuario autenticado nuevamente:", login.logged_user)

        # limpiar ventana
        for widget in list(self.master.winfo_children()):
            try:
                widget.destroy()
            except:
                pass

        # reiniciar app
        self.master.deiconify()
        App(self.master, login.logged_user)

    def _cerrar_aplicacion(self):
        print(">>> Cerrando aplicación definitivamente")

        self.sesion_cerrada = True

        try:
            if self.timer_sesion:
                self.master.after_cancel(self.timer_sesion)
        except:
            pass

        try:
            if self.timer_aviso:
                self.master.after_cancel(self.timer_aviso)
        except:
            pass

        try:
            if self.master.winfo_exists():
                self.master.destroy()
        except:
            pass

    def actividad_usuario(self, event=None):
        self.iniciar_timer_sesion()

    def safe_destroy(self):
        try:
            self.master.destroy()
        except:
            import os
            os._exit(0)

    def on_exit(self):
        """Cierre seguro de la aplicación."""
        try:
            if not messagebox.askyesno("Salir", "¿Desea salir de la aplicación?"):
                return

            # Cancelar tareas programadas
            for attr in dir(self):
                if attr.startswith("_after_"):
                    try:
                        self.master.after_cancel(getattr(self, attr))
                    except:
                        pass

            # Cerrar todas las ventanas
            for widget in list(self.master.winfo_children()):
                try:
                    widget.destroy()
                except:
                    pass

            import matplotlib.pyplot as plt
            plt.close("all")

            # ✅ CORRECTO
            self.safe_destroy()

        except Exception as e:
            print("Error al salir:", e)
            try:
                self.safe_destroy()
            except:
                import os
                os._exit(0)

    def _aplicar_permisos(self):
        if getattr(self, "cerrando", False):
            return 
        rol = self.usuario_actual["rol"].upper()

        # === LEGAJOS ===
        if hasattr(self, "leg_btns"):
            if rol == "BOMBERO":
                for b in ["Guardar", "Modificar", "Eliminar"]:
                    btn = self.leg_btns.get(b)
                    if btn and btn.winfo_exists(): btn.config(state="disabled")
            
            # EL ENCARGADO puede Guardar/Modificar, pero NO Eliminar
            elif rol in ["SUPERVISOR", "ENCARGADO"]:
                if "Eliminar" in self.leg_btns:
                    btn = self.leg_btns.get("Eliminar")
                    if btn and btn.winfo_exists(): btn.config(state="disabled")

        # === CONCEPTOS ===
        # El Encargado puede usar conceptos pero no editarlos (similar al Bombero)
        if rol in ["BOMBERO", "ENCARGADO"] and hasattr(self, "conc_btns"):
            for btn in getattr(self, "conc_btns", {}).values():
                try:
                    if btn and btn.winfo_exists(): btn.config(state="disabled")
                except: pass
            
            # Pero habilitamos Buscar e Imprimir para ambos
            for clave, btn in self.conc_btns.items():
                texto = clave.lower()
                if "buscar" in texto or "imprimir" in texto:
                    btn.config(state="normal")

        # === USUARIOS (Solo ADMIN) ===
        if rol != "ADMIN":
            try:
                idx = self.tab_control.index(self.usuarios_frame)
                self.tab_control.tab(idx, state="hidden")
            except Exception: pass

        # === INFORMES (Solo ENCARGADO, SUPERVISOR y ADMIN) ===
        # El Bombero es el único que no ve la pestaña de Informes generales
        if rol == "BOMBERO":
            try:
                idx = self.tab_control.index(self.informes_frame)
                self.tab_control.tab(idx, state="hidden")
            except Exception: pass

        # === REFUERZO CAMPOS LEGAJO (Solo lectura para Bombero) ===
        if rol == "BOMBERO":
            widgets_bloquear = [
                self.e_apellido, self.e_nombre, self.e_dni, self.cb_grado,
                self.e_cargo, self.e_email, self.e_nro_cel,
                self.cb_situacion, self.chk_autoriza, self.btn_foto
            ]
            for w in widgets_bloquear:
                try: w.config(state="disabled")
                except: pass

        # === ACTIVIDADES (Lógica específica) ===
        if hasattr(self, "_aplicar_permisos_actividades"):
            self._aplicar_permisos_actividades()

        # -------------------------------------------------
        # ADMIN POR DEFECTO
        # -------------------------------------------------
        ensure_default_admin()
    
    def _aplicar_permisos_actividades(self):

        rol = self.usuario_actual["rol"].upper()

        # 🔒 Solo bloquear acceso total si no tiene permiso al módulo
        if rol not in ("BOMBERO", "SUPERVISOR", "ADMIN"):
            for btn in self.act_btns.values():
                btn.config(state="disabled")
            return

        # ❗ NO tocar botones individuales acá
        # El estado lo decide _refrescar_actividades()

    def _determinar_estado_actividad(self, row):

        firma_bombero = row["firma_bombero_fecha"]
        firma_supervisor = row["firma_supervisor_fecha"]
        anulada = row["anulada"]

        # 1️⃣ ANULADA
        if anulada and int(anulada) == 1:
            estado = "ANULADA"

        # 2️⃣ FIRMADA POR SUPERVISOR
        elif firma_supervisor:
            estado = "FIRMADA_SUPERVISOR"

        # 3️⃣ FIRMADA POR BOMBERO
        elif firma_bombero:
            estado = "FIRMADA_BOMBERO"

        # 4️⃣ BORRADOR
        else:
            estado = "BORRADOR"
        return estado

    def _set_foco_inicial_actividades(self):
        try:
            rol = self.usuario_actual.get("rol", "").upper()

            if rol == "ADMIN":
                self.act_btns["Nuevo"].focus_set()
            else:
                self.act_btns["Buscar"].focus_set()

        except Exception as e:
            print("Error foco inicial actividades:", e)

    def _puede_firmar_bombero(self):

        if not self.id_actividad_actual:
            return False

        if self.actividad_anulada:
            return False

        if self.firma_bombero_fecha:
            return False

        rol = self.usuario_actual.get("rol", "").upper()
        legajo_usuario = str(self.usuario_actual.get("legajo", "")).strip()

        # 🔥 USAR SIEMPRE EL MÁS CONFIABLE
        legajo_actividad = str(
            getattr(self, "actividad_legajo", "") or
            getattr(self, "_legajo_original_actividad", "")
        ).strip()

        print("DEBUG FIRMAR BOMBERO:", {
            "usuario": legajo_usuario,
            "actividad": legajo_actividad
        })

        if rol in ("BOMBERO", "SUPERVISOR", "ADMIN") and legajo_usuario == legajo_actividad:
            return True

        return False

    def _puede_aprobar(self):

        rol = self.usuario_actual.get("rol", "").upper()
        estado = getattr(self, "_estado_actividad", None)

        if estado != "FIRMADA_BOMBERO":
            return False

        usuario_legajo = str(self.usuario_actual.get("legajo", "")).strip()
        asignado_actual = str(getattr(self, "actividad_asignado", "")).strip()

        if rol in ("SUPERVISOR", "ADMIN") and usuario_legajo == asignado_actual:
            return True

        return False

    def _puede_anular(self):

        if not getattr(self, "id_actividad_actual", None):
            return False

        rol = self.usuario_actual.get("rol", "").upper()
        estado = getattr(self, "_estado_actividad", None)

        if not estado:
            return False

        # 🔒 Estados no anulables
        if estado in ("ANULADA", "FIRMADA_SUPERVISOR"):
            return False

        # 🔒 Estados no anulables
        if estado in ("ANULADA",):
            return False

        return True

    def _puede_modificar_actividad(self):
        """
        Determina si el botón Modificar debe habilitarse.
        Regla: Si hay firmas (Bombero o Supervisor), el botón se deshabilita.
        """
        # 1. Bloqueos absolutos (ID inexistente o actividad ya anulada)
        if not self.id_actividad_actual or getattr(self, "actividad_anulada", False):
            return False
            
        # 2. Bloqueo por FIRMAS (Lógica de Integridad Ajustada)
        hay_firma_bombero = bool(getattr(self, "firma_bombero_fecha", None))
        hay_firma_supervisor = bool(getattr(self, "firma_supervisor_fecha", None))
        
        # Si ya firmó el SUPERVISOR, la actividad está CERRADA. Nadie edita.
        if hay_firma_supervisor:
            return False

        # 3. Evaluación por Rol y Propiedad
        rol = self.usuario_actual.get("rol", "").upper()
        legajo_logeado = str(self.usuario_actual.get("legajo", "")).strip()
        
        asignado_id = str(getattr(self, "actividad_asignado", "")).split(" - ")[0].strip()
        creador_id = str(getattr(self, "actividad_legajo", "")).strip()

        # --- ADMIN: Edita todo lo que no tenga firma de supervisor ---
        if rol == "ADMIN":
            return True

        # --- SUPERVISOR: Edita si es el autor O el asignado ---
        if rol == "SUPERVISOR":
            # Si él es el autor o el supervisor, puede editar (incluso si el bombero ya firmó)
            return (legajo_logeado == creador_id) or (legajo_logeado == asignado_id)

        # --- BOMBERO: Solo si es su propia actividad y NO la firmó todavía ---
        if rol == "BOMBERO":
            if hay_firma_bombero: 
                return False # El bombero ya no puede tocar lo que firmó
            return (legajo_logeado == creador_id)

        return False

    def _puede_anular_actividad(self):
        """Solo el ADMIN puede anular."""
        if self.usuario_actual.get("rol", "").upper() == "ADMIN":
            return bool(self.id_actividad_actual) and not self.actividad_anulada
        return False

    def _puede_firmar_supervisor(self):
        if not self.id_actividad_actual or getattr(self, "actividad_anulada", False):
            return False
        if getattr(self, "firma_supervisor_fecha", None):
            return False
        if not getattr(self, "firma_bombero_fecha", None):
            return False

        rol = self.usuario_actual.get("rol", "").upper()
        legajo_actual = str(self.usuario_actual.get("legajo", "")).strip()
        
        # IMPORTANTE: Usar el ID limpio del asignado de la actividad
        asig_id = str(getattr(self, "actividad_asignado", "")).split(" - ")[0].strip()

        # Solo habilita si soy Admin/Supervisor Y mi legajo es el asignado
        return (rol in ("ADMIN", "SUPERVISOR")) and (legajo_actual == asig_id)

    def _has_role(self, roles):
        """Devuelve True si el rol actual está en la tupla/lista roles."""
        return self.usuario_actual.get("rol", "").upper() in tuple(r.upper() for r in roles)

    def _protect_ui_actions(self):
        """
        Envuelve métodos sensibles para que verifiquen rol antes de ejecutarse
        y reasigna los comandos de botones a esas versiones protegidas.
        Llamar después de crear todas las UI y de aplicar permisos visuales.
        """
        if getattr(self, "_ui_protected", False):
            return

        self._ui_protected = True
        # --------------------
        # 1) Métodos sensibles + roles permitidos
        # --------------------
        sensitive = {
            # LEGAJOS
            # guardar_legajo SE CONTROLA INTERNAMENTE (rol + modo)
            "eliminar_legajo": ("ADMIN",),

            # CONCEPTOS
            "nuevo_concepto": ("ADMIN", "SUPERVISOR"),
            "guardar_concepto": ("ADMIN", "SUPERVISOR"),
            "eliminar_concepto": ("ADMIN",),

            # USUARIOS
            "nuevo_usuario": ("ADMIN",),
            "guardar_usuario": ("ADMIN",),
            "modificar_usuario": ("ADMIN",),
            "eliminar_usuario": ("ADMIN",),

            # INFORMES
            "exportar_excel": ("ADMIN", "SUPERVISOR"),
            "exportar_pdf": ("ADMIN", "SUPERVISOR"),
            "informe_todos_bomberos": ("ADMIN", "SUPERVISOR"),
            "informe_todas_actividades": ("ADMIN", "SUPERVISOR"),
            "informe_horas_por_periodo": ("ADMIN", "SUPERVISOR"),
        }

        # --------------------
        # 2) Wrapper de protección por rol
        # --------------------
        def make_guard(orig_func, allowed_roles):
            def guarded(*args, **kwargs):
                if not self._has_role(allowed_roles):
                    try:
                        messagebox.showwarning(
                            "Sin permiso",
                            "No tiene permiso para realizar esta acción."
                        )
                    except Exception:
                        print("AVISO: intento de acción sin permiso.")
                    return None

                return orig_func(*args, **kwargs)

            return guarded

        # --------------------
        # 3) Envolver métodos sensibles
        # --------------------
        for name, roles in sensitive.items():
            if hasattr(self, name):
                orig = getattr(self, name)
                if callable(orig):
                    wrapped = make_guard(orig, roles)
                    setattr(self, name, wrapped)

        # --------------------
        # 4) Reasignar comandos de botones
        # --------------------
        btn_to_method = []

        # Legajos
        if hasattr(self, "leg_btns"):
            btn_to_method += [
                (self.leg_btns.get("Guardar"), "guardar_legajo"),
                (self.leg_btns.get("Modificar"), "_habilitar_modificar_legajo"),
                (self.leg_btns.get("Eliminar"), "eliminar_legajo"),
            ]

        # Conceptos
        if hasattr(self, "conc_btns"):
            for key, btn in self.conc_btns.items():
                k = key.lower()
                if "nuevo" in k or "crear" in k:
                    btn_to_method.append((btn, "nuevo_concepto"))
                elif "guardar" in k:
                    btn_to_method.append((btn, "guardar_concepto"))
                elif "modificar" in k or "editar" in k:
                    btn_to_method.append((btn, "habilitar_modificar_concepto"))
                elif "eliminar" in k or "borrar" in k:
                    btn_to_method.append((btn, "eliminar_concepto"))

        # Usuarios
        if hasattr(self, "user_btns"):
            btn_to_method += [
                (self.user_btns.get("Nuevo"), "nuevo_usuario"),
                (self.user_btns.get("Guardar"), "guardar_usuario"),
                (self.user_btns.get("Modificar"), "habilitar_modificar_usuario"),
                (self.user_btns.get("Eliminar"), "eliminar_usuario"),
            ]

        # --------------------
        # 5) Aplicar reassignment real
        # --------------------
        for btn, method_name in btn_to_method:
            if btn and hasattr(self, method_name):
                try:
                    if btn.winfo_exists():
                        btn.config(command=getattr(self, method_name))
                except Exception:
                    pass

        # --------------------
        # 6) Text descripción
        # --------------------
        if hasattr(self, "descripcion") and self.descripcion.winfo_exists():
            try:
                self.descripcion.bind("<Return>", self._descripcion_enter)
                self.descripcion.bind("<Shift-Return>", self._descripcion_shift_enter)
                self.enforce_uppercase_text(self.descripcion)
            except Exception:
                pass

        # --------------------
        # 7) ACTIVIDADES - Protección por estado
        # --------------------

        def wrap_estado_guard(method_name, accion):

            if hasattr(self, method_name):

                original = getattr(self, method_name)

                def guarded(*args, **kwargs):

                    if not self._check_permiso_actividad(accion):
                        try:
                            messagebox.showwarning(
                                "Sin permiso",
                                "No tiene permiso para realizar esta acción."
                            )
                        except Exception:
                            print("AVISO: intento bloqueado por estado.")

                        return None

                    return original(*args, **kwargs)

                setattr(self, method_name, guarded)


        # Aplicar wrappers a actividades
        wrap_estado_guard("guardar_actividad", "editar")
        wrap_estado_guard("firmar_actividad_bombero", "firmar_bombero")
        wrap_estado_guard("firmar_supervisor", "aprobar")
        wrap_estado_guard("anular_actividad", "anular")

    # --------------------
    # Aplicar permisos globales diferidos
    # --------------------

    # -------------------- helpers --------------------
    def enforce_uppercase_var(self, var: StringVar):
        def callback(*args):
            v = var.get()
            if v != v.upper():
                var.set(v.upper())
        var.trace_add("write", lambda *a: callback())

    def _check_permiso_actividad(self, accion):
        """
        Valida permisos de actividades según
        rol + estado lógico actual.
        """

        rol = self.usuario_actual.get("rol", "").upper()
        estado = getattr(self, "_estado_actividad", None)

        if estado is None:
            return False

        # --------------------
        # Validadores
        # --------------------
        if accion == "editar":

            if estado == "ANULADA":
                return False

            if rol == "BOMBERO":
                return estado == "BORRADOR"

            if rol == "SUPERVISOR":
                return estado in ["BORRADOR", "FIRMADA_BOMBERO"]

            if rol == "ADMIN":
                return estado != "ANULADA"

            return False

        # --------------------
        if accion == "firmar_bombero":
            return rol == "BOMBERO" and estado == "BORRADOR"

        # --------------------
        if accion == "aprobar":
            return rol in ["SUPERVISOR", "ADMIN"] and estado == "FIRMADA_BOMBERO"

        # --------------------
        if accion == "anular":
            return rol == "ADMIN" and estado != "ANULADA"

        return False

    def enforce_uppercase_text(self, text_widget: Text):
        def to_upper(event=None):
            content = text_widget.get("1.0", "end-1c")
            upper_content = content.upper()
            if content != upper_content:
                pos = text_widget.index("insert")
                text_widget.delete("1.0", "end")
                text_widget.insert("1.0", upper_content)
                text_widget.mark_set("insert", pos)
        text_widget.bind("<KeyRelease>", to_upper)

    def enforce_lowercase_var(self, var: StringVar):
        def callback(*args):
            v = var.get()
            if v != v.lower():
                var.set(v.lower())
        var.trace_add("write", lambda *a: callback())

    def habilitar_modificar_usuario(self, event=None):

        sel = self.tree_users.selection()
        if not sel:
            messagebox.showwarning("Atención", "Seleccione un usuario de la lista.")
            return
        self.modo_usuario = "modificar"
        item = self.tree_users.item(sel, "values")
        self.user_editing_id = item[0]

        # 🔹 Habilitar campos
        self._estado_campos_usuario("modificar")

        # 🔹 Cargar datos
        self.var_user.set(item[1])
        self.var_pwd.set("")
        self.var_rol.set(item[2].upper())
        self.var_legajo.set(item[3])

        self.cb_rol.update_idletasks()

        # 🔹 Foco correcto
        self.e_pwd.focus_set()
        self.var_legajo.set(item[3] if item[3] else "")
        self._estado_botones_usuario("editando")

    def _mostrar_progreso_firma(self, texto="Procesando..."):
        self._set_estado_visual(texto, fg="#003d80", bg="#e6f0ff")

        # 🔥 NUEVO: usar overlay
        self._mostrar_overlay_cargando(texto)

        self.master.update_idletasks()

    def _ocultar_progreso_firma(self):
        # 🔥 NUEVO: ocultar overlay
        self._ocultar_overlay_cargando()

        self._actualizar_estado_firma()

    def _default_filename(self, base, actividad=None, fecha_desde=None, fecha_hasta=None):
        """
        Genera un nombre de archivo con:
        - base descriptiva (por ej. 'Informe_Actividades')
        - nombre de la actividad (si se pasa)
        - rango de fechas (si se pasa)
        - y la fecha actual solo si NO hay rango de fechas.
        """
        from datetime import datetime
        import re

        partes = [base]

        # 🔹 Actividad opcional
        if actividad:
            actividad_limpia = re.sub(r'[^A-Za-z0-9áéíóúÁÉÍÓÚñÑ _-]', '', actividad)
            actividad_limpia = actividad_limpia.strip().replace(" ", "_")[:40]
            partes.append(actividad_limpia)

        # 🔹 Período de fechas
        if fecha_desde and fecha_hasta:
            try:
                f1 = datetime.strptime(str(fecha_desde), "%Y-%m-%d").strftime("%Y-%m-%d")
                f2 = datetime.strptime(str(fecha_hasta), "%Y-%m-%d").strftime("%Y-%m-%d")
            except Exception:
                f1, f2 = str(fecha_desde), str(fecha_hasta)
            partes.append(f"{f1}_a_{f2}")
        else:
            # Si no hay rango de fechas, incluir la fecha actual
            partes.append(datetime.now().strftime("%Y-%m-%d"))

        return "_".join(partes) + ".pdf"

    def _default_filename_excel(self, base, actividad=None, fecha_desde=None, fecha_hasta=None):
        from datetime import datetime
        import re

        partes = [base]

        # 🔹 Actividad opcional
        if actividad:
            actividad_limpia = re.sub(r'[^A-Za-z0-9áéíóúÁÉÍÓÚñÑ _-]', '', actividad)
            actividad_limpia = actividad_limpia.strip().replace(" ", "_")[:40]
            partes.append(actividad_limpia)

        # 🔹 Período de fechas
        if fecha_desde and fecha_hasta:
            try:
                f1 = datetime.strptime(str(fecha_desde), "%Y-%m-%d").strftime("%Y-%m-%d")
                f2 = datetime.strptime(str(fecha_hasta), "%Y-%m-%d").strftime("%Y-%m-%d")
            except Exception:
                f1, f2 = str(fecha_desde), str(fecha_hasta)
            partes.append(f"{f1}_a_{f2}")
        else:
            # Si no hay rango de fechas, incluir la fecha actual
            partes.append(datetime.now().strftime("%Y-%m-%d"))
        nombre_pdf = self._default_filename(base, actividad, fecha_desde, fecha_hasta)
        return nombre_pdf.replace(".pdf", ".xlsx")

    def _nombre_informe_limpio(self, base, nombre=None, periodo=None, extension=".pdf"):
        """
        Genera un nombre de archivo estándar para informes (PDF o Excel).
        Limpia caracteres especiales, acentos y espacios.
        - base: prefijo, ej. 'Informe_Actividades'
        - nombre: nombre de la actividad, bombero, etc.
        - periodo: texto tipo '2025-10-01_a_2025-10-31'
        - extension: '.pdf' o '.xlsx'
        """
        import unicodedata, re

        partes = [base]

        if nombre:
            limpio = unicodedata.normalize("NFKD", str(nombre))
            limpio = "".join(c for c in limpio if not unicodedata.combining(c))
            limpio = re.sub(r"[^A-Za-z0-9_ -]", "", limpio)
            limpio = limpio.strip().replace(" ", "_")[:40]
            partes.append(limpio)

        if periodo:
            partes.append(periodo.replace(" ", "_"))

        final = "_".join(partes) + extension
        return final

    def _crear_pdf_unificado(
        self,
        buffer_or_path,
        elems,
        titulo="Informe",
        landscape_mode=False,
        estado_actividad=None  # 👈 NUEVO
    ):
        from reportlab.platypus import SimpleDocTemplate, Flowable
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import mm

        try:
            pagesize = landscape(A4) if landscape_mode else A4

            doc = SimpleDocTemplate(
                buffer_or_path,
                pagesize=pagesize,
                leftMargin=20*mm,
                rightMargin=20*mm,
                topMargin=35*mm,
                bottomMargin=25*mm,
            )

            for i, el in enumerate(elems):
                if isinstance(el, list):
                    raise Exception(f"Elemento inválido en elems[{i}]: debe ser un Flowable, no lista.")

            def first_page(canvas, doc_obj):
                self._formato_pdf(
                    canvas,
                    doc_obj,
                    titulo=titulo,
                    landscape_mode=landscape_mode,
                    estado_actividad=estado_actividad
                )

            def later_pages(canvas, doc_obj):
                self._formato_pdf(
                    canvas,
                    doc_obj,
                    titulo=titulo,
                    landscape_mode=landscape_mode,
                    estado_actividad=estado_actividad
                )

            doc.build(
                elems,
                onFirstPage=first_page,
                onLaterPages=later_pages
            )

        except Exception as e:
            # ❌ No más showerror con traceback
            # Solo levanta la excepción para que la maneje el llamador
            raise e

    def convertir_horas_a_decimal(self, valor_horas):
        """
        Convierte horas en varios formatos a valor decimal CORRECTO.
        Método de clase App.
        """
        try:
            if valor_horas is None:
                return 0.0
            
            # Si ya es numérico
            if isinstance(valor_horas, (int, float)):
                return float(valor_horas)
            
            # Convertir a string
            str_valor = str(valor_horas).strip()
            
            # Si está vacío
            if not str_valor:
                return 0.0
            
            # Reemplazar coma por punto
            str_valor = str_valor.replace(',', '.')
            
            # 🔥 CASO 1: Formato HH:MM (5:30, 2:45)
            if ':' in str_valor:
                partes = str_valor.split(':')
                if len(partes) >= 2:
                    horas = float(partes[0]) if partes[0] else 0
                    minutos = float(partes[1]) if partes[1] else 0
                    # CORRECCIÓN: dividir entre 60, no 100
                    return horas + (minutos / 60.0)
            
            # 🔥 CASO 2: Formato decimal (5.5, 2.25)
            try:
                return float(str_valor)
            except:
                # 🔥 CASO 3: Intentar extraer números
                import re
                numeros = re.findall(r'\d+\.?\d*', str_valor)
                if numeros:
                    return float(numeros[0])
                else:
                    return 0.0
                    
        except Exception as e:
            print(f">>> ERROR convirtiendo horas '{valor_horas}': {e}")
            return 0.0    

    def _calcular_promedio_horas_por_bombero(self, df_resumen):
        """Calcular promedio de horas por bombero desde df_resumen"""
        def horas_a_minutos(horas_str):
            try:
                if isinstance(horas_str, str) and ':' in horas_str:
                    h, m = map(int, horas_str.split(':'))
                    return h * 60 + m
                elif isinstance(horas_str, (int, float)):
                    return int(horas_str) * 60
                return 0
            except:
                return 0
        
        # Verificar si df_resumen tiene columna 'Horas'
        if df_resumen is None or df_resumen.empty:
            return "00:00"
        
        # Determinar la columna de horas
        horas_col = None
        for col in df_resumen.columns:
            col_lower = col.lower()
            if 'hora' in col_lower or 'tiempo' in col_lower:
                horas_col = col
                break
        
        # Si no encuentra columna específica, usar la segunda columna
        if horas_col is None and len(df_resumen.columns) > 1:
            horas_col = df_resumen.columns[1]
        elif horas_col is None and len(df_resumen.columns) == 1:
            horas_col = df_resumen.columns[0]
        
        # Calcular total minutos
        total_minutos = 0
        for horas_val in df_resumen[horas_col]:
            total_minutos += horas_a_minutos(str(horas_val))
        
        total_bomberos = len(df_resumen)
        
        if total_bomberos > 0:
            promedio_minutos = total_minutos / total_bomberos
            horas_prom = int(promedio_minutos // 60)
            minutos_prom = int(promedio_minutos % 60)
            return f"{horas_prom:02d}:{minutos_prom:02d}"
        
        return "00:00"

    # ---------------- Helper: crear PDF resumen (usa _crear_pdf_unificado existente) ----------------
    def _crear_pdf_resumen_actividades_total(
        self,
        file_path,
        titulo,
        df_resumen,
        totales
    ):
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle,
            Paragraph, Spacer
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        import os

        styles = getSampleStyleSheet()
        elems = []

        # Espacio para encabezado (logo + título)
        elems.append(Spacer(1, 20))

        # ==========================
        # 📊 TABLA RESUMEN
        # ==========================
        data = [["ACTIVIDAD", "HORAS"]]

        for _, r in df_resumen.iterrows():
            data.append([
                r["Actividad"],
                self._formatear_horas_decimal_a_hhmm(r["Horas"])
            ])

        ancho = 595 - (30 * mm)
        tbl = Table(
            data,
            colWidths=[
                ancho * 0.78,   # 👈 ACTIVIDAD (bien ancha)
                ancho * 0.22    # 👈 HORAS
            ]
        )

        tbl.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.25, colors.gray),

            # 🔹 Encabezados
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E5AAC")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),

            ("ALIGN", (0, 0), (-1, 0), "CENTER"),   # 👈 ENCABEZADOS CENTRADOS
            ("ALIGN", (1, 1), (-1, -1), "CENTER"),  # 👈 HORAS centradas
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))

        elems.append(tbl)
        elems.append(Spacer(1, 15))

        # ==========================
        # 📌 TOTALES
        # ==========================
        estilo = ParagraphStyle(
            "resumen",
            parent=styles["Normal"],
            fontSize=9,
            leftIndent=10,
            spaceAfter=4
        )

        total_actividades = totales.get("total_actividades", 0)
        total_registros = totales.get("total_registros", 0)
        total_bomberos = totales.get("total_bomberos", 0)
        total_horas = totales.get("total_horas", 0)

        # 🔹 Promedio horas por actividad (LÓGICO PARA ESTE INFORME)
        promedio_horas_actividad = (
            total_horas / total_actividades if total_actividades > 0 else 0
        )

        elems.append(Paragraph(
            f"<b>• Actividades distintas:</b> {total_actividades}",
            estilo
        ))

        elems.append(Paragraph(
            f"<b>• Total de registros:</b> {total_registros}",
            estilo
        ))

        elems.append(Paragraph(
            f"<b>• Bomberos distintos participantes:</b> {total_bomberos}",
            estilo
        ))

        elems.append(Paragraph(
            f"<b>• Total de horas:</b> "
            f"{self._formatear_horas_decimal_a_hhmm(total_horas)}",
            estilo
        ))

        elems.append(Paragraph(
            f"<b>• Promedio horas por actividad:</b> "
            f"{self._formatear_horas_decimal_a_hhmm(promedio_horas_actividad)}",
            estilo
        ))

        # ==========================
        # 🧾 CONSTRUIR PDF
        # ==========================
        doc = SimpleDocTemplate(
            file_path,
            pagesize=A4,
            leftMargin=15 * mm,
            rightMargin=15 * mm,
            topMargin=40 * mm,
            bottomMargin=15 * mm
        )

        doc.build(
            elems,
            onFirstPage=lambda c, d:
                self._formato_pdf(c, d, titulo, False)
        )

        return True

    def _formatear_horas_a_hhmm(self, horas_val):
        """
        Convierte horas en cualquier formato a HH:MM
        """
        try:
            # Si es string, limpiarlo
            if isinstance(horas_val, str):
                horas_val = horas_val.strip()
            
            # Intentar convertir a float primero
            horas_float = float(horas_val)
            
            # Convertir decimal a HH:MM
            horas_enteras = int(horas_float)
            minutos = int(round((horas_float - horas_enteras) * 60))
            
            # Ajustar si minutos son 60
            if minutos >= 60:
                horas_enteras += 1
                minutos -= 60
            
            return f"{horas_enteras}:{minutos:02d}"
        
        except (ValueError, TypeError):
            # Si ya está en formato HH:MM o hay error, devolver como está
            return str(horas_val)
            
    def _get_nombre_unificado(self, nombre):
        """
        Devuelve el nombre UNIFICADO y COMPLETO sin divisiones
        """
        if not isinstance(nombre, str):
            return str(nombre)
        
        nombre = nombre.strip()
        
        # DICCIONARIO DE UNIFICACIÓN - Nombres COMPLETOS y CORRECTOS
        unificaciones = {
            # Nombres que aparecen cortados en el PDF
            "14 - TAREA": "14 - TAREA PROGRAMADA",
            "TAREA": "TAREA PROGRAMADA",
            "PROGRAMA": "TAREA PROGRAMADA",  # Para cuando se divida
            "14 - TABEA PROGRA": "14 - TAREA PROGRAMADA",
            "TABEA PROGRAMADA": "TAREA PROGRAMADA",
            
            "CAPACITACIÓN/CURSOCIARLA": "CAPACITACIÓN/CURSO/CHARLA",
            "CAPACITACIÓN/CURS...": "CAPACITACIÓN/CURSO/CHARLA",
            "CAPACITACIÓN Y CURSOS": "CAPACITACIÓN/CURSO/CHARLA",
            
            "EDIFICIO": "EDIFICIO: CONSERVACIÓN/MANTENIMIENTO/REPARACIÓN",
            "CONSERV": "EDIFICIO: CONSERVACIÓN/MANTENIMIENTO/REPARACIÓN",
            "EDIFICIO: CONSERV": "EDIFICIO: CONSERVACIÓN/MANTENIMIENTO/REPARACIÓN",
            
            "EDIFICA": "EDIFICIO: LIMPIEZA/ORDEN",  # Posible error
            "LIMPIEZA": "EDIFICIO: LIMPIEZA/ORDEN",  # Cuando se divide
            "EDIFICIO: LIMPIEZ": "EDIFICIO: LIMPIEZA/ORDEN",
            
            "HABILITACIÓN/VISITA LOCAL COMERCIAL O INSTITUCIÓN": "HABILITACIÓN/VISITA LOCAL COMERCIAL O INSTITUCIÓN",
            "HABILITACIÓN/VISI...": "HABILITACIÓN/VISITA LOCAL COMERCIAL O INSTITUCIÓN",
            "HABILITACIÓN DE VISITAS": "HABILITACIÓN/VISITA LOCAL COMERCIAL O INSTITUCIÓN",
            
            "MÓVILES": "MÓVILES: LIMPIEZA Y ORDEN",
            "MÓVILES: LIMPIEZA": "MÓVILES: LIMPIEZA Y ORDEN",
            
            "REUNIÓN": "REUNIÓN",
            "REUNIONES DE COORDINACIÓN": "REUNIÓN",  # Unificar a REUNIÓN
            
            "SERVICIO/EVENTO/C": "SERVICIO/EVENTO/COLABORACIÓN",
            "SERVICIO / EVENTO / C": "SERVICIO/EVENTO/COLABORACIÓN",
            "SERVICIOR/ENTOC...": "SERVICIO/EVENTO/COLABORACIÓN",
            
            # Nombres de tu lista original
            "MÓVILES: MANTENIMIENTO/REPARACIÓN": "MÓVILES: MANTENIMIENTO/REPARACIÓN",
            "EDIFICIO NUEVO PROYECTO/OBRA": "EDIFICIO NUEVO PROYECTO/OBRA",
            "HERRAMIENTAS Y EQUIPOS: MANTENIMIENTO/REPARACIÓN": "HERRAMIENTAS Y EQUIPOS: MANTENIMIENTO/REPARACIÓN",
            "HERRAMIENTAS Y EQUIPOS: LIMPIEZA/ORDEN": "HERRAMIENTAS Y EQUIPOS: LIMPIEZA/ORDEN",
            "TRÁMITE/VIAJE/COMPRA PROGRAMADA": "TRÁMITE/VIAJE/COMPRA PROGRAMADA",
            "VISITA A INSTITUCIÓN/EMPRESA/ORGANIZACIÓN": "VISITA A INSTITUCIÓN/EMPRESA/ORGANIZACIÓN",
        }
        
        # Buscar coincidencia EXACTA
        if nombre in unificaciones:
            return unificaciones[nombre]
        
        # Buscar coincidencia PARCIAL
        for corto, completo in unificaciones.items():
            if corto in nombre:
                return completo
        
        # Si no encuentra, devolver el nombre original LIMPIO
        return nombre.replace("...", "").strip()

    def _fmt_horas(self, valor):
        """Convierte horas decimales a formato HH:MM."""
        try:
            if valor is None or pd.isna(valor):
                return "0:00"
            
            # Asegurar que sea float
            valor = float(valor)
            
            # Obtener horas enteras y minutos
            horas_enteras = int(valor)
            minutos = int(round((valor - horas_enteras) * 60))
            
            # Ajustar si minutos son 60
            if minutos >= 60:
                horas_enteras += 1
                minutos = 0  # No dejar minutos en 60
                
            return f"{horas_enteras}:{minutos:02d}"
        except Exception:
            return "0:00"

    def _convertir_horas_dataframe(self, df, col_horas='Total_Horas'):
        """Convierte horas en un DataFrame de decimal a HH:MM."""
        if df is None or df.empty:
            return df
        
        df = df.copy()
        
        # Buscar cualquier columna que contenga horas
        for col in df.columns:
            col_lower = str(col).lower()
            if any(keyword in col_lower for keyword in ['hora', 'total_hora', 'horas_total']):
                # Convertir a HH:MM
                df[col] = df[col].apply(
                    lambda x: self._formatear_horas_decimal_a_hhmm(x) 
                    if x is not None and str(x).lower() not in ['nan', 'none', 'null', ''] 
                    else "0:00"
                )
        
        return df

    def _obtener_horas_decimal(self, valor):
        """Convierte cualquier formato de horas a decimal para cálculos (sin pandas)."""
        try:
            if valor is None:
                return 0.0
            
            # Verificar si es NaN
            if isinstance(valor, float):
                import math
                if math.isnan(valor):
                    return 0.0
            
            # Si ya es numérico
            if isinstance(valor, (int, float)):
                return float(valor)
            
            # Si es string
            if isinstance(valor, str):
                valor_str = valor.strip()
                if not valor_str or valor_str.lower() in ['nan', 'none', 'null']:
                    return 0.0
                
                # Si es formato HH:MM
                if ':' in valor_str:
                    partes = valor_str.split(':')
                    if len(partes) >= 2:
                        horas = float(partes[0]) if partes[0] else 0
                        minutos = float(partes[1]) if partes[1] else 0
                        return horas + (minutos / 60.0)
                
                # Intentar como decimal
                try:
                    return float(valor_str.replace(',', '.'))
                except:
                    return 0.0
            
            # Para cualquier otro tipo
            return float(str(valor).replace(',', '.'))
            
        except Exception:
            return 0.0
        
#---------------------------INFORME POR ACTIVIDAD (TODOS) ---------------------------------------
    def informe_resumen_por_actividad(self):
        """Resumen general de actividades (sin selección específica)."""

        try:
            f1 = self.inf_desde.get_date()
            f2 = self.inf_hasta.get_date()
            s1, s2 = f1.strftime("%Y-%m-%d"), f2.strftime("%Y-%m-%d")
        except Exception as e:
            self.ui.show_error("Error", "Debe seleccionar fechas válidas.")
            return
        
        conn = sqlite3.connect(DB_PATH)
        try:
            # CONSULTA: Resumen por actividad con horas en HH:MM
            q = """
                SELECT 
                    COALESCE(a.actividad, 'SIN ACTIVIDAD') AS Actividad,
                    COUNT(DISTINCT a.legajo) AS Total_Bomberos,
                    COUNT(a.id) AS Total_Registros,
                    -- 🔥 Horas en formato HH:MM desde SQL
                    printf('%d:%02d', 
                        CAST(SUM(
                            CASE 
                                WHEN typeof(a.horas) = 'text' AND a.horas LIKE '%:%' THEN
                                    CAST(SUBSTR(a.horas, 1, INSTR(a.horas, ':')-1) AS REAL) + 
                                    CAST(SUBSTR(a.horas, INSTR(a.horas, ':')+1) AS REAL) / 60.0
                                WHEN typeof(a.horas) IN ('real', 'integer') THEN
                                    CAST(a.horas AS REAL)
                                ELSE 0.0
                            END
                        ) AS INTEGER),
                        CAST(ROUND((SUM(
                            CASE 
                                WHEN typeof(a.horas) = 'text' AND a.horas LIKE '%:%' THEN
                                    CAST(SUBSTR(a.horas, 1, INSTR(a.horas, ':')-1) AS REAL) + 
                                    CAST(SUBSTR(a.horas, INSTR(a.horas, ':')+1) AS REAL) / 60.0
                                WHEN typeof(a.horas) IN ('real', 'integer') THEN
                                    CAST(a.horas AS REAL)
                                ELSE 0.0
                            END
                        ) - CAST(SUM(
                            CASE 
                                WHEN typeof(a.horas) = 'text' AND a.horas LIKE '%:%' THEN
                                    CAST(SUBSTR(a.horas, 1, INSTR(a.horas, ':')-1) AS REAL) + 
                                    CAST(SUBSTR(a.horas, INSTR(a.horas, ':')+1) AS REAL) / 60.0
                                WHEN typeof(a.horas) IN ('real', 'integer') THEN
                                    CAST(a.horas AS REAL)
                                ELSE 0.0
                            END
                        ) AS INTEGER)) * 60) AS INTEGER)
                    ) AS Total_Horas,
                    -- Para gráfico: horas en decimal
                    ROUND(SUM(
                        CASE 
                            WHEN typeof(a.horas) = 'text' AND a.horas LIKE '%:%' THEN
                                CAST(SUBSTR(a.horas, 1, INSTR(a.horas, ':')-1) AS REAL) + 
                                CAST(SUBSTR(a.horas, INSTR(a.horas, ':')+1) AS REAL) / 60.0
                            WHEN typeof(a.horas) IN ('real', 'integer') THEN
                                CAST(a.horas AS REAL)
                            ELSE 0.0
                        END
                    ), 2) AS Horas_Decimal
                FROM actividades a
                WHERE (CASE WHEN instr(a.fecha_inicio, '/')>0
                            THEN date(substr(a.fecha_inicio,7,4)||'-'||substr(a.fecha_inicio,4,2)||'-'||substr(a.fecha_inicio,1,2))
                            ELSE date(a.fecha_inicio) END)
                    BETWEEN date(?) AND date(?)
                GROUP BY a.actividad
                HAVING COUNT(a.id) > 0
                ORDER BY Horas_Decimal DESC
            """
            
            df_res = pd.read_sql_query(q, conn, params=(s1, s2))
            
            if df_res.empty:
                self.ui.show_info("Sin datos", "No hay actividades en el período seleccionado.")
                return
            
            # Preparar para gráfico
            df_graf = df_res[['Actividad', 'Horas_Decimal']].copy()
            df_graf.columns = ['Actividad', 'Horas']
            
            # Preparar detalle para tabla
            df_detalle = df_res.copy()
            
            titulo = f"Informe por Actividad ({f1:%d/%m/%Y} - {f2:%d/%m/%Y})"
            
            # Calcular totales
            total_actividades = len(df_res)
            total_bomberos = df_res['Total_Bomberos'].sum()
            total_registros = df_res['Total_Registros'].sum()
            
            # Calcular horas totales en decimal
            total_horas_decimal = df_graf['Horas'].sum()
            promedio_horas_por_actividad = total_horas_decimal / total_actividades if total_actividades > 0 else 0
            
            # Convertir a HH:MM para mostrar
            promedio_horas_display = self._formatear_horas_decimal_a_hhmm(promedio_horas_por_actividad)
            
            # Llamar a función de visualización
            self._mostrar_tabla_y_grafico_filtrado(
                titulo,
                df_graf,
                df_detalle,
                x_col="Bombero",
                y_col="Horas",
                color="green",
                report_type="Informe_Actividades",
                principal_name=actividad_sel,
                tipo_informe="actividad_especifica"  # 🔥 AGREGAR ESTO
            )            
        except Exception as e:
            self.ui.show_error("Error", f"No se pudo generar el informe de actividades:\n{e}")
            import traceback
            traceback.print_exc()
        finally:
            conn.close()

#-----------------INFORME POR ACTIVIDAD ESPECIFICA - CON SELECCIÓN EN COMBO ACTIVIDADES ---------------------------
    def informe_detalle_por_actividad(self):
        """Detalle cuando se selecciona una actividad específica."""
       
        try:
            f1 = self.inf_desde.get_date()
            f2 = self.inf_hasta.get_date()
            s1, s2 = f1.strftime("%Y-%m-%d"), f2.strftime("%Y-%m-%d")
            
            # Obtener actividad seleccionada
            actividad_sel = (self.inf_actividad.get() or "").strip()
            if not actividad_sel:
                messagebox.showwarning("Selección", "Debe seleccionar una actividad.")
                return
                
        except Exception as e:
            self.ui.show_error("Error", f"Datos inválidos: {e}")
            return
        
        conn = sqlite3.connect(DB_PATH)
        try:
            # CONSULTA: Detalle por actividad específica
            q = """
                SELECT 
                    a.id,
                    a.legajo || ' - ' || COALESCE(l.apellido || ' ' || l.nombre, '') AS Bombero,
                    a.actividad,
                    a.area,
                    a.fecha_inicio,
                    a.fecha_fin,
                    a.hora_inicio,
                    a.hora_fin,
                    -- 🔥 Horas en formato HH:MM
                    printf('%d:%02d', 
                        CAST(
                            CASE 
                                WHEN typeof(a.horas) = 'text' AND a.horas LIKE '%:%' THEN
                                    CAST(SUBSTR(a.horas, 1, INSTR(a.horas, ':')-1) AS REAL) + 
                                    CAST(SUBSTR(a.horas, INSTR(a.horas, ':')+1) AS REAL) / 60.0
                                ELSE CAST(COALESCE(a.horas, 0) AS REAL)
                            END
                        AS INTEGER),
                        CAST(ROUND((
                            CASE 
                                WHEN typeof(a.horas) = 'text' AND a.horas LIKE '%:%' THEN
                                    CAST(SUBSTR(a.horas, 1, INSTR(a.horas, ':')-1) AS REAL) + 
                                    CAST(SUBSTR(a.horas, INSTR(a.horas, ':')+1) AS REAL) / 60.0
                                ELSE CAST(COALESCE(a.horas, 0) AS REAL)
                            END
                            - CAST(
                                CASE 
                                    WHEN typeof(a.horas) = 'text' AND a.horas LIKE '%:%' THEN
                                        CAST(SUBSTR(a.horas, 1, INSTR(a.horas, ':')-1) AS REAL) + 
                                        CAST(SUBSTR(a.horas, INSTR(a.horas, ':')+1) AS REAL) / 60.0
                                    ELSE CAST(COALESCE(a.horas, 0) AS REAL)
                                END
                            AS INTEGER)
                        ) * 60) AS INTEGER)
                    ) AS horas,
                    COALESCE(a.descripcion, '') AS descripcion,
                    CASE 
                        WHEN a.asignado IS NOT NULL AND a.asignado != '' 
                        THEN a.asignado || ' - ' || COALESCE(l_asig.apellido, '') || ' ' || COALESCE(l_asig.nombre, '')
                        ELSE '' 
                    END AS asignado,
                    -- Para gráfico: horas en decimal
                    ROUND(
                        CASE 
                            WHEN typeof(a.horas) = 'text' AND a.horas LIKE '%:%' THEN
                                CAST(SUBSTR(a.horas, 1, INSTR(a.horas, ':')-1) AS REAL) + 
                                CAST(SUBSTR(a.horas, INSTR(a.horas, ':')+1) AS REAL) / 60.0
                            ELSE CAST(COALESCE(a.horas, 0) AS REAL)
                        END, 2
                    ) AS Horas_Decimal
                FROM actividades a
                LEFT JOIN legajos l ON a.legajo = l.legajo
                LEFT JOIN legajos l_asig ON a.asignado = l_asig.legajo
                WHERE a.actividad = ?
                AND (CASE WHEN instr(a.fecha_inicio, '/')>0
                            THEN date(substr(a.fecha_inicio,7,4)||'-'||substr(a.fecha_inicio,4,2)||'-'||substr(a.fecha_inicio,1,2))
                            ELSE date(a.fecha_inicio) END)
                    BETWEEN date(?) AND date(?)
                ORDER BY a.fecha_inicio DESC, a.hora_inicio DESC
            """
            
            df_detalle = pd.read_sql_query(q, conn, params=(actividad_sel, s1, s2))
            
            if df_detalle.empty:
                self.ui.show_info("Sin datos", f"No hay registros para la actividad '{actividad_sel}' en el período seleccionado.")
                self.inf_actividad.set("")
                return
            
            # Preparar gráfico: horas por bombero
            df_graf = df_detalle.copy()
            
            # Agrupar por bombero
            df_graf = (
                df_graf.groupby("Bombero", as_index=False)
                .agg({
                    "Horas_Decimal": "sum"
                })
                .rename(columns={"Horas_Decimal": "Horas"})
            )
            
            # Ordenar por horas descendente
            df_graf = df_graf.sort_values("Horas", ascending=False)
            
            titulo = f"Actividad: {actividad_sel} ({f1:%d/%m/%Y} - {f2:%d/%m/%Y})"
            
            # Calcular totales
            total_bomberos = len(df_graf)
            total_registros = len(df_detalle)
            total_horas_decimal = df_graf["Horas"].sum()
            promedio_horas_por_bombero = total_horas_decimal / total_bomberos if total_bomberos > 0 else 0
            
            # Llamar a función de visualización
            self._mostrar_tabla_y_grafico_filtrado(
                titulo,
                df_graf,
                df_detalle,
                x_col="Bombero",
                y_col="Horas",
                color="green",
                report_type="Informe_Actividades",
                principal_name=actividad_sel
            )
            
        except Exception as e:
            self.ui.show_error("Error", f"No se pudo generar el informe de actividad:\n{e}")
            import traceback
            traceback.print_exc()
        finally:
            conn.close()

    def exportar_pdf_bomberos_total(self):
        # 🔹 Revisar si los datos del informe existen
        if not hasattr(self, "_df_resumen_bomberos_total") or self._df_resumen_bomberos_total is None or self._df_resumen_bomberos_total.empty:
            self.ui.show_error("Error", "Primero debe generar el informe.")
            return

        try:
            # Obtener rango de fechas
            f1 = self.inf_desde.get_date()
            f2 = self.inf_hasta.get_date()

            # Sugerir nombre de archivo PDF
            file_path = self._sugerir_nombre_pdf("Informe_Bomberos_Todos", f1, f2)
            if not file_path:
                messagebox.showwarning("Cancelado", "Exportación cancelada.")
                return

            # 🔹 Llamar a la función interna que genera el PDF, con todos los datos
            self._crear_pdf_resumen_bomberos_total(
                file_path=file_path,
                titulo=f"Informe Bomberos Todos - Del {f1:%d/%m/%Y} al {f2:%d/%m/%Y}",
                df_resumen=self._df_resumen_bomberos_total,
                df_detalle=self._df_detalle_bomberos_total,
                totales=self._pdf_totales_bomberos
            )

            self.ui.show_info("Éxito", f"PDF generado correctamente:\n{file_path}")

            # 🔹 Abrir el PDF automáticamente
            os.startfile(os.path.abspath(file_path))

        except Exception as e:
            self.ui.show_error("Error", f"No se pudo generar el PDF:\n{e}")

    def _crear_pdf_resumen_bomberos_total(
        self,
        file_path,
        titulo,
        df_resumen,
        df_detalle,
        totales
    ):
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle,
            Paragraph, Spacer
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        import os

        styles = getSampleStyleSheet()
        elems = []

        # Espacio para encabezado
        elems.append(Spacer(1, 20))

        # ==========================
        # 📊 TABLA RESUMEN
        # ==========================
        data = [["BOMBERO", "HORAS"]]

        for _, r in df_resumen.iterrows():
            nombre_bombero = f"{r['legajo']} - {r['apellido']} {r['nombre']}".strip()
            horas_txt = self._formatear_horas_decimal_a_hhmm(r["Horas_Totales"])
            data.append([nombre_bombero, horas_txt])

        ancho = 595 - (30 * mm)
        tbl = Table(data, colWidths=[ancho * 0.8, ancho * 0.2])

        tbl.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.25, colors.gray),

            # 🔹 ENCABEZADOS
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E5AAC")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),   # 👈 CLAVE
            ("VALIGN", (0, 0), (-1, 0), "MIDDLE"),  # 👈 OPCIONAL pero recomendable

            # 🔹 CUERPO
            ("ALIGN", (1, 1), (-1, -1), "CENTER"),
            ("VALIGN", (0, 1), (-1, -1), "MIDDLE"),
        ]))

        elems.append(tbl)
        elems.append(Spacer(1, 15))

        # ==========================
        # 📌 TOTALES (DEL FORM)
        # ==========================
        estilo = ParagraphStyle(
            "resumen",
            parent=styles["Normal"],
            fontSize=9,
            leftIndent=10,
            spaceAfter=4
        )

        elems.append(Paragraph(
            f"<b>• Total de bomberos:</b> {totales['total_bomberos']}", estilo))
        elems.append(Paragraph(
            f"<b>• Total de registros:</b> {totales['total_registros']}", estilo))
        elems.append(Paragraph(
            f"<b>• Actividades distintas:</b> {totales['actividades_distintas']}", estilo))
        elems.append(Paragraph(
            f"<b>• Total de horas:</b> "
            f"{self._formatear_horas_decimal_a_hhmm(totales['total_horas'])}",
            estilo))
        elems.append(Paragraph(
            f"<b>• Promedio horas/bombero:</b> "
            f"{self._formatear_horas_decimal_a_hhmm(totales['promedio_horas_bombero'])}",
            estilo))

        # ==========================
        # 🧾 CONSTRUIR PDF
        # ==========================
        doc = SimpleDocTemplate(
            file_path,
            pagesize=A4,
            leftMargin=15 * mm,
            rightMargin=15 * mm,
            topMargin=40 * mm,
            bottomMargin=15 * mm
        )

        doc.build(
            elems,
            onFirstPage=lambda c, d:
                self._formato_pdf(c, d, titulo, False)
        )

        return True

        return os.path.exists(file_path)

    #----------------- INFORME POR BOMBERO (TODOS) ----------------------------------
    def informe_resumen_por_bombero(self):
        """RESUMEN GENERAL – TODOS los bomberos (sin seleccionar uno específico)"""

        # ===============================
        # 1️⃣ Fechas
        # ===============================
        try:
            f1 = self.inf_desde.get_date()
            f2 = self.inf_hasta.get_date()
            s1, s2 = f1.strftime("%Y-%m-%d"), f2.strftime("%Y-%m-%d")
        except Exception:
            self.ui.show_error("Error", "Debe seleccionar fechas válidas.")
            return

        conn = sqlite3.connect(DB_PATH)

        # ===============================
        # 2️⃣ Conversión segura de horas
        # ===============================
        def horas_a_decimal(valor):
            if not valor:
                return 0.0
            try:
                txt = str(valor).strip().replace(",", ".")
                if ":" in txt:
                    h, m = txt.split(":")[:2]
                    return float(h) + float(m) / 60.0
                return float(txt)
            except:
                return 0.0

        try:
            # ===============================
            # 3️⃣ Bomberos del período
            # ===============================
            q_bomberos = """
                SELECT DISTINCT
                    a.legajo,
                    COALESCE(l.apellido, '') AS apellido,
                    COALESCE(l.nombre, '') AS nombre
                FROM actividades a
                LEFT JOIN legajos l ON a.legajo = l.legajo
                WHERE a.legajo IS NOT NULL
                AND TRIM(a.legajo) != ''
                AND (
                        CASE WHEN instr(a.fecha_inicio,'/')>0
                        THEN date(substr(a.fecha_inicio,7,4)||'-'||substr(a.fecha_inicio,4,2)||'-'||substr(a.fecha_inicio,1,2))
                        ELSE date(a.fecha_inicio)
                        END
                    ) BETWEEN date(?) AND date(?)
                ORDER BY l.apellido, l.nombre
            """

            df_bomberos = pd.read_sql_query(q_bomberos, conn, params=(s1, s2))

            if df_bomberos.empty:
                self.ui.show_info("Sin datos", "No hay actividades en el período.")
                return

            # ===============================
            # 4️⃣ Cálculo por bombero
            # ===============================
            resumen_rows = []
            detalle_rows = []

            for _, b in df_bomberos.iterrows():
                legajo = str(b["legajo"]).strip()
                apellido = b["apellido"].strip()
                nombre = b["nombre"].strip()

                q_act = """
                    SELECT actividad, horas
                    FROM actividades
                    WHERE legajo = ?
                    AND (
                            CASE WHEN instr(fecha_inicio,'/')>0
                            THEN date(substr(fecha_inicio,7,4)||'-'||substr(fecha_inicio,4,2)||'-'||substr(fecha_inicio,1,2))
                            ELSE date(fecha_inicio)
                            END
                        ) BETWEEN date(?) AND date(?)
                """

                acts = conn.execute(q_act, (legajo, s1, s2)).fetchall()

                horas_total = 0.0
                actividades = set()

                for act, horas in acts:
                    h_dec = horas_a_decimal(horas)
                    horas_total += h_dec
                    actividades.add(act)

                    detalle_rows.append({
                        "Legajo": legajo,
                        "Apellido": apellido,
                        "Nombre": nombre,
                        "Actividad": act,
                        "Horas": round(h_dec, 2)
                    })

                resumen_rows.append({
                    "Bombero": f"{legajo} - {apellido} {nombre}",
                    "Legajo": legajo,
                    "Total_Registros": len(acts),
                    "Actividades_Distintas": len(actividades),
                    "Horas_Totales": round(horas_total, 2)
                })

            # ===============================
            # 5️⃣ DataFrames finales
            # ===============================
            df_res = pd.DataFrame(resumen_rows)
            df_detalle = pd.DataFrame(detalle_rows)

            df_res = df_res.sort_values("Horas_Totales", ascending=False)

            # ===============================
            # 6️⃣ Totales (fuente única)
            # ===============================
            df_res["Horas_Totales"] = df_res["Horas_Totales"].astype(float)

            total_bomberos = len(df_res)
            total_registros = int(df_res["Total_Registros"].sum())
            total_actividades_distintas = int(df_res["Actividades_Distintas"].sum())
            total_horas = float(df_res["Horas_Totales"].sum())

            promedio_horas_bombero = total_horas / total_bomberos if total_bomberos else 0

            # ===============================
            # 7️⃣ Guardar para PDF / Excel
            # ===============================
            self._df_resumen_bomberos_total = df_res.copy()
            self._df_detalle_bomberos_total = df_detalle.copy()

            self._pdf_totales_bomberos = {
                "total_bomberos": total_bomberos,
                "total_registros": total_registros,
                "actividades_distintas": total_actividades_distintas,
                "total_horas": total_horas,
                "promedio_horas_bombero": promedio_horas_bombero,
            }

            # ===============================
            # 8️⃣ Gráfico + formulario
            # ===============================
            df_graf = df_res[["Bombero", "Horas_Totales"]].rename(
                columns={"Horas_Totales": "Horas"}
            )

            titulo = f"Informe por Bombero ({f1:%d/%m/%Y} - {f2:%d/%m/%Y})"

            self._mostrar_form_informe_actividades_todas(
                titulo,
                df_graf,
                df_res,
                x_col="Bombero",
                y_col="Horas",
                report_type="bombero"
            )

        except Exception as e:
            import traceback
            traceback.print_exc()
            self.ui.show_error("Error", f"No se pudo generar el informe:\n{e}")
        finally:
            conn.close()

    def calcular_horas_decimal(horas_str):
        """Función segura para convertir horas a decimal"""
        if horas_str is None or str(horas_str).strip() == '':
            return 0.0
        
        try:
            texto = str(horas_str).strip().replace(',', '.')
            
            # Formato HH:MM
            if ':' in texto:
                partes = texto.split(':')
                if len(partes) >= 2:
                    horas = float(partes[0]) if partes[0] else 0
                    minutos = float(partes[1]) if partes[1] else 0
                    return horas + (minutos / 60.0)
            
            # Formato decimal
            return float(texto)
        except:
            return 0.0

        # Obtener horas para cada bombero
        horas_por_bombero = {}
        for idx, row in df_res.iterrows():
            legajo = str(row['Bombero']).split(' - ')[0].strip()
            
            # Consulta específica para este legajo
            q_horas = """
            SELECT horas FROM actividades 
            WHERE legajo = ? 
            AND (CASE WHEN instr(fecha_inicio, '/')>0
                        THEN date(substr(fecha_inicio,7,4)||'-'||substr(fecha_inicio,4,2)||'-'||substr(fecha_inicio,1,2))
                        ELSE date(fecha_inicio) END)
                BETWEEN date(?) AND date(?)
            """
            
            c = conn.cursor()
            c.execute(q_horas, (legajo, s1, s2))
            registros_horas = c.fetchall()
            
            total_horas = 0.0
            for registro in registros_horas:
                horas_val = registro[0]
                total_horas += calcular_horas_decimal(horas_val)
            
            horas_por_bombero[legajo] = round(total_horas, 2)

        # Actualizar df_res con las horas calculadas
        df_res['Horas_Decimal'] = df_res['Bombero'].apply(
            lambda x: horas_por_bombero.get(str(x).split(' - ')[0].strip(), 0.0)
        )

        # Ordenar por horas
        df_res = df_res.sort_values('Horas_Decimal', ascending=False)

    def exportar_pdf_bombero_individual(self):
        if not hasattr(self, "_df_resumen_bombero_individual"):
            self.ui.show_error("Error", "Primero debe generar el informe.")
            return

        try:
            f1 = self.inf_desde.get_date()
            f2 = self.inf_hasta.get_date()

            file_path = self._sugerir_nombre_pdf(
                f"Informe_Bombero_{self.nombre_bombero_actual}",
                f1,
                f2
            )

            if not file_path:
                return

            titulo = (
                f"Informe de Actividades – {self.nombre_bombero_actual} "
                f"({f1:%d/%m/%Y} a {f2:%d/%m/%Y})"
            )

            self._crear_pdf_resumen_bombero_individual(
                file_path=file_path,
                titulo_pdf=titulo,
                df_resumen=self._df_resumen_bombero_individual,
                totales=self._pdf_totales_bombero_individual
            )

            os.startfile(os.path.abspath(file_path))

        except Exception as e:
            traceback.print_exc()
            self.ui.show_error("Error", f"No se pudo generar el PDF:\n{e}")

    def _crear_pdf_resumen_bombero_individual(
        self,
        file_path,
        titulo_pdf,
        df_resumen,
        totales
    ):
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle,
            Paragraph, Spacer
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm

        styles = getSampleStyleSheet()
        elems = []

        elems.append(Spacer(1, 20))

        # ==========================
        # TABLA RESUMEN POR ACTIVIDAD
        # ==========================
        data = [["ACTIVIDAD", "HORAS"]]

        for _, r in df_resumen.iterrows():
            data.append([
                r["Actividad"],
                self._formatear_horas_decimal_a_hhmm(r["Horas"])
            ])

        ancho = 595 - (30 * mm)
        tbl = Table(data, colWidths=[ancho * 0.75, ancho * 0.25])

        tbl.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.25, colors.gray),

            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E5AAC")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, 0), "MIDDLE"),
            ("ALIGN", (1, 1), (-1, -1), "CENTER"),
        ]))

        elems.append(tbl)
        elems.append(Spacer(1, 15))

        # ==========================
        # TOTALES (FUERA DE LA TABLA)
        # ==========================
        estilo = ParagraphStyle(
            "resumen",
            parent=styles["Normal"],
            fontSize=9,
            leftIndent=10,
            spaceAfter=4
        )

        total_actividades = totales.get("total_actividades", 0)
        total_registros = totales.get("total_registros", 0)
        total_horas = totales.get("total_horas", 0)

        promedio = (
            total_horas / total_actividades
            if total_actividades else 0
        )

        elems.append(Paragraph(
            f"<b>• Total de registros:</b> {total_registros}", estilo
        ))
        elems.append(Paragraph(
            f"<b>• Actividades distintas:</b> {total_actividades}", estilo
        ))
        elems.append(Paragraph(
            f"<b>• Total de horas:</b> "
            f"{self._formatear_horas_decimal_a_hhmm(total_horas)}", estilo
        ))
        elems.append(Paragraph(
            f"<b>• Promedio de horas por actividad:</b> "
            f"{self._formatear_horas_decimal_a_hhmm(promedio)}", estilo
        ))

        doc = SimpleDocTemplate(
            file_path,
            pagesize=A4,
            leftMargin=15 * mm,
            rightMargin=15 * mm,
            topMargin=40 * mm,
            bottomMargin=15 * mm
        )

        doc.build(
            elems,
            onFirstPage=lambda c, d:
                self._formato_pdf(c, d, titulo_pdf, False),
            onLaterPages=lambda c, d:
                self._formato_pdf(c, d, titulo_pdf, False)
        )

    def verificar_horas_legajo_191(self):
        """Función de depuración para verificar las horas del legajo 191"""
        import sqlite3
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        # Consulta detallada para ver todas las horas del legajo 191
        c.execute("""
            SELECT id, fecha_inicio, actividad, horas, 
                CASE 
                    WHEN typeof(horas) = 'text' AND horas LIKE '%:%' THEN
                        CAST(SUBSTR(horas, 1, INSTR(horas, ':')-1) AS REAL) + 
                        CAST(SUBSTR(horas, INSTR(horas, ':')+1) AS REAL) / 60.0
                    ELSE CAST(COALESCE(horas, 0) AS REAL)
                END AS horas_decimal
            FROM actividades 
            WHERE legajo = '191'
            ORDER BY fecha_inicio
        """)
        
        rows = c.fetchall()
        print("=== VERIFICACIÓN DETALLADA LEGAJO 191 ===")
        total_decimal = 0.0
        for row in rows:
            id_, fecha, actividad, horas_str, horas_decimal = row
            print(f"ID {id_}: {fecha} - {actividad} - Horas: '{horas_str}' -> {horas_decimal:.2f}")
            total_decimal += horas_decimal
        
        print(f"\nTotal decimal para 191: {total_decimal:.2f} horas")
        
        conn.close()

        # ---------------- Informe por BOMBERO (específico) - CON SELECCIÓN EN COMBO ----------------
    def informe_detalle_por_bombero(self):
        """Para DETALLE cuando se selecciona un bombero específico."""
        try:
            f1 = self.inf_desde.get_date()
            f2 = self.inf_hasta.get_date()
            s1, s2 = f1.strftime("%Y-%m-%d"), f2.strftime("%Y-%m-%d")
            
            leg_sel = (self.inf_legajo_cb.get() or "").strip()
            if not leg_sel:
                messagebox.showwarning("Selección", "Debe seleccionar un bombero.")
                return
                
            legajo = leg_sel.split(" - ")[0].strip()
            
        except Exception as e:
            self.ui.show_error("Error", f"Datos inválidos: {e}")
            return
        
        conn = sqlite3.connect(DB_PATH)
        try:
            # CONSULTA con horas en HH:MM desde el principio
            q = """
                SELECT 
                    a.id, 
                    a.legajo, 
                    a.actividad, 
                    a.area,
                    a.fecha_inicio, 
                    a.fecha_fin, 
                    a.hora_inicio, 
                    a.hora_fin,
                    -- 🔥 Horas en formato HH:MM directo
                    printf('%d:%02d', 
                        CAST(
                            CASE 
                                WHEN typeof(a.horas) = 'text' AND a.horas LIKE '%:%' THEN
                                    CAST(SUBSTR(a.horas, 1, INSTR(a.horas, ':')-1) AS REAL) + 
                                    CAST(SUBSTR(a.horas, INSTR(a.horas, ':')+1) AS REAL) / 60.0
                                ELSE CAST(COALESCE(a.horas, 0) AS REAL)
                            END
                        AS INTEGER),
                        CAST(ROUND((
                            CASE 
                                WHEN typeof(a.horas) = 'text' AND a.horas LIKE '%:%' THEN
                                    CAST(SUBSTR(a.horas, 1, INSTR(a.horas, ':')-1) AS REAL) + 
                                    CAST(SUBSTR(a.horas, INSTR(a.horas, ':')+1) AS REAL) / 60.0
                                ELSE CAST(COALESCE(a.horas, 0) AS REAL)
                            END
                            - CAST(
                                CASE 
                                    WHEN typeof(a.horas) = 'text' AND a.horas LIKE '%:%' THEN
                                        CAST(SUBSTR(a.horas, 1, INSTR(a.horas, ':')-1) AS REAL) + 
                                        CAST(SUBSTR(a.horas, INSTR(a.horas, ':')+1) AS REAL) / 60.0
                                    ELSE CAST(COALESCE(a.horas, 0) AS REAL)
                                END
                            AS INTEGER)
                        ) * 60) AS INTEGER)
                    ) AS horas,
                    COALESCE(a.descripcion,'') AS descripcion,
                    CASE 
                        WHEN a.asignado IS NOT NULL AND a.asignado != '' 
                        THEN a.asignado || ' - ' || COALESCE(l_asig.apellido, '') || ' ' || COALESCE(l_asig.nombre, '')
                        ELSE '' 
                    END AS asignado
                FROM actividades a
                LEFT JOIN legajos l_asig ON a.asignado = l_asig.legajo
                WHERE a.legajo = ?
                AND (CASE WHEN instr(a.fecha_inicio, '/')>0
                            THEN date(substr(a.fecha_inicio,7,4)||'-'||substr(a.fecha_inicio,4,2)||'-'||substr(a.fecha_inicio,1,2))
                            ELSE date(a.fecha_inicio) END)
                    BETWEEN date(?) AND date(?)
                ORDER BY a.fecha_inicio DESC, a.hora_inicio DESC
            """
            
            df_detalle = pd.read_sql_query(q, conn, params=(legajo, s1, s2))
            
            if df_detalle.empty:
                self.ui.show_info("Sin datos", "No hay actividades para ese bombero en el período seleccionado.")
                self.inf_legajo_cb.set("")
                return
            
            # Extraer nombre del bombero
            nombre_bombero = leg_sel.split(" - ", 1)[1].strip() if " - " in leg_sel else leg_sel.strip()
            
            # Gráfico: horas por actividad
            df_graf = df_detalle.copy()
            df_graf["ActividadCompleta"] = df_graf["actividad"]
            
            # Convertir horas a decimal para gráfico
            df_graf["Horas_Decimal"] = df_graf["horas"].apply(self._obtener_horas_decimal)
            
            # Agrupar por actividad
            df_graf = (
                df_graf.groupby("actividad", as_index=False)
                .agg({
                    "Horas_Decimal": "sum",
                    "ActividadCompleta": "first"
                })
                .rename(columns={"actividad": "Actividad", "Horas_Decimal": "Horas"})
            )
            
            titulo = f"Actividades de {nombre_bombero} ({f1:%d/%m/%Y} - {f2:%d/%m/%Y})"
            
            # Llamar a función de visualización
            self._mostrar_form_informe_bombero_individual(
                titulo=titulo,
                df_graf=df_graf,
                df_tabla=df_detalle,
                nombre_bombero=nombre_bombero,
                x_col="Actividad",
                y_col="Horas",
                color="#4682B4"
            )
            
        except Exception as e:
            self.ui.show_error("Error", f"No se pudo generar el informe:\n{e}")
            import traceback
            traceback.print_exc()
        finally:
            conn.close()

#----------INFORME POR BOMBERO ----------------------------------
    def informe_por_bombero(self):
        """Función principal que decide qué informe mostrar."""
        leg_sel = (self.inf_legajo_cb.get() or "").strip()
        
        if leg_sel:
            # Tiene selección → Detalle
            self.informe_detalle_por_bombero()
        else:
            # Sin selección → Resumen
            self.informe_resumen_por_bombero()

    def _mostrar_form_informe_actividades_todas(
            self, titulo, df_graf, df_tabla=None,
            x_col=None, y_col=None, color="#B22222",
            total_actividades_distintas=0, total_registros=0,
            total_bomberos=0, total_horas=0, cantidad_meses=0,
            promedio_horas_por_registro=0, promedio_registros_por_actividad=0,
            promedio_horas_por_bombero=0, promedio_actividades_por_bombero=0,
            promedio_horas_por_mes=0,
            report_type="actividades_todas"
        ):

        self.tipo_informe_actual = "actividades_todas"

        if df_graf is None or df_graf.empty:
            self.ui.show_info("Sin datos", "No hay datos para mostrar.")
            return

        df_graf = df_graf.copy()
        if x_col is None or x_col not in df_graf.columns:
            x_col = df_graf.columns[0]
        if y_col is None or y_col not in df_graf.columns:
            y_col = df_graf.columns[1] if len(df_graf.columns) > 1 else df_graf.columns[0]
        df_graf[y_col] = pd.to_numeric(df_graf[y_col], errors="coerce").fillna(0)

        ventana = tk.Toplevel(self.master)
        if hasattr(self.master, "_icono_global") and self.master._icono_global:
            ventana.iconphoto(True, self.master._icono_global)
        ventana.title(titulo)
        ventana.state("zoomed")  # Maximizar al abrir
        ventana.configure(bg="white")
        try:
            self.logo_informe = tk.PhotoImage(file="C:/Actividades Bomberos/bomberos.png")
            ventana.iconphoto(False, self.logo_informe)
        except Exception:
            pass
        ventana.grab_set()

        # ---- Panel contenedor principal ----
        container = tk.Frame(ventana, bg="#f0f0f0")
        container.pack(fill="both", expand=True, padx=10, pady=10)

        # ---- Panel izquierdo: gráfico + tabla ----
        left_panel = tk.Frame(container, bg="white", highlightthickness=1, highlightbackground="red")
        left_panel.pack(side="left", fill="both", expand=True, padx=(0,5))

        # Scroll vertical general
        canvas = tk.Canvas(left_panel, bg="white", highlightthickness=0)
        scrollbar = ttk.Scrollbar(left_panel, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        scrollable_frame = tk.Frame(canvas, bg="white")
        canvas.create_window((0,0), window=scrollable_frame, anchor="nw")
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

        # ---- Gráfico (55% altura) ----
        frame_grafico = tk.Frame(scrollable_frame, bg="white", highlightthickness=1, highlightbackground="red")
        frame_grafico.pack(fill="both", expand=True, pady=(0,5))

        try:
            filas = len(df_graf)
            alto_figura = max(4, filas * 0.35)
            fig, ax = plt.subplots(figsize=(10, alto_figura))
            df_plot = df_graf.sort_values(by=y_col, ascending=True)
            etiquetas_y = [re.sub(r'^\d+\s*-\s*', '', str(n))[:27]+"..." if len(str(n))>30 else str(n) for n in df_plot[x_col]]
            barras = ax.barh(etiquetas_y, df_plot[y_col], color=color)
            ax.set_xlabel("Horas Trabajadas", fontsize=8)
            ax.set_title(titulo, fontsize=10, fontweight="bold", pad=10)
            ax.tick_params(axis="y", labelsize=7)
            plt.subplots_adjust(left=0.25)
            max_val = df_plot[y_col].max() if not df_plot.empty else 0
            for bar, width in zip(barras, df_plot[y_col]):
                if width > 0:
                    h_txt = self._formatear_horas_decimal_a_hhmm(width)
                    pos_x = width - (max_val * 0.03) if width > max_val * 0.3 else width + (max_val * 0.01)
                    color_txt = "white" if width > max_val * 0.3 else "black"
                    ax.text(pos_x, bar.get_y()+bar.get_height()/2, h_txt,
                            ha="center" if width > max_val*0.3 else "left",
                            va="center", fontsize=8, fontweight="bold", color=color_txt)
            plt.tight_layout()
            canvas_fig = FigureCanvasTkAgg(fig, master=frame_grafico)
            canvas_fig.draw()
            canvas_fig.get_tk_widget().pack(fill="both", expand=True)
        except Exception as e:
            tk.Label(frame_grafico, text=f"Error al crear gráfico:\n{e}", fg="red", bg="white").pack(expand=True)

        # ---- Tabla (45% altura) ----
        if df_tabla is not None and not df_tabla.empty:
            frame_tabla = tk.Frame(scrollable_frame, bg="white", highlightthickness=1, highlightbackground="red")
            frame_tabla.pack(fill="both", expand=True)
            canvas_tabla = tk.Canvas(frame_tabla, bg="white")
            scroll_y = ttk.Scrollbar(frame_tabla, orient="vertical", command=canvas_tabla.yview)
            scroll_x = ttk.Scrollbar(frame_tabla, orient="horizontal", command=canvas_tabla.xview)
            inner_frame = tk.Frame(canvas_tabla, bg="white")
            canvas_tabla.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
            scroll_y.pack(side="right", fill="y")
            scroll_x.pack(side="bottom", fill="x")
            canvas_tabla.pack(side="left", fill="both", expand=True)
            canvas_tabla.create_window((0,0), window=inner_frame, anchor="nw")
            inner_frame.bind("<Configure>", lambda e: canvas_tabla.configure(scrollregion=canvas_tabla.bbox("all")))

            for col in df_tabla.columns:
                if "hora" in col.lower():
                    df_tabla[col] = df_tabla[col].apply(lambda x: self._formatear_horas_decimal_a_hhmm(x) if isinstance(x,(int,float)) else x)

            style = ttk.Style()
            style.theme_use("default")
            style.configure("Treeview", font=("Arial",8), rowheight=24)
            style.configure("Treeview.Heading", font=("Arial",10,"bold"), foreground="red", anchor="center")
            tree = ttk.Treeview(inner_frame, columns=list(df_tabla.columns), show="headings", height=10)
            for col in df_tabla.columns:
                tree.heading(col, text=col, anchor="center")
                tree.column(col, width=160, anchor="center")
            for _, row in df_tabla.iterrows():
                tree.insert("", "end", values=row.tolist())
            tree.pack(fill="both", expand=True)

        # ---- Panel derecho: botones + totales ----
        right_panel = tk.Frame(container, bg="#f5f5f5", highlightthickness=1, highlightbackground="red", width=220)
        right_panel.pack(side="right", fill="y", padx=(5,0), pady=10)
        right_panel.pack_propagate(False)

        # Botones con colores
        style_btn = ttk.Style()
        style_btn.configure("BtnPDF.TButton", foreground="red", font=("Arial",10,"bold"))
        style_btn.configure("BtnExcel.TButton", foreground="green", font=("Arial",10,"bold"))

        ttk.Button(right_panel, text="📄 Exportar PDF", style="BtnPDF.TButton",
                command=self.exportar_pdf_actividades_total
                ).pack(fill="x", padx=15, pady=5)
        ttk.Button(right_panel, text="📊 Exportar Excel", style="BtnExcel.TButton",
        command=lambda: self._exportar_informe_excel(titulo, df_tabla)).pack(fill="x", padx=15, pady=5)
        ttk.Separator(right_panel, orient="horizontal").pack(fill="x", padx=10, pady=10)
        ttk.Button(right_panel, text="❌ Cerrar", command=ventana.destroy).pack(fill="x", padx=15, pady=5)

        # ---- Totales ----
        resumen_frame = tk.Frame(right_panel, bg="#e8e8e8", highlightthickness=1, highlightbackground="red")
        resumen_frame.pack(fill="x", pady=20, padx=5)
        lineas_resumen = []
        if total_actividades_distintas > 0: lineas_resumen.append(f"Actividades distintas: {total_actividades_distintas}")
        if total_registros > 0: lineas_resumen.append(f"Total registros: {total_registros}")
        if total_bomberos > 0: lineas_resumen.append(f"Bomberos participantes: {total_bomberos}")
        if isinstance(total_horas,(int,float)) and total_horas>0: lineas_resumen.append(f"Total horas: {self._formatear_horas_decimal_a_hhmm(total_horas)}")
        if promedio_horas_por_bombero > 0: lineas_resumen.append(f"Promedio horas/bombero: {self._formatear_horas_decimal_a_hhmm(promedio_horas_por_bombero)}")
        if promedio_actividades_por_bombero > 0: lineas_resumen.append(f"Promedio actividades/bombero: {promedio_actividades_por_bombero:.1f}")
        if promedio_horas_por_mes > 0: lineas_resumen.append(f"Promedio horas/mes: {promedio_horas_por_mes:.1f}")
        for linea in lineas_resumen:
            tk.Label(resumen_frame, text=linea, font=("Arial",10,"bold"), bg="#e8e8e8", fg="black", anchor="w").pack(fill="x", padx=5, pady=2)

        # =========================================================
        # 🔐 GUARDAR DATOS PARA EXPORTAR (PDF / EXCEL)
        # =========================================================
        self._df_tabla_actual = df_tabla.copy() if df_tabla is not None else pd.DataFrame()
        self._df_graf_actual = df_graf.copy()
        self._titulo_actual = titulo

        self._pdf_totales_actividades = {
            "total_actividades": total_actividades_distintas,
            "total_bomberos": total_bomberos,
            "total_registros": total_registros,
            "total_horas": total_horas
        }

        self.tipo_informe_actual = "actividades_total"

        ventana.protocol("WM_DELETE_WINDOW", lambda: (plt.close("all"), ventana.destroy()))

    def _mostrar_form_informe_bomberos_todos(
        self, titulo, df_graf, df_tabla=None,
        x_col=None, y_col=None, color="#B22222",
        total_actividades_distintas=0, total_registros=0,
        total_bomberos=0, total_horas=0,
        promedio_horas_por_bombero=0, promedio_actividades_por_bombero=0,
        promedio_horas_por_mes=0
    ):

        self.tipo_informe_actual = "bomberos_todos"

        if df_graf is None or df_graf.empty:
            self.ui.show_info("Sin datos", "No hay datos para mostrar.")
            return

        df_graf = df_graf.copy()
        if x_col is None or x_col not in df_graf.columns:
            x_col = df_graf.columns[0]
        if y_col is None or y_col not in df_graf.columns:
            y_col = df_graf.columns[1] if len(df_graf.columns) > 1 else df_graf.columns[0]
        df_graf[y_col] = pd.to_numeric(df_graf[y_col], errors="coerce").fillna(0)

        ventana = tk.Toplevel(self.master)
        if hasattr(self.master, "_icono_global") and self.master._icono_global:
            ventana.iconphoto(True, self.master._icono_global)
        ventana.title(titulo)
        ventana.state("zoomed")
        ventana.configure(bg="white")
        try:
            self.logo_informe = tk.PhotoImage(file="C:/Actividades Bomberos/bomberos.png")
            ventana.iconphoto(False, self.logo_informe)
        except Exception:
            pass
        ventana.grab_set()

        # ---- Panel contenedor principal ----
        container = tk.Frame(ventana, bg="#f0f0f0")
        container.pack(fill="both", expand=True, padx=10, pady=10)

        # ---- Panel izquierdo: gráfico + tabla ----
        left_panel = tk.Frame(container, bg="white", highlightthickness=1, highlightbackground="red")
        left_panel.pack(side="left", fill="both", expand=True, padx=(0,5))

        canvas = tk.Canvas(left_panel, bg="white", highlightthickness=0)
        scrollbar = ttk.Scrollbar(left_panel, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        scrollable_frame = tk.Frame(canvas, bg="white")
        canvas.create_window((0,0), window=scrollable_frame, anchor="nw")
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

        # ---- Gráfico ----
        frame_grafico = tk.Frame(scrollable_frame, bg="white", highlightthickness=1, highlightbackground="red")
        frame_grafico.pack(fill="both", expand=True, pady=(0,5))

        try:
            filas = len(df_graf)
            alto_figura = max(4, filas * 0.35)
            fig, ax = plt.subplots(figsize=(10, alto_figura))
            df_plot = df_graf.sort_values(by=y_col, ascending=True)
            etiquetas_y = [re.sub(r'^\d+\s*-\s*', '', str(n))[:27]+"..." if len(str(n))>30 else str(n) for n in df_plot[x_col]]
            barras = ax.barh(etiquetas_y, df_plot[y_col], color=color)
            ax.set_xlabel("Horas Trabajadas", fontsize=8)
            ax.set_title(titulo, fontsize=10, fontweight="bold", pad=10)
            ax.tick_params(axis="y", labelsize=7)
            plt.subplots_adjust(left=0.25)
            max_val = df_plot[y_col].max() if not df_plot.empty else 0
            for bar, width in zip(barras, df_plot[y_col]):
                if width > 0:
                    h_txt = self._formatear_horas_decimal_a_hhmm(width)
                    pos_x = width - (max_val * 0.03) if width > max_val * 0.3 else width + (max_val * 0.01)
                    color_txt = "white" if width > max_val * 0.3 else "black"
                    ax.text(pos_x, bar.get_y()+bar.get_height()/2, h_txt,
                            ha="center" if width > max_val*0.3 else "left",
                            va="center", fontsize=8, fontweight="bold", color=color_txt)
            plt.tight_layout()
            canvas_fig = FigureCanvasTkAgg(fig, master=frame_grafico)
            canvas_fig.draw()
            canvas_fig.get_tk_widget().pack(fill="both", expand=True)
        except Exception as e:
            tk.Label(frame_grafico, text=f"Error al crear gráfico:\n{e}", fg="red", bg="white").pack(expand=True)

        # ---- Tabla ----
        if df_tabla is not None and not df_tabla.empty:
            frame_tabla = tk.Frame(scrollable_frame, bg="white", highlightthickness=1, highlightbackground="red")
            frame_tabla.pack(fill="both", expand=True)
            canvas_tabla = tk.Canvas(frame_tabla, bg="white")
            scroll_y = ttk.Scrollbar(frame_tabla, orient="vertical", command=canvas_tabla.yview)
            scroll_x = ttk.Scrollbar(frame_tabla, orient="horizontal", command=canvas_tabla.xview)
            inner_frame = tk.Frame(canvas_tabla, bg="white")
            canvas_tabla.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
            scroll_y.pack(side="right", fill="y")
            scroll_x.pack(side="bottom", fill="x")
            canvas_tabla.pack(side="left", fill="both", expand=True)
            canvas_tabla.create_window((0,0), window=inner_frame, anchor="nw")
            inner_frame.bind("<Configure>", lambda e: canvas_tabla.configure(scrollregion=canvas_tabla.bbox("all")))

            for col in df_tabla.columns:
                if "hora" in col.lower():
                    df_tabla[col] = df_tabla[col].apply(lambda x: self._formatear_horas_decimal_a_hhmm(x) if isinstance(x,(int,float)) else x)

            style = ttk.Style()
            style.theme_use("default")
            style.configure("Treeview", font=("Arial",8), rowheight=24)
            style.configure("Treeview.Heading", font=("Arial",10,"bold"), foreground="red", anchor="center")
            tree = ttk.Treeview(inner_frame, columns=list(df_tabla.columns), show="headings", height=10)
            for col in df_tabla.columns:
                tree.heading(col, text=col, anchor="center")
                tree.column(col, width=160, anchor="center")
            for _, row in df_tabla.iterrows():
                tree.insert("", "end", values=row.tolist())
            tree.pack(fill="both", expand=True)

        # Guardar tabla y título actuales para Excel
        self._df_tabla_actual = df_tabla.copy() if df_tabla is not None else pd.DataFrame()
        self._titulo_actual = titulo

        # ---- Panel derecho: botones + totales ----
        right_panel = tk.Frame(container, bg="#f5f5f5", highlightthickness=1, highlightbackground="red", width=260)
        right_panel.pack(side="right", fill="y", padx=(5,0), pady=10)
        right_panel.pack_propagate(False)

        style_btn = ttk.Style()
        style_btn.configure("BtnPDF.TButton", foreground="red", font=("Arial",10,"bold"))
        style_btn.configure("BtnExcel.TButton", foreground="green", font=("Arial",10,"bold"))

        # PDF usa datos de informe ya generado
        ttk.Button(right_panel, text="📄 Exportar PDF", style="BtnPDF.TButton",
                command=self.exportar_pdf_bomberos_total).pack(fill="x", padx=15, pady=5)
        # Excel usa tabla actual
        ttk.Button(right_panel, text="📊 Exportar Excel", style="BtnExcel.TButton",
                command=lambda: self._exportar_informe_excel(titulo, df_tabla)).pack(fill="x", padx=15, pady=5)
        ttk.Separator(right_panel, orient="horizontal").pack(fill="x", padx=10, pady=10)
        ttk.Button(right_panel, text="❌ Cerrar", command=ventana.destroy).pack(fill="x", padx=15, pady=5)

        # ---- Totales ----
        resumen_frame = tk.Frame(right_panel, bg="#e8e8e8", highlightthickness=1, highlightbackground="red")
        resumen_frame.pack(fill="x", pady=20, padx=5)
        lineas_resumen = []
        if total_actividades_distintas > 0: lineas_resumen.append(f"Actividades distintas: {total_actividades_distintas}")
        if total_registros > 0: lineas_resumen.append(f"Total registros: {total_registros}")
        if total_bomberos > 0: lineas_resumen.append(f"Bomberos participantes: {total_bomberos}")
        if isinstance(total_horas,(int,float)) and total_horas>0: lineas_resumen.append(f"Total horas: {self._formatear_horas_decimal_a_hhmm(total_horas)}")
        if promedio_horas_por_bombero>0: lineas_resumen.append(f"Promedio horas/bombero: {self._formatear_horas_decimal_a_hhmm(promedio_horas_por_bombero)}")
        if promedio_actividades_por_bombero>0: lineas_resumen.append(f"Promedio actividades/bombero: {promedio_actividades_por_bombero:.1f}")
        if promedio_horas_por_mes>0: lineas_resumen.append(f"Promedio horas/mes: {promedio_horas_por_mes:.1f}")
        for linea in lineas_resumen:
            tk.Label(
                resumen_frame,
                text=linea,
                font=("Arial",10,"bold"),
                bg="#e8e8e8",
                fg="black",
                anchor="w",
                wraplength=240
            ).pack(fill="x", padx=5, pady=2)

        ventana.protocol("WM_DELETE_WINDOW", lambda: (plt.close("all"), ventana.destroy()))

    #-----------------------------INFORME HORAS POR PERIODO (TODOS)---------------------------------
    def informe_horas_por_periodo(self):
        """Informe de horas totales por período (agrupado por mes) - Diseño TODOS."""
       
        self.tipo_informe_actual = "horas_periodo"

        try:
            f1 = self.inf_desde.get_date()
            f2 = self.inf_hasta.get_date()
            s1, s2 = f1.strftime("%Y-%m-%d"), f2.strftime("%Y-%m-%d")
        except Exception as e:
            self.ui.show_error("Error", "Debe seleccionar fechas válidas.")
            return
        
        conn = sqlite3.connect(DB_PATH)
        try:
            # PRIMERO: Verificar qué datos realmente hay en la base
            # Consulta 1: Ver todas las fechas distintas en el período
            q_test = """
                SELECT DISTINCT 
                    fecha_inicio,
                    CASE 
                        WHEN instr(fecha_inicio, '/')>0 THEN
                            substr(fecha_inicio,7,4)||'-'||substr(fecha_inicio,4,2)||'-'||substr(fecha_inicio,1,2)
                        ELSE fecha_inicio
                    END as fecha_convertida
                FROM actividades
                WHERE fecha_inicio IS NOT NULL AND fecha_inicio != ''
                LIMIT 10
            """
            
            test_dates = pd.read_sql_query(q_test, conn)
            
            # Consulta 2: Ver el rango real de fechas
            q_range = """
                SELECT 
                    MIN(
                        CASE 
                            WHEN instr(fecha_inicio, '/')>0 THEN
                                substr(fecha_inicio,7,4)||'-'||substr(fecha_inicio,4,2)||'-'||substr(fecha_inicio,1,2)
                            ELSE fecha_inicio
                        END
                    ) as fecha_min,
                    MAX(
                        CASE 
                            WHEN instr(fecha_inicio, '/')>0 THEN
                                substr(fecha_inicio,7,4)||'-'||substr(fecha_inicio,4,2)||'-'||substr(fecha_inicio,1,2)
                            ELSE fecha_inicio
                        END
                    ) as fecha_max
                FROM actividades
                WHERE fecha_inicio IS NOT NULL AND fecha_inicio != ''
            """
            
            date_range = pd.read_sql_query(q_range, conn)
            
            # Consulta 3: Verificar si hay datos en el período solicitado
            q_check = """
                SELECT COUNT(*) as total
                FROM actividades
                WHERE (
                    CASE 
                        WHEN instr(fecha_inicio, '/')>0 THEN
                            date(substr(fecha_inicio,7,4)||'-'||substr(fecha_inicio,4,2)||'-'||substr(fecha_inicio,1,2))
                        ELSE date(fecha_inicio)
                    END
                ) BETWEEN date(?) AND date(?)
            """
            
            c = conn.cursor()
            c.execute(q_check, (s1, s2))
            total_registros = c.fetchone()[0]
            
            if total_registros == 0:
                self.ui.show_info("Sin datos", f"No hay actividades entre {f1:%d/%m/%Y} y {f2:%d/%m/%Y}")
                return
            
            # Consulta 4: Ver meses específicos que hay
            q_months = """
                SELECT DISTINCT 
                    strftime('%Y-%m', 
                        CASE 
                            WHEN instr(fecha_inicio, '/')>0 THEN
                                substr(fecha_inicio,7,4)||'-'||substr(fecha_inicio,4,2)||'-'||substr(fecha_inicio,1,2)
                            ELSE fecha_inicio
                        END
                    ) AS Mes
                FROM actividades
                WHERE (
                    CASE 
                        WHEN instr(fecha_inicio, '/')>0 THEN
                            date(substr(fecha_inicio,7,4)||'-'||substr(fecha_inicio,4,2)||'-'||substr(fecha_inicio,1,2))
                        ELSE date(fecha_inicio)
                    END
                ) BETWEEN date(?) AND date(?)
                ORDER BY Mes
            """
            
            months_data = pd.read_sql_query(q_months, conn, params=(s1, s2))
            
            # AHORA la consulta principal (MODIFICADA para mejor debug)
            q = """
                SELECT 
                    strftime('%Y-%m', 
                        CASE 
                            WHEN instr(fecha_inicio, '/')>0 THEN
                                substr(fecha_inicio,7,4)||'-'||substr(fecha_inicio,4,2)||'-'||substr(fecha_inicio,1,2)
                            ELSE fecha_inicio
                        END
                    ) AS Mes,
                    COUNT(DISTINCT legajo) AS Total_Bomberos,
                    COUNT(id) AS Total_Registros,
                    printf('%d:%02d', 
                        CAST(SUM(
                            CASE 
                                WHEN typeof(horas) = 'text' AND horas LIKE '%:%' THEN
                                    CAST(SUBSTR(horas, 1, INSTR(horas, ':')-1) AS REAL) + 
                                    CAST(SUBSTR(horas, INSTR(horas, ':')+1) AS REAL) / 60.0
                                WHEN typeof(horas) IN ('real', 'integer') THEN
                                    CAST(horas AS REAL)
                                ELSE 0.0
                            END
                        ) AS INTEGER),
                        CAST(ROUND((SUM(
                            CASE 
                                WHEN typeof(horas) = 'text' AND horas LIKE '%:%' THEN
                                    CAST(SUBSTR(horas, 1, INSTR(horas, ':')-1) AS REAL) + 
                                    CAST(SUBSTR(horas, INSTR(horas, ':')+1) AS REAL) / 60.0
                                WHEN typeof(horas) IN ('real', 'integer') THEN
                                    CAST(horas AS REAL)
                                ELSE 0.0
                            END
                        ) - CAST(SUM(
                            CASE 
                                WHEN typeof(horas) = 'text' AND horas LIKE '%:%' THEN
                                    CAST(SUBSTR(horas, 1, INSTR(horas, ':')-1) AS REAL) + 
                                    CAST(SUBSTR(horas, INSTR(horas, ':')+1) AS REAL) / 60.0
                                WHEN typeof(horas) IN ('real', 'integer') THEN
                                    CAST(horas AS REAL)
                                ELSE 0.0
                            END
                        ) AS INTEGER)) * 60) AS INTEGER)
                    ) AS Total_Horas,
                    ROUND(SUM(
                        CASE 
                            WHEN typeof(horas) = 'text' AND horas LIKE '%:%' THEN
                                CAST(SUBSTR(horas, 1, INSTR(horas, ':')-1) AS REAL) + 
                                CAST(SUBSTR(horas, INSTR(horas, ':')+1) AS REAL) / 60.0
                            WHEN typeof(horas) IN ('real', 'integer') THEN
                                CAST(horas AS REAL)
                            ELSE 0.0
                        END
                    ), 2) AS Horas_Decimal
                FROM actividades
                WHERE fecha_inicio IS NOT NULL AND fecha_inicio != ''
                AND (
                    CASE 
                        WHEN instr(fecha_inicio, '/')>0 THEN
                            date(substr(fecha_inicio,7,4)||'-'||substr(fecha_inicio,4,2)||'-'||substr(fecha_inicio,1,2))
                        ELSE date(fecha_inicio)
                    END
                ) BETWEEN date(?) AND date(?)
                GROUP BY strftime('%Y-%m', 
                    CASE 
                        WHEN instr(fecha_inicio, '/')>0 THEN
                            substr(fecha_inicio,7,4)||'-'||substr(fecha_inicio,4,2)||'-'||substr(fecha_inicio,1,2)
                        ELSE fecha_inicio
                    END
                )
                ORDER BY Mes
            """
            
            df_res = pd.read_sql_query(q, conn, params=(s1, s2))
            
            if df_res.empty:
                self.ui.show_info("Sin datos", "No hay actividades en el período seleccionado.")
                return

            # AGREGAR MESES FALTANTES (si hay saltos)
            # Generar todos los meses entre las fechas seleccionadas
            from dateutil.relativedelta import relativedelta

            all_months = []
            current = datetime(f1.year, f1.month, 1)
            end = datetime(f2.year, f2.month, 1)

            while current <= end:
                all_months.append(current.strftime('%Y-%m'))
                current += relativedelta(months=1)

            # Crear DataFrame con todos los meses
            df_all_months = pd.DataFrame({'Mes': all_months})

            # Hacer merge con los datos encontrados
            df_res = pd.merge(df_all_months, df_res, on='Mes', how='left')

            # Rellenar valores nulos
            df_res['Total_Bomberos'] = df_res['Total_Bomberos'].fillna(0).astype(int)
            df_res['Total_Registros'] = df_res['Total_Registros'].fillna(0).astype(int)
            df_res['Total_Horas'] = df_res['Total_Horas'].fillna('0:00')
            df_res['Horas_Decimal'] = df_res['Horas_Decimal'].fillna(0.0)

            # Formatear nombres de meses
            meses_esp = {
                '01': 'Ene', '02': 'Feb', '03': 'Mar', '04': 'Abr',
                '05': 'May', '06': 'Jun', '07': 'Jul', '08': 'Ago',
                '09': 'Sep', '10': 'Oct', '11': 'Nov', '12': 'Dic'
            }

            df_res['Mes_Display'] = df_res['Mes'].apply(
                lambda x: f"{meses_esp.get(x.split('-')[1], x.split('-')[1])}-{x.split('-')[0]}"
                if '-' in str(x) and len(x.split('-')) >= 2 else str(x)
            )

            # --- PREPARAR DATOS para el nuevo diseño TODOS ---
            # DataFrame para tabla principal
            df_tabla = df_res.copy()
            df_tabla['Promedio_Horas_Bombero'] = df_tabla.apply(
                lambda row: round(row['Horas_Decimal'] / row['Total_Bomberos'], 2) 
                if row['Total_Bomberos'] > 0 else 0, 
                axis=1
            )

            # DataFrame para gráfico (líneas de tendencia)
            df_graf = df_res[['Mes_Display', 'Horas_Decimal']].copy()
            df_graf.columns = ['Mes', 'Horas']

            # Calcular estadísticas para el PDF
            total_meses = len(df_res)
            total_bomberos_acum = df_res['Total_Bomberos'].sum()
            total_registros = df_res['Total_Registros'].sum()
            total_horas_decimal = df_graf['Horas'].sum()
            promedio_horas_mensual = total_horas_decimal / total_meses if total_meses > 0 else 0

            # Encontrar mes pico y mes valle
            if not df_res.empty:
                idx_max = df_res['Horas_Decimal'].idxmax()
                idx_min = df_res['Horas_Decimal'].idxmin()
                mes_pico = df_res.loc[idx_max, 'Mes_Display']
                horas_pico = df_res.loc[idx_max, 'Horas_Decimal']
                mes_valle = df_res.loc[idx_min, 'Mes_Display']
                horas_valle = df_res.loc[idx_min, 'Horas_Decimal']
                
                # Calcular variación si hay más de 1 mes
                variacion = 0
                if len(df_res) > 1:
                    horas_primero = df_res.iloc[0]['Horas_Decimal']
                    horas_ultimo = df_res.iloc[-1]['Horas_Decimal']
                    if horas_primero > 0:
                        variacion = ((horas_ultimo - horas_primero) / horas_primero) * 100
            else:
                mes_pico = mes_valle = "N/A"
                horas_pico = horas_valle = 0
                variacion = 0

            # Guardar datos para PDF
            self._df_resumen_horas_periodo = df_tabla.copy()
            self._df_graf_horas_periodo = df_graf.copy()

            self._pdf_totales_horas_periodo = {
                "total_meses": total_meses,
                "total_bomberos_acum": total_bomberos_acum,
                "total_registros": total_registros,
                "total_horas": total_horas_decimal,
                "promedio_horas_mensual": promedio_horas_mensual,
                "mes_pico": mes_pico,
                "horas_pico": horas_pico,
                "mes_valle": mes_valle,
                "horas_valle": horas_valle,
                "variacion_tendencia": round(variacion, 2)
            }

            titulo = f"Horas por Período ({f1:%d/%m/%Y} - {f2:%d/%m/%Y})"
            tot = self._pdf_totales_horas_periodo

            self._df_totales_horas_periodo = pd.DataFrame([
                ["Total meses", tot["total_meses"]],
                ["Total registros", tot["total_registros"]],
                ["Total horas (decimal)", round(tot["total_horas"], 2)],
                ["Promedio horas mensual", round(tot["promedio_horas_mensual"], 2)],
                ["Mes pico", f'{tot["mes_pico"]} ({tot["horas_pico"]} h)'],
                ["Mes valle", f'{tot["mes_valle"]} ({tot["horas_valle"]} h)'],
                ["Variación tendencia (%)", tot["variacion_tendencia"]],
            ], columns=["Concepto", "Valor"])

            # =====================================================
            # Guardar datos para EXPORTAR EXCEL
            # =====================================================
            self._df_excel_horas_periodo = df_tabla.copy()

            # Llamar a la NUEVA función de visualización
            self._mostrar_form_informe_horas_periodo(
                titulo=titulo,
                df_graf=df_graf,
                df_tabla=df_tabla,
                x_col="Mes",
                y_col="Horas",
                color="#4682B4"
            )
        except Exception as e:
            self.ui.show_error("Error", f"No se pudo generar el informe de horas por período:\n{e}")
            import traceback
            traceback.print_exc()
        finally:
            conn.close()

    def _crear_botones_informe(self, parent_frame, titulo, df=None, export_pdf_func=None):

        # ----- FUNCIONES DE LOS BOTONES -----
        def exportar_excel():
            if df is not None and not df.empty:
                try:
                    # Llama a tu función existente de exportar Excel
                    self.exportar_excel(df, titulo)
                except Exception as e:
                    self.ui.show_error("Error Excel", f"No se pudo exportar a Excel:\n{e}")
            else:
                self.ui.show_info("Sin datos", "No hay datos para exportar a Excel.")

    # ----------------------------------------------------------
    def exportar_pdf_horas_periodo(self):
        """Exporta el informe de horas por período a PDF."""
        
        if not hasattr(self, "_df_excel_horas_periodo"):
            self.ui.show_error("Error", "Primero debe generar el informe.")
            return
        
        try:
            # Obtener fechas
            try:
                f1 = self.inf_desde.get_date()
                f2 = self.inf_hasta.get_date()
            except:
                f1 = f2 = None
            
            # Generar nombre sugerido
            if f1 and f2:
                sugerido = f"Horas_Periodo_{f1:%Y%m%d}_a_{f2:%Y%m%d}.pdf"
            else:
                sugerido = "Horas_Periodo.pdf"
            
            # Pedir ubicación para guardar
            file_path = self.ui.ask_save_file(
                defaultextension=".pdf",
                initialfile=sugerido,
                filetypes=[("PDF files", "*.pdf")],
                title="Guardar Informe Horas por Período"
            )
            
            if not file_path:
                return
            
            # Título para el PDF
            if f1 and f2:
                titulo_pdf = f"Horas por Período ({f1:%d/%m/%Y} - {f2:%d/%m/%Y})"
            else:
                titulo_pdf = "Informe de Horas por Período"
            
            # Verificar si tenemos la función para crear PDF
            if hasattr(self, "_crear_pdf_resumen_horas_periodo"):
                # Generar PDF
                success = self._crear_pdf_resumen_horas_periodo(
                    file_path=file_path,
                    titulo_pdf=titulo_pdf,
                    df_resumen=self._df_resumen_horas_periodo,
                    totales=getattr(self, "_pdf_totales_horas_periodo", {})
                )
                
                if success:
                    self.ui.show_info(
                        "Éxito", 
                        f"Informe exportado correctamente:\n{os.path.basename(file_path)}"
                    )
                    # Abrir el PDF
                    try:
                        os.startfile(os.path.abspath(file_path))
                    except:
                        pass
                else:
                    self.ui.show_error("Error", "No se pudo generar el PDF.")
            else:
                # Función no implementada todavía
                messagebox.showwarning(
                    "Funcionalidad pendiente",
                    "La función para generar PDF de horas por período aún no está implementada.\n\n"
                    "Se implementará en la próxima actualización."
                )
                
        except Exception as e:
            import traceback
            traceback.print_exc()
            self.ui.show_error("Error", f"No se pudo exportar el PDF:\n{e}")

    def exportar_excel_horas_periodo(self):
        """Exporta el informe de horas por período a Excel."""
        if not hasattr(self, "_df_resumen_horas_periodo"):
            self.ui.show_error("Error", "Primero debe generar el informe.")
            return

        try:
            # Obtener fechas
            try:
                f1 = self.inf_desde.get_date()
                f2 = self.inf_hasta.get_date()
                titulo = f"Horas por Período ({f1:%d/%m/%Y} - {f2:%d/%m/%Y})"
            except:
                titulo = "Horas por Período"

            # -------- TOTALES --------
            tot = self._pdf_totales_horas_periodo

            # Definir tipo de informe (CLAVE para la hoja 2)
            self.tipo_informe_actual = "horas_periodo"
            self._pdf_totales_horas_periodo = tot

            # ---------- LIMPIEZA DE DATAFRAME ----------
            df = self._df_resumen_horas_periodo.copy()

            # Quedarse SOLO con las columnas necesarias
            columnas_finales = {
                "Mes_Display": "Mes",
                "Horas_Decimal": "Horas",
                "Total_Registros": "Registros",
                "Total_Bomberos": "Bomberos"
            }

            # Filtrar columnas existentes
            columnas_usar = [c for c in columnas_finales if c in df.columns]
            df = df[columnas_usar]

            # Renombrar
            df.rename(columns=columnas_finales, inplace=True)

            # --------- AGREGAR PROMEDIO ---------
            if "Horas" in df.columns and "Bomberos" in df.columns:
                # Evitar división por cero
                df["Promedio_Horas"] = df.apply(
                    lambda row: row["Horas"] / row["Bomberos"] if row["Bomberos"] else 0,
                    axis=1
                )
                # Redondear a 2 decimales
                df["Promedio_Horas"] = df["Promedio_Horas"].round(2)

            # Guardar DF limpio
            self._df_resumen_horas_periodo = df

            # ---------- EXPORTAR ----------
            self._exportar_informe_excel(
                titulo,
                self._df_resumen_horas_periodo
            )

        except Exception as e:
            self.ui.show_error("Error", f"No se pudo exportar a Excel:\n{e}")

    def _crear_pdf_resumen_horas_periodo(self, file_path, titulo_pdf, df_resumen, totales):
        """
        Crea el PDF del informe de horas por período.
        Diseño similar al de 'Bomberos TODOS' pero con gráfico de líneas.
        """
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle,
            Paragraph, Spacer, Image as RLImage
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm, inch
        from reportlab.graphics.shapes import Drawing
        from reportlab.graphics.charts.linecharts import HorizontalLineChart
        import matplotlib.pyplot as plt
        import tempfile, os
        
        # Verificar que tenemos datos
        if df_resumen is None or df_resumen.empty:
            return False
        
        try:
            # Preparar datos para el gráfico
            meses = df_resumen['Mes_Display'].tolist()
            horas_decimal = df_resumen['Horas_Decimal'].tolist()
            
            # Crear gráfico matplotlib temporal para incluir en PDF
            temp_chart = None
            try:
                fig, ax = plt.subplots(figsize=(8, 4))
                ax.plot(meses, horas_decimal, marker='o', linewidth=2, 
                    color='#4682B4', markersize=6)
                ax.fill_between(meses, horas_decimal, alpha=0.2, color='#4682B4')
                ax.set_xlabel('Mes', fontweight='bold')
                ax.set_ylabel('Horas Totales', fontweight='bold')
                ax.set_title('Evolución Mensual de Horas', fontweight='bold')
                ax.grid(True, alpha=0.3, linestyle='--')
                
                if len(meses) > 6:
                    plt.xticks(rotation=45, ha='right')
                
                plt.tight_layout()
                
                # Guardar temporalmente
                temp_dir = tempfile.gettempdir()
                temp_chart_path = os.path.join(temp_dir, f"chart_horas_{os.getpid()}.png")
                plt.savefig(temp_chart_path, dpi=150, bbox_inches='tight')
                plt.close('all')
                
                temp_chart = temp_chart_path
            except Exception as e:
                temp_chart = None
            
            # Configurar documento
            styles = getSampleStyleSheet()
            elems = []
            
            # Espacio para encabezado
            elems.append(Spacer(1, 20))
            
            # ==========================
            # 📊 TABLA RESUMEN
            # ==========================
            data = [["MES", "HORAS TOTALES", "BOMBEROS", "REGISTROS", "PROMEDIO H/B"]]
            
            for _, r in df_resumen.iterrows():
                horas_fmt = self._formatear_horas_decimal_a_hhmm(r["Horas_Decimal"])
                promedio = r.get("Promedio_Horas_Bombero", 0)
                promedio_fmt = self._formatear_horas_decimal_a_hhmm(promedio)
                
                data.append([
                    r["Mes_Display"],
                    horas_fmt,
                    str(int(r["Total_Bomberos"])),
                    str(int(r["Total_Registros"])),
                    promedio_fmt
                ])
            
            # Calcular anchos de columna
            ancho_total = 595 - (30 * mm)  # A4 menos márgenes
            ancho_cols = [
                ancho_total * 0.25,  # Mes
                ancho_total * 0.20,  # Horas
                ancho_total * 0.15,  # Bomberos
                ancho_total * 0.15,  # Registros
                ancho_total * 0.25   # Promedio
            ]
            
            tbl = Table(data, colWidths=ancho_cols)
            
            tbl.setStyle(TableStyle([
                ("GRID", (0, 0), (-1, -1), 0.25, colors.gray),
                
                # 🔹 ENCABEZADOS
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E5AAC")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("VALIGN", (0, 0), (-1, 0), "MIDDLE"),
                
                # 🔹 CUERPO
                ("ALIGN", (1, 1), (4, -1), "CENTER"),  # Centrar columnas 1-4
                ("ALIGN", (0, 1), (0, -1), "LEFT"),     # Alinear izquierda la columna Mes
                ("VALIGN", (0, 1), (-1, -1), "MIDDLE"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
            ]))
            
            elems.append(tbl)
            elems.append(Spacer(1, 15))
            
            # ==========================
            # 📈 GRÁFICO DE LÍNEAS
            # ==========================
            if temp_chart and os.path.exists(temp_chart):
                try:
                    elems.append(Paragraph(
                        "<b>EVOLUCIÓN MENSUAL DE HORAS</b>",
                        ParagraphStyle(
                            "ChartTitle",
                            parent=styles["Heading3"],
                            fontSize=11,
                            spaceAfter=6
                        )
                    ))
                    
                    chart_img = RLImage(temp_chart, width=400, height=200)
                    elems.append(chart_img)
                    elems.append(Spacer(1, 10))
                    
                except Exception as e:
                    pass
            
            # ==========================
            # 📊 ESTADÍSTICAS
            # ==========================
            elems.append(Paragraph(
                "<b>ESTADÍSTICAS DEL PERÍODO</b>",
                ParagraphStyle(
                    "StatsTitle",
                    parent=styles["Heading3"],
                    fontSize=11,
                    spaceBefore=10,
                    spaceAfter=6
                )
            ))
            
            estilo_stats = ParagraphStyle(
                "stats",
                parent=styles["Normal"],
                fontSize=9,
                leftIndent=10,
                spaceAfter=4
            )
            
            # Estadísticas principales
            elems.append(Paragraph(
                f"<b>• Total de meses analizados:</b> {totales.get('total_meses', 0)}",
                estilo_stats
            ))
            elems.append(Paragraph(
                f"<b>• Horas totales del período:</b> "
                f"{self._formatear_horas_decimal_a_hhmm(totales.get('total_horas', 0))}",
                estilo_stats
            ))
            elems.append(Paragraph(
                f"<b>• Registros totales:</b> {totales.get('total_registros', 0)}",
                estilo_stats
            ))
            elems.append(Paragraph(
                f"<b>• Bomberos distintos (acumulado):</b> {totales.get('total_bomberos_acum', 0)}",
                estilo_stats
            ))
            elems.append(Paragraph(
                f"<b>• Promedio mensual de horas:</b> "
                f"{self._formatear_horas_decimal_a_hhmm(totales.get('promedio_horas_mensual', 0))}",
                estilo_stats
            ))
            
            # Estadísticas comparativas
            elems.append(Spacer(1, 5))
            elems.append(Paragraph("<b>ANÁLISIS COMPARATIVO</b>", estilo_stats))
            
            elems.append(Paragraph(
                f"<b>• Mes con mayor actividad:</b> {totales.get('mes_pico', 'N/A')} "
                f"({self._formatear_horas_decimal_a_hhmm(totales.get('horas_pico', 0))})",
                estilo_stats
            ))
            elems.append(Paragraph(
                f"<b>• Mes con menor actividad:</b> {totales.get('mes_valle', 'N/A')} "
                f"({self._formatear_horas_decimal_a_hhmm(totales.get('horas_valle', 0))})",
                estilo_stats
            ))
            
            variacion = totales.get('variacion_tendencia', 0)
            if variacion != 0:
                tendencia = "creciente ↗" if variacion > 0 else "decreciente ↘"
                elems.append(Paragraph(
                    f"<b>• Tendencia del período:</b> {tendencia} ({abs(variacion):.1f}%)",
                    estilo_stats
                ))
            
            # ==========================
            # 🧾 CONSTRUIR PDF
            # ==========================
            doc = SimpleDocTemplate(
                file_path,
                pagesize=A4,
                leftMargin=15 * mm,
                rightMargin=15 * mm,
                topMargin=40 * mm,
                bottomMargin=15 * mm
            )
            
            doc.build(
                elems,
                onFirstPage=lambda c, d:
                    self._formato_pdf(c, d, titulo_pdf, False),
                onLaterPages=lambda c, d:
                    self._formato_pdf(c, d, titulo_pdf, False)
            )
            # 🧹 Eliminar gráfico temporal DESPUÉS de generar el PDF
            if temp_chart and os.path.exists(temp_chart):
                try:
                    os.remove(temp_chart)
                except Exception as e:
                    pass
            
            return True
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return False
    
    def imprimir_legajo(self):

        nro_legajo = self.var_id.get().strip()
        if not nro_legajo:
            self.ui.show_error("Error", "Debe indicar un legajo antes de exportar.")
            return

        sugerido = self._default_informe_filename(f"Legajo_{nro_legajo}")
        path = self.ui.ask_save_file(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            initialfile=sugerido
        )
        if not path:
            return

        try:
            pdf_bytes = self._generar_pdf_legajo_en_memoria()

            with open(path, "wb") as f:
                f.write(pdf_bytes)

            os.startfile(path)

        except Exception as e:
            traceback.print_exc()
            self.ui.show_error("Error", f"No se pudo generar el PDF:\n{e}")

    # ----------------------------------------------------------
    def imprimir_listado_legajos(self):
        """Genera un PDF con el listado completo de legajos."""
        sugerido = self._default_filename("Listado_Legajos")
        file_path = self.ui.ask_save_file(
            defaultextension=".pdf",
            initialfile=sugerido,
            filetypes=[("PDF files", "*.pdf")],
            title="Guardar Listado de Legajos"
        )
        if not file_path:
            return

        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("SELECT legajo, apellido, nombre, grado, cargo, situacion, fecha_alta, fecha_baja FROM legajos ORDER BY apellido, nombre")
        rows = c.fetchall()
        conn.close()

        elems = []
        styles = getSampleStyleSheet()
        elems.append(Paragraph("<b>Listado General de Legajos</b>", styles["Title"]))
        elems.append(Spacer(1, 10))

        data = [["Legajo", "Apellido", "Nombre", "Grado", "Cargo", "Situación", "Alta", "Baja"]] + list(rows)
        table = Table(data, repeatRows=1, hAlign="CENTER", colWidths=[45, 90, 90, 70, 100, 80, 60, 60])
        table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        elems.append(table)

        self._crear_pdf_unificado(file_path, elems, "Listado de Legajos")
        self.ui.show_info("Éxito", f"Listado de legajos exportado:\n{file_path}")

    # ----------------------------------------------------------
    def imprimir_listado_conceptos(self):
        """Genera un PDF con todos los conceptos registrados."""
        sugerido = self._default_filename("Listado_Conceptos")
        file_path = self.ui.ask_save_file(
            defaultextension=".pdf",
            initialfile=sugerido,
            filetypes=[("PDF files", "*.pdf")],
            title="Guardar Listado de Conceptos"
        )
        if not file_path:
            return

        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("SELECT concepto, puntos, detalle FROM conceptos ORDER BY concepto")
        rows = c.fetchall()
        conn.close()

        elems = []
        styles = getSampleStyleSheet()
        elems.append(Paragraph("<b>Listado de Conceptos</b>", styles["Title"]))
        elems.append(Spacer(1, 10))

        data = [["Concepto", "Puntos", "Detalle"]] + rows
        table = Table(data, repeatRows=1, hAlign="CENTER", colWidths=[150, 60, 250])
        table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        elems.append(table)

        self._crear_pdf_unificado(file_path, elems, "Listado de Conceptos")
        self.ui.show_info("Éxito", f"Listado de conceptos exportado:\n{file_path}")

    # ----------------------------------------------------------
    def imprimir_listado_actividades(self):
        """Genera un PDF con todas las actividades registradas (modo apaisado)."""
        sugerido = self._default_filename("Listado_Actividades")
        file_path = self.ui.ask_save_file(
            defaultextension=".pdf",
            initialfile=sugerido,
            filetypes=[("PDF files", "*.pdf")],
            title="Guardar Listado de Actividades"
        )
        if not file_path:
            return

        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("""
            SELECT legajo, asignado, actividad, area, fecha_inicio, fecha_fin, hora_inicio, hora_fin, descripcion
            FROM actividades ORDER BY fecha_inicio DESC
        """)
        rows = c.fetchall()
        conn.close()

        elems = []
        styles = getSampleStyleSheet()
        elems.append(Paragraph("<b>Listado General de Actividades</b>", styles["Title"]))
        elems.append(Spacer(1, 10))

        data = [["Legajo", "Asignado", "Actividad", "Área", "Inicio", "Fin", "H.Inicio", "H.Fin", "Descripción"]] + rows
        table = Table(data, repeatRows=1, hAlign="CENTER",
                      colWidths=[45, 80, 80, 80, 60, 60, 45, 45, 160])
        table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.4, colors.black),
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        elems.append(table)

        self._crear_pdf_unificado(file_path, elems, "Listado de Actividades", landscape_mode=True)
        self.ui.show_info("Éxito", f"Listado de actividades exportado:\n{file_path}")

    def _formato_pdf(self, canvas, doc, titulo, landscape_mode, estado_actividad=None):
        """
        Dibuja encabezado y pie de página con logo, título, usuario, fecha y líneas decorativas.
        AHORA SOPORTA PÁGINAS HORIZONTALES Y VERTICALES.
        """
        from datetime import datetime
        from reportlab.lib.units import mm
        from reportlab.lib import colors

        width, height = doc.pagesize
        c = canvas

        # === Datos base ===
        try:
            logo_path = resource_path("bomberos.png")
        except Exception:
            logo_path = "bomberos.png"

        usuario = self.usuario_actual.get("username", "DESCONOCIDO")
        fecha = datetime.now().strftime("%d/%m/%Y %H:%M")

        c.saveState()

        # === AJUSTES PARA PÁGINAS HORIZONTALES ===
        if landscape_mode:
            # Subimos los títulos para que haya más aire
            logo_y = height - 55              # El logo un poco más arriba
            titulo_principal_y = height - 25  # Nombre del Cuartel
            subtitulo_y = height - 48         # LISTADO DE CONCEPTOS (Subió de 70 a 48)
            linea_y = height - 65             # La línea roja (Subió de 80 a 65)
            usuario_y = height - 75           # Usuario y Fecha (Subió de 92 a 75)
            pie_linea_y = 20
            pie_texto_y = 12
        else:
            # En vertical (valores originales)
            logo_y = height - 70
            titulo_principal_y = height - 30
            subtitulo_y = height - 50
            linea_y = height - 80
            usuario_y = height - 92
            pie_linea_y = 25
            pie_texto_y = 15

        # === LOGO ===
        try:
            c.drawImage(logo_path, 25, logo_y, width=42, height=42,
                        preserveAspectRatio=True, mask="auto")
        except Exception:
            pass

        # === TÍTULO INSTITUCIONAL ===
        c.setFont("Helvetica-Bold", 10)
        c.setFillColor(colors.black)
        c.drawCentredString(
            width / 2,
            titulo_principal_y,
            "SOCIEDAD BOMBEROS VOLUNTARIOS DE ALMAFUERTE"
        )

        # === NOMBRE DEL SISTEMA ===
        c.setFont("Helvetica", 8)
        c.drawCentredString(
            width / 2,
            titulo_principal_y - 15,
            "SIAB - Sistema Informático Automatizado de Bomberos"
        )

        # === TÍTULO DEL INFORME ===
        if titulo:
            c.setFont("Helvetica-Bold", 9)
            c.setFillColor(colors.black)
            c.drawCentredString(width / 2, subtitulo_y - 10, titulo.upper())

        # === ESTADO DE LA ACTIVIDAD ===
        if estado_actividad:
            c.setFont("Helvetica-Bold", 7)

            estado_upper = estado_actividad.upper()

            if estado_upper == "ANULADA":
                c.setFillColor(colors.red)

            elif estado_upper == "FIRMADA_SUPERVISOR":
                c.setFillColor(colors.green)

            elif estado_upper == "FIRMADA_BOMBERO":
                c.setFillColor(colors.blue)

            elif estado_upper == "BORRADOR":
                c.setFillColor(colors.red)

            else:
                c.setFillColor(colors.black)

            c.drawCentredString(
                width / 2,
                subtitulo_y - 18,   # más separación visual
                f"ESTADO: {estado_upper}"
            )

        # === LÍNEA DECORATIVA ROJA ===
        c.setStrokeColor(colors.red)
        c.setLineWidth(1.5)
        c.line(20, linea_y, width - 20, linea_y)

        # === Usuario + Fecha ===
        c.setFont("Helvetica", 8)
        c.setFillColor(colors.black)
        c.drawString(25, usuario_y, f"Usuario: {usuario}")
        c.drawRightString(width - 25, usuario_y, f"Fecha: {fecha}")

        c.restoreState()

        # === PIE DE PÁGINA ===
        c.saveState()
        c.setStrokeColor(colors.red)
        c.setLineWidth(1.5)
        c.line(20, pie_linea_y, width - 20, pie_linea_y)
        c.setFont("Helvetica", 8)
        c.setFillColor(colors.black)
        c.drawCentredString(width / 2, pie_texto_y, f"Página {doc.page}")
        c.restoreState()

    def _formato_pdf_simple(self, canvas, doc, titulo="Informe"):
        """
        Formato liviano para informes específicos (ej: bombero individual)
        Incluye:
            ✔ logo
            ✔ línea decorativa
            ✔ título del informe
            ✔ número de página
        Excluye:
            ❌ usuario
            ❌ fecha
            ❌ encabezado institucional grande
        """
        from reportlab.lib import colors

        width, height = doc.pagesize
        c = canvas

        try:
            logo_path = resource_path("bomberos.png")
        except Exception:
            logo_path = "bomberos.png"

        c.saveState()

        # === LOGO ===
        try:
            c.drawImage(
                logo_path,
                25,
                height - 60,
                width=36,
                height=36,
                preserveAspectRatio=True,
                mask="auto"
            )
        except Exception:
            pass

        # === TÍTULO DEL INFORME (más chico, una sola vez) ===
        c.setFont("Helvetica-Bold", 10)
        c.setFillColor(colors.black)
        c.drawCentredString(width / 2, height - 35, titulo)

        # === LÍNEA DECORATIVA ===
        c.setStrokeColor(colors.red)
        c.setLineWidth(1.2)
        c.line(20, height - 55, width - 20, height - 55)

        c.restoreState()

        # === PIE DE PÁGINA ===
        c.saveState()
        c.setStrokeColor(colors.red)
        c.setLineWidth(1.2)
        c.line(20, 25, width - 20, 25)
        c.setFont("Helvetica", 8)
        c.setFillColor(colors.black)
        c.drawCentredString(width / 2, 15, f"Página {doc.page}")
        c.restoreState()

    def _copiar_foto_si_corresponde(self):
        if not self.foto_path or not os.path.exists(self.foto_path):
            return ""
        try:
            nombre = os.path.basename(self.foto_path)
            destino = os.path.join(FOTOS_DIR, nombre)
            if os.path.abspath(self.foto_path) != os.path.abspath(destino):
                shutil.copy2(self.foto_path, destino)
            return os.path.relpath(destino, APP_DIR)
        except:
            return ""

    def init_usuarios(self):
        Label(self.usuarios_frame, text="GESTIÓN DE USUARIOS",
            font=("Arial", 18, "bold"), fg="white", bg="red").place(x=350, y=10)

        self.modo_usuario = "nuevo"
        # --- Campos ---
        Label(self.usuarios_frame, text="Usuario:", fg="white", bg="red").place(x=20, y=60)
        self.var_user = StringVar()
        self.enforce_uppercase_var(self.var_user)
        self.e_user = Entry(self.usuarios_frame, textvariable=self.var_user,
                            font=("Arial", 11), state="disabled")
        self.e_user.place(x=150, y=60, width=200)

        self.e_user.bind(
            "<FocusOut>",
            lambda e: self.var_user.set(self.var_user.get().strip().upper())
        )
        Label(self.usuarios_frame, text="Contraseña:", fg="white", bg="red").place(x=20, y=100)
        self.var_pwd = StringVar()
        self.e_pwd = Entry(self.usuarios_frame, textvariable=self.var_pwd, show="*",
                        font=("Arial", 11), state="disabled")
        self.e_pwd.place(x=150, y=100, width=200)

        Label(self.usuarios_frame, text="Rol:", fg="white", bg="red").place(x=20, y=140)
        self.var_rol = StringVar()
        self.cb_rol = ttk.Combobox(self.usuarios_frame, textvariable=self.var_rol,
                                values=["BOMBERO", "SUPERVISOR", "ADMIN"],
                                font=("Arial", 11), state="disabled")
        self.cb_rol.place(x=150, y=140, width=200)
        self.cb_rol.bind("<<ComboboxSelected>>", self._on_change_rol_usuario)

        # --- Legajo ---
        Label(self.usuarios_frame, text="Legajo:", fg="white", bg="red").place(x=20, y=180)
        self.var_legajo = StringVar()
        self.e_legajo_user = Entry(self.usuarios_frame, textvariable=self.var_legajo,
                                font=("Arial", 11), state="disabled")
        self.e_legajo_user.place(x=150, y=180, width=200)

        # --- Botones ---
        botones = [
            "Nuevo", "Guardar", "Modificar",
            "Dar de baja", "Resetear Clave", "Generar Acceso",
            "Envío Masivo", "Crear desde Legajos", "Limpiar"
        ]

        comandos = [
            self.nuevo_usuario,
            self.guardar_usuario,
            self.habilitar_modificar_usuario,
            self.eliminar_usuario,
            self.resetear_clave_usuario,
            self.generar_acceso_usuario,
            self.envio_masivo_accesos,
            self.crear_usuarios_desde_legajos,
            self.limpiar_usuario
        ]

        self.user_btns = {}

        base_x = 400
        base_y = 60
        botones_por_fila = 4
        espacio_x = 150
        espacio_y = 45

        tooltips = {
            "Nuevo": "Prepara el formulario para cargar un nuevo usuario.",
            "Guardar": "Guarda el usuario nuevo o los cambios realizados.",
            "Modificar": "Habilita la edición del usuario seleccionado.",
            "Dar de baja": "Da de baja o reactiva el usuario seleccionado.",
            "Resetear Clave": "Genera una nueva contraseña provisoria para el usuario.",
            "Generar Acceso": "Genera contraseña y envía el acceso al email del usuario cargado con doble click.",
            "Envío Masivo": "Genera y envía accesos a todos los usuarios seleccionados (marcados de azul).",
            "Crear desde Legajos": "Crea automáticamente usuarios a partir de los legajos activos.",
            "Limpiar": "Limpia los campos del formulario."
        }


        for i, (txt, cmd) in enumerate(zip(botones, comandos)):

            fila = i // botones_por_fila
            col = i % botones_por_fila

            x = base_x + (col * espacio_x)
            y = base_y + (fila * espacio_y)

            b = Button(self.usuarios_frame, text=txt, command=cmd, width=16, takefocus=True)
            b.place(x=x, y=y)

            b.bind("<Return>", lambda e, btn=b: btn.invoke())

            self.user_btns[txt] = b

            # 🔹 TOOLTIP AQUÍ
            ToolTip(b, tooltips.get(txt, ""))

        # --- Estilo Treeview ---
        style = ttk.Style()

        style.configure("Treeview",
            rowheight=24,
            font=("Arial", 10)
        )

        style.map(
            "Treeview",
            background=[("selected", "#4a90e2")],  # azul visible
            foreground=[("selected", "white")]
        )        
        style.configure("Treeview.Heading", font=("Arial", 10, "bold"))
        style.configure("Treeview", rowheight=24)

        # --- Tabla ---
        cols = ("ID", "Usuario", "Rol", "Legajo", "Estado")
        self.tree_users = ttk.Treeview(
            self.usuarios_frame,
            columns=cols,
            show="headings",
            height=12,
            selectmode="extended"   # ← permite múltiples usuarios
        )
        # Colores de estado
        self.tree_users.tag_configure("primer_acceso", background="#fff3cd")
        self.tree_users.tag_configure("inactivo", background="#f8d7da")

        self.tree_users.heading("ID", text="ID")
        self.tree_users.heading("Usuario", text="Usuario")
        self.tree_users.heading("Rol", text="Rol")
        self.tree_users.heading("Legajo", text="Legajo")
        self.tree_users.heading("Estado", text="Estado")

        self.tree_users.column("ID", width=50, anchor="center", stretch=False)
        self.tree_users.column("Usuario", width=220, anchor="w")
        self.tree_users.column("Rol", width=120, anchor="center")
        self.tree_users.column("Legajo", width=90, anchor="center")
        self.tree_users.column("Estado", width=80, anchor="center")

        self.tree_users.bind("<<TreeviewSelect>>", self._on_select_usuario)

        # --- Scrollbar vertical ---
        scroll_y = ttk.Scrollbar(self.usuarios_frame, orient="vertical", command=self.tree_users.yview)
        self.tree_users.configure(yscrollcommand=scroll_y.set)

        # Posición
        self.tree_users.place(x=20, y=235, width=720, height=160)
        scroll_y.place(x=745, y=235, height=160)
        
        self.tree_users.bind("<<TreeviewSelect>>", self._actualizar_seleccion_usuarios, add="+")
        self.tree_users.bind("<Double-1>", self.on_user_double_click)

        self.lbl_sel_usuarios = Label(
            self.usuarios_frame,
            text="Seleccionados: 0",
            fg="white",
            bg="red",
            font=("Arial", 10, "bold")
        )

        # 🔹 Navegación con Enter lineal

        self.e_user.bind("<Return>", lambda e: self.e_pwd.focus_set())
        self.e_pwd.bind("<Return>", lambda e: self.cb_rol.focus_set())
        self.cb_rol.bind("<Return>", lambda e: (self.e_legajo_user.focus_set(), "break"))

        # 🔹 Último campo → ir a botón Guardar
        self.e_legajo_user.bind(
            "<Return>",
            lambda e: self.user_btns["Guardar"].focus_set()
        )

        # 🔹 Enter en Guardar → ejecutar guardar
        self.user_btns["Guardar"].bind(
            "<Return>",
            lambda e: self.guardar_usuario()
        )
        self.cargar_grilla_usuarios()
        self._estado_campos_usuario("inicial")
        self._estado_botones_usuario("inicial")
        print("ALTO FRAME:", self.usuarios_frame.winfo_height())

    def _on_select_usuario(self, event):

        seleccionado = self.tree_users.focus()
        if not seleccionado:
            return

        valores = self.tree_users.item(seleccionado, "values")
        estado = valores[4].strip().upper()

        self._estado_campos_usuario("cargado")
        self._estado_botones_usuario("cargado")

        boton_baja = self.user_btns.get("Dar de baja")

        if boton_baja:
            if estado in ("ACTIVO", "PENDIENTE"):
                boton_baja.config(
                    text="Dar de baja",
                    bg="#ffcccc"
                )
            else:
                boton_baja.config(
                    text="Reactivar",
                    bg="#ccffcc"
                )

        boton_acceso = self.user_btns.get("Generar Acceso")

        if boton_acceso:
            if estado == "ACTIVO":
                boton_acceso.config(state="normal")
            else:
                boton_acceso.config(state="disabled")

    def _salto_desde_rol(self):
        self.e_legajo_user.focus_set()

        if rol == "BOMBERO":
            self.e_legajo_user.focus_set()
        else:
            self.user_btns["Guardar"].focus_set()

    def resetear_clave_usuario(self):
        seleccionado = self.tree_users.focus()
        if not seleccionado:
            messagebox.showwarning("Aviso", "Seleccione un usuario.")
            return

        datos = self.tree_users.item(seleccionado)
        user_id = datos["values"][0]
        username = datos["values"][1]

        confirmar = messagebox.askyesno(
            "Confirmar",
            f"¿Resetear contraseña de {username}?"
        )

        if not confirmar:
            return

        # 🔐 Generar contraseña segura
        nueva_clave = self._generar_password_seguro()

        import sqlite3
        conn = sqlite3.connect(DB_PATH)

        try:
            # 🔐 Guardar en base (CORRECTO)
            self.actualizar_password_en_bd(
                conn,
                user_id,
                nueva_clave,
                forzar_cambio=True
            )

            conn.commit()

        finally:
            conn.close()

        # 🔥 Mostrar ventana (esto lo dejamos intacto ✔)
        self.mostrar_password_generada(nueva_clave)

        # 🔄 Limpiar selección y reset UI
        for item in self.tree_users.selection():
            self.tree_users.selection_remove(item)

        self.ui.show_info(
            "Contraseña reseteada",
            f"Nueva contraseña generada para {username}"
        )

        self._estado_botones_usuario("inicial")

    def mostrar_password_generada(self, password):

        win = Toplevel(self.master)
        win.title("Nueva contraseña generada")
        win.geometry("400x200")
        win.resizable(False, False)
        win.grab_set()

        Label(win, text="Nueva contraseña generada:",
            font=("Arial", 11, "bold")).pack(pady=10)

        entry = Entry(win, width=35, justify="center")
        entry.pack(pady=10)
        entry.insert(0, password)

        # Permitir copiar
        entry.select_range(0, 'end')
        entry.focus_set()

        def copiar():
            win.clipboard_clear()
            win.clipboard_append(password)

        Button(win, text="Copiar", command=copiar, width=15).pack(pady=5)
        Button(win, text="Cerrar", command=win.destroy, width=15).pack(pady=5)

    def _generar_password_seguro(self, largo=8):
        import random

        mayus = "ABCDEFGHJKLMNPQRSTUVWXYZ"
        minus = "abcdefghijkmnopqrstuvwxyz"
        nums = "23456789"

        # Asegurar al menos uno de cada tipo
        password = [
            random.choice(mayus),
            random.choice(minus),
            random.choice(nums),
        ]

        resto = mayus + minus + nums

        while len(password) < largo:
            password.append(random.choice(resto))

        random.shuffle(password)

        return "".join(password)

    def actualizar_password_en_bd(self, conn, user_id, password, forzar_cambio=True):

        c = conn.cursor()

        nuevo_hash = hash_password(password)

        c.execute("""
            UPDATE usuarios
            SET password_hash=?,
                debe_cambiar_password=?
            WHERE id=?
        """, (nuevo_hash, int(forzar_cambio), user_id))

    def guardar_nueva_password(self, user_id, nueva_password):

        nuevo_hash = hash_password(nueva_password)

        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()

        # 🔴 PASO 5 VA EXACTAMENTE AQUÍ
        c.execute("""
            UPDATE usuarios
            SET password_hash=?,
                debe_cambiar_password=0
            WHERE id=?
        """, (nuevo_hash, user_id))

        conn.commit()
        conn.close()

        self.ui.show_info("Éxito", "Contraseña actualizada correctamente")

        self.top.destroy()   # cerrar ventana cambio

    def generar_acceso_usuario(self):

        import sqlite3
        from datetime import datetime

        seleccionado = self.tree_users.focus()

        if not seleccionado:
            messagebox.showwarning("Aviso", "Seleccione un usuario.")
            return

        datos = self.tree_users.item(seleccionado)

        user_id = datos["values"][0]

        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()

        c.execute("""
            SELECT l.email
            FROM usuarios u
            LEFT JOIN legajos l ON u.legajo = l.legajo
            WHERE u.id = ?
        """, (user_id,))

        row = c.fetchone()

        if not row:
            print("❌ No se encontró email")
            return

        print("EMAIL:", row)

        row = c.fetchone()
        conn.close()

        if not row or not row[0]:
            messagebox.showwarning(
                "Sin email",
                "El usuario no tiene email cargado en el legajo."
            )
            return

        destino = row[0]

        username = datos["values"][1]
        email = destino

        confirmar = messagebox.askyesno(
            "Confirmar",
            f"¿Generar nueva contraseña y enviar acceso a {username}?"
        )

        if not confirmar:
            return

        # 🔐 Generar contraseña segura
        nueva_clave = self._generar_password_seguro()

        # 🔐 Guardar contraseña en BD (forzar cambio)
        self.actualizar_password_en_bd(
            conn,
            user_id,
            nueva_clave,
            forzar_cambio=True
        )

        if "@" not in email:
            self.ui.show_error("Email inválido", f"El email '{email}' no es válido.")
            return

        # 📧 Enviar correo
        enviado = False
        try:
            self.enviar_mail_acceso(username, nueva_clave, email)
            enviado = True
        except Exception as e:
            print("Error enviando correo:", e)
            self.ui.show_error("Error", f"No se pudo enviar el correo:\n{e}")

        # 🔔 Registrar notificación del sistema
        try:

            conn = sqlite3.connect(DB_PATH)
            c = conn.cursor()

            c.execute("""
                INSERT INTO notificaciones
                (actividad_id, tipo, destinatario, asunto, fecha_envio, estado)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (
                None,
                "ACCESO_USUARIO",
                email,
                "Generación de acceso al sistema",
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "OK" if enviado else "ERROR"
            ))

            conn.commit()
            conn.close()

        except Exception as e:
            print("Error registrando notificación:", e)

        if enviado:
            self.ui.show_info(
                "Acceso generado",
                f"Se generó acceso para {username} y se envió a {email}"
            )
        # 🔄 Limpiar selección y reset UI
        for item in self.tree_users.selection():
            self.tree_users.selection_remove(item)

        self._estado_botones_usuario("inicial")            

    def habilitar_busqueda_combo(self, combo):

        combo._search_index = -1
        combo._last_key = None

        def on_key(event):
            if combo.cget("state") == "disabled":
                return

            key = event.char.lower()
            if not key.isprintable():
                return

            valores = combo["values"]
            if not valores:
                return

            # 🔥 reiniciar si cambia letra
            if combo._last_key != key:
                combo._search_index = -1

            combo._last_key = key

            for i in range(len(valores)):
                idx = (combo._search_index + 1 + i) % len(valores)

                texto = str(valores[idx]).lower()

                # 🔥 CLAVE: buscar después del " - "
                if " - " in texto:
                    texto_busqueda = texto.split(" - ", 1)[1]
                else:
                    texto_busqueda = texto

                if key in texto_busqueda:
                    combo.current(idx)
                    combo._search_index = idx
                    break

            return "break"

        combo.bind("<KeyPress>", on_key)

    def enviar_mail_acceso(self, usuario, password, destino):

        import smtplib
        from email.message import EmailMessage

        asunto = "SIAB - Acceso al Sistema de Actividades"

        mensaje = f"""
        Sistema Informático de Actividades - Bomberos Almafuerte

        Se ha generado su acceso al sistema.

        Usuario: {usuario}
        Contraseña provisoria: {password}

        IMPORTANTE:
        La contraseña NO contiene los siguientes caracteres:
        - Número 0 (cero)
        - Letra O (o mayúscula)
        - Letra I (i mayúscula)
        - Letra l (ele minúscula)

        Para ingresar utilice la aplicación del sistema SIAB.
        Por seguridad, al ingresar deberá cambiar su contraseña.

        Mucha Suerte!! Saludos
        Administración SIAB
        """

        msg = EmailMessage()

        msg["Subject"] = asunto
        msg["From"] = self.smtp_user
        msg["To"] = destino

        msg.set_content(mensaje)

        with smtplib.SMTP_SSL(self.smtp_server, self.smtp_port) as smtp:
            smtp.login(self.smtp_user, self.smtp_pass)
            smtp.send_message(msg)

    def mostrar_procesando(self, texto="Procesando..."):

        self.win_proceso = tk.Toplevel(self.master)
        self.win_proceso.title("Espere")
        self.win_proceso.geometry("250x100")
        self.win_proceso.resizable(False, False)

        self.win_proceso.transient(self.master)

        Label(
            self.win_proceso,
            text=texto,
            font=("Arial", 10)
        ).pack(pady=20)

        self.win_proceso.update()

    def cerrar_procesando(self):
        if hasattr(self, "win_proceso"):
            self.win_proceso.destroy()

    def envio_masivo_accesos(self):

        seleccionados = self.tree_users.selection()

        if not seleccionados:
            messagebox.showwarning("Aviso", "Seleccione uno o más usuarios.")
            return

        confirmar = messagebox.askyesno(
            "Confirmar envío masivo",
            f"Se generará y enviará acceso a {len(seleccionados)} usuarios.\n\n¿Continuar?"
        )

        if not confirmar:
            return

        # 🔵 Mostrar ventana de proceso
        self.mostrar_procesando("Enviando accesos...")

        enviados = 0
        omitidos = 0
        errores = 0
        detalle = []

        try:

            conn = sqlite3.connect(DB_PATH)
            c = conn.cursor()
            print("Usuarios seleccionados:", len(seleccionados))

            for item in seleccionados:

                datos = self.tree_users.item(item)
                user_id = datos["values"][0]
                username = datos["values"][1]

                print("Procesando:", username)

                try:

                    c.execute("""
                        SELECT 
                            u.debe_cambiar_password,
                            u.activo,
                            l.email
                        FROM usuarios u
                        LEFT JOIN legajos l ON u.legajo = l.legajo
                        WHERE u.id = ?
                    """, (user_id,))

                    row = c.fetchone()

                    if not row:
                        errores += 1
                        detalle.append(f"{username} → ❌ Error (no encontrado)")
                        continue

                    debe_cambiar, activo, email = row

                    if activo == 0:
                        omitidos += 1
                        detalle.append(f"{username} → ⚠ Omitido (inactivo)")
                        continue

                    if debe_cambiar == 0:
                        omitidos += 1
                        detalle.append(f"{username} → ⚠ Omitido (ya activo)")
                        continue

                    if not email or "@" not in email:
                        errores += 1
                        detalle.append(f"{username} → ❌ Error (sin email)")
                        continue

                    nueva_clave = self._generar_password_seguro()

                    self.actualizar_password_en_bd(
                        conn,
                        user_id,
                        nueva_clave,
                        forzar_cambio=True
                    )

                    self.enviar_mail_acceso(username, nueva_clave, email)

                    detalle.append(f"{username} → ✅ Enviado")

                    enviados += 1

                except Exception as e:
                    print("Error envío:", e)
                    errores += 1
                    detalle.append(f"{username} → ❌ Error ({str(e)[:40]})")

            conn.commit()
            conn.close()

        finally:
            # 🔵 Cerrar ventana de proceso SIEMPRE
            self.cerrar_procesando()

        # 🔄 Limpiar selección y reset UI
        for item in self.tree_users.selection():
            self.tree_users.selection_remove(item)

        self._estado_botones_usuario("inicial")

        self.ui.show_info(
            "Resultado del envío",
            f"""Resultado del envío

        Correos enviados: {enviados}
        Usuarios omitidos: {omitidos}
        Errores: {errores}

        DETALLE:
        ------------------------
        {chr(10).join(detalle)}
        """
        )

    def crear_usuarios_desde_legajos(self):

        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        c = conn.cursor()

        seleccionados = self.tree_legajos.selection()
        usar_filtro = bool(seleccionados)

        confirmar = messagebox.askyesno(
            "Confirmar generación de accesos",
            """Se crearán usuarios para TODOS los legajos sin acceso.

        Se generará usuario y contraseña provisoria.

        ¿Desea continuar?"""
        )

        if not confirmar:
            return

        # DEBUG
        c.execute("SELECT COUNT(*) FROM legajos")
        print("TOTAL LEGAJOS:", c.fetchone()[0])

        c.execute("SELECT COUNT(*) FROM usuarios")
        print("TOTAL USUARIOS:", c.fetchone()[0])

        # 🔍 QUERY CORRECTA
        c.execute("""
            SELECT 
                l.legajo,
                l.apellido,
                l.nombre
            FROM legajos l
            LEFT JOIN usuarios u 
                ON CAST(l.legajo AS INTEGER) = u.legajo
            WHERE u.id IS NULL
        """)

        rows = c.fetchall()

        print("FILAS ENCONTRADAS:", len(rows))

        if not rows:
            self.ui.show_info(
                "Usuarios",
                "No hay legajos disponibles para crear usuarios."
            )
            conn.close()
            return

        creados = 0

        for r in rows:

            apellido = (r["apellido"] or "").strip().upper()
            nombre = (r["nombre"] or "").strip().upper()

            if not apellido:
                continue

            inicial = nombre[0] if nombre else ""
            username = f"{apellido}{inicial}"

            # evitar duplicados
            c.execute(
                "SELECT id FROM usuarios WHERE username = ?",
                (username,)
            )

            if c.fetchone():
                username = f"{apellido}{inicial}{r['legajo']}"

            c.execute("""
                INSERT INTO usuarios
                (username, password_hash, rol, legajo, activo, debe_cambiar_password)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (
                username,
                "",
                "BOMBERO",
                int(r["legajo"]) if r["legajo"] else None,
                1,
                1
            ))

            creados += 1

        conn.commit()
        conn.close()

        self.cargar_grilla_usuarios()

        self.ui.show_info(
            "Usuarios creados",
            f"Se crearon {creados} usuarios automáticamente."
        )

    def _estado_botones_usuario(self, estado):

        if not hasattr(self, "user_btns"):
            return

        for b in list(self.user_btns.values()):
            try:
                if b and b.winfo_exists():
                    b.config(state="disabled")
            except:
                pass

        def activar(nombre):
            try:
                btn = self.user_btns.get(nombre)
                if btn and btn.winfo_exists():
                    btn.config(state="normal")
            except:
                pass

        if estado == "inicial":
            activar("Nuevo")
            activar("Envío Masivo")
            activar("Crear desde Legajos")

            btn = self.user_btns.get("Nuevo")
            if btn and btn.winfo_exists():
                btn.focus_set()

        elif estado == "nuevo":
            activar("Guardar")
            activar("Limpiar")

        elif estado == "cargado":
            activar("Modificar")
            activar("Dar de baja")
            activar("Resetear Clave")
            activar("Limpiar")
            activar("Generar Acceso")
            activar("Envío Masivo")
            activar("Crear desde Legajos")

        elif estado == "editando":
            activar("Guardar")
            activar("Limpiar")
        print(">>> ESTADO USUARIOS:", estado)    

    def _on_change_rol_usuario(self, event=None):
        # En versión simple, el rol NO afecta el legajo.
        # Solo limpiamos si querés mantener consistencia.
        pass

    def cargar_grilla_usuarios(self):
        for i in self.tree_users.get_children():
            self.tree_users.delete(i)

        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()

        c.execute("""
            SELECT id, username, rol,
                IFNULL(legajo, '') as legajo,
                CASE WHEN activo = 1 THEN 'ACTIVO'
                        ELSE 'BAJA'
                END as estado,
                debe_cambiar_password
            FROM usuarios
            ORDER BY username
        """)

        for i, row in enumerate(c.fetchall()):

            tag = "par" if i % 2 == 0 else "impar"

            estado = row[4].strip().upper()
            primer_acceso = row[5] == 1

            if estado == "ACTIVO" and primer_acceso:
                estado = "PENDIENTE"

            fila = list(row[:5])
            fila[4] = estado

            tags = [tag]

            if row[4] == "BAJA":
                tags.append("baja")
            elif primer_acceso:
                tags.append("primer_acceso")

            self.tree_users.insert(
                "",
                "end",
                values=fila,
                tags=tuple(tags)
            )
        conn.close()

        self.tree_users.tag_configure("par", background="#f9f9f9")
        self.tree_users.tag_configure("impar", background="#eaeaea")
        self.tree_users.tag_configure(
            "baja",
            foreground="gray",
            background="#f0f0f0"
        )
        self.tree_users.tag_configure(
            "primer_acceso",
            background="#fff3cd",
            foreground="#856404"
        )

    def nuevo_usuario(self, *args):
        self.modo_usuario = "nuevo"
        self.var_user.set("")
        self.var_pwd.set("")
        self.var_rol.set("BOMBERO")
        self._estado_campos_usuario("nuevo")
        self._on_change_rol_usuario()
        self._estado_botones_usuario("nuevo")
        self.e_user.focus_set()

    def _enter_siguiente(self, event):
        event.widget.tk_focusNext().focus()
        return "break"

    def guardar_usuario(self, event=None):

        usuario = self.var_user.get().strip()
        password = self.var_pwd.get().strip()
        rol = self.var_rol.get().strip()
        legajo = self.var_legajo.get().strip()

        # --- Validaciones obligatorias ---
        if not usuario:
            messagebox.showwarning("Faltan datos", "Debe ingresar un usuario.")
            self.e_user.focus_set()
            return

        if not password and self.modo_usuario == "nuevo":
            messagebox.showwarning("Faltan datos", "Debe ingresar una contraseña.")
            self.e_pwd.focus_set()
            return

        if not rol:
            messagebox.showwarning("Faltan datos", "Debe seleccionar un rol.")
            self.cb_rol.focus_set()
            return

        if rol in ["BOMBERO", "SUPERVISOR"] and not legajo:
            messagebox.showwarning("Faltan datos", "Debe ingresar un legajo.")
            self.e_legajo_user.focus_set()
            return
    
        if self.modo_usuario == "modificar":
            self.modificar_usuario()
            return

        # 🔹 INSERT normal (nuevo)
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()

        try:
            username = self.var_user.get().strip().upper()
            password = self.var_pwd.get().strip()
            rol = self.var_rol.get().strip()
            legajo = self.var_legajo.get().strip()

            # 🔐 Generar hash correcto
            password_hash = hash_password(password)

            c.execute("""
                INSERT INTO usuarios (username, password_hash, rol, legajo, activo)
                VALUES (?, ?, ?, ?, 1)
            """, (
                username,
                password_hash,
                rol,
                legajo if legajo else None
            ))
            conn.commit()

        finally:
            conn.close()

        self.limpiar_usuario()
        self.cargar_grilla_usuarios()

    def _estado_campos_usuario(self, estado):

        if estado == "nuevo":
            self.e_user.config(state="normal")
            self.e_pwd.config(state="normal")
            self.cb_rol.config(state="readonly")
            self.e_legajo_user.config(state="normal")

        elif estado == "modificar":
            self.e_user.config(state="disabled")   # normalmente no se modifica el username
            self.e_pwd.config(state="normal")
            self.cb_rol.config(state="readonly")
            self.e_legajo_user.config(state="normal")

        else:  # inicial o después de guardar
            self.e_user.config(state="disabled")
            self.e_pwd.config(state="disabled")
            self.cb_rol.config(state="disabled")
            self.e_legajo_user.config(state="disabled")

    def modificar_usuario(self, event=None):

        if not hasattr(self, "user_editing_id"):
            self.ui.show_error("Error", "No hay usuario en edición")
            return

        user_id = self.user_editing_id
        user = self.var_user.get().strip()
        pwd = self.var_pwd.get().strip()
        rol = self.var_rol.get().strip().upper()
        legajo = self.var_legajo.get().strip()

        # 🔹 Validaciones
        if not user or not rol:
            self.ui.show_error("Error", "Complete usuario y rol")
            return

        if not legajo:
            self.ui.show_error("Error", "El legajo es obligatorio")
            return

        try:
            conn = sqlite3.connect(DB_PATH)
            c = conn.cursor()

            if pwd:
                pwd_hash = hash_password(pwd)
                c.execute("""
                    UPDATE usuarios
                    SET username=?, password_hash=?, rol=?, legajo=?
                    WHERE id=?
                """, (user, pwd_hash, rol, legajo, user_id))
            else:
                c.execute("""
                    UPDATE usuarios
                    SET username=?, rol=?, legajo=?
                    WHERE id=?
                """, (user, rol, legajo, user_id))

            conn.commit()
            conn.close()

            self.ui.show_info("OK", "Usuario modificado")

            self.cargar_grilla_usuarios()
            self.limpiar_usuario()
            del self.user_editing_id

        except sqlite3.IntegrityError:
            self.ui.show_error("Error", "Ya existe un usuario con ese nombre.")
        except Exception as e:
            self.ui.show_error("Error", str(e))

    def _habilitar_modificar_legajo(self, event=None):
        self.modo_legajo = "modificar"

        try:
            self._recreate_date_entries()
        except Exception:
            pass

        self._actualizar_estado_campos()

        for w in [
            self.e_apellido, self.e_nombre, self.e_cargo, self.e_dni,
            self.e_email, self.e_nro_cel
        ]:
            try:
                w.config(state="normal")
            except Exception:
                pass

        self.cb_grado.config(state="readonly")
        self.cb_situacion.config(state="readonly")
        self.chk_autoriza.config(state="normal")

        self._actualizar_estado_fecha_baja()

        self.btn_foto.config(state="normal")
        self.e_legajo.config(state="disabled")

        self._estado_botones_legajo("modificar")

        self.e_apellido.focus_set()

    def limpiar_usuario(self):
        self.modo_usuario = "nuevo"

        self.var_user.set("")
        self.var_pwd.set("")
        self.var_rol.set("")
        self.var_legajo.set("")  # 🔥 ESTA LÍNEA FALTABA

        self._estado_campos_usuario("inicial")
        self._estado_botones_usuario("inicial")

        self.e_user.focus_set()

    def eliminar_usuario(self, event=None):

        sel = self.tree_users.selection()
        if not sel:
            self.ui.show_error("Error", "Seleccione un usuario")
            return

        item = self.tree_users.item(sel[0], "values")
        user_id = item[0]
        username = item[1]

        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()

        # 🔎 Consultar estado real en la BD
        c.execute("SELECT activo FROM usuarios WHERE id = ?", (user_id,))
        row = c.fetchone()

        if not row:
            conn.close()
            self.ui.show_error("Error", "Usuario no encontrado")
            return

        estado = row[0]

        if estado == 1:

            if not messagebox.askyesno("Confirmar", f"¿Dar de baja al usuario {username}?"):
                conn.close()
                return

            c.execute("UPDATE usuarios SET activo = 0 WHERE id = ?", (user_id,))
            self.ui.show_info("OK", "Usuario dado de baja")

        else:

            if not messagebox.askyesno("Confirmar", f"¿Reactivar al usuario {username}?"):
                conn.close()
                return

            c.execute("UPDATE usuarios SET activo = 1 WHERE id = ?", (user_id,))
            self.ui.show_info("OK", "Usuario reactivado")

        conn.commit()
        conn.close()

        self.cargar_grilla_usuarios()
        self.limpiar_usuario()

    def on_user_double_click(self, event):

        item = self.tree_users.identify_row(event.y)
        if not item:
            return

        self._on_select_usuario(event)

        row = self.tree_users.item(item, "values")

        self.var_user.set(row[1])
        self.var_pwd.set("")
        self.var_rol.set(row[2])

        self._estado_botones_usuario("cargado")

    # ===================== pestaña LEGAJO =====================
    def only_time_validate(self, value):
        if value == "":
            return True  # permitir borrar
        if not all(c.isdigit() or c == ":" for c in value):
            return False  # bloquear letras y símbolos
        if len(value) > 5:
            return False
        if ":" in value:
            partes = value.split(":")
            if len(partes) > 2:
                return False
            h = partes[0]
            m = partes[1] if len(partes) > 1 else ""
            if h and (not h.isdigit() or int(h) > 23):
                return False
            if m and (not m.isdigit() or int(m) > 59):
                return False
        else:
            if not value.isdigit() or int(value) > 23:
                return False
        return True

    def _buscar_grado_por_letra(self, event):
            letra = event.char.upper()
            if not letra.isalpha():
                return
            for valor in self.cb_grado["values"]:
                if valor.upper().startswith(letra):
                    self.cb_grado.set(valor)
                    break        

    def init_legajo(self):
        Label(self.legajo_frame, text="LEGAJOS", font=("Arial", 18, "bold"),
            fg="white", bg="red").place(x=400, y=6)   # x ajustado a la derecha

        # N° Legajo
        Label(self.legajo_frame, text="N° Legajo:", fg="white", bg="red").place(x=20, y=50)
        vcmd_legajo = (self.legajo_frame.register(lambda P: P.isdigit() or P == ""), "%P")
        self.e_legajo = Entry(self.legajo_frame, textvariable=self.var_id, font=("Arial", 11),
                            validate="key", validatecommand=vcmd_legajo)
        self.e_legajo.place(x=150, y=50, width=160)
        self.e_legajo.bind("<Return>", lambda e: self.buscar_legajo_enter())

        # Apellido
        Label(self.legajo_frame, text="Apellido:", fg="white", bg="red").place(x=20, y=80)
        def validar_texto(P):
            return all(c.isalpha() or c.isspace() for c in P)

        vcmd_texto = (self.legajo_frame.register(validar_texto), "%P")
        self.e_apellido = Entry(self.legajo_frame, textvariable=self.var_apellido,
                        font=("Arial", 11), state="disabled",
                        validate="key", validatecommand=vcmd_texto)
        self.e_apellido.place(x=150, y=80, width=200)
        self.e_apellido.bind("<Return>", lambda e: self.e_nombre.focus_set())

        # Nombre
        def validar_texto(P):
            return all(c.isalpha() or c.isspace() for c in P)

        vcmd_texto = (self.legajo_frame.register(validar_texto), "%P")
        Label(self.legajo_frame, text="Nombre:", fg="white", bg="red").place(x=20, y=110)
        self.e_nombre = Entry(self.legajo_frame, textvariable=self.var_nombre,
                      font=("Arial", 11), state="disabled",
                      validate="key", validatecommand=vcmd_texto)
        self.e_nombre.place(x=150, y=110, width=200)
        self.e_nombre.bind("<Return>", lambda e: self.e_dni.focus_set())

        # DNI
        Label(self.legajo_frame, text="DNI:", fg="white", bg="red").place(x=20, y=140)
        vcmd_dni = (self.legajo_frame.register(lambda P: P.isdigit() and len(P) <= 8 or P == ""), "%P")
        self.e_dni = Entry(self.legajo_frame, textvariable=self.var_dni, font=("Arial", 11),
                        state="disabled",
                        validate="key", validatecommand=vcmd_dni)
        self.e_dni.place(x=150, y=140, width=200)
        self.e_dni.bind("<Return>", lambda e: self.cb_grado.focus_set())

        # Grado
        Label(self.legajo_frame, text="Grado:", fg="white", bg="red").place(x=20, y=170)
        self.cb_grado = ttk.Combobox(self.legajo_frame, textvariable=self.var_grado,
                                    font=("Arial", 11), state="disabled", values=self.GRADOS)
        self.cb_grado.place(x=150, y=170, width=200)
        self.cb_grado.bind("<Return>", lambda e: self.e_cargo.focus_set())
        self.cb_grado.bind("<Key>", self._buscar_grado_por_letra)

        # Cargo
        Label(self.legajo_frame, text="Cargo/Función:", fg="white", bg="red").place(x=20, y=200)
        self.e_cargo = Entry(self.legajo_frame, textvariable=self.var_cargo, font=("Arial", 11),
                            state="disabled")
        self.e_cargo.place(x=150, y=200, width=200)
        self.e_cargo.bind("<Return>", lambda e: self.e_email.focus_set())

        # Email
        Label(self.legajo_frame, text="Email:", fg="white", bg="red").place(x=20, y=230)
        self.var_email = StringVar()
        self.e_email = Entry(self.legajo_frame, textvariable=self.var_email, font=("Arial", 11), state="disabled")
        self.e_email.place(x=150, y=230, width=200)

        self.e_email.bind("<Return>", lambda e: self.e_nro_cel.focus_set())

        # Cel
        Label(self.legajo_frame, text="Cel:", fg="white", bg="red").place(x=20, y=260)
        vcmd_cel = (self.legajo_frame.register(lambda P: P.isdigit() or P == ""), "%P")
        self.e_nro_cel = Entry(
            self.legajo_frame,
            textvariable=self.var_nro_cel,
            font=("Arial", 11),
            state="disabled",
            validate="key",
            validatecommand=vcmd_cel
        )
        self.e_nro_cel.place(x=150, y=260, width=200)
        self.e_nro_cel.bind("<Return>", lambda e: self.cb_situacion.focus_set())

        # Situación
        Label(self.legajo_frame, text="Situación:", fg="white", bg="red").place(x=20, y=290)
        self.cb_situacion = ttk.Combobox(self.legajo_frame, state="disabled", font=("Arial", 11),
                                        values=[
                                            "ACTIVO", "LICENCIA", "BAJA", "CUERPO AUXILIAR",
                                            "RETIRO EFECTIVO", "MIEMBRO COMISIÓN DIRECTIVA",
                                            "MIEMBRO COMISIÓN DE EVENTOS", "CON SANCIÓN",
                                            "PERMISO ESPECIAL", "OTRO"
                                        ])
        # después de crear y ubicar self.cb_situacion:
        self.cb_situacion.place(x=150, y=290, width=200)
        # después de self.cb_situacion.place(...)
        self.cb_situacion.bind("<<ComboboxSelected>>", self._on_situacion_change)
        self.cb_situacion.bind("<Return>", self._on_situacion_enter)
        self.cb_situacion.bind("<KeyRelease-Return>", self._on_situacion_enter)
        # usar KeyRelease para que event.char esté disponible y la búsqueda por letra sea más fiable
        self.cb_situacion.bind("<KeyRelease>", self._buscar_opcion_combo)
        self.cb_situacion.bind("<FocusIn>", self._on_situacion_focus)

        self.cb_situacion.bind("<<ComboboxSelected>>", 
                               lambda e: print("DEBUG >> ComboboxSelected valor:", 
                                               self.cb_situacion.get()))
        # Autoriza
        self.var_autoriza = StringVar(value="NO")
        self.chk_autoriza = Checkbutton(self.legajo_frame, text="Autoriza",
                                        variable=self.var_autoriza,
                                        onvalue="SI", offvalue="NO",
                                        bg="red", fg="white", selectcolor="red")
        self.chk_autoriza.place(x=20, y=320)

        # Fecha Baja
        Label(self.legajo_frame, text="Baja/Lic/Sanción:", fg="white", bg="red").place(x=20, y=350)
        self.e_fecha_baja = DateEntry(self.legajo_frame, date_pattern='dd/mm/yyyy',
                                    state="disabled")
                # Guardar geometría para re-colocar fácilmente
        self._fecha_baja_place = {"x":150, "y":350, "width":200}
        # ocultar al inicio (se mostrará sólo cuando corresponda)
        try:
            self.e_fecha_baja.place_forget()
        except Exception:
            pass

        # Foto
        self.lbl_foto = Label(self.legajo_frame, text="", bg="white")
        self.lbl_foto.place(x=400, y=50, width=120, height=120)
        self.btn_foto = Button(self.legajo_frame, text="Cargar Foto 4x4",
                            command=self.cargar_foto, state="disabled")
        self.btn_foto.place(x=400, y=180, width=120)

        # Botones Legajo en dos columnas
        botones = [
            "Guardar", "Modificar",
            "Eliminar", "Limpiar/Cancelar",
            "Exportar PDF", "Ver en Pantalla",
            "Imprimir Directo", "Listado Excel"
        ]

        comandos = [
            self.guardar_legajo, self._habilitar_modificar_legajo,
            self.eliminar_legajo, self.limpiar_legajo,
            self.imprimir_legajo, self.ver_legajo_en_pantalla,
            self.imprimir_legajo_directo, self.exportar_legajos_excel
        ]

        self.leg_btns = {}
        for i, (txt, cmd) in enumerate(zip(botones, comandos)):
            col = i % 2              # 0 = columna izquierda, 1 = derecha
            row = i // 2             # avanza cada 2 botones
            x = 600 + col * 150      # 600 para la primera, 750 para la segunda
            y = 50 + row * 40   #DISTANCIA DE BOTONES
            b = Button(self.legajo_frame, text=txt, command=cmd, width=16)
            b.place(x=x, y=y)
            self.leg_btns[txt] = b
            

        # Grilla de legajos
        cols = ("Legajo", "Apellido", "Nombre", "Grado", "Cargo/Función", "Situación")
        self.tree_legajos = ttk.Treeview(
            self.legajo_frame,
            columns=cols,
            show="headings",
            height=6
        )

        self.tree_legajos.bind(
            "<<TreeviewSelect>>",
            self._on_legajo_selected
        )

        # Quitar binds previos y volver a asignar
        try:
            self.tree_legajos.unbind("<Double-1>")
        except Exception:
            pass
        self.tree_legajos.bind("<Double-1>", self.on_legajo_double_click)

        for c in cols:
            self.tree_legajos.heading(c, text=c)
            self.tree_legajos.column(c, width=100, stretch=True)

        # --- Scrollbar vertical (igual a Usuarios) ---
        scroll_y_legajos = ttk.Scrollbar(
            self.legajo_frame,
            orient="vertical",
            command=self.tree_legajos.yview
        )
        self.tree_legajos.configure(yscrollcommand=scroll_y_legajos.set)

        self.tree_legajos.bind(
            "<<TreeviewSelect>>",
            self._on_legajo_selected
        )

        # Posición
        self.tree_legajos.place(x=370, y=230, width=630, height=150)
        scroll_y_legajos.place(x=1005, y=230, height=150)

        # --- PANEL DE CALIFICACIONES (NUEVO) ---
        # Lo ubicamos a la derecha de la foto para aprovechar el espacio
        self.frame_notas = LabelFrame(self.legajo_frame, text=" Puntajes de Capacitación (Playón) ", 
                                     bg="red", fg="yellow", font=("Arial", 10, "bold"))
        self.frame_notas.place(x=540, y=50, width=480, height=170)

        cols_notas = ("fecha", "tipo", "tema", "nota")
        self.tree_notas = ttk.Treeview(self.frame_notas, columns=cols_notas, show="headings", height=5)
        
        self.tree_notas.heading("fecha", text="Fecha")
        self.tree_notas.heading("tipo", text="Tipo")
        self.tree_notas.heading("tema", text="Tema/Posta")
        self.tree_notas.heading("nota", text="Nota")

        self.tree_notas.column("fecha", width=85, anchor="center")
        self.tree_notas.column("tipo", width=100, anchor="w")
        self.tree_notas.column("tema", width=220, anchor="w")
        self.tree_notas.column("nota", width=50, anchor="center")

        # Scrollbar para las notas
        scroll_notas = ttk.Scrollbar(self.frame_notas, orient="vertical", command=self.tree_notas.yview)
        self.tree_notas.configure(yscrollcommand=scroll_notas.set)
        
        self.tree_notas.pack(side=LEFT, padx=5, pady=5, fill=BOTH, expand=True)
        scroll_notas.pack(side=RIGHT, fill=Y)

        # Reubicación de botones para que no se superpongan (ajuste opcional de x/y)
        # Si ves que los botones tapan el panel, ajusta 'y' en el bucle de botones.

       # Estado inicial
        self._estado_botones_legajo("inicial")
        self.cargar_grilla_legajos()

    def _on_situacion_focus(self, event=None):
        print("DEBUG >> Combo Situación recibió foco. Valor actual:", self.cb_situacion.get())

    def _actualizar_seleccion_usuarios(self, event=None):

        seleccion = self.tree_users.selection()
        cantidad = len(seleccion)

        self.lbl_sel_usuarios.config(
            text=f"Seleccionados: {cantidad}"
        )

    def _actualizar_estado_campos(self):
        """Habilita o deshabilita los campos según el modo_legajo"""
        if self.modo_legajo in ("nuevo", "modificar"):
            # habilitar
            self.chk_autoriza.config(state="normal")
            # si querés, agregá aquí más campos con .config(state="normal")
        else:
            # deshabilitar
            self.chk_autoriza.config(state="disabled")
            # lo mismo: acá ponés los otros en disabled si corresponde

    def ver_legajo_en_pantalla(self):
        """Muestra el legajo en pantalla usando un PDF temporal (vista previa)."""
#        from tkinter.messagebox import showerror
        import tempfile, os, time, threading

        try:
            pdf_bytes = self._generar_pdf_legajo_en_memoria()

            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as f:
                f.write(pdf_bytes)
                temp_path = f.name

            os.startfile(temp_path)  # Abre el visor PDF

            # Borrado automático después de unos segundos
            def borrar_luego():
                time.sleep(20)
                try:
                    os.remove(temp_path)
                except:
                    pass

            threading.Thread(target=borrar_luego, daemon=True).start()

        except Exception as e:
            self.ui.show_error("Error", f"No se pudo mostrar el legajo:\n{e}")

    def _on_legajo_selected(self, event=None):
        # 1. Verificar si hay selección
        seleccion = self.tree_legajos.selection()
        if not seleccion:
            self._estado_botones_legajo("inicial")
            return

        # 2. Cambiar estado de botones
        self._estado_botones_legajo("seleccionado")

        # 3. Obtener el legajo de la fila seleccionada (asumiendo que es la primera columna)
        item_id = seleccion[0]
        valores = self.tree_legajos.item(item_id, "values")
        if not valores:
            return
            
        legajo_nro = valores[0]

        # 4. Cargar los datos en los campos de texto (Entry)
        # Esto dispara tu lógica de búsqueda para completar Apellido, Nombre, etc.
        self.var_id.set(legajo_nro)
        self.buscar_legajo_enter() 

        # 5. ¡NUEVA LÓGICA DE PUNTAJES!
        # Cada vez que tocas un bombero, refrescamos su tabla de notas de MySQL
        if hasattr(self, "actualizar_tabla_notas"):
            self.actualizar_tabla_notas(legajo_nro)

    def _estado_campos_legajo(self, estado):
        rol = self.usuario_actual.get("rol", "").upper()

        campos = [
            self.e_apellido, self.e_nombre, self.e_cargo, self.e_dni,
            self.e_email, self.e_nro_cel,
            self.cb_grado, self.cb_situacion,
            self.chk_autoriza,
            self.btn_foto,
            self.e_fecha_baja
        ]

        # 🚫 BOMBERO: SIEMPRE LECTURA
        if rol == "BOMBERO":
            for w in campos:
                try:
                    w.config(state="disabled")
                except Exception:
                    pass
            return

        # -------- ROLES CON PERMISOS (SUPERVISOR / ADMIN) --------
        if estado == "lectura":
            for w in campos:
                try:
                    w.config(state="disabled")
                except Exception:
                    pass

        elif estado == "modificar":
            for w in campos:
                try:
                    w.config(state="normal")
                except Exception:
                    pass

    def _estado_botones_legajo(self, estado):
        rol = self.usuario_actual.get("rol", "").upper()
        print(f">>> ESTADO BOTONES LEGAJO = {estado} | ROL = {rol}")

        # Apagar todo siempre
        for b in self.leg_btns.values():
            b.config(state="disabled")

        # ---------- ESTADO INICIAL ----------
        if estado == "inicial":
            return

        # ---------- MODO NUEVO ----------
        if estado == "nuevo":
            for k in ("Guardar", "Limpiar/Cancelar"):
                self.leg_btns[k].config(state="normal")
            return

        # ---------- SOLO FILA SELECCIONADA ----------
        if estado == "seleccionado":
            self.leg_btns["Limpiar/Cancelar"].config(state="normal")
            return

        # ---------- LEGAJO CARGADO / LECTURA ----------
        if estado in ("cargado", "lectura"):

            for k in (
                "Modificar",
                "Exportar PDF",
                "Ver en Pantalla",
                "Imprimir Directo",
                "Limpiar/Cancelar",
            ):
                self.leg_btns[k].config(state="normal")

            if rol in ("SUPERVISOR", "ADMIN"):
                self.leg_btns["Listado Excel"].config(state="normal")

            if rol == "ADMIN":
                self.leg_btns["Eliminar"].config(state="normal")

            return

        # ---------- MODO MODIFICAR ----------
        if estado == "modificar":
            for k in ("Guardar", "Limpiar/Cancelar"):
                self.leg_btns[k].config(state="normal")
            return

    def _on_situacion_change(self, event=None):
        rol = self.usuario_actual.get("rol", "").upper()
        if rol == "BASICO":
            return "break"

        """
        Handler único para cambios en Situación (mouse o teclado).
        Habilita/deshabilita Autoriza y delega la lógica de Fecha Baja.
        """
        self._actualizar_estado_fecha_baja()
        situacion = self.cb_situacion.get().upper().strip()

        # Situaciones que habilitan autoriza SIEMPRE (nuevo o modificar)
        habilita_autoriza = situacion in (
            "ACTIVO", "CUERPO AUXILIAR", "RETIRO EFECTIVO",
            "MIEMBRO COMISIÓN DIRECTIVA", "MIEMBRO COMISIÓN DE EVENTOS",
            "PERMISO ESPECIAL", "OTRO"
        )

        if habilita_autoriza:
            self.chk_autoriza.config(state="normal")
        else:
            self.var_autoriza.set("NO")
            self.chk_autoriza.config(state="disabled")

        # Delegar todo lo relativo a Fecha Baja (visibilidad/estado/foco)
        self._actualizar_estado_fecha_baja()

    def buscar_legajo(self):
        nro = self.var_id.get().strip()
        if not nro:
            return

        # Si viene en formato "123 - APELLIDO NOMBRE", tomar la parte antes del " - "
        if " - " in nro:
            nro = nro.split(" - ")[0].strip()

        try:
            conn = sqlite3.connect(DB_PATH)
            conn.row_factory = sqlite3.Row 
            c = conn.cursor()
            c.execute("""SELECT 
                legajo, apellido, nombre, dni, grado, cargo, email, nro_cel, foto, situacion, autoriza, fecha_baja
                FROM legajos WHERE legajo=?""", (nro,))
            row = c.fetchone()
            ...
            if row:  # existe el legajo
                self.var_id.set(row[0])
                self.var_apellido.set(row[1])
                self.var_nombre.set(row[2])
                self.var_dni.set(row[3])
                self.var_grado.set(row[4])
                self.var_cargo.set(row[5])
                self.var_email.set(row[6] or "")
                self.var_nro_cel.set(row[7] or "")

                # modo actual → vista/cargado
                self.modo_legajo = "cargado"

                # Foto
                if row[8]:
                    self._mostrar_foto(row[8])
                else:
                    self.lbl_foto.config(image="", text="")

                # Situación
                situacion = (row[9] or "").upper().strip()
                self.cb_situacion.set(situacion)

                # Autoriza: cargar desde la DB (SI/NO) pero NO sobreescribirlo después
                valor_aut = (row[10] or "NO").upper()
                self.var_autoriza.set("SI" if valor_aut == "SI" else "NO")

                # Fecha de baja: sólo setear valor; visibilidad/estado lo maneja _actualizar_estado_fecha_baja()
                try:
                    self.e_fecha_baja.config(state="normal")
                    self.e_fecha_baja.delete(0, "end")
                except Exception:
                    pass

                if row[11]:
                    fecha_str = row[11]
                    fecha_dt = None
                    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
                        try:
                            fecha_dt = datetime.strptime(fecha_str, fmt).date()
                            break
                        except ValueError:
                            continue
                    if not fecha_dt:
                        fecha_dt = date.today()
                    try:
                        self.e_fecha_baja.set_date(fecha_dt)
                    except Exception:
                        pass
                else:
                    try:
                        self.e_fecha_baja.set_date(date(1900, 1, 1))
                    except Exception:
                        pass

                # Aplicar visibilidad/estado (ya no sobreescribirá autoriza)
                self._actualizar_estado_fecha_baja()

                # Botones
                self._estado_botones_legajo("cargado")
            else:
                # 🔒 Si NO existe el legajo → controlar permisos
                if not self._has_role(("ADMIN", "SUPERVISOR")):
                    messagebox.showwarning(
                        "Sin permiso",
                        "No tiene permiso para crear legajos.\n"
                        "Consulte con un supervisor o administrador."
                    )
                    self.limpiar_legajo()
                    return

                # ADMIN / SUPERVISOR → preguntar si crear
                if messagebox.askyesno("Confirmar", f"El legajo {nro} no existe. ¿Desea crearlo?"):
                    self.modo_legajo = "nuevo"
                    self._actualizar_estado_campos()
                    self.limpiar_legajo()
                    self.var_id.set(nro)
                    self.cb_situacion.set("ACTIVO")
                    self._estado_botones_legajo("nuevo")

                    # Activar campos iniciales
                    for w in [self.e_apellido, self.e_nombre, self.e_cargo, self.e_dni, self.e_email, self.e_nro_cel]:
                        w.config(state="normal")

                    self.cb_grado.config(state="readonly")
                    self.cb_situacion.config(state="readonly")
                    self.chk_autoriza.config(state="normal")
                    self.btn_foto.config(state="normal")

                    self.e_fecha_baja.delete(0, "end")
                    self.e_fecha_baja.config(state="disabled")

                    self.e_apellido.focus_set()
                else:
                    self.limpiar_legajo()

        except Exception as e:
            self.ui.show_error("Error", f"No se pudo buscar el legajo: {e}")

    def buscar_legajo_enter(self, event=None):
        """Atajo para poder usar <Return> en el Entry del N° Legajo"""
        self.buscar_legajo()

    def cargar_foto(self):
        file_path = self.ui.ask_open_file(
            title="Seleccionar Foto 4x4",
            filetypes=[("Imágenes", "*.jpg *.jpeg *.png *.bmp *.gif")]
        )
        if not file_path:
            return
        try:
            img = Image.open(file_path)
            img.thumbnail((120, 120), RESAMPLE)
            self.foto_img = ImageTk.PhotoImage(img)
            self.lbl_foto.config(image=self.foto_img, text="")
            self.foto_path = file_path
        except Exception as e:
            self.ui.show_error("Error", f"No se pudo cargar la foto: {e}")

    def _copiar_foto_si_corresponde(self):
        if not self.foto_path or not os.path.exists(self.foto_path):
            return ""
        try:
            nombre = os.path.basename(self.foto_path)
            destino = os.path.join(FOTOS_DIR, nombre)
            if os.path.abspath(self.foto_path) != os.path.abspath(destino):
                shutil.copy2(self.foto_path, destino)
            return os.path.relpath(destino, APP_DIR)
        except:
            return ""

    def _mostrar_foto(self, path_relativo):
        """Muestra la foto en el label a partir de la ruta guardada en la DB."""
        try:
            # Si no hay ruta guardada
            if not path_relativo or path_relativo in ("None", "NULL"):
                self.lbl_foto.config(image="", text="SIN FOTO")
                self.foto_path = None
                return

            # Ruta absoluta desde APP_DIR
            path = os.path.join(APP_DIR, path_relativo)
            if not os.path.exists(path):
                # probar también con FOTOS_DIR + basename
                alt_path = os.path.join(FOTOS_DIR, os.path.basename(path_relativo))
                if os.path.exists(alt_path):
                    path = alt_path
                else:
                    self.lbl_foto.config(image="", text="SIN FOTO")
                    self.foto_path = None
                    return

            img = Image.open(path)
            img.thumbnail((120, 120), RESAMPLE)
            self.foto_img = ImageTk.PhotoImage(img)
            self.lbl_foto.config(image=self.foto_img, text="")
            self.foto_path = path

        except Exception as e:
            self.ui.show_error("Error", f"No se pudo mostrar la foto: {e}")
            self.lbl_foto.config(image="", text="SIN FOTO")
            self.foto_path = None

    def nuevo_legajo(self):
        # 🔒 Control de permisos
        if not self._has_role(("ADMIN", "SUPERVISOR")):
            messagebox.showwarning(
                "Sin permiso",
                "No tiene permiso para crear legajos.\n"
                "Consulte con un supervisor o administrador."
            )
            return

        self.modo_legajo = "nuevo"
        self._actualizar_estado_campos()

        # Habilita campos editables
        for w in [self.e_apellido, self.e_nombre, self.e_cargo, self.e_dni]:
            w.config(state="normal")

        self.cb_grado.config(state="readonly")
        self.cb_situacion.config(state="readonly")
        self.btn_foto.config(state="normal")
        self.e_legajo.config(state="normal")

        # Estado inicial del campo fecha_baja
        self.e_fecha_baja.delete(0, "end")
        self.e_fecha_baja.config(state="disabled")

        # Checkbox autoriza
        self.var_autoriza.set("NO")
        self.chk_autoriza.config(state="normal")

        # Aplicar lógica según situación
        self._actualizar_estado_fecha_baja()

        # 🔵 IMPORTANTE: habilitar botones de modo nuevo
        self._estado_botones_legajo("nuevo")

        self.e_legajo.focus_set()
        
    def limpiar_legajo(self):
        self.modo_legajo = "nuevo"   
        self._actualizar_estado_campos()
        # Habilitar los campos primero (solo para limpiar)
        for w in [self.e_apellido, self.e_nombre, self.cb_grado, self.e_cargo, self.e_dni, self.cb_situacion, self.e_fecha_baja,self.e_email, self.e_nro_cel]:
            try:
                w.config(state="normal")
            except Exception:
                pass
        # Limpiar los datos
        self.var_id.set("")
        self.var_apellido.set("")
        self.var_nombre.set("")
        self.var_dni.set("")
        self.var_grado.set("")
        self.var_cargo.set("")
        self.var_email.set("")
        self.var_nro_cel.set("")
        self.lbl_foto.config(image="", text="")
        self.foto_path = None
        self.cb_situacion.set("")
        self.e_fecha_baja.config(state="normal")
        self.e_fecha_baja.delete(0, "end")   # <-- borra la fecha
        self.e_fecha_baja.config(state="disabled")
        self.var_autoriza.set("NO")
        self.chk_autoriza.config(state="disabled")
        # Deshabilitar los campos de nuevo
        for w in [self.e_apellido, self.e_nombre, self.cb_grado, self.e_cargo, self.e_dni, self.cb_situacion, self.e_fecha_baja,self.e_email, self.e_nro_cel]:
            try:
                w.config(state="disabled")
            except Exception:
                pass

        self.btn_foto.config(state="disabled")
        self.e_legajo.config(state="normal")
        self.e_legajo.focus_set()

        self._estado_botones_legajo("inicial")

    def _habilitar_modificar_legajo(self):
        if not self.tree_legajos.selection():
            messagebox.showwarning(
                "Atención",
                "Debe seleccionar un legajo primero."
            )
            return

        rol = self.usuario_actual.get("rol", "").upper()
        legajo_usuario = str(self.usuario_actual.get("legajo"))

        item = self.tree_legajos.selection()[0]
        valores = self.tree_legajos.item(item, "values")
        legajo_seleccionado = str(valores[0])

        # 🔒 BOMBERO: solo su propio legajo
        if rol == "BOMBERO" and legajo_seleccionado != legajo_usuario:
            messagebox.showwarning(
                "Acceso denegado",
                "Solo puede modificar su propio legajo."
            )
            return

        self.modo_legajo = "modificar"

        try:
            self._recreate_date_entries()
        except Exception:
            pass

        self._actualizar_estado_campos()

        # ---------- CAMPOS EDITABLES ----------
        campos_editables = [
            self.e_apellido,
            self.e_nombre,
            self.e_dni,
            self.e_email,
            self.e_nro_cel,
            self.btn_foto
        ]

        for w in campos_editables:
            try:
                w.config(state="normal")
            except Exception:
                pass

        # ---------- RESTRICCIONES BOMBERO ----------
        if rol == "BOMBERO":
            self.e_cargo.config(state="disabled")
            self.cb_grado.config(state="disabled")
            self.cb_situacion.config(state="disabled")
            self.chk_autoriza.config(state="disabled")
            self.e_fecha_baja.config(state="disabled")
        else:
            # SUPERVISOR / ADMIN
            self.e_cargo.config(state="normal")
            self.cb_grado.config(state="readonly")
            self.cb_situacion.config(state="readonly")
            self.chk_autoriza.config(state="normal")
            self._actualizar_estado_fecha_baja()

        self.e_legajo.config(state="disabled")

        self.leg_btns["Guardar"].config(state="normal")
        self.leg_btns["Limpiar/Cancelar"].config(state="normal")

        self.e_apellido.focus_set()

        print(
            f"DEBUG >> Modificar habilitado | ROL={rol} | "
            f"LEGAJO={legajo_seleccionado}"
        )

    def cargar_legajos_combobox(self):
        if not hasattr(self, "usuario_actual") or not self.usuario_actual:
            return
            
        rol_usuario = self.usuario_actual.get("rol", "").upper()
        legajo_propio = str(self.usuario_actual.get("legajo", ""))
        
        try:
            conn = get_db_connection()
            # Usamos dictionary=True para mantener la compatibilidad con r['columna']
            c = conn.cursor(dictionary=True)
            
            # 1. Obtenemos legajos activos (Sintaxis MySQL)
            query = """
                SELECT legajo, apellido, nombre
                FROM legajos
                WHERE UPPER(situacion) NOT IN ('BAJA', 'SANCIÓN', 'CON SANCIÓN', 'LICENCIA')
                ORDER BY apellido, nombre
            """
            c.execute(query)
            rows = c.fetchall()
            
            values = []
            for r in rows:
                legajo_id = str(r['legajo']).strip()
                # Buscamos datos del usuario en el diccionario local para filtrar jerarquías
                user_data = self.usuarios_dict.get(legajo_id, {})
                rol_iterado = user_data.get("rol", "").upper()

                # --- FILTRO DE JERARQUÍA ACTUALIZADO ---
                if rol_usuario in ("SUPERVISOR", "ENCARGADO"):
                    # Supervisor y Encargado NO ven a los ADMIN. 
                    # Tampoco ven a otros de su mismo rango (excepto a ellos mismos).
                    if rol_iterado in ("ADMIN", "SUPERVISOR", "ENCARGADO") and legajo_id != legajo_propio:
                        continue
                
                texto = f"{legajo_id} - {r['apellido']} {r['nombre']}".upper()
                values.append(texto)

            # --- ASIGNACIÓN A COMBOS ---
            # Caso Bombero: Solo él mismo
            if rol_usuario == "BOMBERO":
                apellido = self.usuario_actual.get("apellido", "")
                nombre = self.usuario_actual.get("nombre", "")
                valor_propio = f"{legajo_propio} - {apellido} {nombre}".upper()

                if hasattr(self, 'cb_legajo_sel'):
                    self.cb_legajo_sel.config(state="normal")
                    self.cb_legajo_sel["values"] = [valor_propio]
                    self.cb_legajo_sel.set(valor_propio)
                    self.cb_legajo_sel.config(state="disabled")
                return 

            # Caso ADMIN / SUPERVISOR / ENCARGADO
            if hasattr(self, 'cb_legajo_sel'):
                self.cb_legajo_sel.config(state="readonly")
                self.cb_legajo_sel["values"] = [""] + values
                self.cb_legajo_sel.set("")
    
            if hasattr(self, 'inf_legajo_cb'):
                self.inf_legajo_cb["values"] = [""] + values
                self.inf_legajo_cb.set("")

            c.close()
            conn.close()
        except Exception as e:
            print(f"Error en cargar_legajos_combobox (MySQL): {e}")
            
    def cargar_grilla_legajos(self):
        # Limpiar grilla
        for i in self.tree_legajos.get_children():
            self.tree_legajos.delete(i)

        if not hasattr(self, "usuario_actual") or not self.usuario_actual:
            return

        rol = self.usuario_actual.get("rol", "").upper()
        legajo_propio = self.usuario_actual.get("legajo")

        # Dentro de cargar_grilla_legajos(self):
        try:
            conn = get_db_connection()
            c = conn.cursor(dictionary=True)

            # Vamos a traer TODO sin filtros para verificar que la conexión es real
            query = "SELECT legajo, apellido, nombre, grado, cargo, situacion FROM legajos ORDER BY apellido ASC"
            c.execute(query)
            
            rows = c.fetchall()
            print(f"DEBUG: MySQL devolvió {len(rows)} filas.") # Esto debe decir 97 ahora

            for r in rows:
                self.tree_legajos.insert(
                    "",
                    "end",
                    values=(
                        r["legajo"],
                        r["apellido"],
                        r["nombre"],
                        r["grado"],
                        r["cargo"],
                        r["situacion"]
                    )
                )
            # ... resto del código

            c.close()
            conn.close()
            
            # Ejecutar ajuste de columnas si existe el método
            if hasattr(self, "_ajustar_columnas_legajos"):
                self.master.after(100, self._ajustar_columnas_legajos)

        except Exception as e:
            print(f"Error en cargar_grilla_legajos (MySQL): {e}")

    def _ajustar_columnas_legajos(self):
        """Ajusta el ancho de las columnas de la grilla de legajos."""

        if not hasattr(self, "tree_legajos"):
            return

        if not self.tree_legajos.winfo_exists():
            return

        for col in self.tree_legajos["columns"]:
            max_len = len(col)
            for item in self.tree_legajos.get_children():
                texto = str(self.tree_legajos.set(item, col))
                if len(texto) > max_len:
                    max_len = len(texto)

            ancho = max(80, min(200, max_len * 8))
            self.tree_legajos.column(col, width=ancho, stretch=True)

    def eliminar_legajo(self, event=None):
        nro = self.var_id.get().strip().upper()
        if not nro: return
        if not messagebox.askyesno("Confirmar","¿Eliminar legajo?"): return
        conn = sqlite3.connect(DB_PATH); c = conn.cursor()
        try:
            c.execute("DELETE FROM legajos WHERE legajo=?",(nro,))
            conn.commit()
            self.ui.show_info("OK","Legajo eliminado")
        finally:
            conn.close()
        self.cargar_legajos_combobox()
        self.limpiar_legajo();
        # Forzar ocultar fecha_baja después de eliminar
        try:
            self.e_fecha_baja.place_forget()
            self.e_fecha_baja.config(state="disabled")
        except Exception:
            pass
        self.cargar_grilla_legajos()
        self.tree_legajos.selection_remove(self.tree_legajos.selection())
        self._estado_botones_legajo("inicial")

    def on_legajo_double_click(self, event):
        tree = event.widget
        sel = tree.selection()
        if not sel:
            return

        item = sel[0]
        values = tree.item(item, "values")
        if not values:
            return

        leg = str(values[0])

        # BOMBERO solo su propio legajo
        if self.usuario_actual["rol"].upper() == "BOMBERO":
            if leg != str(self.usuario_actual.get("legajo")):
                messagebox.showwarning(
                    "Acceso restringido",
                    "Solo puede visualizar su propio legajo."
                )
                return

        # cargar datos
        self.var_id.set(leg)
        self.buscar_legajo()

        # modo lectura
        self.modo_legajo = "lectura"
        self._estado_campos_legajo("lectura")
        self._estado_botones_legajo("cargado")

    def guardar_legajo(self):
        nro = self.var_id.get().strip().upper()
        if not nro:
            self.ui.show_error("Error", "Debe ingresar N° de Legajo"); return
        if not self.var_apellido.get().strip():
            self.ui.show_error("Error", "Debe ingresar Apellido"); return
        if not self.var_nombre.get().strip():
            self.ui.show_error("Error", "Debe ingresar Nombre"); return
        if not self.var_grado.get().strip():
            self.ui.show_error("Error", "Debe seleccionar Grado"); return

        # -------- Fecha baja --------
        if self.e_fecha_baja.cget("state") in ("normal", "readonly"):
            try:
                fecha_dt = self.e_fecha_baja.get_date()
                fecha_baja = "" if fecha_dt == date(1900, 1, 1) else fecha_dt.strftime("%d/%m/%Y")
            except Exception:
                fecha_baja = ""
        else:
            fecha_baja = ""

        rol = self.usuario_actual.get("rol", "").upper()

        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()

        try:
            foto_rel = self._copiar_foto_si_corresponde()

            # =====================================================
            # 🔒 BLINDAJE PARA USUARIO BÁSICO
            # =====================================================
            if rol == "BASICO":
                c.execute(
                    "SELECT situacion, autoriza FROM legajos WHERE legajo=?",
                    (nro,)
                )
                row = c.fetchone()

                situacion = row[0] if row else self.cb_situacion.get()
                autoriza_db = row[1] if row else self.var_autoriza.get()
            else:
                situacion = self.cb_situacion.get()
                autoriza_db = self.var_autoriza.get()

            # =====================================================
            # UPDATE / INSERT
            # =====================================================
            if self.modo_legajo == "modificar":

                c.execute("""UPDATE legajos SET
                    apellido=?, nombre=?, dni=?, grado=?, cargo=?, email=?,
                    nro_cel=?, foto=?, situacion=?, autoriza=?, fecha_baja=?
                    WHERE legajo=?""",
                    (
                        self.var_apellido.get(),
                        self.var_nombre.get(),
                        self.var_dni.get(),
                        self.var_grado.get(),
                        self.var_cargo.get(),
                        self.var_email.get(),
                        self.var_nro_cel.get(),
                        foto_rel,
                        situacion,
                        autoriza_db,
                        fecha_baja,
                        nro
                    )
                )

            elif self.modo_legajo == "nuevo":

                c.execute("SELECT 1 FROM legajos WHERE legajo=?", (nro,))
                if c.fetchone():
                    self.ui.show_error("Error", f"El legajo {nro} ya existe")
                    return

                c.execute("""INSERT INTO legajos
                    (legajo, apellido, nombre, dni, grado, cargo,
                    email, nro_cel, foto, situacion, autoriza, fecha_baja)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (
                        nro,
                        self.var_apellido.get(),
                        self.var_nombre.get(),
                        self.var_dni.get(),
                        self.var_grado.get(),
                        self.var_cargo.get(),
                        self.var_email.get(),
                        self.var_nro_cel.get(),
                        foto_rel,
                        situacion,
                        autoriza_db,
                        fecha_baja
                    )
                )

            else:
                self.ui.show_error("Error", "Modo inválido. No se puede guardar.")
                return

            conn.commit()
            self.ui.show_info("Legajo guardado", "OK")

            # Reset visual
            self.modo_legajo = "nuevo"
            self._estado_botones_legajo("inicial")
            self.cargar_grilla_legajos()

        except Exception as e:
            self.ui.show_error("Error", f"No se pudo guardar: {e}")

        finally:
            conn.close()
            self.cargar_legajos_combobox()
            self.limpiar_legajo()
            self.cargar_asignados_autorizados()
            self.e_legajo.focus_set()

    def _on_situacion_enter(self, event=None):
        """
        Enter en Situación → aplica lógica y mueve foco (versión robusta).
        """
        rol = self.usuario_actual.get("rol", "").upper()
        if rol == "BASICO":
            return "break"
        
        self.cb_situacion.bind("<Return>", self._on_situacion_change)

        situ = self.cb_situacion.get().upper().strip()
        if self.modo_legajo in ("nuevo", "modificar"):
            if situ in ("BAJA", "LICENCIA", "CON SANCIÓN", "SANCIÓN"):
                # mover momentáneamente el foco fuera del combobox para evitar que lo recupere
                try:
                    if event is not None and getattr(event, "widget", None):
                        self.master.focus_set()
                    else:
                        self.master.focus_set()
                except Exception:
                    pass

                # intentar enfocar el entry interno de DateEntry varias veces
                def intentar(n=0):
                    try:
                        if hasattr(self.e_fecha_baja, "entry"):   # tkcalendar nuevo
                            self.e_fecha_baja.entry.focus_force()
                        elif hasattr(self.e_fecha_baja, "_entry"):  # fallback
                            self.e_fecha_baja._entry.focus_force()
                        else:
                            self.e_fecha_baja.focus_force()
                    except Exception:
                        try:
                            self.e_fecha_baja.focus_set()
                        except Exception:
                            pass

                    # reintentar algunas veces por si otro handler pisa el foco
                    if n < 4:
                        self.master.after(50, lambda: intentar(n + 1))

                # pequeña demora antes de la primera tentativa
                self.master.after(20, intentar)
            else:
                # si no corresponde fecha_baja, ir al checkbox Autoriza
                self.master.after_idle(lambda: self.chk_autoriza.focus_set())

        return "break"   # evita que el combobox recupere el foco por la cadena de bindings

    def _forzar_foco_fecha_baja(self):
        """Intenta dar foco al campo interno de DateEntry."""
        try:
            if hasattr(self.e_fecha_baja, "entry"):   # versión moderna de tkcalendar
                self.e_fecha_baja.entry.focus_set()
            elif hasattr(self.e_fecha_baja, "_entry"):  # fallback
                self.e_fecha_baja._entry.focus_set()
            else:
                self.e_fecha_baja.focus_set()
        except Exception:
            try:
                self.e_fecha_baja.focus_force()
            except Exception:
                pass

    def _actualizar_estado_fecha_baja(self, event=None):
        """
        Muestra/oculta y habilita/inhabilita la Fecha Baja.
        Controla además el checkbox Autoriza según situación, modo y rol.
        """
        situacion = (self.cb_situacion.get() or "").upper().strip()
        modo_editable = self.modo_legajo in ("nuevo", "modificar")
        rol = self.usuario_actual.get("rol", "").upper()

        visibles = {"BAJA", "LICENCIA", "CON SANCIÓN", "SANCIÓN"}

        # -------------------------------
        # SITUACIÓN: BLOQUEO POR ROL
        # -------------------------------
        if rol == "BASICO":
            try:
                self.cb_situacion.config(state="disabled")
            except Exception:
                pass
        else:
            try:
                # readonly para evitar texto libre
                self.cb_situacion.config(state="readonly")
            except Exception:
                pass

        # -------------------------------
        # FECHA BAJA
        # -------------------------------
        if situacion in visibles:
            # Mostrar campo
            try:
                self.e_fecha_baja.place(**self._fecha_baja_place)
            except Exception:
                try:
                    self.e_fecha_baja.place(x=150, y=350, width=200)
                except Exception:
                    pass

            if modo_editable:
                try:
                    self.e_fecha_baja.config(state="normal")
                except Exception:
                    pass

                # Setear fecha por defecto si está vacía
                try:
                    fecha_dt = self.e_fecha_baja.get_date()
                except Exception:
                    fecha_dt = None

                if not fecha_dt or fecha_dt == date(1900, 1, 1):
                    try:
                        self.e_fecha_baja.set_date(date.today())
                    except Exception:
                        pass

                # Forzar foco
                self.master.after(50, self._forzar_foco_fecha_baja)
            else:
                try:
                    self.e_fecha_baja.config(state="disabled")
                except Exception:
                    pass
        else:
            # Ocultar y resetear
            try:
                self.e_fecha_baja.place_forget()
            except Exception:
                pass
            try:
                self.e_fecha_baja.set_date(date(1900, 1, 1))
            except Exception:
                pass
            try:
                self.e_fecha_baja.config(state="disabled")
            except Exception:
                pass

        # -------------------------------
        # CHECKBOX AUTORIZA (CONTROL POR ROL)
        # -------------------------------
        habilita_autoriza = situacion in (
            "ACTIVO", "CUERPO AUXILIAR", "RESERVA ACTIVA",
            "MIEMBRO COMISIÓN DIRECTIVA", "MIEMBRO COMISIÓN DE EVENTOS",
            "PERMISO ESPECIAL", "OTRO"
        )

        if rol in ("ADMIN", "SUPERVISOR") and modo_editable and habilita_autoriza:
            try:
                self.chk_autoriza.config(state="normal")
            except Exception:
                pass
        else:
            try:
                # Solo lectura para BÁSICO o cuando no corresponde
                self.chk_autoriza.config(state="disabled")
            except Exception:
                pass

    def _generar_pdf_legajo_en_memoria(self):
        from reportlab.platypus import Paragraph, Spacer, Image as RLImage, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
        from datetime import datetime
        import os, io

        nro_legajo = self.var_id.get().strip()
        if not nro_legajo:
            raise ValueError("Debe indicar un legajo antes de imprimir.")

        styles = getSampleStyleSheet()
        elems = []

        # --- Título ---
        title_style = ParagraphStyle('TitleSmall', parent=styles['Title'], fontSize=14, alignment=TA_CENTER)
        elems.append(Paragraph(f"REPORTE DE LEGAJO Nº {nro_legajo}", title_style))
        elems.append(Spacer(1, 12))

        # --- Datos ---
        labels = [
            "Apellido", "Nombre", "Grado", "Cargo/Función",
            "DNI", "Teléfono", "Email", "Situación", "Fecha Baja"
        ]

        valores = [
            self.var_apellido.get(),
            self.var_nombre.get(),
            self.var_grado.get(),
            self.var_cargo.get(),
            self.var_dni.get(),
            self.var_nro_cel.get(),    # ✅ correcto
            self.var_email.get(),
            self.cb_situacion.get(),
            self.e_fecha_baja.get()
        ]

        left_rows = []
        for lab, val in zip(labels, valores):
            if lab == "Fecha Baja" and val in ("01/01/1900", "1900-01-01", "", None):
                continue
            left_rows.append([Paragraph(f"<b>{lab}:</b>", styles["Normal"]), Paragraph(val or "", styles["Normal"])])

        left_table = Table(left_rows, colWidths=[110, 260])
        left_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))

        # --- Foto ---
        foto_rl = None
        if getattr(self, "foto_path", None) and os.path.exists(self.foto_path):
            try:
                foto_rl = RLImage(self.foto_path, width=120, height=140)
            except:
                foto_rl = None

        if foto_rl:
            main_tbl = Table([[left_table, foto_rl]], colWidths=[400, 120])
            main_tbl.setStyle(TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ]))
            elems.append(main_tbl)
        else:
            elems.append(left_table)

        # --- Crear PDF en memoria ---
        buffer = io.BytesIO()
        self._crear_pdf_unificado(buffer, elems, "")
        return buffer.getvalue()

    def imprimir_legajo_directo(self):
#        from tkinter.messagebox import showerror, showinfo
        import tempfile, os

        try:
            pdf_bytes = self._generar_pdf_legajo_en_memoria()

            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as f:
                f.write(pdf_bytes)
                temp_path = f.name

            # Enviar directo a la impresora (Windows)
            os.startfile(temp_path, "print")

            self.ui.show_info("Impresión", "El legajo fue enviado a la impresora.")

        except Exception as e:
            import traceback
            traceback.print_exc()
            self.ui.show_error("Error", f"No se pudo imprimir el legajo:\n{e}")

    def imprimir_listado_legajos(self):
        """Genera el listado de legajos con encabezado y pie unificados."""
        try:
            conn = sqlite3.connect(DB_PATH)
            c = conn.cursor()
            c.execute("""
                SELECT legajo, apellido, nombre, grado, cargo, dni, situacion, fecha_baja
                FROM legajos
                ORDER BY apellido, nombre
            """)
            rows = c.fetchall()
            conn.close()
            if not rows:
                self.ui.show_error("Error", "No hay legajos.")
                return

            # --- Nombre automático ---
            sugerido = self._default_informe_filename("Listado_Legajos")
            path = self.ui.ask_save_file(
                defaultextension=".pdf",
                filetypes=[("PDF","*.pdf")],
                initialfile=sugerido
            )
            if not path:
                return

            styles = getSampleStyleSheet()
            elems = []

            # --- Título ---
            elems.append(Paragraph("LISTADO DE LEGAJOS", styles["Title"]))
            elems.append(Spacer(1, 10))

            # --- Tabla ---
            headers = ["Legajo", "Apellido", "Nombre", "Grado", "Cargo/Función", "DNI", "Situación", "Fecha Baja"]
            data = [headers] + [list(r) for r in rows]

            tbl = Table(data, repeatRows=1)
            tbl.setStyle(TableStyle([
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("BACKGROUND", (0, 0), (-1, 0), colors.gray),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
            ]))
            elems.append(tbl)

            # --- Crear PDF con encabezado/pie estándar ---
            self._crear_pdf_unificado(path, elems, "Listado de Legajos")

            try:
                os.startfile(path)  # 🔹 abre el PDF automáticamente
            except Exception as e:
                messagebox.showwarning(
                    "Aviso",
                    f"El PDF se creó correctamente pero no se pudo abrir.\n{e}"
                )
        except Exception as e:
            self.ui.show_error("Error", str(e))

    def cargar_actividades_desde_conceptos(self):
        try:
            conn = sqlite3.connect(DB_PATH)
            c = conn.cursor()
            c.execute("SELECT id, concepto FROM conceptos ORDER BY concepto")
            conceptos = c.fetchall()
            conn.close()

            # ✅ Solo BD, sin mezclas
            lista = [f"{id_} - {nombre}" for id_, nombre in conceptos]

            self.actividad['values'] = lista

            # ✅ Mapas consistentes
            self.conceptos_map = {id_: nombre for id_, nombre in conceptos}
            self.conceptos_text_to_id = {f"{id_} - {nombre}": id_ for id_, nombre in conceptos}

        except Exception as e:
            print(f"Error cargando conceptos: {e}")
            self.actividad['values'] = []

    def exportar_legajos_excel(self):
        """Exporta los legajos a Excel (detalle ampliado con todas las columnas)."""
        try:
            conn = sqlite3.connect(DB_PATH)
            c = conn.cursor()
            c.execute("""
                SELECT 
                    legajo, apellido, nombre, grado, cargo, dni, situacion, nro_cel, email, fecha_baja
                FROM legajos
                ORDER BY apellido, nombre
            """)
            rows = c.fetchall()
            conn.close()

            if not rows:
                self.ui.show_error("Error", "No hay legajos para exportar.")
                return

            hoy = date.today().strftime("%Y%m%d")
            default_name = f"Listado_Legajos_{hoy}.xlsx"

            path = self.ui.ask_save_file(
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")],
                initialfile=default_name
            )
            if not path:
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "Legajos"

            # --- Encabezados coherentes con el PDF institucional ---
            headers = [
                "N° Legajo", "Apellido", "Nombre", "Grado", "Cargo / Función",
                "DNI", "Situación", "Celular", "Email", "Fecha Baja"
            ]
            ws.append(headers)

            # --- Estilo de encabezado ---
            header_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            center = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center
                cell.border = thin_border

            # --- Agregar filas ---
            for row in rows:
                row = list(row)
                # Formatear fecha baja
                fecha = row[-1]
                if fecha:
                    try:
                        if isinstance(fecha, str):
                            fecha_dt = datetime.strptime(fecha, "%Y-%m-%d")
                        else:
                            fecha_dt = fecha
                        row[-1] = fecha_dt.strftime("%d/%m/%Y")
                    except Exception:
                        row[-1] = str(fecha)
                else:
                    row[-1] = ""
                ws.append(row)

            # --- Ajustar anchos automáticamente ---
            for col in ws.columns:
                max_len = max(len(str(cell.value)) for cell in col if cell.value)
                ws.column_dimensions[col[0].column_letter].width = max_len + 3

            # --- Centrar algunas columnas ---
            for col_letter in ["A", "D", "F", "G", "J"]:
                for cell in ws[col_letter]:
                    cell.alignment = center

            wb.save(path)
            self.ui.show_info("OK", f"Listado de legajos exportado correctamente:\n{path}")

            # 🔹 Abrir automáticamente el Excel
            try:
                import os
                os.startfile(path)
            except Exception as e:
                messagebox.showwarning(
                    "Aviso",
                    f"El archivo se creó correctamente pero no se pudo abrir automáticamente.\n{e}"
                )
        except Exception as e:
            self.ui.show_error("Error", f"No se pudo exportar el listado:\n{e}")

    def _refrescar_actividades(self):
            # 1. Reset de botones: todos apagados al empezar el chequeo
            for btn in self.act_btns.values():
                btn.config(state="disabled")

            rol = self.usuario_actual.get("rol", "").upper()
            estado_ui = getattr(self, "_estado_ui", "inicial")
            
            # 2. Obtener permisos base según Rol y Estado de la Pantalla
            permisos = PERMISOS_ACTIVIDADES.get(rol, {}).get(estado_ui, ())

            # 3. Variables de control para las reglas de negocio
            firmada = bool(getattr(self, "firma_bombero_fecha", None))
            
            # 4. Único bucle para aplicar reglas y estilos
            for nombre in permisos:
                # Por defecto habilitamos si está en el diccionario, 
                # pero pasamos por el filtro de seguridad de tus funciones:
                habilitar = True
                estilo = "normal"

                if nombre == "Guardar":
                    # Se habilita si es nuevo O si se puede modificar la actividad actual
                    if estado_ui == "nuevo":
                        habilitar = True
                    else:
                        habilitar = self._puede_modificar_actividad()

                # --- FILTROS DE SEGURIDAD BASADOS EN TUS FUNCIONES ---
                if nombre == "Modificar":
                    habilitar = self._puede_modificar_actividad()
                
                elif nombre == "Firmar actividad":
                    habilitar = self._puede_firmar_bombero()
                    estilo = "positivo"
                
                elif nombre == "Firmar Supervisor":
                    habilitar = self._puede_firmar_supervisor()
                    estilo = "positivo"
                
                elif nombre == "Anular":
                    habilitar = self._puede_anular_actividad()
                
                # --- BOTONES QUE SIEMPRE DEBEN ESTAR VIVOS SI ESTÁN EN EL DICT ---
                elif nombre in ("Pendientes de Firma", "Buscar", "Nuevo", "Limpiar/Cancelar"):
                    habilitar = True
                    if nombre == "Pendientes de Firma" and self._tiene_pendientes_firma():
                        estilo = "positivo"

                # Aplicar el estado final
                self._set_btn_estado(nombre, habilitar, estilo)

            # 5. Bloqueo de campos (Combos, fechas, etc.) según el estado
            self._set_estado_campos_actividades(estado_ui)

    def init_actividades(self):
        # -------------------------------------------------
        # TÍTULO
        # -------------------------------------------------
        Label(
            self.actividades_frame,
            text="ACTIVIDADES (Carga/Edición)",
            font=("Arial", 16, "bold"),
            fg="white",
            bg="red"
        ).place(x=20, y=6)

        # -------------------------------------------------
        # ESTADO VISUAL DE FIRMA (debajo del título)
        # -------------------------------------------------
        self.frm_estado_firma = Frame(
            self.actividades_frame,
            bg=self.actividades_frame["bg"],
            bd=1,
            relief="solid"
        )
        self.frm_estado_firma.place(x=20, y=40, width=1000, height=36)

        self.lbl_estado_firma = Label(
            self.frm_estado_firma,
            text="",
            font=("Arial", 10, "bold"),
            anchor="center"
        )
        self.lbl_estado_firma.pack(fill="both", expand=True, padx=4, pady=2)

        self.lbl_estado_anulada = Label(
            self.actividades_frame,
            text="",
            font=("Arial", 14, "bold"),
            fg="white",
            bg="red"
        )
        self.lbl_estado_anulada.place(x=420, y=6)

        # -------------------------------------------------
        # ESTILO COMBOBOX
        # -------------------------------------------------
        style = ttk.Style()

        style.configure(
            "Custom.TCombobox",
            font=("Arial", 11),
            foreground="black",
            fieldbackground="white",
            background="white"
        )

        style.configure(
            "Focus.TCombobox",
            fieldbackground="#fff59d",
            background="white"
        )

        style.map(
            "Focus.TCombobox",
            fieldbackground=[("readonly", "#fff59d")]
        )

        # -------------------------------------------------
        # VARIABLES DE LAYOUT
        # -------------------------------------------------
        y0, dy = 100, 28   # empieza debajo del estado

        # -------------------------------------------------
        # N° ACTIVIDAD
        # -------------------------------------------------
        Label(self.actividades_frame, text="N° Actividad:", fg="white", bg="red").place(x=20, y=y0)
        self.e_id_actividad = Entry(
            self.actividades_frame,
            textvariable=self.var_id_actividad,
            font=("Arial", 11),
            state="readonly",
            fg="black",
            bg="white"
        )
        self.e_id_actividad.place(x=150, y=y0, width=160)
        self.e_id_actividad.bind("<Return>", lambda e: self._focus_next_widget(e.widget))

        self.lbl_info = Label(
            self.actividades_frame,
            text="",
            font=("Arial", 11, "bold"),
            fg="black",
            bg="red"
        )
        self.lbl_info.place(x=330, y=y0 + 2)

        # -------------------------------------------------
        # LEGAJO
        # -------------------------------------------------
# -------------------------------------------------
        # LEGAJO (REEMPLAZA ESTA PARTE)
        # -------------------------------------------------
        Label(self.actividades_frame, text="Legajo:", fg="white", bg="red").place(x=20, y=y0 + dy)
        self.cb_legajo_sel = ttk.Combobox(
            self.actividades_frame,
            font=("Arial", 11),
            state="readonly",
            style="Custom.TCombobox"
        )
        self.cb_legajo_sel.place(x=150, y=y0 + dy, width=240)
        
        # Habilitar búsqueda (la función que ya tienes)
        self.habilitar_busqueda_combo(self.cb_legajo_sel)

        # BINDS CRUCIALES PARA TECLADO Y MOUSE:
        # Usamos una función intermedia (wrapper) para que haga todo junto
        self.cb_legajo_sel.bind("<<ComboboxSelected>>", self._on_legajo_changed_completo)
        self.cb_legajo_sel.bind("<Return>", self._on_legajo_changed_completo)
        self.cb_legajo_sel.bind("<FocusOut>", self._on_legajo_changed_completo)

        # -------------------------------------------------
        # ASIGNADO
        # -------------------------------------------------
        Label(self.actividades_frame, text="Asignado por:", fg="white", bg="red").place(x=20, y=y0 + 2 * dy)
        self.cb_asignado = ttk.Combobox(
            self.actividades_frame,
            font=("Arial", 11),
            state="readonly",
            style="Custom.TCombobox"
        )
        self.cb_asignado.place(x=150, y=y0 + 2 * dy, width=360)
        self.cb_asignado.bind("<Return>", lambda e: self._focus_next_widget(e.widget))

        self.cb_asignado.bind("<<ComboboxSelected>>", self._validar_asignado)
        self.cb_asignado.bind("<FocusOut>", self._validar_asignado)
        self.habilitar_busqueda_combo(self.cb_asignado)

        self.cargar_asignados_autorizados()
        self._ajustar_combo_asignado_por_contexto()

        # -------------------------------------------------
        # ACTIVIDAD
        # -------------------------------------------------
        Label(self.actividades_frame, text="Actividad:", fg="white", bg="red").place(x=20, y=y0 + 3 * dy)
        self.actividad = ttk.Combobox(
            self.actividades_frame,
            font=("Arial", 11),
            state="readonly",
            style="Custom.TCombobox"
        )
        self.actividad.place(x=150, y=y0 + 3 * dy, width=360)
        self.actividad.bind("<Return>", lambda e: self._focus_next_widget(e.widget))
        self.habilitar_busqueda_combo(self.actividad)

        # -------------------------------------------------
        # ÁREA
        # -------------------------------------------------
        Label(self.actividades_frame, text="Área/Dpto:", fg="white", bg="red").place(x=20, y=y0 + 4 * dy)
        self.area = ttk.Combobox(
            self.actividades_frame,
            values=self.AREAS,
            state="readonly",
            font=("Arial", 11),
            style="Custom.TCombobox"
        )
        self.area.place(x=150, y=y0 + 4 * dy, width=360)
        self.area.bind("<Return>", lambda e: self._forzar_foco_dateentry(self.fecha_inicio))
        self.habilitar_busqueda_combo(self.area)

        # -------------------------------------------------
        # FECHAS
        # -------------------------------------------------
        Label(self.actividades_frame, text="Fecha Inicio:", fg="white", bg="red").place(x=20, y=y0 + 5 * dy)
        self.fecha_inicio = DateEntry(
            self.actividades_frame,
            date_pattern='dd/mm/yyyy',
            state="disabled",
            font=("Arial", 11)
        )
        self.fecha_inicio.place(x=150, y=y0 + 5 * dy)
        self.fecha_inicio.bind("<Return>", lambda e: self._focus_next_widget(e.widget))

        Label(self.actividades_frame, text="Fecha fin:", fg="white", bg="red").place(x=20, y=y0 + 6 * dy)
        self.fecha_fin = DateEntry(
            self.actividades_frame,
            date_pattern='dd/mm/yyyy',
            state="disabled",
            font=("Arial", 11)
        )
        self.fecha_fin.place(x=150, y=y0 + 6 * dy)

        def _on_fecha_fin_enter(event=None):
            self.hora_inicio.focus_set()
            return "break"

        self.fecha_fin.bind("<Return>", _on_fecha_fin_enter)

        # -------------------------------------------------
        # HORAS
        # -------------------------------------------------
        Label(self.actividades_frame, text="Hora inicio:", fg="white", bg="red").place(x=20, y=y0 + 7 * dy)

        vcmd = (self.master.register(self.validar_hora), "%P")

        self.hora_inicio = tk.Entry(
            self.actividades_frame,
            validate="key",
            validatecommand=vcmd,
            font=("Arial", 11),
            fg="black",
            bg="white"
        )
        self.hora_inicio.place(x=150, y=y0 + 7 * dy, width=160)
        self.hora_inicio.bind("<Return>", self._on_hora_inicio_enter)

        Label(self.actividades_frame, text="Hora fin:", fg="white", bg="red").place(x=20, y=y0 + 8 * dy)

        self.hora_finalizacion = tk.Entry(
            self.actividades_frame,
            validate="key",
            validatecommand=vcmd,
            font=("Arial", 11),
            fg="black",
            bg="white"
        )
        self.hora_finalizacion.place(x=150, y=y0 + 8 * dy, width=160)
        self.hora_finalizacion.bind("<Return>", self._on_hora_fin_enter)
        # Normalización y cálculo automático
        self.hora_inicio.bind("<FocusOut>", self._on_hora_focus_out)
        self.hora_finalizacion.bind("<FocusOut>", self._on_hora_focus_out)

        Label(self.actividades_frame, text="Total Horas:", fg="white", bg="red").place(x=330, y=y0 + 8 * dy)
        self.var_horas = StringVar()
        self.e_horas = Entry(
            self.actividades_frame,
            textvariable=self.var_horas,
            font=("Arial", 11),
            state="readonly",
            fg="grey",
            bg="white"
        )
        self.e_horas.place(x=415, y=y0 + 8 * dy, width=100)

        # -------------------------------------------------
        # DESCRIPCIÓN
        # -------------------------------------------------
        Label(self.actividades_frame, text="Descripción:", fg="white", bg="red").place(x=20, y=y0 + 9 * dy)

        self.descripcion = Text(
            self.actividades_frame,
            width=50,
            height=2,          # antes 4 → ahora más chico
            font=("Arial", 11),
            bd=0,
            relief="flat",
            highlightthickness=0
        )
        self.descripcion.place(x=150, y=y0 + 9 * dy + 0)   # sube un poco
        self.descripcion.bind("<Return>", self._descripcion_enter)
        self.descripcion.bind("<Shift-Return>", self._descripcion_shift_enter)
        self.enforce_uppercase_text(self.descripcion)

        # -------------------------------------------------
        # BOTONES (lado derecho, como antes)
        # -------------------------------------------------
        botones = [
            "Nuevo",
            "Guardar",
            "Modificar",
            "Anular",
            "Limpiar/Cancelar",
            "Buscar",
            "Exportar PDF",
            "Imprimir Actividad",
            "Impr.Listado Actividad",
            "Ver historial",
            "Firmar actividad",
            "Firmar Supervisor",
            "No firmar ahora",
            "Pendientes de Firma"
        ]

        comandos = [
            self.nuevo_actividad,
            self.guardar_actividad,
            self._habilitar_modificar_actividad,
            self._anular_actividad,
            self.limpiar_actividad,
            self.buscar_actividad,
            self._exportar_pdf_actividad,
            self.imprimir_actividad,
            self.imprimir_listado_actividades,
            self.ver_historial_actividad,
            self.firmar_actividad_bombero,
            self.firmar_actividad_supervisor,
            self.no_firmar_ahora,
            self.ver_pendientes_firma 
        ]

        self.act_btns = {}

        # Layout lateral derecho (igual que antes)
        btn_x0 = 600
        btn_y0 = y0
        dx = 230
        dy_btn = 40

        for i, (txt, cmd) in enumerate(zip(botones, comandos)):
            col = i % 2
            fila = i // 2

            # color de borde según función
            if txt == "Firmar actividad":
                borde = "#2E8B57"      # verde
            elif txt == "Firmar Supervisor":
                borde = "#1E5AA8"      # azul
            elif txt == "No firmar ahora":
                borde = "#6C757D"      # gris
            else:
                borde = "#B22222"      # rojo

            cont = Frame(self.actividades_frame, bg=borde, padx=2, pady=2)

            b = Button(
                cont,
                text=txt,
                command=cmd,
                width=22,
                relief="flat",
                font=("Arial", 10)
            )

            b.pack(fill="both", expand=True)

            cont.place(
                x=btn_x0 + col * dx,
                y=btn_y0 + fila * dy_btn
            )

            self._aplicar_estilo_boton(b, habilitado=False)
            self.act_btns[txt] = b

        # ENTER ejecuta botones
        for b in self.act_btns.values():
            b.bind("<Return>", self._button_enter)

        # Activar navegación con Enter
        self._configurar_navegacion_actividades()

        # ESTADO INICIAL (lógica + estilo visual)
        self._estado_ui = "inicial"
        self.actividad_anulada = False
        self._refrescar_actividades()

        print("ROL:", self.usuario_actual.get("rol"))
        for nombre, btn in self.act_btns.items():
            print(nombre, "=>", btn["state"])

        # 🔥 RESALTADO TIPO EXCEL
        widgets_foco = [
            self.cb_legajo_sel,
            self.cb_asignado,
            self.actividad,
            self.area,
            self.fecha_inicio,
            self.fecha_fin,
            self.hora_inicio,
            self.hora_finalizacion,
            self.descripcion,
        ]

        for w in widgets_foco:
            try:
                w.bind("<FocusIn>", lambda e: self.activar_resaltado_foco(e.widget))
                w.bind("<FocusOut>", lambda e: self.desactivar_resaltado_foco(e.widget))
            except:
                pass

        self.actualizar_contador_pendientes()
        self.master.after(200, self._set_foco_inicial_actividades) 
        # 🔥 ÚLTIMA PALABRA SOBRE ASIGNADO
        try:
            self._ajustar_combo_asignado_por_contexto()
            print("🔥 POST FINAL ASIGNADO:", self.cb_asignado.cget("state"))
        except Exception as e:
            print("ERROR FINAL ASIGNADO:", e)      
        # 🔥 FOCO INICIAL SEGÚN ROL
        # Al final de la función de botones, SIEMPRE llamar a los campos

    def _on_legajo_changed_completo(self, event=None):
            """Maneja el cambio de legajo por cualquier método (mouse/teclado)"""
            # 1. Validamos asignación
            self._validar_asignado()
            
            # 2. Actualizamos la lista de la combo de abajo
            self._ajustar_combo_asignado_por_contexto()
            
            # 3. Si fue un 'Enter', pasamos al siguiente campo
            if event and event.keysym == 'Return':
                self._focus_next_widget(self.cb_legajo_sel)

    def _buscar_opcion_combo(self, event):

        combo = event.widget

        if not isinstance(combo, ttk.Combobox):
            return

        if event.keysym in ("Return", "Tab", "Up", "Down"):
            return

        texto = combo.get().lower()

        valores = combo["values"]

        for v in valores:
            if texto in str(v).lower():
                combo.set(v)
                combo.icursor(len(texto))
                break

    def _on_hora_inicio_enter(self, event):
        self._on_hora_focus_out(event)
        self.hora_finalizacion.focus_set()
        return "break"

    def _on_hora_fin_enter(self, event):
        self._on_hora_focus_out(event)
        self.descripcion.focus_set()
        return "break"

    def _estado_visual_firma(self, modo):

        firma_bombero_fecha = getattr(self, "firma_bombero_fecha", None)
        firma_supervisor_fecha = getattr(self, "firma_supervisor_fecha", None)
        # 🔄 Recalcular permisos antes de pintar
        puede_bombero = self._puede_firmar_bombero()
        puede_supervisor = self._puede_firmar_supervisor()

        # Reset
        for b in self.act_btns.values():
            self._aplicar_estilo_boton(b, False)

        # ------------------------
        # ESTADO INICIAL
        # ------------------------
        if modo == "inicial":
            self._aplicar_estilo_boton(self.act_btns["Nuevo"], True)
            self._aplicar_estilo_boton(self.act_btns["Buscar"], True)

            self.lbl_estado_firma.config(text="", fg="black")

        # ------------------------
        # ACTIVIDAD CARGADA (sin firmar)
        # ------------------------
        elif modo == "cargado":

            # 🔵 BOTONES BASE (siempre disponibles)
            for k in ("Nuevo", "Buscar", "Pendientes de Firma"):
                if k in self.act_btns:
                    self._aplicar_estilo_boton(self.act_btns[k], True)

            # 🔵 ACCIONES SOBRE LA ACTIVIDAD
            for k in ("Modificar", "Exportar PDF", "Imprimir Actividad"):
                if k in self.act_btns:
                    self._aplicar_estilo_boton(self.act_btns[k], True)

            # 🟢 BOTONES DE FIRMA (en verde)
            btn = self.act_btns["Firmar actividad"]

            if self._puede_firmar_bombero():
                self._aplicar_estilo_boton(btn, True, "positivo")
                btn.update_idletasks()  # 🔥 fuerza visual
            else:
                self._aplicar_estilo_boton(btn, False)
            print("COLOR BOTON:", btn.cget("bg"))

            # ⚪ Neutro
            puede_no_firmar = not bool(firma_bombero_fecha)

            self._aplicar_estilo_boton(
                self.act_btns["No firmar ahora"],
                puede_no_firmar
            )

            # 🔵 Cancelar / limpiar también activo
            if "Limpiar/Cancelar" in self.act_btns:
                self._aplicar_estilo_boton(self.act_btns["Limpiar/Cancelar"], True)

            self.lbl_estado_firma.config(
                text="🟡 Actividad cargada – pendiente de firma del bombero",
                fg="goldenrod"
            )

        # ------------------------
        # MODO EDICIÓN
        # ------------------------
        elif modo == "edicion":

            for k in ("Guardar", "Cancelar", "Limpiar/Cancelar"):
                if k in self.act_btns:
                    self._aplicar_estilo_boton(self.act_btns[k], True)

            if firma_bombero_fecha or firma_supervisor_fecha:
                self.lbl_estado_firma.config(
                    text="⚠ Modificando actividad – las firmas serán invalidadas",
                    fg="#E67E22"
                )
            else:
                self.lbl_estado_firma.config(
                    text="✏️ Modificando actividad",
                    fg="#E67E22"
                )
        # ------------------------
        # FIRMADA POR BOMBERO
        # ------------------------
        elif modo == "firmado_bombero":

            # Siempre habilitados
            for k in ("Exportar PDF", "Imprimir Actividad"):
                self._aplicar_estilo_boton(self.act_btns[k], True)
            # 🔵 BOTONES BASE
            for k in ("Nuevo", "Buscar", "Pendientes de Firma"):
                if k in self.act_btns:
                    self._aplicar_estilo_boton(self.act_btns[k], True)

            if self._puede_firmar_supervisor():

                self._aplicar_estilo_boton(
                    self.act_btns["Firmar Supervisor"], True, "positivo"
                )

                self.lbl_estado_firma.config(
                    text="🟢 Firmada por bombero – pendiente de supervisor",
                    fg="green"
                )

            else:
                self._aplicar_estilo_boton(
                    self.act_btns["Firmar Supervisor"], False
                )

                ToolTip(
                    self.act_btns["Firmar Supervisor"],
                    "La actividad debe ser firmada por el supervisor asignado o un ADMIN."
                )

                self.lbl_estado_firma.config(
                    text="🟢 Firmada por bombero – requiere aprobación externa",
                    fg="green"
                )

        # ------------------------
        # FIRMADA POR SUPERVISOR
        # ------------------------
        elif modo == "firmado_supervisor":
            for k in ("Exportar PDF", "Imprimir Actividad", "Impr.Listado Actividad"):
                self._aplicar_estilo_boton(self.act_btns[k], True)

            self.lbl_estado_firma.config(
                text="🔵 Actividad firmada y cerrada",
                fg="#1E5AA8"
            )

        # ------------------------
        # ACTIVIDAD ANULADA
        # ------------------------
        elif modo == "anulada":

            for k in ("Exportar PDF", "Imprimir Actividad"):
                self._aplicar_estilo_boton(self.act_btns[k], True)

            self.lbl_estado_firma.config(
                text="⛔ Actividad anulada",
                fg="red"
            )

    def _on_hora_focus_out(self, event=None):
        widget = event.widget
        valor = widget.get().strip()

        if not valor:
            return

        # 🔴 BLOQUEAR 24
        if valor.startswith("24"):
            self.ui.show_error(
                "Hora inválida",
                "La hora 24 no es válida.\nUse 23:59 o 00:00."
            )
            widget.focus_set()
            return

        # 🟢 SI ES SOLO NÚMERO → FORMATEAR DIRECTO Y SALIR
        if valor.isdigit():
            h = int(valor)

            if 0 <= h <= 23:
                widget.delete(0, "end")
                widget.insert(0, f"{h:02d}:00")

                # 🔴 IMPORTANTE: NO seguir normalizando
                self.calcular_duracion()
                return
            else:
                self.ui.show_error(
                    "Hora inválida",
                    "Ingrese una hora válida entre 0 y 23."
                )
                widget.focus_set()
                return

        # 🟢 RESTO DE CASOS → usar tu lógica existente
        self._normalizar_entry_hora(widget)
        self.calcular_duracion()

    def _set_estado_visual(self, texto, fg, bg):
        """Aplica estilo visual al estado de firma."""
        try:
            self.lbl_estado_firma.config(text=texto, fg=fg, bg=bg)
            self.frm_estado_firma.config(bg=bg)
        except Exception:
            pass

    def _anular_actividad(self):
            # -------------------------------------------------
            # 🔒 VALIDACIÓN DE ROL: Solo ADMIN puede anular
            # -------------------------------------------------
            rol_actual = self.usuario_actual.get("rol", "").upper()
            
            if rol_actual != "ADMIN":
                self.ui.show_error(
                    "Acceso Denegado", 
                    "Solo un ADMINISTRADOR puede anular actividades.\n\n"
                    "Si esto es un error, contacte al soporte técnico."
                )
                return

            # --- Validación de actividad cargada ---
            if not getattr(self, "id_actividad_actual", None):
                self.ui.show_error("Error", "No hay actividad cargada.")
                return

            # --- Pedir motivo ---
            motivo = self._pedir_motivo_anulacion()
            if not motivo:
                messagebox.showwarning(
                    "Anulación",
                    "Debe ingresar un motivo para anular la actividad."
                )
                return

            confirmar = messagebox.askyesno(
                "Confirmar",
                "⚠ Está por ANULAR una actividad.\n\n¿Desea continuar?"
            )

            if not confirmar:
                return

            # -------------------------------------------------
            # PROCESO DE ANULACIÓN
            # -------------------------------------------------
            from datetime import datetime
            ahora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # Armamos el nombre y legajo para el registro
            nombre_admin = f"{self.usuario_actual.get('apellido', '').strip()} {self.usuario_actual.get('nombre', '').strip()}".strip()
            legajo_admin = self.usuario_actual.get('legajo', 'S/L')
            usuario_anula_str = f"{nombre_admin} (Leg. {legajo_admin})"

            try:
                conn = sqlite3.connect("siab.db")
                cursor = conn.cursor()
                
                cursor.execute("""
                    UPDATE actividades
                    SET anulada = 1,
                        motivo_anulacion = ?,
                        usuario_anula = ?,
                        fecha_anulacion = ?
                    WHERE id = ?
                """, (
                    motivo,
                    usuario_anula_str,
                    ahora,
                    self.id_actividad_actual
                ))

                conn.commit()
                conn.close()

                # Notificación en hilo separado
                threading.Thread(
                    target=self._worker_notificacion_firma,
                    args=(self.id_actividad_actual, "ANULADA"),
                    daemon=True
                ).start()

                self.ui.show_info("OK", f"Actividad N° {self.id_actividad_actual} anulada correctamente.")

                # 🔥 Recargar y refrescar
                self.buscar_actividad_por_id(self.id_actividad_actual)
                self.actualizar_contador_pendientes()
                self._refrescar_actividades()

            except Exception as e:
                self.ui.show_error("Error de Base de Datos", f"No se pudo anular: {e}")

    def ver_pendientes_firma(self):

        # Evitar múltiples ventanas
        if hasattr(self, "win_pendientes") and self.win_pendientes.winfo_exists():
            self.win_pendientes.lift()
            self.win_pendientes.focus_force()
            return

        rol = self.usuario_actual["rol"].upper()
        legajo = str(self.usuario_actual.get("legajo", "")).strip()
        username = self.usuario_actual.get("username", "").strip().upper()

        self.win_pendientes = Toplevel(self.master)
        win = self.win_pendientes

        # --- HACER VENTANA MODAL ---
        win.transient(self.master)   # Asociada a la principal
        win.grab_set()             # Bloquea el form principal
        win.focus_force()          # Forzar foco
        win.lift()                 # Traer al frente

        win.title("Pendientes de Firma")
        win.geometry("1100x500")
        win.configure(bg="white")

        # -------------------------------------------------
        # FRAME SUPERIOR (FILTROS)
        # -------------------------------------------------
        frm_top = Frame(win, bg="white")
        frm_top.pack(fill="x", padx=10, pady=5)

        # ------------------------------
        # TITULO DINÁMICO
        # ------------------------------
        titulo_txt = "PENDIENTES"

        if rol == "BOMBERO":
            leg = str(self.usuario_actual.get("legajo", "")).strip()
            ape = self.usuario_actual.get("apellido", "").upper()
            nom = self.usuario_actual.get("nombre", "").upper()
            titulo_txt = f"PENDIENTES – LEGAJO {leg} - {ape} {nom}"

        Label(
            frm_top,
            text=titulo_txt,
            font=("Arial", 14, "bold"),
            bg="white"
        ).pack(side="left")

        lbl_contador = Label(frm_top, text="",
                            font=("Arial", 11, "bold"),
                            bg="white")

        cb_filtro = None

        if rol in ("SUPERVISOR", "ADMIN"):
            opciones = [
                "Todas",
                "Pendientes de bombero",
                "Pendientes de supervisor",
                "Asignadas a mí"
            ]

            cb_filtro = ttk.Combobox(
                frm_top,
                values=opciones,
                state="readonly",
                width=25
            )

            # 🔥 DEFAULT INTELIGENTE CORREGIDO
            # Priorizamos que vea lo que tiene que firmar él mismo (Asignadas a mí)
            if rol in ("SUPERVISOR", "ADMIN"):
                cb_filtro.set("Asignadas a mí")
            else:
                cb_filtro.set("Todas") 

            # Eliminé la línea que pisaba el valor con "Pendientes de supervisor"
            cb_filtro.pack(side="right", padx=10)

        # -------------------------------------------------
        # CARGA DE DATOS
        # -------------------------------------------------
        def cargar_datos():

            conn = sqlite3.connect(DB_PATH)
            conn.row_factory = sqlite3.Row
            c = conn.cursor()

            filtro_valor = cb_filtro.get() if cb_filtro else "Todas"
            # -------------------
            # CONSULTA BASE
            # -------------------
            c.execute("""
                SELECT a.id,
                    a.legajo,
                    lb.apellido,
                    lb.nombre,
                    a.actividad,
                    a.fecha_inicio,
                    a.asignado,
                    ls.apellido,
                    ls.nombre,
                    a.firma_bombero_usuario,
                    a.firma_bombero_fecha,
                    a.firma_supervisor_usuario,
                    a.firma_supervisor_fecha,
                    a.estado 
                FROM actividades a
                LEFT JOIN legajos lb ON a.legajo = lb.legajo
                LEFT JOIN legajos ls ON a.asignado = ls.legajo
                WHERE a.anulada = 0
            """)
            rows = c.fetchall()
            conn.close()

            # -------------------
            # FILTRO BOMBERO (solo sus actividades)
            # -------------------
            if rol == "BOMBERO":
                rows = [r for r in rows if str(r[1]) == legajo]

            # -------------------
            # FILTRO SUPERVISOR
            # -------------------
            if rol in ("SUPERVISOR", "ADMIN") and cb_filtro:
                filtro = cb_filtro.get()

                if filtro == "Pendientes de bombero":
                    rows = [r for r in rows if r[9] is None]

                elif filtro == "Pendientes de supervisor":
                    rows = [
                        r for r in rows
                        if r[9] and not r[11]
                    ]

                elif filtro == "Borradores":
                    rows = [r for r in rows if r[13] == "BORRADOR"]    

                elif filtro == "Asignadas a mí":
                    rows = [r for r in rows if str(r[6]) == legajo]

            def orden_estado(r):
                # r[9] es firma_bombero_usuario
                # r[11] es firma_supervisor_usuario
                firma_b_ok = bool(r[9] and str(r[9]).strip())
                firma_s_ok = bool(r[11] and str(r[11]).strip())

                # 1. PENDIENTE DE SUPERVISOR (Bombero ya firmó, falta el jefe)
                if firma_b_ok and not firma_s_ok:
                    return 0
                
                # 2. PENDIENTE DE BOMBERO (Todavía es borrador)
                elif not firma_b_ok:
                    return 1
                
                # 3. FIRMADAS POR AMBOS (Cerradas)
                else:
                    return 2

            # Aplicar el ordenamiento a la lista de filas
            rows = sorted(rows, key=orden_estado)

            # 🔴 SOLO ordenar si el filtro es "Todas"
            if not cb_filtro or cb_filtro.get() == "Todas":
                rows = sorted(rows, key=orden_estado)

            self.rows_actuales = rows
            lbl_contador.config(text=f"Total: {len(rows)}")
            # -------------------
            # LIMPIAR TREE
            # -------------------
            for i in tree.get_children():
                tree.delete(i)

            # -------------------
            # INSERTAR
            # -------------------
            for r in rows:

                estado = r[13]  # 🔴 AGREGAR ACÁ

                bombero_txt = f"{r[1]} - {(r[2] or '').upper()}"
                supervisor_txt = ""

                if r[6]:
                    supervisor_txt = f"{r[6]} - {(r[7] or '').upper()}"

                firma_b_usuario = r[9]
                firma_s_usuario = r[11]

                firma_b_ok = bool(firma_b_usuario and str(firma_b_usuario).strip())
                firma_s_ok = bool(firma_s_usuario and str(firma_s_usuario).strip())

                firma_b = "❌" if not firma_b_ok else "✅"
                firma_s = "❌" if not firma_s_ok else "✅"

                if not firma_b_ok:
                    estado_txt = "🔵 BORRADOR"
                elif not firma_s_ok:
                    estado_txt = "🟡 PENDIENTE"
                else:
                    estado_txt = "🟢 CERRADA"                

                if not firma_b_ok:
                    tag = "borrador"

                elif firma_b_ok and not firma_s_ok:
                    tag = "pend_supervisor"

                else:
                    tag = "cerrada"

                tree.insert(
                    "",
                    "end",
                    values=(
                        r[0],
                        bombero_txt,
                        r[4],
                        r[5],
                        supervisor_txt,
                        firma_b,
                        firma_s
                    ),
                    tags=(tag,)
                )

        ttk.Button(
            frm_top,
            text="Refrescar",
            command=cargar_datos
        ).pack(side="right", padx=5)
        # -------------------------------------------------
        # TREEVIEW
        # -------------------------------------------------
        columnas = (
            "id",
            "bombero",
            "actividad",
            "fecha_inicio",
            "asignado",
            "firma_bombero",
            "firma_supervisor"
        )
        tree = ttk.Treeview(win, columns=columnas, show="headings")
        tree.pack(fill="both", expand=True, padx=10, pady=5)

        # -------------------------------------------------
        # LEYENDA DE COLORES
        # -------------------------------------------------
        frm_leyenda = Frame(win, bg="white")
        frm_leyenda.pack(fill="x", padx=10, pady=(0, 5))

        Label(
            frm_leyenda,
            text="   🔵 Borrador (sin firma bombero)   ",
            bg="#D0E7FF"
        ).pack(side="left", padx=5)

        Label(frm_leyenda, text="   Firmada por bombero (Pendiente Supervisor)   ",
            bg="#FFD43B",
            font=("Arial", 9, "bold")).pack(side="left", padx=5)

        Label(frm_leyenda, text="   Firmada y cerrada   ",
            bg="#69DB7C",
            font=("Arial", 9, "bold")).pack(side="left", padx=5)

        headings = {
            "id": "ID",
            "bombero": "Bombero",
            "actividad": "Actividad",
            "fecha_inicio": "Fecha",
            "asignado": "Supervisor",
            "firma_bombero": "F. Bombero",
            "firma_supervisor": "F. Supervisor"
        }

        tree.heading("id", text="ID")
        tree.column("id", width=60, anchor="center")

        tree.heading("bombero", text="Bombero")
        tree.column("bombero", width=190, anchor="w")

        tree.heading("actividad", text="Actividad")
        tree.column("actividad", width=280, anchor="w")

        tree.heading("fecha_inicio", text="Fecha")
        tree.column("fecha_inicio", width=110, anchor="center")

        tree.heading("asignado", text="Asignado")
        tree.column("asignado", width=220, anchor="w")

        tree.heading("firma_bombero", text="F. Bombero")
        tree.column("firma_bombero", width=100, anchor="center")

        tree.heading("firma_supervisor", text="F. Supervisor")
        tree.column("firma_supervisor", width=110, anchor="center")

        # Tags de colores
        tree.tag_configure("borrador", background="#D0E7FF")       # 🔵
        tree.tag_configure("pend_supervisor", background="#FFD43B") # 🟡
        tree.tag_configure("cerrada", background="#69DB7C")         # 🟢

        self.rows_actuales = []

        btn_exportar = ttk.Button(
            frm_top,
            text="Exportar PDF",
            command=self.exportar_pdf_pendientes
        )
        btn_exportar.pack(side="right", padx=5)
        lbl_contador.pack(side="right", padx=10)

        # Evento filtro
        if cb_filtro:
            cb_filtro.bind("<<ComboboxSelected>>", lambda e: cargar_datos())

        # Doble click abre actividad
        def abrir_actividad(event):
            sel = tree.selection()
            if not sel:
                return
            item = tree.item(sel[0])
            id_act = item["values"][0]

            cerrar_pendientes()
            self.cargar_actividad_por_id(id_act)

        def cerrar_pendientes():
            if win.winfo_exists():
                win.grab_release()
                win.destroy()

        win.protocol("WM_DELETE_WINDOW", cerrar_pendientes)

        tree.bind("<Double-1>", abrir_actividad)
        cargar_datos()

    def _mostrar_overlay_cargando(self, mensaje="Procesando..."):
        self.overlay = tk.Toplevel(self.master)

        if hasattr(self.master, "icono_global") and self.master.icono_global:
            self.overlay.iconphoto(True, self.master.icono_global)

        self.overlay.transient(self.master)
        self.overlay.grab_set()
        self.overlay.resizable(False, False)
        self.overlay.title("Procesando")
        self.overlay.configure(bg="#2c3e50")

        self.overlay.config(cursor="watch")

        self.overlay.update_idletasks()

        ancho = 320
        alto = 120

        x = self.master.winfo_rootx() + (self.master.winfo_width() // 2) - (ancho // 2)
        y = self.master.winfo_rooty() + (self.master.winfo_height() // 2) - (alto // 2)

        self.overlay.geometry(f"{ancho}x{alto}+{x}+{y}")

        self.lbl_overlay = tk.Label(
            self.overlay,
            text=mensaje,
            fg="white",
            bg="#2c3e50",
            font=("Segoe UI", 11, "bold"),
            cursor="watch"
        )
        self.lbl_overlay.pack(pady=15)

        self.lbl_anim = tk.Label(
            self.overlay,
            text="● ○ ○",
            fg="white",
            bg="#2c3e50",
            font=("Segoe UI", 16),
            cursor="watch"
        )
        self.lbl_anim.pack()

        self._anim_index = 0
        self._animar_overlay()

    def _animar_overlay(self):
        frames = [
            ("● ○ ○", "#2ecc71"),
            ("○ ● ○", "#27ae60"),
            ("○ ○ ●", "#1e8449"),
            ("○ ● ○", "#27ae60"),
        ]

        if hasattr(self, "overlay") and self.overlay.winfo_exists():
            texto, color = frames[self._anim_index]

            self.lbl_anim.config(text=texto, fg=color)
            self._anim_index = (self._anim_index + 1) % len(frames)

            self.overlay.after(250, self._animar_overlay)

    def _ocultar_overlay_cargando(self):
        if hasattr(self, "overlay") and self.overlay.winfo_exists():
            self.overlay.grab_release()
            self.overlay.destroy()

    def exportar_pdf_tabla(self, datos, headers, titulo="INFORME"):

        if not datos:
            messagebox.showwarning("Atención", "No hay datos para exportar.")
            return

        fecha_actual = datetime.now().strftime("%Y-%m-%d")
        sugerido = f"{titulo.replace(' ', '_')}_{fecha_actual}.pdf"

        file_path = self.ui.ask_save_file(
            defaultextension=".pdf",
            initialfile=sugerido,
            filetypes=[("PDF files", "*.pdf")],
            title="Guardar PDF"
        )

        if not file_path:
            return

        styles = getSampleStyleSheet()
        elems = []

        # ===== TÍTULO =====
        legajo = self.usuario_actual["legajo"]
        apellido = self.usuario_actual["apellido"].strip()
        nombre = self.usuario_actual["nombre"].strip()

        titulo_completo = f"{titulo} - LEGAJO {legajo} - {apellido} {nombre}"

        elems.append(Paragraph(f"<b>{titulo_completo}</b>", styles["Title"]))
        elems.append(Spacer(1, 12))

        # ===== TABLA =====
        data = [headers] + list(datos)

        tabla = Table(data, repeatRows=1)

        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#a50000")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))

        elems.append(tabla)

        try:
            # 🔥 TODA CREACIÓN PASA POR ACÁ
            self._crear_pdf_unificado(file_path, elems, titulo_completo)
            self.ui.show_info("Éxito", f"PDF guardado en:\n{file_path}")
        except Exception as e:
            self.ui.show_error("Error", f"No se pudo exportar el PDF:\n{e}")

    def exportar_pdf_pendientes(self):

        if not self.rows_actuales:
            messagebox.showwarning("Atención", "No hay datos.")
            return

        fecha_actual = datetime.now().strftime("%Y-%m-%d")
        file_path = self.ui.ask_save_file(
            defaultextension=".pdf",
            initialfile=f"Pendientes_{fecha_actual}.pdf",
            filetypes=[("PDF files", "*.pdf")]
        )

        if not file_path:
            return

        datos_pdf = []

        for r in self.rows_actuales:

            bombero_txt = f"{r[1]} - {(r[2] or '').upper()}"

            supervisor_txt = ""
            if r[6]:
                supervisor_txt = f"{r[6]} - {(r[7] or '').upper()}"

            firma_b = "SI" if r[9] else "NO"
            firma_s = "SI" if r[11] else "NO"

            datos_pdf.append((
                r[0],
                bombero_txt,
                r[4],
                r[5],
                supervisor_txt,
                firma_b,
                firma_s
            ))

        headers = [
            "ID",
            "Bombero",
            "Actividad",
            "Fecha",
            "Supervisor",
            "F. Bombero",
            "F. Supervisor"
        ]

        # ---------------------------------
        # CÁLCULO DE TOTALES
        # ---------------------------------
        total = len(self.rows_actuales)
        sin_bombero = 0
        pend_supervisor = 0
        cerradas = 0

        for r in self.rows_actuales:
            firma_b = r[9]
            firma_s = r[11]

            if not firma_b:
                sin_bombero += 1
            elif not firma_s:
                pend_supervisor += 1
            else:
                cerradas += 1

        resumen = [
            f"Total de registros: {total}",
            f"Sin firma de bombero: {sin_bombero}",
            f"Pendientes de supervisor: {pend_supervisor}",
            f"Cerradas: {cerradas}",
        ]
        pdf = PDFManager(
            usuario_actual=self.usuario_actual,
            logo_path="logo.png"  # ajustá ruta si hace falta
        )

        # ---------------------------------
        # TÍTULO DINÁMICO
        # ---------------------------------
        legajo = self.usuario_actual.get("legajo", "")
        apellido = self.usuario_actual.get("apellido", "").upper()
        nombre = self.usuario_actual.get("nombre", "").upper()

        titulo_secundario = (
            f"ACTIVIDADES PENDIENTES DE FIRMAR - "
            f"LEGAJO {legajo} {apellido} {nombre}"
        )

        pdf = PDFManager(
            usuario_actual=self.usuario_actual,
            logo_path="logo.png"
        )

        pdf.exportar_tabla(
            file_path,
            datos_pdf,
            headers,
            titulo=titulo_secundario,
            resumen=resumen
        )

    def actualizar_contador_pendientes(self):

        import sqlite3

        rol = self.usuario_actual["rol"].upper()
        legajo = str(self.usuario_actual.get("legajo", "")).strip()
        username = self.usuario_actual.get("username", "").strip().upper()

        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        c = conn.cursor()

        # -----------------------------
        # BOMBERO
        # -----------------------------
        if rol == "BOMBERO":
            c.execute("""
                SELECT COUNT(*)
                FROM actividades
                WHERE anulada = 0
                AND legajo = ?
                AND firma_supervisor_fecha IS NULL
            """, (legajo,))

        # -----------------------------
        # SUPERVISOR / ADMIN
        # -----------------------------
        else:
            c.execute("""
                SELECT COUNT(*)
                FROM actividades
                WHERE anulada = 0
                AND firma_supervisor_fecha IS NULL
            """)

        total = c.fetchone()[0]
        conn.close()

        if "Pendientes de Firma" in self.act_btns:
            self.act_btns["Pendientes de Firma"].config(
                text=f"Pendientes de Firma ({total})"
            )

    def _hay_pendientes_firma(self):
        import sqlite3

        try:
            conn = sqlite3.connect(DB_PATH)
            c = conn.cursor()

            rol = self.usuario_actual.get("rol", "").upper()
            legajo = str(self.usuario_actual.get("legajo", "")).strip()

            if rol == "BOMBERO":
                c.execute("""
                    SELECT 1 FROM actividades
                    WHERE legajo = ?
                    AND firma_bombero_fecha IS NULL
                    AND anulada = 0
                    LIMIT 1
                """, (legajo,))

            elif rol in ("SUPERVISOR", "ADMIN"):
                c.execute("""
                    SELECT 1 FROM actividades
                    WHERE asignado = ?
                    AND firma_bombero_fecha IS NOT NULL
                    AND firma_supervisor_fecha IS NULL
                    AND anulada = 0
                    LIMIT 1
                """, (legajo,))

            else:
                conn.close()
                return False

            result = c.fetchone()
            conn.close()

            return result is not None

        except Exception:
            return False

    def _pedir_motivo_anulacion(self):
        import tkinter as tk

        top = tk.Toplevel(self.master)
        top.title("Motivo de Anulación")
        top.geometry("400x150")
        top.grab_set()
        top.protocol("WM_DELETE_WINDOW", lambda: None)

        tk.Label(top, text="Ingrese motivo (obligatorio):").pack(pady=10)

        motivo_var = tk.StringVar()

        entry = tk.Entry(top, textvariable=motivo_var, width=50)
        entry.pack()
        entry.focus()

        # 🔥 Forzar mayúsculas mientras escribe
        entry.bind(
            "<KeyRelease>",
            lambda e: motivo_var.set(motivo_var.get().upper())
        )
        resultado = {"valor": None}

        def aceptar(event=None):
            texto = motivo_var.get().strip().upper()
            if not texto:
                return
            resultado["valor"] = texto
            top.destroy()

        def cancelar():
            top.destroy()

        entry.bind("<Return>", aceptar)

        tk.Button(top, text="Aceptar", command=aceptar).pack(side="left", padx=50, pady=20)
        tk.Button(top, text="Cancelar", command=cancelar).pack(side="right", padx=50, pady=20)

        self.master.wait_window(top)

        return resultado["valor"]

    def _bloquear_por_anulada(self):
        """
        Bloquea únicamente acciones que modifican la actividad.
        No altera botones estructurales.
        """

        # Acciones que modifican
        acciones_a_bloquear = [
            "Guardar",
            "Modificar",
            "Firmar actividad",
            "Firmar Supervisor",
            "Anular"
        ]

        for nombre in acciones_a_bloquear:
            if nombre in self.act_btns:
                self.act_btns[nombre].config(state="disabled")

    def _bloquear_texto(self, txt):
        self._set_text_readonly(txt, True)

    def _desbloquear_texto(self, txt):
        self._set_text_readonly(txt, False)

    def _crear_boton_con_borde(self, parent, texto, comando, color_borde, habilitado=True):
        cont = Frame(parent, bg=color_borde, padx=2, pady=2)

        btn = Button(
            cont,
            text=texto,
            command=comando,
            width=18,
            relief="flat",
            font=("Arial", 10, "bold" if habilitado else "normal")
        )

        btn.pack(fill="both", expand=True)

        self._aplicar_estilo_boton(btn, color_borde, habilitado)

        return cont, btn

    def ver_historial_actividad(self):

        if not getattr(self, "id_actividad_actual", None):
            self.ui.show_error("Historial", "Debe seleccionar una actividad.")
            return

        id_act = self.id_actividad_actual

        try:
            conn = sqlite3.connect(DB_PATH)
            conn.row_factory = sqlite3.Row
            c = conn.cursor()

            c.execute("""
                SELECT
                    h.fecha,
                    COALESCE(l.apellido || ' ' || l.nombre, u.username) as usuario,
                    h.campo,
                    h.valor_anterior,
                    h.valor_nuevo
                FROM actividades_historial h
                LEFT JOIN usuarios u ON u.id = h.usuario_id
                LEFT JOIN legajos l ON l.legajo = u.legajo
                WHERE h.actividad_id = ?
                ORDER BY h.fecha
            """, (id_act,))

            rows = c.fetchall()
            tiene_historial = len(rows) > 0
            conn.close()

        except Exception as e:
            self.ui.show_error("Error", f"No se pudo cargar el historial:\n{e}")
            return

        # -------------------------------------------------
        # Ventana
        # -------------------------------------------------
        win = Toplevel(self.master)
        win.title(f"Historial de actividad N° {id_act}")
        win.geometry("800x400")
        win.configure(bg="white")

        frame = Frame(win, bg="white")
        frame.pack(fill=BOTH, expand=True, padx=10, pady=10)

        cols = ("fecha", "usuario", "campo", "anterior", "nuevo")

        tree = ttk.Treeview(frame, columns=cols, show="headings")

        tree.heading("fecha", text="Fecha")
        tree.heading("usuario", text="Usuario")
        tree.heading("campo", text="Campo")
        tree.heading("anterior", text="Valor anterior")
        tree.heading("nuevo", text="Valor nuevo")

        tree.column("fecha", width=140)
        tree.column("usuario", width=100)
        tree.column("campo", width=120)
        tree.column("anterior", width=200)
        tree.column("nuevo", width=200)

        tree.tag_configure("horas", background="#D9EDF7")       # celeste
        tree.tag_configure("actividad", background="#FCF8E3")   # amarillo
        tree.tag_configure("asignacion", background="#FBE5D6")  # naranja
        tree.tag_configure("descripcion", background="#EEEEEE") # gris
        tree.tag_configure("firmas", background="#F8D7DA")      # rojo claro

        # Scroll
        scroll = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scroll.set)

        scroll.pack(side=RIGHT, fill=Y)
        tree.pack(side=LEFT, fill=BOTH, expand=True)

        # Cargar datos con colores según el campo modificado
        for r in rows:
            fecha, usuario, campo, anterior, nuevo = r
            campo_txt = str(campo).lower()

            # Detectar tipo de cambio
            if "hora" in campo_txt or "fecha" in campo_txt:
                tag = "horas"
            elif "actividad" in campo_txt or "area" in campo_txt:
                tag = "actividad"
            elif "legajo" in campo_txt or "asignado" in campo_txt:
                tag = "asignacion"
            elif "descripcion" in campo_txt:
                tag = "descripcion"
            elif "firma" in campo_txt:
                tag = "firmas"

                # Texto especial para firmas anuladas
                anterior = "Firmada"
                nuevo = "ANULADA por modificación"

            else:
                tag = ""

            anterior_fmt = self._formatear_valor_historial(campo, anterior)
            nuevo_fmt = self._formatear_valor_historial(campo, nuevo)

            tree.insert(
                "",
                "end",
                values=(fecha, usuario, campo, anterior_fmt, nuevo_fmt),
                tags=(tag,)
            )

        # Botón Exportar PDF
        Button(
            win,
            text="Exportar historial a PDF",
            command=lambda: self.exportar_historial_pdf(id_act),
            bg="#1E5AA8",
            fg="white",
            font=("Arial", 10, "bold")
        ).pack(pady=10)

        # Botón cerrar
        Button(
            win,
            text="Cerrar",
            command=win.destroy,
            width=12
        ).pack(pady=5)

    def _formatear_valor_historial(self, campo, valor):
        if valor is None:
            return ""

        valor = str(valor).strip()

        # Si ya viene completo → no tocar
        if " - " in valor:
            return valor

        # Si es actividad o concepto → buscar nombre
        if campo in ("actividad", "concepto_id") and valor.isdigit():
            try:
                conn = sqlite3.connect(DB_PATH)
                c = conn.cursor()

                c.execute("""
                    SELECT descripcion
                    FROM conceptos
                    WHERE id = ?
                """, (valor,))

                row = c.fetchone()
                conn.close()

                if row:
                    return f"{valor} - {row[0]}"
            except:
                pass

        return valor

    def exportar_historial_pdf(self, id_act):

        file = self.ui.ask_save_file(
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
            title="Guardar historial como PDF",
            initialfile=f"Historial_Actividad_{id_act}.pdf"
        )
        if not file:
            return

        try:
            conn = sqlite3.connect(DB_PATH)
            c = conn.cursor()
            c.execute("""
                SELECT fecha, usuario_id, campo, valor_anterior, valor_nuevo
                FROM actividades_historial
                WHERE actividad_id = ?
                ORDER BY fecha
            """, (id_act,))
            rows = c.fetchall()
            conn.close()
        except Exception as e:
            self.ui.show_error("Error", str(e))
            return

        # Evitar PDF vacío
        if not rows:
            messagebox.showwarning("Historial", "La actividad no tiene historial.")
            return

        styles = getSampleStyleSheet()

        title = Paragraph(
            f"<b>Historial de la Actividad N° {id_act}</b>",
            styles["Title"]
        )

        data = [["Fecha", "Usuario", "Campo", "Antes", "Después"]]
        data.extend(rows)

        table = Table(data, colWidths=[100, 60, 120, 120, 120])

        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#a50000")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("ALIGN", (1, 1), (-1, -1), "LEFT"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]))

        elements = [title, Spacer(1, 15), table]

        try:
            self._crear_pdf_unificado(
                file,
                elements,
                titulo=f"Historial Actividad N° {id_act}"
            )
            self.ui.show_info("OK", "Historial exportado a PDF.")
        except Exception as e:
            self.ui.show_error("Error", f"No se pudo exportar el PDF:\n{e}")

    def _aplicar_estilo_boton(self, btn, habilitado, tipo="normal"):

        if not habilitado:
            btn.config(
                state="disabled",
                bg="#DDDDDD",
                fg="#888888",
                activebackground="#DDDDDD"
            )
            return

        if tipo == "positivo":
            bg = "#2E7D32"
            active = "#1B5E20"

        elif tipo == "peligro":
            bg = "#C62828"
            active = "#8E0000"

        else:
            bg = "#1565C0"
            active = "#0D47A1"

        btn.config(
            state="normal",
            bg=bg,
            fg="white",
            activebackground=active
        )

    def _set_btn_estado(self, nombre, habilitado=True, tipo="normal"):
        btn = self.act_btns.get(nombre)
        if not btn:
            return

        # 🔥 SIEMPRE aplicar estilo completo
        self._aplicar_estilo_boton(btn, habilitado, tipo)

    def no_firmar_ahora(self):
        self._estado_ui = "cargado"
        self._refrescar_actividades()

    def imprimir_actividad(self):
        import os
        import tempfile
#        from tkinter.messagebox import showerror

        id_act = self.var_id_actividad.get().strip()
        if not id_act:
            self.ui.show_error("Error", "Debe cargar primero una actividad.")
            return

        try:
            tmp_dir = tempfile.gettempdir()
            tmp_path = os.path.join(tmp_dir, f"Actividad_{id_act}.pdf")

            self._crear_pdf_actividad(tmp_path)

            # 👀 Abrir en visor PDF (NO imprime)
            os.startfile(tmp_path)

        except Exception as e:
            self.ui.show_error("Error", f"No se pudo generar la vista previa:\n{e}")

    def _generar_pdf_bytes_actividad(self, id_act):
        import tempfile, os

        tmp_dir = tempfile.gettempdir()
        tmp_path = os.path.join(tmp_dir, f"Actividad_{id_act}_tmp.pdf")

        # Usamos tu PDF real
        self._crear_pdf_actividad(tmp_path)

        # Leemos los bytes
        with open(tmp_path, "rb") as f:
            pdf_bytes = f.read()

        return pdf_bytes

    def obtener_estado_firma_texto(self):
        """
        Devuelve un texto corto y consistente sobre el estado de firma
        para usar en UI, mails e informes.
        """
        if self.firma_supervisor_fecha:
            return "ACTIVIDAD FIRMADA POR SUPERVISOR"

        if self.firma_bombero_fecha:
            return "ACTIVIDAD FIRMADA POR BOMBERO (PENDIENTE DE SUPERVISIÓN)"

        return "ACTIVIDAD SIN FIRMAR"

    def _actualizar_estado_firma(self):
        rol = self.usuario_actual.get("rol", "").upper()

        if not getattr(self, "id_actividad_actual", None):
            self._set_estado_visual("", fg="black", bg=self.actividades_frame["bg"])
            return

        if getattr(self, "actividad_anulada", False):
            self._bloquear_por_anulada()
            return

        if getattr(self, "modo_actividad", None) in ("nuevo", "editando"):
            self._set_estado_visual(
                "✏️ EDICIÓN EN CURSO",
                fg="#003d80",
                bg="#e6f0ff"
            )
            return

        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        c = conn.cursor()
        c.execute("""
            SELECT firma_bombero_fecha,
                firma_supervisor_fecha,
                asignado,
                legajo
            FROM actividades
            WHERE id = ?
        """, (self.id_actividad_actual,))
        row = c.fetchone()
        conn.close()

        if not row:
            self._set_estado_visual("", fg="black", bg=self.actividades_frame["bg"])
            return

        firma_bombero_fecha, firma_supervisor_fecha, supervisor_legajo, legajo_actividad = row

        self.firma_bombero_fecha = firma_bombero_fecha
        self.firma_supervisor_fecha = firma_supervisor_fecha
        self.actividad_asignado = supervisor_legajo
        self.actividad_legajo = legajo_actividad

        firmado_bombero = bool(firma_bombero_fecha)
        firmado_supervisor = bool(firma_supervisor_fecha)

        if firmado_supervisor:
            self._set_estado_visual(
                "🟢 ACTIVIDAD FIRMADA POR EL SUPERVISOR",
                fg="#0a4d1a",
                bg="#e6ffe6"
            )
            return

        if firmado_bombero:
            self._set_estado_visual(
                "🟡 PENDIENTE DE SUPERVISIÓN",
                fg="#7a5a00",
                bg="#fff6cc"
            )
            return

        self._set_estado_visual(
            "🔴 ACTIVIDAD SIN FIRMAR",
            fg="#7a0000",
            bg="#ffe5e5"
        )
        
    def _get_df_actividades(
        self,
        legajo=None,
        fecha_desde=None,
        fecha_hasta=None
    ):
        """
        Fuente única de verdad para actividades.

        Params:
            legajo: int | None
            fecha_desde: date | None
            fecha_hasta: date | None

        Returns:
            pandas.DataFrame
        """
        import sqlite3
        import pandas as pd

        query = """
            SELECT
                a.id,
                a.legajo,
                COALESCE(l.apellido, '') AS apellido,
                COALESCE(l.nombre, '') AS nombre,
                a.actividad,
                a.area,
                a.fecha_inicio,
                a.fecha_fin,
                a.hora_inicio,
                a.hora_fin,
                a.horas,
                a.descripcion,
                a.asignado
            FROM actividades a
            LEFT JOIN legajos l ON a.legajo = l.legajo
            WHERE 1=1
        """

        params = []

        # ---- filtros dinámicos ----
        if legajo is not None:
            query += " AND a.legajo = ?"
            params.append(legajo)

        if fecha_desde is not None and fecha_hasta is not None:
            query += " AND a.fecha_inicio BETWEEN ? AND ?"
            params.extend([fecha_desde, fecha_hasta])

        query += " ORDER BY a.fecha_inicio DESC, a.id DESC"

        # ---- ejecutar ----
        conn = sqlite3.connect(DB_PATH)
        try:
            df = pd.read_sql_query(query, conn, params=params)
        finally:
            conn.close()

        return df

    def _notificar_evento_actividad(self, id_act, evento):
        """
        Centraliza las notificaciones formales del sistema.
        evento: "firma_bombero" | "firma_supervisor"
        """

        parent = self.master if hasattr(self, "master") else None

        try:
            conn = sqlite3.connect(DB_PATH)
            c = conn.cursor()

            c.execute("""
                SELECT legajo, asignado, actividad, area,
                    fecha_inicio, hora_inicio, hora_fin,
                    descripcion, horas
                FROM actividades
                WHERE id = ?
            """, (id_act,))

            row = c.fetchone()
            conn.close()

            if not row:
                messagebox.showwarning(
                    "Notificación",
                    "No se encontró la actividad para notificar.",
                    parent=parent
                )
                return

            (
                legajo,
                asig,
                actividad_txt,
                area_val,
                fecha_txt,
                hi,
                hf,
                descripcion_txt,
                horas_trabajadas
            ) = row

        except Exception as e:
            print("Error leyendo actividad para notificar:", e)
            self.ui.show_error(
                "Error",
                f"No se pudo leer la actividad para notificar:\n{e}",
                parent=parent
            )
            return

        print("TIPO EVENTO MAIL:", tipo_evento)
        # -------- Envío de correos --------
        try:
            ok = self._enviar_correos_actividad(
                id_act=id_act,
                legajo=legajo,
                asig=asig,
                actividad_txt=actividad_txt,
                descripcion_txt=descripcion_txt,
                area_val=area_val,
                hi=hi,
                hf=hf,
                horas_trabajadas=horas_trabajadas,
                fecha_txt=fecha_txt,
                tipo_evento=evento 
            )

            # Si la función devuelve False → fallo
            if ok is False:
                messagebox.showwarning(
                    "Notificación",
                    "No se pudo enviar el correo de notificación.",
                    parent=parent
                )
            else:
                self.ui.show_info(
                    "Notificación",
                    "Correo de notificación enviado correctamente.",
                    parent=parent
                )

        except Exception as e:
            self.ui.show_error(
                "Error de notificación",
                f"No se pudo enviar el correo:\n{e}",
                parent=parent
            )

    def _enviar_correos_actividad(
            self,
            id_act, legajo, asig,
            actividad_txt, descripcion_txt,
            area_val, hi, hf,
            horas_trabajadas, fecha_txt,
            tipo_evento=None
        ):
            print("🔥 PROCESANDO ENVÍO DE MAIL - EVENTO:", tipo_evento)
            enviados = []
            
            try:
                conn = sqlite3.connect(DB_PATH)
                c = conn.cursor()

                # 1. OBTENER DATOS DEL BOMBERO
                c.execute("SELECT apellido, nombre, email FROM legajos WHERE legajo = ?", (legajo,))
                row_bombero = c.fetchone()
                if not row_bombero:
                    print("⚠️ No se encontró el bombero.")
                    return False, []
                
                nombre_bombero = f"{row_bombero[1].strip().title()} {row_bombero[0].strip().title()}"

                # 2. OBTENER ESTADO Y FIRMAS DE LA ACTIVIDAD
                c.execute("""
                    SELECT firma_bombero_fecha, firma_supervisor_fecha,
                        usuario_anula, fecha_anulacion, motivo_anulacion
                    FROM actividades WHERE id = ?
                """, (id_act,))
                row_act = c.fetchone()
                if not row_act: return False, []
                
                firma_bombero, firma_supervisor, usuario_anula, fecha_anulacion, motivo_anulacion = row_act
                estado_real = "FIRMADA_SUPERVISOR" if firma_supervisor else ("FIRMADA_BOMBERO" if firma_bombero else "BORRADOR")

                # 3. IDENTIFICAR MODIFICADOR (Usuario logueado)
                modificador_txt = f"{self.usuario_actual.get('nombre', '').title()} {self.usuario_actual.get('apellido', '').title()}"

                # 4. PREPARAR LISTA DE DESTINATARIOS (Personas)
                nombre_bombero = f"{row_bombero[1].strip()} {row_bombero[0].strip()}"
                modificador_txt = f"{self.usuario_actual.get('nombre', '').strip()} {self.usuario_actual.get('apellido', '').strip()}"

                # --- LÓGICA DE DESTINATARIOS FILTRADA ---
                personas = []

                
                # El BOMBERO siempre recibe (Anulación, Modificación y Firmas)
                if row_bombero[2]:
                    personas.append({
                        "email": row_bombero[2], 
                        "rol": "BOMBERO", 
                        "nombre": row_bombero[1].strip(), 
                        "apellido": row_bombero[0].strip()
                    })
                
                # El SUPERVISOR NO recibe si es Anulación o Modificación (pedido del usuario)
                # Solo recibe en registros nuevos o firmas normales
                if asig and tipo_evento not in ["ANULADA", "MODIFICADA"]:
                    c.execute("SELECT apellido, nombre, email FROM legajos WHERE legajo = ?", (asig,))
                    row_sup = c.fetchone()
                    if row_sup and row_sup[2]:
                        personas.append({
                            "email": row_sup[2], 
                            "rol": "SUPERVISOR", 
                            "nombre": row_sup[1].strip(), 
                            "apellido": row_sup[0].strip()
                        })

                # 5. GENERAR PDF (Una sola vez)
                try:
                    pdf_bytes = self._generar_pdf_bytes_actividad(id_act)
                    nombre_pdf = f"Actividad_{id_act}.pdf"
                except:
                    pdf_bytes = None

                # --- 6. TRADUCCIÓN DE ESTADOS PARA EL ASUNTO ---
                traducciones = {
                    "BORRADOR": "Nueva Actividad",
                    "FIRMADA_BOMBERO": "Pendiente de Supervisión",
                    "FIRMADA_SUPERVISOR": "Actividad Finalizada",
                    "MODIFICADA": "Modificación de Datos",
                    "ANULADA": "Actividad Anulada",
                    "FIRMA_BOMBERO": "Firma de Bombero Registrada",
                    "FIRMA_SUPERVISOR": "Actividad Cerrada y Firmada"
                }
                
                # Determinamos la etiqueta del asunto
                etiqueta = tipo_evento if tipo_evento else estado_real
                texto_asunto = traducciones.get(etiqueta, etiqueta).upper()

                # ASUNTO DEL MAIL
                asunto = f"SIAB - Actividad Nº {id_act} - {texto_asunto}"
                actividad_limpia = actividad_txt.split(" - ", 1)[1] if " - " in actividad_txt and actividad_txt.split(" - ", 1)[0].isdigit() else actividad_txt

                # ----------------------------------------------------------
                # 7. ENVÍO DE CORREOS
                # ----------------------------------------------------------
                with smtplib.SMTP_SSL(self.smtp_server, self.smtp_port) as smtp:
                    smtp.login(self.smtp_user, self.smtp_pass)

                    for p in personas:
                        # --- INICIALIZACIÓN Y SEGURIDAD ---
                        texto_base = "Notificación de sistema SIAB"
                        mensaje_estado = ""
                        detalles_auditoria = ""
                        estilo_banner = "background: #f8f9fa; border-left: 4px solid #6c757d; color: #333;"
                        ahora_f = datetime.now().strftime('%d/%m/%Y %H:%M')

                        # --- LÓGICA DE EVENTOS (QUIÉN, CUÁNDO, MOTIVO) ---
                        if tipo_evento == "MODIFICADA":
                            texto_base = "Tu actividad ha sido modificada administrativamente."
                            mensaje_estado = "⚠️ FIRMAS ANULADAS - REQUIERE FIRMA NUEVAMENTE"
                            detalles_auditoria = f"<b>Modificado por:</b> {modificador_txt}<br><b>Fecha:</b> {ahora_f}"
                            estilo_banner = "background: #fff3cd; border-left: 4px solid #ffc107; color: #856404;"

                        elif tipo_evento == "ANULADA":
                            texto_base = "La actividad ha sido anulada y ya no es válida."
                            mensaje_estado = "🚫 ACTIVIDAD ANULADA"
                            detalles_auditoria = f"<b>Anulada por:</b> {usuario_anula}<br><b>Fecha:</b> {fecha_anulacion}<br><b>Motivo:</b> {motivo_anulacion}"
                            estilo_banner = "background: #f8d7da; border-left: 4px solid #dc3545; color: #721c24;"

                        else:
                            # FLUJO NORMAL DE FIRMAS
                            if p["rol"] == "SUPERVISOR":
                                # Si el supervisor aún no firmó
                                if not firma_supervisor:
                                    texto_base = "Nueva actividad pendiente de su firma."
                                    mensaje_estado = "📋 ACTIVIDAD PENDIENTE DE SU FIRMA"
                                else:
                                    texto_base = "La actividad ha sido cerrada correctamente."
                                    mensaje_estado = "✅ FIRMADA POR USTED"
                            else:
                                # ROL BOMBERO
                                if not firma_supervisor:
                                    texto_base = "Tu actividad ha sido registrada y enviada a control."
                                    mensaje_estado = "⏳ ESPERANDO FIRMA DE SUPERVISOR"
                                else:
                                    texto_base = "¡Excelente! Tu actividad ha sido aprobada y finalizada."
                                    mensaje_estado = "✅ ACTIVIDAD APROBADA"

                        # --- HTML COMPACTO ---
                        nombre_formateado = f"{p['nombre']} {p['apellido']}".title()
                        cuerpo_html = f"""
                        <html>
                        <body style="font-family: Arial, sans-serif; margin:0; padding:10px;">
                            <div style="max-width: 450px; border: 1px solid #eee; border-radius: 6px; overflow: hidden; margin:auto;">
                                <div style="background: #b30000; color: white; padding: 10px; text-align: center; font-size: 15px;">
                                    <strong>SIAB - Actividad Nº {id_act}</strong>
                                </div>
                                <div style="padding: 15px;">
                                    <p style="margin: 0 0 10px 0; font-size: 14px;">Hola <b>{nombre_formateado}</b>,</p>
                                    <p style="margin: 0 0 10px 0; font-size: 13px; color: #444; line-height: 1.5;">{texto_base}</p>
                                    <div style="{estilo_banner} padding: 12px; font-size: 12px; margin-bottom: 15px;">
                                        <b style="font-size: 13px;">{mensaje_estado}</b><br>
                                        <div style="margin-top: 6px; padding-top: 6px; border-top: 1px dotted rgba(0,0,0,0.1); line-height: 1.4;">
                                            {detalles_auditoria}
                                        </div>
                                    </div>
                                    <table style="width: 100%; font-size: 12px; border-collapse: collapse;">
                                        <tr><td style="padding: 4px;"><b>Actividad:</b></td><td>{actividad_limpia}</td></tr>
                                        <tr><td style="padding: 4px;"><b>Fecha:</b></td><td>{fecha_txt} ({hi} a {hf})</td></tr>
                                    </table>
                                    <p style="font-size: 11px; color: #888; margin-top: 15px;"><b>Descripción:</b> {descripcion_txt[:100]}...</p>
                                </div>
                            </div>
                        </body>
                        </html>
                        """

                        # OBJETO EMAIL
                        msg = EmailMessage()
                        msg["Subject"] = asunto
                        msg["From"] = self.smtp_user
                        msg["To"] = p["email"]
                        msg.set_content(mensaje_estado)
                        msg.add_alternative(cuerpo_html, subtype="html")

                        if pdf_bytes:
                            msg.add_attachment(pdf_bytes, maintype="application", subtype="pdf", filename=nombre_pdf)

                        # INTENTO DE ENVÍO Y REGISTRO
                        try:
                            smtp.send_message(msg)
                            enviados.append(p["email"])
                            # Registrar OK en DB
                            c.execute("INSERT INTO notificaciones (actividad_id, tipo, destinatario, asunto, fecha_envio, estado) VALUES (?, ?, ?, ?, ?, ?)",
                                    (id_act, tipo_evento or estado_real, p["email"], asunto, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "OK"))
                        except Exception as err_mail:
                            print(f"❌ Error enviando a {p['email']}: {err_mail}")
                            c.execute("INSERT INTO notificaciones (actividad_id, tipo, destinatario, asunto, fecha_envio, estado) VALUES (?, ?, ?, ?, ?, ?)",
                                    (id_act, tipo_evento or estado_real, p["email"], asunto, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), f"ERROR: {str(err_mail)}"))
                
                conn.commit()
                conn.close()
                return True, enviados

            except Exception as e:
                print("❌ ERROR GENERAL EN MAIL:", e)
                return False, []
    
    def _formatear_nombre(self, apellido, nombre):
        return f"{nombre.strip().title()} {apellido.strip().title()}"

    def firmar_actividad_bombero(self):
        rol = self.usuario_actual.get("rol", "").upper()
        legajo_usuario = str(self.usuario_actual.get("legajo", "")).strip()
        legajo_actividad = str(getattr(self, "actividad_legajo", "")).strip()

        if not legajo_actividad:
            self.ui.show_error("Error", "No se pudo determinar el titular de la actividad.")
            return

        # Puede firmar:
        # - Bombero titular
        # - Supervisor si la actividad es propia
        es_propia = legajo_usuario == legajo_actividad

        if not (
            es_propia and rol in ("BOMBERO", "SUPERVISOR", "ADMIN")
        ):
            showwarning(
                "Acceso no permitido",
                "Solo el titular de la actividad puede firmarla."
            )
            return
            messagebox.showwarning(
                "Acceso no permitido",
                "Solo el titular de la actividad puede firmarla."
            )
            return

        if not self.id_actividad_actual:
            self._actualizar_estado_firma()
            return

        if getattr(self, "_estado_ui", None) == "editando":
            messagebox.showwarning(
                "Debe guardar primero",
                "Está modificando la actividad.\n\nDebe GUARDAR antes de poder firmarla."
            )
            self._actualizar_estado_firma()
            return

        from datetime import datetime
        import sqlite3
        from tkinter.messagebox import askyesno, showerror
        import threading

        confirmar = messagebox.askyesno(
            "Confirmar firma",
            "¿Confirma que los datos de la actividad son correctos?\n\n"
            "Una vez firmada, no podrá modificarla."
        )
        if not confirmar:
            self._actualizar_estado_firma()
            return

        self._procesando_firma = True
        self._bloqueo_fuerte_actividades(True)
        self._mostrar_progreso_firma("⏳ Firmando actividad...")

        fecha_firma = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        usuario = self.usuario_actual.get("username")

        if getattr(self, "actividad_anulada", False):
            self._procesando_firma = False
            self._ocultar_progreso_firma()
            self._bloqueo_fuerte_actividades(False)
            self.ui.show_error("Error", "No se puede firmar una actividad anulada.")
            return

        try:
            conn = sqlite3.connect(DB_PATH)
            c = conn.cursor()

            # Traer datos necesarios en una sola consulta
            c.execute("""
                SELECT legajo, firma_bombero_fecha, anulada
                FROM actividades
                WHERE id = ?
            """, (self.id_actividad_actual,))

            row = c.fetchone()

            if not row:
                raise Exception("No se pudo verificar la actividad.")

            legajo_creador, firma_existente, anulada = row

            legajo_actual = str(self.usuario_actual.get("legajo", "")).strip()

            # 🔒 VALIDACIÓN REAL (BD manda)
            if str(legajo_creador).strip() != legajo_actual:
                raise PermissionError(
                    "Solo el titular de la actividad puede firmarla."
                )
            legajo_sel = ""

            if self.cb_legajo_sel.get():
                legajo_sel = self.cb_legajo_sel.get().split(" - ")[0].strip()

            # 🔍 DEBUG CLAVE
            print("===================================")
            print("ROL:", rol)
            print("LEGAJO ACTUAL:", legajo_actual)
            print("LEGAJO SELECCIONADO:", legajo_sel)
            print("ES PROPIA?:", legajo_sel == legajo_actual)
            print("VALOR COMBO LEGAJO:", self.cb_legajo_sel.get())
            print("===================================")

            # Ya firmada
            if firma_existente:
                raise Exception("La actividad ya se encuentra firmada.")

            # Anulada
            if anulada:
                raise Exception("No se puede firmar una actividad anulada.")

            # UPDATE
            c.execute("""
                UPDATE actividades
                SET firma_bombero_usuario = ?, firma_bombero_fecha = ?
                WHERE id = ?
            """, (usuario, fecha_firma, self.id_actividad_actual))

            conn.commit()
            conn.close()

            self.firma_bombero_fecha = fecha_firma

            threading.Thread(
                target=self._worker_notificacion_firma,
                args=(self.id_actividad_actual, "firma_bombero"),
                daemon=True
            ).start()

            self._estado_ui = "cargado"

            # 🔥 RECARGAR DESDE BD POR ID
            self.cargar_actividad_por_id(self.id_actividad_actual)

            # Refrescar UI
            self._refrescar_actividades()

        except PermissionError as pe:
            self._procesando_firma = False
            self._ocultar_progreso_firma()
            self._bloqueo_fuerte_actividades(False)
            self.ui.show_error("Acceso no permitido", str(pe))

        except Exception as e:
            self._procesando_firma = False
            self._ocultar_progreso_firma()
            self._bloqueo_fuerte_actividades(False)
            self.ui.show_error("Error", f"No se pudo firmar la actividad:\n{e}")

        self.actualizar_contador_pendientes()

    def firmar_actividad_supervisor(self):
        if not self._puede_firmar_supervisor():
            messagebox.showwarning(
                "Acceso no permitido",
                "Solo el supervisor asignado puede aprobar esta actividad."
            )
            return

        if not self.id_actividad_actual:
            self._actualizar_estado_firma()
            return

        if getattr(self, "_estado_ui", None) == "editando":
            messagebox.showwarning(
                "Debe guardar primero",
                "Está modificando la actividad.\n\nDebe GUARDAR antes de aprobarla."
            )
            self._actualizar_estado_firma()
            return

        from datetime import datetime
        import sqlite3
        from tkinter.messagebox import askyesno, showerror
        import threading

        confirmar = messagebox.askyesno(
            "Confirmar aprobación",
            "¿Aprueba definitivamente esta actividad?\n\n"
            "Una vez firmada no podrá modificarse."
        )
        if not confirmar:
            self._actualizar_estado_firma()
            return

        self._procesando_firma = True
        self._bloqueo_fuerte_actividades(True)
        self._mostrar_progreso_firma("⏳ Aprobando actividad...")

        fecha_firma = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        usuario = self.usuario_actual.get("username")

        try:
            conn = sqlite3.connect(DB_PATH)
            c = conn.cursor()

            # Traer datos necesarios en una sola consulta
            c.execute("""
                SELECT firma_bombero_fecha,
                    firma_supervisor_fecha,
                    anulada,
                    asignado
                FROM actividades
                WHERE id = ?
            """, (self.id_actividad_actual,))

            row = c.fetchone()

            if not row:
                raise Exception("No se pudo verificar la actividad.")

            firma_bombero_fecha, firma_supervisor_fecha, anulada, asignado = row

            # 🔒 No puede aprobar si está anulada
            if anulada:
                raise Exception("No se puede aprobar una actividad anulada.")

            # 🔒 No puede aprobar si no está firmada por bombero
            if not firma_bombero_fecha:
                raise Exception("La actividad debe estar firmada por el bombero antes de ser firmada por el supervisor.")

            # 🔒 Ya firmada por el supervisor
            if firma_supervisor_fecha:
                raise Exception("La actividad ya se encuentra firmada por el supervisor.")

            # 🔒 Solo puede aprobar si está asignado como supervisor
            legajo_actual = str(self.usuario_actual.get("legajo", "")).strip()

            # UPDATE aprobación
            c.execute("""
                UPDATE actividades
                SET firma_supervisor_usuario = ?, firma_supervisor_fecha = ?
                WHERE id = ?
            """, (usuario, fecha_firma, self.id_actividad_actual))

            conn.commit()
            conn.close()

            # 🔥 RECARGAR DESDE DB
            self.cargar_actividad_por_id(self.id_actividad_actual)

            threading.Thread(
                target=self._worker_notificacion_firma,
                args=(self.id_actividad_actual, "firma_supervisor"),
                daemon=True
            ).start()

        except PermissionError as pe:
            self._procesando_firma = False
            self._ocultar_progreso_firma()
            self._bloqueo_fuerte_actividades(False)
            self.ui.show_error("Acceso no permitido", str(pe))

        except Exception as e:
            self._procesando_firma = False
            self._ocultar_progreso_firma()
            self._bloqueo_fuerte_actividades(False)
            self.ui.show_error("Error", f"No se pudo aprobar la actividad:\n{e}")

        self.actualizar_contador_pendientes()

    def _bloquear_todos_botones_actividades(self):
        for b in self.act_btns.values():
            self._aplicar_estilo_boton(b, False)

    def _bloqueo_fuerte_actividades(self, bloquear=True):

        for nombre, b in self.act_btns.items():

            # 🔹 Nunca bloquear botones de navegación
            if nombre in {"limpiar", "cancelar", "nuevo", "buscar"}:
                continue

            if bloquear:
                b.config(state="disabled")
            else:
                # No restauramos aquí, eso lo maneja
                # _set_estado_botones_actividades()
                pass
        
    def _worker_notificacion_firma(self, id_actividad, tipo_evento):
        """
        Ejecuta el envío de notificaciones en segundo plano.
        """
        try:
            ok, destinatarios = self._notificar_evento_actividad(
                id_actividad,
                tipo_evento
            )

        except Exception as e:
            ok = False
            destinatarios = []
            error = str(e)
            self.master.after(
                0,
                lambda: self._finalizar_firma(False, error, [])
            )
            return

        # Volver al hilo principal para actualizar la interfaz
        self.master.after(
            0,
            lambda: self._finalizar_firma(ok, tipo_evento, destinatarios)
        )

    def _finalizar_firma(self, ok, tipo_evento, destinatarios):

        self._procesando_firma = False
        self._ocultar_progreso_firma()
        self._actualizar_estado_firma()
        self._estado_ui = "cargado"
        self._refrescar_actividades()

        if not ok:
            showerror(
                "Error",
                "No se pudo enviar la notificación por correo."
            )
            return

        # Mensaje según tipo de firma
        if tipo_evento == "firma_bombero":
            titulo = "Actividad firmada"
            mensaje = "La actividad fue firmada correctamente."
        elif tipo_evento == "firma_supervisor":
            titulo = "Actividad firmada por el supervisor"
            mensaje = "La actividad fue firmada por el supervisor correctamente."
        else:
            titulo = "Actividad actualizada"
            mensaje = "La actividad fue actualizada."

        # Agregar info de correos
        if destinatarios:
            mensaje += "\n\nNotificación enviada a:\n"
            mensaje += "\n".join(destinatarios)

        self.ui.show_info(titulo, mensaje)

    def _notificar_evento_actividad(self, id_actividad, tipo_evento):
        import sqlite3

        conn = None

        try:
            conn = sqlite3.connect(DB_PATH)
            c = conn.cursor()

            c.execute("""
                SELECT id, fecha_inicio, descripcion, legajo
                FROM actividades
                WHERE id = ?
            """, (id_actividad,))

            row = c.fetchone()

            if not row:
                return False, []

            id_act, fecha, descripcion, legajo = row

            # -------------------------
            # Datos UI
            # -------------------------
            try:
                asig = self.cb_asignado.get().split(" - ")[0]
            except:
                asig = None

            actividad_txt = self.actividad.get()
            descripcion_txt = self.descripcion.get("1.0", "end").strip()
            area_val = self.area.get()
            hi = self.hora_inicio.get()
            hf = self.hora_finalizacion.get()
            horas_trabajadas = self.var_horas.get()
            fecha_txt = self.fecha_inicio.get()

            # 🔥 LLAMADA
            ok, destinatarios = self._enviar_correos_actividad(
                id_act,
                legajo,
                asig,
                actividad_txt,
                descripcion_txt,
                area_val,
                hi,
                hf,
                horas_trabajadas,
                fecha_txt,
                tipo_evento=tipo_evento 
            )

            return ok, destinatarios

        except Exception as e:
            print("ERROR en notificación:", e)
            return False, []

        finally:
            if conn:
                conn.close()

    def _tiene_pendientes_firma(self):
        """Consulta si existen registros que requieran la firma del usuario actual"""
        try:
            rol_actual = self.usuario_actual.get("rol", "").upper()
            legajo_actual = str(self.usuario_actual.get("legajo", "")).strip()
            
            conn = sqlite3.connect(DB_PATH)
            c = conn.cursor()
            
            # --- CORRECCIÓN PUNTO 3: Nombres de columnas reales ---
            if rol_actual == "BOMBERO":
                # El Bombero firma en 'firma_bombero_fecha'
                query = """
                    SELECT COUNT(*) FROM actividades 
                    WHERE legajo = ? 
                    AND (firma_bombero_fecha IS NULL OR firma_bombero_fecha = '')
                    AND (anulada = 0 OR anulada IS NULL)
                """
                c.execute(query, (legajo_actual,))
            else:
                # El Supervisor firma en 'firma_supervisor_fecha' 
                # y su legajo está en la columna 'asignado'
                query = """
                    SELECT COUNT(*) FROM actividades 
                    WHERE asignado = ? 
                    AND (firma_supervisor_fecha IS NULL OR firma_supervisor_fecha = '')
                    AND (firma_bombero_fecha IS NOT NULL AND firma_bombero_fecha != '')
                    AND (anulada = 0 OR anulada IS NULL)
                """
                c.execute(query, (legajo_actual,))
            
            count = c.fetchone()[0]
            conn.close()
            return count > 0

        except Exception as e:
            # Esto te avisará en consola si algo más falla, pero no romperá la UI
            print(f"DEBUG SQL PENDIENTES: {e}")
            return False

    def _set_estado_campos_actividades(self, estado_ui):
            rol = self.usuario_actual.get("rol", "").upper()
            
            # 1. Definir estado base según el momento de la UI
            if estado_ui in ("nuevo", "editando"):
                estado_base = "readonly"
                estado_texto = "normal"
            else:
                estado_base = "disabled"
                estado_texto = "disabled"

            # 2. Aplicar estado general a los campos de datos
            widgets_datos = [
                self.actividad, self.area, self.fecha_inicio, 
                self.fecha_fin, self.hora_inicio, self.hora_finalizacion
            ]
            
            for w in widgets_datos:
                try:
                    # Si es un Entry o similar va 'normal', si es Combo va 'readonly'
                    s = estado_base if hasattr(w, 'set') else estado_texto
                    w.config(state=s)
                except: pass

            self._set_text_readonly(self.descripcion, estado_texto == "disabled")

            # 3. 🔥 MANEJO DE JERARQUÍA (Legajo y Asignado)
            # Delegamos a la función experta para que no haya choques
            if estado_ui in ("nuevo", "editando"):
                self._ajustar_combo_asignado_por_contexto()

    def _get_legajo_display(self, legajo):
        """
        Devuelve 'legajo - APELLIDO Nombre'
        o solo legajo si no encuentra datos
        """
        try:
            conn = sqlite3.connect(DB_PATH)
            cur = conn.cursor()
            cur.execute(
                "SELECT apellido, nombre FROM legajos WHERE legajo = ?",
                (legajo,)
            )
            row = cur.fetchone()
            conn.close()

            if row:
                apellido, nombre = row
                return f"{legajo} - {apellido} {nombre}"
        except Exception as e:
            print("WARN _get_legajo_display:", e)

        return legajo

    def _set_estado_botones_actividades(self, estado):

        # 🔒 Motor de botones migrado a _refrescar_actividades()
        # Esta función queda anulada para evitar conflictos
        return

        rol = self.usuario_actual.get("rol", "").upper()
        print(">>> SET ESTADO ACTIVIDADES:", estado, "ROL:", rol)

        # 🔥 PRIORIDAD ABSOLUTA: MODO EDICIÓN
        if estado == "editando":
            print(">>> Modo edición activo - control manual de botones")

            for b in self.act_btns.values():
                self._aplicar_estilo_boton(b, False)

            for nombre in ("Guardar", "Limpiar/Cancelar", "Cancelar"):
                if nombre in self.act_btns:
                    self._aplicar_estilo_boton(self.act_btns[nombre], True)

            return

        # 🔒 SI ESTÁ FIRMADA POR EL SUPERVISOR
        if getattr(self, "firma_supervisor_fecha", None):
            print(">>> Actividad firmada por el supervisor - bloqueo parcial")

            # Primero apagar todos
            for b in self.act_btns.values():
                self._aplicar_estilo_boton(b, False)

            # 🔓 Botones de navegación SIEMPRE habilitados
            for nombre in ("Nuevo", "Buscar", "Limpiar/Cancelar", "Pendientes de Firma"):
                if nombre in self.act_btns:
                    self._aplicar_estilo_boton(self.act_btns[nombre], True)

            # 🔓 ADMIN puede anular
            if rol == "ADMIN" and "Anular" in self.act_btns:
                self._aplicar_estilo_boton(self.act_btns["Anular"], True)

            return

        # 🔒 Si está procesando firma, bloquear todo
        if getattr(self, "_procesando_firma", False):
            for b in self.act_btns.values():
                self._aplicar_estilo_boton(b, False)
            return

        # 1️⃣ Apagar todos los botones
        for b in self.act_btns.values():
            self._aplicar_estilo_boton(b, False)

        # 2️⃣ Obtener permisos desde matriz
        permisos_rol = PERMISOS_ACTIVIDADES.get(rol, {})
        botones_habilitados = permisos_rol.get(estado, ())

        # 3️⃣ Habilitar solo los definidos en la matriz
        for nombre in botones_habilitados:

            if nombre not in self.act_btns:
                continue

            # 🔒 REGLA ESPECIAL: SUPERVISOR solo modifica lo que asignó
            if nombre == "Modificar":
                puede = self._puede_modificar_actividad()
                self._aplicar_estilo_boton(self.act_btns["Modificar"], puede)
                continue

                puede_modificar = False

                if rol == "ADMIN":
                    puede_modificar = True

                elif rol == "SUPERVISOR" and getattr(self, "id_actividad_actual", None):

                    import sqlite3
                    conn = sqlite3.connect(DB_PATH)
                    c = conn.cursor()
                    c.execute("""
                        SELECT asignado
                        FROM actividades
                        WHERE id = ?
                    """, (self.id_actividad_actual,))
                    row = c.fetchone()
                    conn.close()

                    if row:
                        asignado = (row[0] or "").strip().upper()
                        usuario = str(self.usuario_actual.get("legajo", "")).strip().upper()

                        if asignado == usuario:
                            puede_modificar = True

                self._aplicar_estilo_boton(self.act_btns["Modificar"], puede_modificar)
                continue

            # 🚫 Nunca permitir firmar en modo edición
            if estado == "editando" and nombre in (
                "Firmar actividad",
                "Firmar Supervisor"
            ):
                continue

            # 🔒 REGLA ESPECIAL: Firma Supervisor (Opción A)
            if nombre == "Firmar Supervisor":

                puede_aprobar = self._puede_firmar_supervisor()

                self._aplicar_estilo_boton(
                    self.act_btns["Firmar Supervisor"],
                    puede_aprobar
                )

                continue
            self._aplicar_estilo_boton(self.act_btns[nombre], True)

        # 🔎 Botón Pendientes siempre visible para roles operativos
        if rol in ("ADMIN", "SUPERVISOR", "BOMBERO"):
            if "Pendientes de Firma" in self.act_btns:
                self._aplicar_estilo_boton(
                    self.act_btns["Pendientes de Firma"],
                    True
                )

        # 🔴 SI ACTIVIDAD ANULADA
        if getattr(self, "actividad_anulada", False):
            print(">>> Actividad anulada - bloqueo parcial")

            bloquear = [
                "Modificar",
                "Guardar",
                "Firmar actividad",
                "Firmar Supervisor",
                "No firmar ahora",
            ]

            for nombre in bloquear:
                if nombre in self.act_btns:
                    self._aplicar_estilo_boton(self.act_btns[nombre], False)

        print(">>> BOTONES ACTIVOS:")
        for k, b in self.act_btns.items():
            print(k, "=>", b["state"])

    def _set_btn(self, nombre, habilitar, color_borde="#B22222"):
        if nombre not in self.act_btns:
            return

        btn = self.act_btns[nombre]

        if habilitar:
            btn.config(state="normal")
            self._style_btn(btn, color_borde, habilitado=True)
        else:
            btn.config(state="disabled")
            self._style_btn(btn, color_borde, habilitado=False)

    def _set_text_readonly(self, txt, readonly: bool):
        if readonly:
            txt.config(
                state="disabled",
                bg="#f0f0f0",
                fg="#808080",
                insertbackground="#808080",
                cursor="arrow",
                highlightthickness=0,
                bd=0,
                relief="flat"
            )
        else:
            txt.config(
                state="normal",
                bg="white",
                fg="black",
                insertbackground="black",
                cursor="xterm",
                highlightthickness=0,
                bd=1,
                relief="sunken"
            )

    def _descripcion_shift_enter(self, event):
           self.btn_guardar.focus_set()

    def _button_enter(self, event):
        try:
            event.widget.invoke()
        except Exception:
            pass
        return "break"

    def _on_tab_changed(self, event):

        if getattr(self, "cerrando", False):
            return

        try:
            self._aplicar_permisos()
        except:
            return

        tab_actual = event.widget.tab(event.widget.select(), "text")

        if tab_actual == "Usuarios":
            self._estado_botones_usuario("inicial")

        try:
            tab = event.widget.select()
            tab_widget = event.widget.nametowidget(tab)

            if tab_widget == self.actividades_frame:

                self._set_estado_campos_actividades("inicial")

                if getattr(self, "modo_actividad", "nuevo") == "nuevo":
                    self.master.after(
                        120,
                        lambda: (
                            not getattr(self, "cerrando", False)
                            and self.act_btns["Nuevo"].focus_set()
                        )
                    )

        except Exception as e:
            print("WARN foco pestaña actividades:", e)

    def _validar_asignado(self, event=None):
        """Valida que el asignado no sea el mismo legajo. Evita mensajes repetidos usando un flag."""
        try:
            legajo = (self.cb_legajo_sel.get().split(" - ")[0].strip()) if self.cb_legajo_sel.get() else ""
            asignado = (self.cb_asignado.get().split(" - ")[0].strip()) if self.cb_asignado.get() else ""
        except Exception:
            legajo = (self.cb_legajo_sel.get().split(" - ")[0].strip()) if hasattr(self, 'cb_legajo_sel') and self.cb_legajo_sel.get() else ""
            asignado = (self.cb_asignado.get().split(" - ")[0].strip()) if hasattr(self, 'cb_asignado') and self.cb_asignado.get() else ""

        # reset flag si valores vacíos o cambian
        last = getattr(self, '_last_asignado_error', None)
        if not legajo or not asignado:
            self._last_asignado_error = None
            return

        if legajo == asignado:
            if last == (legajo, asignado):
                return
            self._last_asignado_error = (legajo, asignado)
            self.ui.show_error("Error", "El asignado no puede ser la misma persona del legajo.")
            try:
                self.cb_asignado.set("")
            except Exception:
                pass
            try:
                self.master.after(300, lambda: setattr(self, '_last_asignado_error', None))
            except Exception:
                self._last_asignado_error = None

    def _ajustar_combo_asignado_por_contexto(self):
        rol_actual = self.usuario_actual.get("rol", "").upper()
        legajo_actual = str(self.usuario_actual.get("legajo", "")).strip()
        nombre_completo = f"{legajo_actual} - {self.usuario_actual.get('apellido','')} {self.usuario_actual.get('nombre','')}"
        estado_ui = getattr(self, "_estado_ui", "")

        # --- BLOQUEO PARA ACTIVIDADES EXISTENTES ---
        # Si la actividad ya tiene ID (ya existe en DB), los combos NO se tocan
        if hasattr(self, "id_actividad_actual") and self.id_actividad_actual:
            if hasattr(self, "cb_legajo_sel"):
                self.cb_legajo_sel.config(state="disabled")
            if hasattr(self, "cb_asignado"):
                self.cb_asignado.config(state="disabled")
            return  # Salimos inmediatamente

        # --- LÓGICA PARA NUEVAS ACTIVIDADES ---
        
        # 1. GESTIÓN DEL LEGAJO (QUIÉN REALIZA)
        if rol_actual == "BOMBERO":
            self.cb_legajo_sel.config(state="normal")
            self.cb_legajo_sel.set(nombre_completo)
            self.cb_legajo_sel.config(state="disabled")
            es_propia = True
        else:
            # Admin o Supervisor pueden elegir a otros (solo si es nuevo)
            self.cb_legajo_sel.config(state="readonly" if estado_ui == "nuevo" else "disabled")
            legajo_sel_raw = self.cb_legajo_sel.get()
            legajo_sel = legajo_sel_raw.split(" - ")[0].strip() if legajo_sel_raw else ""
            es_propia = (not legajo_sel or legajo_sel == legajo_actual)

        # 2. CARGAR LISTA BASE (Filtra según jerarquía)
        self.cargar_asignados_autorizados()
        
        # 3. LÓGICA DE ASIGNACIÓN (QUIÉN SUPERVISA)
        self.cb_asignado.config(state="normal")

        if rol_actual == "ADMIN":
            if es_propia:
                # Si carga para él, debe elegir a otro Admin/Supervisor de la lista
                self.cb_asignado.config(state="readonly" if estado_ui in ("nuevo", "editando") else "disabled")
            else:
                # Si carga para otro, se auto-asigna y se bloquea
                self.cb_asignado.set(f"{nombre_completo}") # Quitamos el (Admin) para evitar fallos de split
                self.cb_asignado.config(state="disabled")

        elif rol_actual == "SUPERVISOR":
            if es_propia:
                self.cb_asignado.config(state="readonly" if estado_ui in ("nuevo", "editando") else "disabled")
            else:
                # Si carga para un Bombero, se auto-asigna y se bloquea
                self.cb_asignado.set(f"{nombre_completo}")
                self.cb_asignado.config(state="disabled")

        elif rol_actual == "BOMBERO":
            self.cb_asignado.config(state="readonly" if estado_ui in ("nuevo", "editando") else "disabled")

        # 4. SEGURIDAD FINAL: Si el estado global de la UI no es de edición, bloqueamos todo
        if estado_ui not in ("nuevo", "editando"):
            self.cb_legajo_sel.config(state="disabled")
            self.cb_asignado.config(state="disabled")

        print(f">>> AJUSTE FINAL: Rol={rol_actual} | ID={getattr(self, 'id_actividad_actual', None)} | UI={estado_ui}")
            
    def _configurar_navegacion_actividades(self):
        """
        Sistema de navegación con Enter para ACTIVIDADES.
        No interfiere con binds existentes.
        Solo actúa si el widget no tiene manejo propio.
        """

        # Orden lógico de navegación
        self._orden_actividades = [
            self.cb_legajo_sel,
            self.cb_asignado,
            self.actividad,
            self.area,
            self.fecha_inicio,
            self.fecha_fin,
            self.hora_inicio,
            self.hora_finalizacion,
            self.descripcion,
            self.act_btns["Guardar"]
        ]

        # Binds seguros (sin romper los existentes)
        for i, widget in enumerate(self._orden_actividades):

            def handler(event, idx=i, w=widget):
                # Si el widget tiene un bind propio, no intervenir
                if self._widget_tiene_bind_propio(w):
                    return None

                return self._focus_siguiente_actividad(idx)

            widget.bind("<Return>", handler, add="+")

    def _widget_tiene_bind_propio(self, widget):
        """
        Detecta si el widget ya tiene un bind de Enter
        distinto al sistema de navegación.
        """
        try:
            bind = widget.bind("<Return>")
            if bind and "_focus_siguiente_actividad" not in str(bind):
                return True
        except Exception:
            pass
        return False

    def _focus_siguiente_actividad(self, idx):
        """Mueve el foco al siguiente campo habilitado o al botón Guardar."""
        try:
            total = len(self._orden_actividades)
            i = idx + 1

            while i < total:
                w = self._orden_actividades[i]

                try:
                    estado = str(w.cget("state"))
                except Exception:
                    estado = "normal"

                if estado not in ("disabled",):
                    w.focus_set()

                    # Si es hora fin, calcular duración
                    if w == self.hora_finalizacion:
                        self.calcular_duracion()

                    return "break"

                i += 1

            # Si no hay más campos habilitados → ir a Guardar
            if hasattr(self, "act_btns") and "Guardar" in self.act_btns:
                btn = self.act_btns["Guardar"]
                if str(btn.cget("state")) != "disabled":
                    btn.focus_set()

        except Exception as e:
            pass

        return "break"

    def _anular_firmas_por_modificacion(self):
        import sqlite3
        from datetime import datetime

        try:
            conn = sqlite3.connect(DB_PATH)
            c = conn.cursor()

            id_act = self.id_actividad_actual
            usuario_id = self.usuario_actual.get("id")

            # Registrar en historial
            ahora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            if self.firma_bombero_fecha:
                c.execute("""
                    INSERT INTO actividades_historial
                    (actividad_id, fecha, usuario_id, campo, valor_anterior, valor_nuevo)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (
                    id_act, ahora, usuario_id,
                    "firma_bombero_fecha",
                    self.firma_bombero_fecha,
                    None
                ))

            if self.firma_supervisor_fecha:
                c.execute("""
                    INSERT INTO actividades_historial
                    (actividad_id, fecha, usuario_id, campo, valor_anterior, valor_nuevo)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (
                    id_act, ahora, usuario_id,
                    "firma_supervisor_fecha",
                    self.firma_supervisor_fecha,
                    None
                ))

            # Anular firmas en actividad
            c.execute("""
                UPDATE actividades
                SET firma_bombero_usuario = NULL,
                    firma_bombero_fecha = NULL,
                    firma_supervisor_usuario = NULL,
                    firma_supervisor_fecha = NULL
                WHERE id = ?
            """, (id_act,))

            conn.commit()
            conn.close()

            print(f">>> Firmas anuladas por modificación en actividad {id_act}")

        except Exception as e:
            print("ERROR al anular firmas:", e)

    def calcular_duracion(self, event=None):
        """Calcula la diferencia entre hora fin y hora inicio CORREGIDA."""
        try:
            h_ini = self.hora_inicio.get().strip()
            h_fin = self.hora_finalizacion.get().strip()
            
            if not h_ini or not h_fin:
                self._horas_decimal = None
                self.var_horas.set("")
                return
            
            # Función de conversión CORREGIDA
            def convertir_a_decimal(hora_str):
                """Convierte HH:MM a decimal CORRECTAMENTE."""
                if not hora_str:
                    return 0.0
                
                hora_str = str(hora_str).strip().replace(',', '.')
                
                # Formato HH:MM
                if ':' in hora_str:
                    partes = hora_str.split(':')
                    if len(partes) >= 2:
                        try:
                            horas = float(partes[0]) if partes[0] else 0
                            minutos = float(partes[1]) if partes[1] else 0
                            # CORRECCIÓN: minutos / 60.0 (no / 100)
                            return horas + (minutos / 60.0)
                        except:
                            return 0.0
                
                # Formato decimal (ya está bien)
                try:
                    return float(hora_str)
                except:
                    return 0.0
            
            # Convertir a decimal
            ini_dec = convertir_a_decimal(h_ini)
            fin_dec = convertir_a_decimal(h_fin)
            
            # Calcular diferencia (manejar día siguiente)
            diferencia = fin_dec - ini_dec
            if diferencia < 0:
                diferencia += 24.0
            
            if diferencia < 0:
                self._horas_decimal = None
                self.var_horas.set("")
                return
            
            # Guardar valor decimal CORRECTO
            self._horas_decimal = round(diferencia, 2)
            
            # Mostrar en formato HH:MM
            horas_enteras = int(self._horas_decimal)
            minutos = int(round((self._horas_decimal - horas_enteras) * 60))
            
            # Ajustar si minutos son 60
            if minutos >= 60:
                horas_enteras += 1
                minutos -= 60
            
            display = f"{horas_enteras}:{minutos:02d}"
            self.var_horas.set(display)
             
        except Exception as e:
            print(f"ERROR CALCULAR >> {e}")
            self._horas_decimal = None
            self.var_horas.set("")

    def _habilitar_modificar_actividad(self):

        # 🔥 ÚNICO CONTROL REAL
        if not self._puede_modificar_actividad():
            print("⛔ SIN PERMISO PARA MODIFICAR")
            return

        self.modo_actividad = "modificar"

        actividad_actual = self.actividad.get()

        try:
            self.cargar_actividades_desde_conceptos()

            if actividad_actual:
                self.actividad.set(actividad_actual)

            self.actividad.config(state="readonly")

        except Exception as e:
            print("ERROR recargando combo actividad:", e)

        self._estado_ui = "editando"
        self._refrescar_actividades()

        try:
            if actividad_actual:
                self.actividad.set(actividad_actual)

            self.actividad.config(state="readonly")

        except:
            pass

        self._estado_visual_firma("edicion")

        print("ASIGNADO STATE:", self.cb_asignado.cget("state"))

    def nuevo_actividad(self):
            """Prepara el formulario para cargar una NUEVA actividad"""
            from datetime import date
            rol = self.usuario_actual.get("rol", "").upper()
            
            # 1) ESTADOS INICIALES
            self.modo_actividad = "nuevo"
            self._estado_ui = "nuevo"
            self._estado_actividad = "BORRADOR"
            self.id_actividad_actual = None
            self._reset_estado_actividad()

            # 2) ID PROVISORIO
            try:
                self.var_id_actividad.set(str(self._next_actividad_id()))
            except Exception:
                self.var_id_actividad.set("")

            # 3) LIMPIAR CAMPOS BÁSICOS
            for cb in (self.cb_legajo_sel, self.cb_asignado, self.actividad, self.area):
                try: 
                    cb.config(state="normal") # Habilitamos temporalmente para limpiar
                    cb.set("")
                except: pass

            # 4) FECHAS Y HORAS
            for w in (self.fecha_inicio, self.fecha_fin, self.hora_inicio, self.hora_finalizacion):
                try:
                    w.config(state="normal")
                    if hasattr(w, 'delete'): w.delete(0, END)
                except: pass
            
            self.fecha_inicio.set_date(date.today())
            self.fecha_fin.set_date(date.today())

            # 5) DESCRIPCIÓN Y ACTIVIDADES (Habilitar para el nuevo registro)
            try:
                self._set_text_readonly(self.descripcion, False)
                self.descripcion.delete("1.0", END)
                self.cargar_actividades_desde_conceptos()
                # Ahora sí los ponemos en modo edición
                self.actividad.config(state="readonly")
                self.area.config(state="readonly")
            except: pass

            # 6) LÓGICA DE ROLES PARA LEGAJO
            if rol == "BOMBERO":
                legajo = str(self.usuario_actual.get("legajo", "")).strip()
                apellido = self.usuario_actual.get("apellido", "")
                nombre = self.usuario_actual.get("nombre", "")
                texto_bombero = f"{legajo} - {apellido} {nombre}".upper()
                
                self.cb_legajo_sel.set(texto_bombero)
                self.cb_legajo_sel.config(state="disabled")
                self.master.after(200, self.cb_asignado.focus_set)
            else:
                self.cb_legajo_sel.config(state="readonly")
                self.cargar_legajos_combobox() 
                self.master.after(200, self.cb_legajo_sel.focus_set)

            # 7) REFUERZO FINAL
            self._refrescar_actividades()
            
            # Esto cargará los supervisores y bloqueará/habilitará 'Asignado por'
            self.master.after(150, self._aplicar_ajustes_finales_nuevo)

            self.lbl_estado_firma.config(
                text="🟡 MODO NUEVO - Cargando actividad",
                bg="#fff3cd",
                fg="#856404"
            )

    def _aplicar_ajustes_finales_nuevo(self):
        """Función auxiliar para asegurar que los combos queden perfectos"""
        self.cargar_asignados_autorizados()
        self._ajustar_combo_asignado_por_contexto()

    def _reset_estado_actividad(self):
        self.firma_bombero_fecha = None
        self.firma_bombero_usuario = None
        self.firma_supervisor_fecha = None
        self.firma_supervisor_usuario = None
        self.actividad_anulada = False

    def _aplicar_estado_asignado(self, asignados):
        
        if not asignados:
            estado = "disabled"
        
        elif len(asignados) == 1:
            # 🔥 CASO SUPERVISOR → solo admin disponible
            estado = "readonly"
        
        else:
            estado = "readonly"

        print("FINAL ASIGNADO:", estado)
        self.combo_asignado.config(state=estado)

    def validar_hora(self, P):
        """
        Validación suave durante escritura:
        - permite escritura progresiva
        - solo números y :
        - máximo 5 caracteres
        """
        if P == "":
            return True

        if len(P) > 5:
            return False

        for c in P:
            if not (c.isdigit() or c == ":"):
                return False

        return True

    def validar_campos_actividad(self):
        errores = []

        # --- Campos obligatorios ---
        if not self.cb_legajo_sel.get().strip():
            errores.append("Debe seleccionar un Legajo.")

        if not self.cb_asignado.get().strip():
            errores.append("Debe seleccionar quién asigna.")

        if not self.actividad.get().strip():
            errores.append("Debe seleccionar una Actividad.")

        if not self.area.get().strip():
            errores.append("Debe seleccionar un Área/Dpto.")

        if not self.fecha_inicio.get():
            errores.append("Debe ingresar Fecha Inicio.")

        if not self.fecha_fin.get():
            errores.append("Debe ingresar Fecha Fin.")

        # --- Horas obligatorias ---
        hi = self.hora_inicio.get().strip()
        hf = self.hora_finalizacion.get().strip()

        if not hi:
            errores.append("Debe ingresar Hora Inicio.")
        elif hi == "Hora inválida":
            errores.append("Hora Inicio inválida.")

        if not hf:
            errores.append("Debe ingresar Hora Fin.")
        elif hf == "Hora inválida":
            errores.append("Hora Fin inválida.")

        # --- Validación fuerte HH:MM ---
        import re
        patron = r"^([01]\d|2[0-3]):[0-5]\d$"

        if hi and hi != "Hora inválida" and not re.match(patron, hi):
            errores.append("Hora Inicio debe ser HH:MM válida.")

        if hf and hf != "Hora inválida" and not re.match(patron, hf):
            errores.append("Hora Fin debe ser HH:MM válida.")

        # --- Descripción obligatoria ---
        desc = self.descripcion.get("1.0", "end").strip()
        if not desc:
            errores.append("Debe ingresar una Descripción.")

        return errores

    def autoformato_hora(self, event):
        widget = event.widget
        texto = widget.get().strip()

        if not texto:
            widget.config(fg="black")
            return

        # 🔧 Normalizar
        partes = texto.split(":")

        try:
            hora = partes[0].zfill(2)

            if len(partes) == 1:
                minuto = "00"
            else:
                minuto = partes[1].zfill(2)

            h = int(hora)
            m = int(minuto)

            if not (0 <= h <= 23 and 0 <= m <= 59):
                raise ValueError

            widget.delete(0, tk.END)
            widget.insert(0, f"{h:02d}:{m:02d}")
            widget.config(fg="black")

        except Exception:
            widget.delete(0, tk.END)
            widget.insert(0, "Hora inválida")
            widget.config(fg="red")
            return

        # ⏱️ Recalcular duración si ambas están bien
        self.calcular_duracion()

    def _recalcular_estado_actividad(self):

        # Si está anulada
        if getattr(self, "actividad_anulada", False):
            self._estado_actividad = "ANULADA"
            return

        # Firmas
        firmada_bombero = getattr(self, "firma_bombero_fecha", None)
        firmada_supervisor = getattr(self, "firma_supervisor_fecha", None)

        if firmada_supervisor:
            self._estado_actividad = "FIRMADA_SUPERVISOR"
        elif firmada_bombero:
            self._estado_actividad = "FIRMADA_BOMBERO"
        else:
            self._estado_actividad = "BORRADOR"

    def combo_autoselect(self, event):
        widget = event.widget
        value = widget.get().upper()  # o lower según tus datos
        values = [v for v in widget["values"]]

        if value:  # si hay algo escrito
            # buscar primer valor que empiece con esa letra
            for v in values:
                if v.upper().startswith(value):
                    widget.set(v)
                    break

    def limpiar_actividad(self):
            """Limpia el formulario y vuelve al estado inicial"""
            # 1. Resetear todas las variables de firma y estado interno
            self._reset_estado_actividad()
            
            # 2. Reset visual de etiquetas de estado
            self.lbl_estado_anulada.config(
                text="", 
                bg=self.actividades_frame.cget("bg")
            )
            self.lbl_info.config(text="")
            
            # 3. Reset de IDs y Modos
            self.modo_actividad = None
            self.id_actividad_actual = None
            self._estado_ui = "inicial"
            self._legajo_original_actividad = None
            self.actividad_asignado = None
            self._estado_actividad = None

            # 4. Limpiar Widgets (Combos y Entries)
            # Hacemos una lista para no repetir código
            combos = (self.cb_legajo_sel, self.cb_asignado, self.actividad, self.area)
            for cb in combos:
                try:
                    cb.config(state="normal")
                    cb.set("")
                except: pass

            try: self.var_id_actividad.set("")
            except: pass

            # 5. Fechas y Horas
            from datetime import date
            try:
                for f in (self.fecha_inicio, self.fecha_fin):
                    f.config(state="normal")
                    f.set_date(date.today())
                
                for h in (self.hora_inicio, self.hora_finalizacion):
                    h.config(state="normal")
                    h.delete(0, END)
                
                self.e_horas.config(state="normal")
                self.var_horas.set("")
            except: pass

            # 6. Descripción
            try:
                self._set_text_readonly(self.descripcion, False)
                self.descripcion.delete("1.0", END)
            except: pass

            # ---------------------------------------------------------
            # 7. 🔥 LA CLAVE: DELEGAR EL BLOQUEO FINAL
            # ---------------------------------------------------------
            # En lugar de poner .config(state="disabled") acá uno por uno,
            # dejamos que _refrescar_actividades haga su trabajo.
            
            self._refrescar_actividades()

    def _forzar_foco_dateentry(self, dateentry_widget):
        """Intenta enfocar internamente el DateEntry para que su calendario funcione."""
        try:
            # tkcalendar nuevo tiene .entry o ._entry interno según versión
            if hasattr(dateentry_widget, "entry"):
                dateentry_widget.entry.focus_force()
            elif hasattr(dateentry_widget, "_entry"):
                dateentry_widget._entry.focus_force()
            else:
                dateentry_widget.focus_force()
        except Exception:
            try:
                dateentry_widget.focus_set()
            except Exception:
                pass

    def _recreate_date_entries(self):
        """Recrea los DateEntry de fecha_inicio y fecha_fin para resetear su estado/calendario."""
        try:
            info1 = self.fecha_inicio.place_info() if hasattr(self, "fecha_inicio") else None
            info2 = self.fecha_fin.place_info() if hasattr(self, "fecha_fin") else None
        except Exception:
            info1 = info2 = None

        # destruir existentes
        try:
            if hasattr(self, "fecha_inicio"):
                self.fecha_inicio.destroy()
        except Exception:
            pass
        try:
            if hasattr(self, "fecha_fin"):
                self.fecha_fin.destroy()
        except Exception:
            pass

        # recrear widgets
        try:
            from tkcalendar import DateEntry
            self.fecha_inicio = DateEntry(self.actividades_frame, date_pattern="dd/mm/yyyy",
                                          state="normal", font=("Arial", 11))
            self.fecha_inicio.place(**info1 if info1 else {"x": 150, "y": 46 + 5 * 28})
            self.fecha_inicio.bind("<Return>", lambda e: self._focus_next_widget(e.widget))

            self.fecha_fin = DateEntry(self.actividades_frame, date_pattern="dd/mm/yyyy",
                                       state="normal", font=("Arial", 11))
            self.fecha_fin.place(**info2 if info2 else {"x": 150, "y": 46 + 6 * 28})
            # Al presionar Enter en Fecha Fin, pasar el foco a Hora Inicio
            self.fecha_fin.bind("<Return>", lambda e: self.hora_inicio.focus_set())
        except Exception:
            # fallback sin calendario
            self.fecha_inicio = Entry(self.actividades_frame, font=("Arial", 11))
            self.fecha_inicio.place(**info1 if info1 else {"x": 150, "y": 46 + 5 * 28})
            self.fecha_fin = Entry(self.actividades_frame, font=("Arial", 11))
            self.fecha_fin.place(**info2 if info2 else {"x": 150, "y": 46 + 6 * 28})

    def _actualizar_estado_fechas_actividad(self, editable):
        import sys
        import os
        import tempfile
        """
        Si editable==True: habilita las fechas, setea a HOY si estaban vacías y fuerza foco en inicio.
        Si editable==False: deshabilita las fechas (dejando valor válido).
        """
        try:
            if editable:
                try:
                    self.fecha_inicio.config(state="normal")
                    self.fecha_fin.config(state="normal")
                except Exception:
                    pass

                # setear a hoy si vienen vacías o con valor neutro
                try:
                    val = ""
                    try:
                        val = self.fecha_inicio.get()
                    except Exception:
                        val = ""
                    if not val or val in ("", "01/01/1900", "1900-01-01"):
                        self.fecha_inicio.set_date(date.today())
                except Exception:
                    pass

                try:
                    val = ""
                    try:
                        val = self.fecha_fin.get()
                    except Exception:
                        val = ""
                    if not val or val in ("", "01/01/1900", "1900-01-01"):
                        self.fecha_fin.set_date(date.today())
                except Exception:
                    pass

                # intentar dar foco al entry interno (mejora la interacción con el calendario)
                try:
                    self.master.after(40, lambda: self._forzar_foco_dateentry(self.fecha_inicio))
                except Exception:
                    pass

            else:
                try:
                    self.fecha_inicio.config(state="disabled")
                    self.fecha_fin.config(state="disabled")
                except Exception:
                    pass
        except Exception:
                pass

    def abrir_pdf_temporal(self, pdf_bytes, titulo="PDF Temporal"):
        """Abre un PDF desde memoria en un archivo temporal y lo lanza en el sistema."""
        import tempfile, os, sys, subprocess

        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
                tmp.write(pdf_bytes)
                tmp_path = tmp.name

            if sys.platform.startswith("win"):
                os.startfile(tmp_path)
            elif sys.platform.startswith("darwin"):
                subprocess.call(["open", tmp_path])
            else:
                subprocess.call(["xdg-open", tmp_path])
        except Exception as e:
#            from tkinter.messagebox import showwarning
            messagebox.showwarning("Aviso", f"No se pudo abrir el PDF temporal: {e}")

    def _preparar_elementos_pdf(
        self, id_act, legajo, legajo_ap, legajo_nom, legajo_dni,
        asig, asignado_ap, asignado_nom, asignado_dni,
        actividad_txt, descripcion_txt, area_val,
        f_inicio, f_fin, hi, hf, horas_trabajadas,
        modo=None
    ):
        """
        Devuelve los elementos para PDF temporal con:
        - Tabla de datos en el mismo orden que registros antiguos
        - Subtítulo: REGISTRO DE ACTIVIDAD Nº XX
        - Firmas con nombre, apellido y DNI centrados
        """
        from reportlab.platypus import Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.platypus import Image, Spacer
        styles = getSampleStyleSheet()
        elems = []

        try:
            elems.append(Image("Bomberos.png", width=120, height=60))
            elems.append(Spacer(1, 12))
        except:
            pass

        # 🔴 Marca si fue modificada por ADMIN
        if modo == "modificar":
            # 🔒 BLOQUEAR MODIFICACIÓN SI ESTÁ FIRMADA POR SUPERVISOR
            if getattr(self, "firma_supervisor_fecha", None):
 #               from tkinter.messagebox import showerror
                showerror(
                    "No permitido",
                    "La actividad ya fue firmada por el Supervisor.\n\n"
                    "No puede ser modificada.\n"
                    "Solo un ADMIN puede anularla."
                )
                conn.close()
                return            
            aviso = (
                "<b>⚠ DOCUMENTO GENERADO TRAS MODIFICACIÓN ADMINISTRATIVA</b><br/>"
                "Las firmas anteriores fueron invalidadas."
            )
            p = Paragraph(aviso, styles["Normal"])
            elems.append(p)
            elems.append(Spacer(1, 12))

        # --- Subtítulo ---
        elems.append(Spacer(1, 10))
        subtitulo_style = ParagraphStyle('Subtitulo', parent=styles['Heading2'], alignment=1)  # centrado
        elems.append(Paragraph(f"REGISTRO DE ACTIVIDAD Nº {id_act}", subtitulo_style))
        elems.append(Spacer(1, 12))

        # --- Tabla de datos ---
        labels = ["ID", "Legajo", "Apellido y Nombre", "Asignado por", "Actividad", "Área",
                "Fecha Inicio", "Fecha Fin", "Hora Inicio", "Hora Fin", "Total Hs", "Descripción"]
        valores = [id_act, legajo, f"{legajo_ap} {legajo_nom}",
                f"{asig} - {asignado_ap} {asignado_nom}" if asig else "",
                actividad_txt, area_val,
                f_inicio, f_fin, hi, hf, f"{horas_trabajadas:.2f}", descripcion_txt]

        data = []
        for lab, val in zip(labels, valores):
            data.append([Paragraph(f"<b>{lab}:</b>", styles["Normal"]), Paragraph(str(val), styles["Normal"])])

        tbl = Table(data, colWidths=[120, 380])
        tbl.setStyle(TableStyle([
            ("VALIGN", (0,0), (-1,-1), "TOP"),
            ("LINEBELOW", (0,0), (-1,-1), 0.25, colors.grey),
            ("WORDWRAP", (0,0), (-1,-1), True),
        ]))
        elems.append(tbl)

        elems.append(Spacer(1, 40))

        # --- Firmas centradas debajo de los nombres ---
        firma_tbl = Table([
            [
                Paragraph(
                    f"<para alignment='center'>__________________________<br/><b>Firma del Bombero</b><br/>{legajo_ap} {legajo_nom}<br/>DNI: {legajo_dni}<br/>Legajo: {legajo}</para>",
                    styles["Normal"]
                ),
                Paragraph(
                    f"<para alignment='center'>__________________________<br/><b>Firma del Responsable</b><br/>{asignado_ap} {asignado_nom}<br/>DNI: {asignado_dni}<br/>Legajo: {asig}</para>",
                    styles["Normal"]
                )
            ]
        ], colWidths=[250, 250])
        firma_tbl.setStyle(TableStyle([("VALIGN", (0,0), (-1,-1), "BOTTOM"), ("TOPPADDING", (0,0), (-1,-1), 20)]))
        elems.append(firma_tbl)

        return elems

    def generar_texto_mail_estado(self, accion):
        """
        accion: 'REGISTRADA' o 'MODIFICADA'
        """
        estado = self.obtener_estado_firma_texto()

        texto = []
        texto.append(f"La actividad {accion}.")
        texto.append("")
        texto.append(f"Estado actual: {estado}.")
        texto.append("")

        if not self.firma_supervisor_fecha:
            texto.append(
                "Recuerde que esta actividad debe ser revisada y firmada por un Supervisor."
            )

        return "\n".join(texto)

    def guardar_actividad(self):
            hubo_cambio_sensible = False  # Valor por defecto
            errores = self.validar_campos_actividad()
            if errores:
                self.ui.show_error(
                    "Campos obligatorios",
                    "No se puede guardar por los siguientes motivos:\n\n" +
                    "\n".join(f"• {e}" for e in errores)
                )
                return

            # -------------------------------------------------
            # Datos base
            # -------------------------------------------------
            rol_actual = self.usuario_actual.get("rol", "").upper()
            legajo_actual = str(self.usuario_actual.get("legajo", "")).strip()
            id_actual = str(self.usuario_actual.get("id", "")).strip()
            modo = getattr(self, "modo_actividad", "nuevo")
         
            # Determinación de legajo según rol
            if rol_actual == "BOMBERO":
                legajo = legajo_actual
            else:
                legajo = self.cb_legajo_sel.get().split(" - ")[0].strip() if self.cb_legajo_sel.get() else ""

            # BLOQUEO: supervisor no puede cargar actividad para ADMIN
            if rol_actual == "SUPERVISOR" and self._es_admin(legajo):
                self.ui.show_error("Operación no permitida", "Un SUPERVISOR no puede cargar actividades para un ADMIN.")
                return
        
            if not legajo:
                self.ui.show_error("Error", "Debe seleccionar un Legajo.")
                return

            actividad_sel = self.actividad.get().strip().upper()
            asig_texto = self.cb_asignado.get()
            
            if not actividad_sel:
                self.ui.show_error("Error", "Debe seleccionar una Actividad.")
                return
            if not asig_texto:
                self.ui.show_error("Error", "Debe seleccionar un Asignado.")
                return

            # Procesar IDs y textos
            asig = asig_texto.split(" - ")[0].strip()
            concepto_id = None
            if " - " in actividad_sel:
                try:
                    concepto_id = int(actividad_sel.split(" - ")[0])
                except:
                    concepto_id = None

            if legajo == asig:
                self.ui.show_error("Error", "El asignado no puede ser la misma persona del legajo.")
                return

            # Lógica de asignación para MODO NUEVO
            if modo != "modificar":
                es_propia = (legajo == legajo_actual)
                if rol_actual == "SUPERVISOR":
                    if es_propia:
                        if asig == legajo_actual:
                            self.ui.show_error("Asignación inválida", "No puede autoasignarse una actividad propia.")
                            return
                        if "(ADMIN)" not in asig_texto.upper():
                            self.ui.show_error("Asignado inválido", "Las actividades propias deben asignarse a ADMIN.")
                            return
                    else:
                        asig = legajo_actual
                elif rol_actual == "ADMIN":
                    if es_propia:
                        if asig == legajo_actual:
                            self.ui.show_error("Asignación inválida", "Un ADMIN no puede autoasignarse.")
                            return
                        if not self._es_admin(asig):
                            self.ui.show_error("Asignado inválido", "Debe asignar a otro ADMIN.")
                            return
                    else:
                        asig = legajo_actual

            # Datos del formulario
            area_val = self.area.get().strip().upper()
            descripcion_txt = self.descripcion.get("1.0", "end").strip().upper()
            fecha_ini = self.fecha_inicio.get()
            fecha_fin = self.fecha_fin.get()
            hi = self.hora_inicio.get().strip()
            hf = self.hora_finalizacion.get().strip()

            try:
                self.calcular_duracion()
                horas_trabajadas = getattr(self, "_horas_decimal", 0)
            except:
                horas_trabajadas = 0

            if horas_trabajadas <= 0:
                self.ui.show_error("Error", "La duración debe ser mayor a cero.")
                return

            conn = None
            try:
                conn = sqlite3.connect(DB_PATH)
                conn.row_factory = sqlite3.Row # Para manejar diccionarios de datos
                c = conn.cursor()
                id_act = None

                if modo == "modificar":
                    id_act = int(self.var_id_actividad.get())
                    
                    # Obtener datos anteriores para historial y validación
                    c.execute("SELECT * FROM actividades WHERE id = ?", (id_act,))
                    fila_anterior = c.fetchone()

                    if not fila_anterior:
                        self.ui.show_error("Error", "No se encontró la actividad.")
                        return

                    # Validación de permisos
                    f_bomb = fila_anterior["firma_bombero_fecha"]
                    f_sup = fila_anterior["firma_supervisor_fecha"]
                    estado = getattr(self, "_estado_actividad", "BORRADOR")
                    
                    permitido = False
                    if estado == "BORRADOR":
                        if legajo_actual in (str(fila_anterior["creado_por"]), str(fila_anterior["legajo"])) or id_actual == str(fila_anterior["creado_por"]):
                            permitido = True
                    elif f_bomb and not f_sup:
                        if rol_actual == "ADMIN" or legajo_actual == str(fila_anterior["asignado"]):
                            permitido = True
                    
                    if not permitido:
                        self.ui.show_error("Acceso Denegado", "No tiene permisos para modificar esta actividad.")
                        return

                    # Detectar cambios sensibles
                    datos_nuevos = {
                        "actividad": actividad_sel, "area": area_val, 
                        "hora_inicio": hi, "hora_fin": hf, "descripcion": descripcion_txt
                    }
                    
                    if any(str(fila_anterior[k]).strip().upper() != str(datos_nuevos[k]).strip().upper() for k in datos_nuevos):
                        hubo_cambio_sensible = True
                    actividad_estaba_firmada = (f_bomb is not None)

                    if actividad_estaba_firmada and hubo_cambio_sensible:
                        if not self.ui.ask_yes_no("Actividad firmada", "Se anularán las firmas por los cambios realizados. ¿Continuar?"):
                            return

                    # Registrar en historial y Update
                    fecha_hist = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    hubo_cambios = False
                    
                    # Campos a comparar para el historial completo
                    campos_a_verificar = ["legajo", "actividad", "area", "hora_inicio", "hora_fin", "descripcion", "horas"]
                    for campo in campos_a_verificar:
                        val_ant = str(fila_anterior[campo] or "").strip().upper()
                        # Mapeo simple para obtener el valor nuevo correspondiente
                        val_new = str(locals().get(campo) if campo != 'actividad' else actividad_sel).strip().upper()
                        
                        if val_ant != val_new:
                            hubo_cambios = True
                            c.execute("""INSERT INTO actividades_historial (actividad_id, campo, valor_anterior, valor_nuevo, usuario_id, fecha)
                                        VALUES (?, ?, ?, ?, ?, ?)""", (id_act, campo, val_ant, val_new, self.usuario_actual["id"], fecha_hist))

                    if not hubo_cambios:
                        self.ui.show_info("Sin cambios", "No se detectaron modificaciones.")
                        return

                    # Ejecutar UPDATE
                    sql_update = """UPDATE actividades SET legajo=?, asignado=?, actividad=?, concepto_id=?, area=?, 
                                    fecha_inicio=?, fecha_fin=?, hora_inicio=?, hora_fin=?, descripcion=?, horas=?, 
                                    modificado_por=?, fecha_modificacion=? """
                    params = [legajo, asig, actividad_sel, concepto_id, area_val, fecha_ini, fecha_fin, hi, hf, 
                            descripcion_txt, horas_trabajadas, legajo_actual, fecha_hist]

                    if hubo_cambio_sensible:
                        # Agregamos la limpieza de los 4 campos de firma
                        sql_update += ", firma_bombero_fecha=NULL, firma_bombero_usuario=NULL, firma_supervisor_fecha=NULL, firma_supervisor_usuario=NULL "

                    sql_update += " WHERE id = ?"
                    params.append(id_act)

                    c.execute(sql_update, params)
                    conn.commit() # Aseguramos el commit aquí mismo

                    # --- Notificación ---
                    if rol_actual in ("SUPERVISOR", "ADMIN") and actividad_estaba_firmada and hubo_cambio_sensible:
                        # 1. Mostrar tu overlay de carga
                        self._mostrar_overlay_cargando("Procesando cambios...")
                        
                        d = {
                            "id": id_act, "leg": legajo, "asig": asig,
                            "act": actividad_sel, "desc": descripcion_txt,
                            "area": area_val, "hi": hi, "hf": hf,
                            "hs": horas_trabajadas, "f": fecha_ini
                        }

                        def ejecutar_envio():
                            try:
                                # Enviar mail con tu función original
                                self._enviar_correos_actividad(
                                    d["id"], d["leg"], d["asig"],
                                    d["act"], d["desc"], d["area"],
                                    d["hi"], d["hf"], d["hs"], d["f"],
                                    tipo_evento="MODIFICADA"
                                )
                                
                                # 2. Actualizar el Label de la interfaz (el que está debajo del título)
                                self.lbl_estado_firma.config(
                                    text=f"ACTIVIDAD N° {d['id']} MODIFICADA POR SUPERVISIÓN - FIRMAS ANULADAS",
                                    fg="white", bg="#d35400" # Naranja oscuro para advertir el cambio
                                )
                                
                                # 3. Cartel de aviso
                                self.ui.show_info("Éxito", f"Actividad N° {d['id']} MODIFICADA.\nSe enviaron los correos de anulación de firmas.")
                                
                            finally:
                                # 4. Quitar overlay
                                self._ocultar_overlay_cargando()

                        self.master.after(200, ejecutar_envio)

                    else:
                        # Caso de modificación común (sin cambios sensibles o hecha por el propio bombero)
                        self.lbl_estado_firma.config(
                            text=f"ACTIVIDAD N° {id_act} ACTUALIZADA CORRECTAMENTE",
                            fg="white", bg="#27ae60" # Verde éxito
                        )
                        self.ui.show_info("Éxito", f"Actividad N° {id_act} MODIFICADA correctamente.")

                    conn.commit()

                else: # MODO NUEVO
                    fecha_carga = datetime.now().strftime("%Y-%m-%d")
                    fecha_hist = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    c.execute("""INSERT INTO actividades (legajo, asignado, actividad, concepto_id, area, fecha_inicio, 
                                fecha_fin, hora_inicio, hora_fin, descripcion, fecha_carga, usuario_id, horas, creado_por, fecha_creacion)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                            (legajo, asig, actividad_sel, concepto_id, area_val, fecha_ini, fecha_fin, hi, hf, 
                            descripcion_txt, fecha_carga, self.usuario_actual["id"], horas_trabajadas, legajo_actual, fecha_hist))
                    id_act = c.lastrowid
                    self.ui.show_info("Registro exitoso", f"Actividad N° {id_act} registrada.")

                conn.commit()

            except Exception as e:
                traceback.print_exc()
                self.ui.show_error("Error", f"No se pudo guardar:\n{e}")
                return
            finally:
                if conn: conn.close()

            # -------------------------------------------------
            # Post-guardado: Actualizar UI
            # -------------------------------------------------
            self.id_actividad_actual = id_act
            self.var_id_actividad.set(str(id_act))
            self.modo_actividad = None
            self._estado_ui = "cargado"
            
            self.actualizar_contador_pendientes()
            self._recalcular_estado_actividad()
            self._actualizar_estado_firma()
            self._set_estado_campos_actividades("cargado")
            self._refrescar_actividades()

            # --- AGREGAR ESTE BLOQUE DE SEGURIDAD AQUÍ ---
            if hasattr(self, "cb_legajo_sel"):
                self.cb_legajo_sel.config(state="disabled")
            if hasattr(self, "cb_asignado"):
                self.cb_asignado.config(state="disabled")
            # ----------------------------------------------

            # CAMBIO AQUÍ: Solo poner "ACTIVIDAD GUARDADA" si no acabamos de poner el cartel naranja/verde arriba
            if not hubo_cambio_sensible:
                self.lbl_estado_firma.config(text=f"🔵 ACTIVIDAD N° {id_act} GUARDADA", bg="#d1ecf1", fg="#0c5460")

    def _set_foco_post_guardar(self):
        try:
            # Solo si estamos en estado cargado
            if getattr(self, "_estado_ui", "") == "cargado":
                self.cb_asignado.focus_set()
        except Exception:
            pass
    
    def buscar_actividad_por_id(self, id_act):
        self.cargar_actividad_por_id(id_act)

    def _get_bombero_por_legajo(self, legajo):
        try:
            cur = self.con.cursor()   # 👈 USÁ TU CONEXIÓN REAL
            cur.execute(
                "SELECT legajo, apellido, nombre FROM legajos WHERE legajo = ?",
                (legajo,)
            )
            return cur.fetchone()
        except Exception as e:
            print("WARN _get_bombero_por_legajo:", e)
            return None

    def mostrar_actividades_del_bombero(self):
        legajo = self.usuario_actual.get("legajo")

        try:
            conn = sqlite3.connect(DB_PATH)
            conn.row_factory = sqlite3.Row 
            c = conn.cursor()

            base_sql = self._sql_actividades_base()
            sql = base_sql + " WHERE a.legajo = ? ORDER BY a.id DESC"

            c.execute(sql, (legajo,))
            rows = c.fetchall()
            conn.close()

            self._rows_actividades_bombero = rows
            self.limpiar_actividad()

            if not rows:
                self.ui.show_info("Actividades", "No tenés actividades registradas.")
                self._estado_ui = "inicial"
                self.actividad_anulada = False
                self._refrescar_actividades()
                self.master.after(120, lambda: self.act_btns["Nuevo"].focus_set())
                return

            self.mostrar_resultados_en_treeview(rows)
            self._estado_ui = "inicial"
            self._refrescar_actividades()

        except Exception as e:
            self.ui.show_error("Error", f"No se pudieron cargar las actividades: {e}")

    def _es_admin(self, legajo):
        import sqlite3

        legajo = str(legajo).strip()

        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()

        c.execute("SELECT rol FROM usuarios WHERE legajo = ?", (legajo,))
        row = c.fetchone()
        conn.close()

        if not row:
            print("⚠ Usuario no encontrado para legajo:", legajo)
            return False

        return row[0].upper() == "ADMIN"

    def buscar_actividad(self):
        rol = self.usuario_actual.get("rol", "").upper()

        if rol == "BOMBERO":
            self.mostrar_actividades_del_bombero()
            return

        legajo_usuario = self.usuario_actual.get("legajo")

        self.cargar_asignados_autorizados()

        criterio = simpledialog.askstring(
            "Buscar",
            "Ingrese N° de Actividad o texto (Actividad / Descripción / Fecha):"
        )
        if not criterio:
            return

        criterio_raw = criterio.strip()
        print("CRITERIO RAW:", repr(criterio_raw))

        base_sql = self._sql_actividades_base()

        # ==========================================================
        # 1) BÚSQUEDA RÁPIDA POR ID
        # ==========================================================
        try:
            print("Intentando convertir a int...")
            id_buscar = int(criterio_raw)
            print("Convertido correctamente:", id_buscar)

            conn = sqlite3.connect(DB_PATH)
            conn.row_factory = sqlite3.Row
            c = conn.cursor()

            if rol == "BOMBERO":
                sql = base_sql + " WHERE a.id = ? AND a.legajo = ?"
                params = (id_buscar, legajo_usuario)
            else:
                sql = base_sql + " WHERE a.id = ?"
                params = (id_buscar,)

            c.execute(sql, params)
            row = c.fetchone()
            conn.close()

            if row:
                self._cargar_actividad_row(row)

                # 🔒 BLOQUEAR CAMPOS
                self.modo_actividad = None
                self._estado_ui = "cargado"
                self._refrescar_actividades()
                self.ui.show_info("Resultado", "Actividad cargada correctamente.")
            else:
                self.ui.show_info("Sin resultados", "No se encontró ninguna actividad con ese criterio.")

            return

        except ValueError as e:
            print("NO ES NÚMERO:", e)
            # Continúa a búsqueda por texto

        except Exception as e:
            self.ui.show_error("Error", f"No se pudo buscar por ID: {e}")
            return

        # ==========================================================
        # 2) BÚSQUEDA POR TEXTO
        # ==========================================================
        criterio_like = f"%{criterio_raw.upper()}%"

        try:
            conn = sqlite3.connect(DB_PATH)
            conn.row_factory = sqlite3.Row
            c = conn.cursor()

            if rol == "BOMBERO":
                sql = base_sql + """
                    WHERE (
                        UPPER(a.actividad) LIKE ?
                        OR UPPER(a.descripcion) LIKE ?
                        OR a.fecha_inicio LIKE ?
                        OR a.fecha_fin LIKE ?
                    )
                    AND a.legajo = ?
                    ORDER BY a.id DESC
                """
                params = (
                    criterio_like,
                    criterio_like,
                    criterio_raw,
                    criterio_raw,
                    legajo_usuario
                )

            else:
                sql = base_sql + """
                    WHERE (
                        CAST(a.legajo AS TEXT) LIKE ?
                        OR UPPER(a.actividad) LIKE ?
                        OR UPPER(a.descripcion) LIKE ?
                        OR UPPER(l.apellido) LIKE ?
                        OR UPPER(l.nombre) LIKE ?
                        OR a.fecha_inicio LIKE ?
                        OR a.fecha_fin LIKE ?
                    )
                    ORDER BY a.id DESC
                """
                params = (
                    criterio_like,
                    criterio_like,
                    criterio_like,
                    criterio_like,
                    criterio_like,
                    criterio_raw,
                    criterio_raw
                )

            c.execute(sql, params)
            rows = c.fetchall()
            conn.close()

        except Exception as e:
            self.ui.show_error("Error", f"No se pudo buscar: {e}")
            return

        # ==========================================================
        # 3) RESULTADOS (Integración con Treeview)
        # ==========================================================
        if not rows:
            self.ui.show_info("Resultados", "No se encontraron actividades con ese criterio.")
            self.limpiar_actividad()
            return

        if len(rows) > 1:
            # Aquí llamamos a la función que antes fallaba
            self.mostrar_resultados_en_treeview(rows)
            # Cambiamos a estado inicial para que no queden botones prendidos erróneamente
            self._estado_ui = "inicial"
            self._refrescar_actividades()
            return

        # Si hay exactamente uno, lo cargamos directo
        self._cargar_actividad_row(rows[0])
        self.modo_actividad = None
        self._estado_ui = "cargado"
        self._refrescar_actividades()
        self._estado_visual_firma("cargado")
        self.ui.show_info("Resultados", "Actividad cargada correctamente.")

    def nuevo_actividad(self):
        """Prepara el formulario para cargar una NUEVA actividad"""
        from datetime import date
        rol = self.usuario_actual.get("rol", "").upper()
        nivel = self.usuario_actual.get("nivel", "").upper()

        # -----------------------------
        # 1) MODO
        # -----------------------------
        self.modo_actividad = "nuevo"

        self._estado_ui = "nuevo"

        # 🔥 CLAVE: estado lógico real
        self._estado_actividad = "BORRADOR"
        self.id_actividad_actual = None
        # -----------------------------
        # 3) LIMPIAR CAMPOS
        # -----------------------------
        for cb in (self.cb_legajo_sel, self.cb_asignado, self.actividad):
            try: cb.set("")
            except: pass

        try: self.area.set("")
        except: pass

        # -----------------------------
        # 4) FECHAS
        # -----------------------------
        try:
            self.fecha_inicio.config(state="normal")
            self.fecha_fin.config(state="normal")
            self.fecha_inicio.set_date(date.today())
            self.fecha_fin.set_date(date.today())
        except Exception:
            pass

        # -----------------------------
        # 5) HORAS (editables)
        # -----------------------------
        for w in (self.hora_inicio, self.hora_finalizacion):
            try:
                w.config(state="normal")
                w.delete(0, END)
            except:
                pass

        # -----------------------------
        # 6) DESCRIPCIÓN
        # -----------------------------
        try:
            self._set_text_readonly(self.descripcion, False)   # habilitar edición
            self.descripcion.delete("1.0", END)
        except Exception:
            pass

        # -----------------------------
        # 7) COMBO ACTIVIDAD
        # -----------------------------
        try:
            self.cargar_actividades_desde_conceptos()
            self.actividad.config(state="readonly")
        except Exception as e:
            print("WARN combo actividad:", e)

        # -----------------------------
        # 8) LEGAJO SEGÚN ROL
        # -----------------------------
        if rol == "BOMBERO":
            legajo = str(self.usuario_actual.get("legajo", "")).strip()
            apellido = self.usuario_actual.get("apellido", "")
            nombre = self.usuario_actual.get("nombre", "")

            if apellido or nombre:
                texto = f"{legajo} - {apellido} {nombre}"
            else:
                texto = legajo  # fallback realista

            self.cb_legajo_sel.config(state="readonly")
            self.cb_legajo_sel["values"] = [texto]
            self.cb_legajo_sel.set(texto)
            self.cb_legajo_sel.config(state="disabled")

        # 9) ESTADO GLOBAL (CLAVE)
        # -----------------------------
        self._estado_ui = "nuevo"
        self._refrescar_actividades()

        # -----------------------------
        # FOCO INICIAL CORRECTO
        # -----------------------------
        try:
            rol = self.usuario_actual.get("rol", "").upper()

            if rol == "BOMBERO":
                # BOMBERO: legajo automático → foco en ASIGNADO
                self.cb_legajo_sel.config(state="disabled")
#                self.cb_asignado.config(state="readonly")
                self.master.after(150, self.cb_asignado.focus_set)
            else:
                # ADMIN / otros: foco en legajo
                self.cb_legajo_sel.config(state="readonly")
                self.master.after(150, self.cb_legajo_sel.focus_set)

        except Exception as e:
            print("WARN foco nuevo_actividad:", e)

        try:
            self.master.after(120, lambda: self.cb_asignado.focus_set())
        except Exception:
            pass

        # 🔒 Refuerzo final BOMBERO BÁSICO (anti pisado de estado)
        if rol == "BOMBERO" and nivel == "BASICO":
            legajo = str(self.usuario_actual.get("legajo", "")).strip()
            self.master.after(
                50,
                lambda l=legajo: (
                    self.cb_legajo_sel.config(state="readonly"),
                    self.cb_legajo_sel.set(l),
                    self.cb_legajo_sel.config(state="disabled")
                )
            )

        # 🔥 Estado único oficial
        self._estado_ui = "nuevo"
        self._refrescar_actividades()

        # 🔥 REFORZAR ID (ANTI PISADO)
        try:
            self.var_id_actividad.set(str(self._next_actividad_id()))
        except Exception:
            self.var_id_actividad.set("")
        self.var_id_actividad.set(str(self._next_actividad_id()))
        print("ID SETEADO:", self.var_id_actividad.get())

    def _next_actividad_id(self):
        import sqlite3

        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()

        c.execute("SELECT MAX(id) FROM actividades")
        row = c.fetchone()

        conn.close()

        max_id = row[0] if row and row[0] else 0
        return max_id + 1

    def cargar_asignados_autorizados(self):
        rol_usuario = str(self.usuario_actual.get("rol", "")).upper()
        legajo_usuario = str(self.usuario_actual.get("legajo", "")).strip()
        
        # Obtenemos el ID de la actividad que estamos viendo actualmente
        id_actual = getattr(self, "id_actividad_actual", None)
        valores = []

        for legajo_id, data in self.usuarios_dict.items():
            legajo_iterado = str(legajo_id).strip()
            rol_iterado = str(data.get("rol", "")).upper()
            nombre_iterado = data.get("nombre", "")

            # --- CAMBIO AQUÍ: LA REGLA DE ORO ---
            # Solo aplicamos el bloqueo de auto-asignación si es una actividad NUEVA.
            # Si la actividad ya existe (id_actual no es None), permitimos que aparezcas
            # para que el Combo pueda mostrar tu nombre correctamente.
            if not id_actual: 
                if legajo_iterado == legajo_usuario:
                    continue

            # --- FILTROS POR ROL ---
            if rol_usuario == "ADMIN":
                if rol_iterado != "ADMIN": continue
            elif rol_usuario == "SUPERVISOR":
                # Un supervisor debe poder verse a sí mismo si ya es el asignado
                if rol_iterado != "ADMIN" and legajo_iterado != legajo_usuario: 
                    continue
            elif rol_usuario == "BOMBERO":
                if rol_iterado != "SUPERVISOR": continue

            texto = f"{legajo_iterado} - {nombre_iterado} ({rol_iterado.capitalize()})"
            valores.append(texto)

        valores.sort()
        self.cb_asignado["values"] = valores
        
        # Si no hay actividad cargada, limpiamos
        if not id_actual:
            self.cb_asignado.set("")

    def activar_resaltado_foco(self, widget):
        try:
            # Entry / Text
            try:
                if not hasattr(widget, "_bg_original"):
                    widget._bg_original = widget.cget("background")
                widget.configure(background="#fff59d")
            except:
                pass

            # Combobox
            if isinstance(widget, ttk.Combobox):
                widget.configure(style="Focus.TCombobox")

        except:
            pass

    def desactivar_resaltado_foco(self, widget):
        try:
            # Entry / Text
            if hasattr(widget, "_bg_original"):
                try:
                    widget.configure(background=widget._bg_original)
                except:
                    pass

            # Combobox
            if isinstance(widget, ttk.Combobox):
                widget.configure(style="Custom.TCombobox")

        except:
            pass

    def mostrar_resultados_en_treeview(self, rows):
        """Muestra ventana con Treeview para elegir actividad cuando hay varios resultados."""
        
        # --- SOLUCIÓN AL ERROR 'ventana/win is not defined' ---
        # Creamos la ventana de nivel superior (Toplevel) usando el master de la clase
        win = tk.Toplevel(self.master) 
        win.title("Resultados de búsqueda")
        win.geometry("950x450")
        
        # Intentar aplicar el ícono global si el master lo tiene
        if hasattr(self.master, "_icono_global") and self.master._icono_global:
            win.iconphoto(True, self.master._icono_global)

        # Hacemos que la ventana sea modal (opcional, para que no toquen lo de atrás)
        win.transient(self.master)
        win.grab_set()

        frame = tk.Frame(win)
        frame.pack(fill="both", expand=True, padx=10, pady=10)

        # Definición de columnas
        cols = (
            "ID", "Legajo", "Apellido", "Nombre", "Actividad", "Área",
            "Fecha Inicio", "Fecha Fin", "Hora Inicio", "Hora Fin",
            "Horas", "Asignado Por"
        )

        self.tree_res = ttk.Treeview(frame, columns=cols, show="headings", height=12)

        # Scrollbars vinculadas
        scroll_y = ttk.Scrollbar(frame, orient="vertical", command=self.tree_res.yview)
        scroll_x = ttk.Scrollbar(frame, orient="horizontal", command=self.tree_res.xview)
        self.tree_res.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)

        # Layout con grid para los scrolls
        self.tree_res.grid(row=0, column=0, sticky="nsew")
        scroll_y.grid(row=0, column=1, sticky="ns")
        scroll_x.grid(row=1, column=0, sticky="ew")

        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(0, weight=1)

        # --- LEYENDA DE COLORES (Usando 'win' correctamente) ---
        leyenda_frame = tk.Frame(win)
        leyenda_frame.pack(fill="x", pady=5)

        colores = [
            ("#f5b7b1", "Pendiente"),
            ("#ffe08a", "Firmada Bombero"),
            ("#b7e4c7", "Firmada Supervisor"),
            ("#e6e6e6", "Anulada")
        ]

        for color, texto in colores:
            tk.Label(leyenda_frame, text="   ", bg=color, width=3).pack(side="left", padx=(10,2))
            lbl = tk.Label(leyenda_frame, text=texto)
            if texto == "Anulada":
                lbl.config(fg="#c0392b", font=("TkDefaultFont", 9, "bold"))
            lbl.pack(side="left", padx=(0,15))

        # Configuración de anchos de columna
        for c in cols:
            self.tree_res.heading(c, text=c)
            self.tree_res.column(c, width=100, minwidth=50)

        # --- CARGA DE DATOS ---
        for r in rows:
            # Determinación de tags para colores (índices según tu Row de SQLite)
            firma_bom_ok = bool(r[15]) # firma_bombero_fecha
            firma_sup_ok = bool(r[17]) # firma_supervisor_fecha
            es_anulada = r[18] == 1

            if es_anulada: tag = "anulada"
            elif firma_sup_ok: tag = "firmada_supervisor"
            elif firma_bom_ok: tag = "firmada_parcial"
            else: tag = "pendiente"

            # Valores a mostrar (ajustar índices si cambiaste el SELECT)
            # Usamos una estructura segura para evitar IndexError
            try:
                valores = (r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7], r[8], r[9], r[10], r[12])
                self.tree_res.insert("", "end", values=valores, tags=(tag,))
            except Exception as e:
                print(f"Error insertando fila: {e}")

        # Configuración de tags de colores
        self.tree_res.tag_configure("firmada_supervisor", background="#b7e4c7")
        self.tree_res.tag_configure("firmada_parcial", background="#ffe08a")
        self.tree_res.tag_configure("pendiente", background="#f5b7b1")
        self.tree_res.tag_configure("anulada", background="#e6e6e6", foreground="#c0392b")

        # --- EVENTO SELECCIÓN ---
        def on_select(event=None):
            selection = self.tree_res.selection()
            if not selection: return
            
            item_values = self.tree_res.item(selection[0], "values")
            id_seleccionado = item_values[0]
            
            # Llamamos a tu función de carga por ID
            self.cargar_actividad_por_id(id_seleccionado)
            win.destroy()

        self.tree_res.bind("<Double-1>", on_select)
        
        # Botón cerrar opcional
        tk.Button(win, text="Cerrar", command=win.destroy, width=15).pack(pady=10)

    def cargar_actividad_por_id(self, id_actividad):
        try:
            conn = sqlite3.connect(DB_PATH)
            conn.row_factory = sqlite3.Row
            c = conn.cursor()

            c.execute("""
                SELECT *
                FROM actividades
                WHERE id = ?
            """, (id_actividad,))

            row = c.fetchone()
            
            if not row:
                conn.close()
                self.ui.show_error("Error", "No se encontró la actividad.")
                return

            # --- NUEVO: Buscar nombre del Supervisor para el Combo ---
            asignado_nro = row["asignado"]
            texto_asignado = str(asignado_nro)
            
            if asignado_nro:
                c.execute("SELECT apellido, nombre FROM legajos WHERE legajo = ?", (asignado_nro,))
                res_sup = c.fetchone()
                if res_sup:
                    # Formato: "121 - APELLIDO NOMBRE"
                    texto_asignado = f"{asignado_nro} - {res_sup['apellido']} {res_sup['nombre']}"
            
            conn.close()

            # -------------------------------------------------
            # CARGAR DATOS EN FORMULARIO
            # -------------------------------------------------
            self._cargar_actividad_row(row)
            
            # Forzamos al combo a mostrar el nombre y no solo el número
            if hasattr(self, 'combo_asignado'):
                self.combo_asignado.set(texto_asignado)

            # -------------------------------------------------
            # SINCRONIZAR ESTADO LÓGICO REAL
            # -------------------------------------------------
            self.id_actividad_actual = row["id"] # Asegúrate de tener el ID actual
            self.actividad_anulada = bool(row["anulada"])
            self.firma_bombero_fecha = row["firma_bombero_fecha"]
            self.firma_supervisor_fecha = row["firma_supervisor_fecha"]
            self.actividad_asignado = str(row["asignado"]).strip() # Guardamos legajo como string para comparar
            self._legajo_original_actividad = str(row["creado_por"]).strip()
            
            # Recalcular permite que el sistema sepa quién es el dueño/asignado
            self._recalcular_estado_actividad()

            # -------------------------------------------------
            # ESTADO UI
            # -------------------------------------------------
            self._estado_ui = "cargado"

            # Aplicar estado visual y botones
            self._estado_visual_firma("cargado") 
            self._refrescar_actividades() 

        except Exception as e:
            print(f"Error al cargar actividad {id_actividad}: {e}")
            self.ui.show_error("Error", f"No se pudo cargar la actividad: {e}")

    def _autoajustar_columnas_tree(self, tree, max_width=300):
        f = font.nametofont("TkDefaultFont")

        for col in tree["columns"]:
            # ancho mínimo: el título
            max_ancho = f.measure(col)

            for item in tree.get_children():
                valor = str(tree.set(item, col))
                ancho = f.measure(valor)
                if ancho > max_ancho:
                    max_ancho = ancho

            # padding + límite
            tree.column(col, width=min(max_ancho + 20, max_width))

    def _cargar_actividad_row(self, r):
        """
        Carga en el formulario de actividades los datos de la fila obtenida de la DB.
        r debe tener: id, legajo, actividad, area, fecha_inicio, fecha_fin,
                    hora_inicio, hora_fin, horas, descripcion, asignado, concepto_id,
                    firma_bombero_usuario, firma_bombero_fecha,
                    firma_supervisor_usuario, firma_supervisor_fecha
        """

        print("TIPO DE r:", type(r))
        print("LEN DE r:", len(r))
        self.tab_control.select(self.actividades_frame)

        # ---------------------------
        # Desempaquetar (una sola vez)
        # ---------------------------
        # --- Mapear solo las primeras 17 columnas necesarias ---
        id_ = r["id"]
        legajo = r["legajo"]
        actividad_txt = r["actividad"]
        area_txt = r["area"]
        f_ini = r["fecha_inicio"]
        f_fin = r["fecha_fin"]
        h_ini = r["hora_inicio"]
        h_fin = r["hora_fin"]
        horas_trabajadas = r["horas"]
        descripcion_txt = r["descripcion"]
        asignado = r["asignado"]
        concepto_id = r["concepto_id"]
        firma_bombero_usuario = r["firma_bombero_usuario"]
        firma_bombero_fecha = r["firma_bombero_fecha"]
        firma_supervisor_usuario = r["firma_supervisor_usuario"]
        firma_supervisor_fecha = r["firma_supervisor_fecha"]
        anulada = r["anulada"]

        print("firma_bombero_usuario:", firma_bombero_usuario)
        print("firma_bombero_fecha:", firma_bombero_fecha)
        print("firma_supervisor_usuario:", firma_supervisor_usuario)
        print("firma_supervisor_fecha:", firma_supervisor_fecha)
        print("anulada:", anulada)

        # --- 1. Guardar firmas y estados básicos ---
        self.id_actividad_actual = id_
        self.var_id_actividad.set(str(id_))
        self.firma_bombero_usuario = firma_bombero_usuario
        self.firma_bombero_fecha = firma_bombero_fecha
        self.firma_supervisor_usuario = firma_supervisor_usuario
        self.firma_supervisor_fecha = firma_supervisor_fecha
        self.actividad_anulada = bool(anulada)
        self.actividad_actual = r

        # --- 2. Normalización del Asignado (CLAVE PARA EL ADMIN) ---
        # Extraemos solo el número (ej: de "140 - FERRUCCI" o de 140 sacamos "140")
        asig_limpio = str(asignado).split("-")[0].split(" ")[0].strip()
        
        self.asig_actual = asig_limpio
        self.actividad_asignado = asig_limpio 
        
        # Guardamos el legajo del autor de la actividad
        self.actividad_legajo = str(legajo).strip()

        # --- 3. Determinar estado lógico ---
        row_dict = {
            "firma_bombero_fecha": firma_bombero_fecha,
            "firma_supervisor_fecha": firma_supervisor_fecha,
            "anulada": self.actividad_anulada
        }
        self._estado_actividad = self._determinar_estado_actividad(row_dict)
        # ---------------------------
        # Refrescar combos
        # ---------------------------
        self.cargar_legajos_combobox()
        self.cargar_asignados_autorizados()

        # --- Combo legajo ---
        legajo_str = ""
        for v in (self.cb_legajo_sel["values"] or []):
            if str(legajo) == str(v).split(" - ")[0]:
                legajo_str = v
                break
        self.cb_legajo_sel.set(legajo_str if legajo_str else str(legajo))

        # Traer apellido y nombre
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        c = conn.cursor()
        c.execute("SELECT apellido, nombre FROM legajos WHERE legajo=?", (legajo,))
        row = c.fetchone()
        conn.close()
        self.var_apn_actividades.set((f"{row[0]} {row[1]}" if row else "").upper())

        # --- Combo actividad ---
        actividad_a_mostrar = ""

        if concepto_id and hasattr(self, 'conceptos_map') and concepto_id in self.conceptos_map:
            actividad_a_mostrar = f"{concepto_id} - {self.conceptos_map[concepto_id]}"

        if not actividad_a_mostrar and actividad_txt:
            actividad_a_mostrar = actividad_txt.upper()

        if hasattr(self, "conceptos_map"):
            self.actividad["values"] = [
                f"{cid} - {txt}" for cid, txt in self.conceptos_map.items()
            ]
        self.actividad.set(actividad_a_mostrar)

        # --- Área ---
        self.area.config(state="readonly")
        self.area.set((area_txt or "").upper())

        # --- Fechas ---
        try:
            if parse_ddmmyyyy(f_ini):
                self.fecha_inicio.config(state="normal")
                self.fecha_inicio.set_date(parse_ddmmyyyy(f_ini))
                self.fecha_inicio.config(state="disabled")
        except:
            pass

        try:
            if parse_ddmmyyyy(f_fin):
                self.fecha_fin.config(state="normal")
                self.fecha_fin.set_date(parse_ddmmyyyy(f_fin))
                self.fecha_fin.config(state="disabled")
        except:
            pass

        # --- Horas ---
        self.hora_inicio.config(state="normal")
        self.hora_inicio.delete(0, END)
        self.hora_inicio.insert(0, (h_ini or "").upper())
        self.hora_inicio.config(state="disabled")

        self.hora_finalizacion.config(state="normal")
        self.hora_finalizacion.delete(0, END)
        self.hora_finalizacion.insert(0, (h_fin or "").upper())
        self.hora_finalizacion.config(state="disabled")

        # --- Horas trabajadas ---
        try:
            horas_val = horas_trabajadas or 0
            if horas_val:
                horas_enteras = int(horas_val)
                minutos = int(round((horas_val - horas_enteras) * 60))
                if minutos >= 60:
                    horas_enteras += 1
                    minutos -= 60
                display = f"{horas_enteras}:{minutos:02d}"
                print(f"DEBUG CARGAR >> {horas_val} → {display}")
                self.var_horas.set(display)
            else:
                self.var_horas.set("0")
        except Exception as e:
            print(f"ERROR convirtiendo horas al cargar: {e}")
            self.var_horas.set("0")

        # --- Descripción ---
        try:
            self._set_text_readonly(self.descripcion, False)
            self.descripcion.delete("1.0", END)
            self.descripcion.insert("1.0", (descripcion_txt or "").upper())
            self._set_text_readonly(self.descripcion, True)
        except Exception:
            pass

        # --- Combo asignado ---
        if hasattr(self, "cb_asignado"):
            self.cb_asignado.config(state="normal")
            
            # 1. Tomamos el valor de la base de datos
            val_id = str(r["asignado"]).split("-")[0].strip()
            
            # 2. Buscamos el nombre en el dict global
            usuarios = getattr(self, "usuarios_dict", {})
            user_info = usuarios.get(val_id)
            
            if user_info:
                texto_final = f"{val_id} - {user_info.get('nombre', '')}"
            else:
                texto_final = val_id

            # 3. Inyectamos el valor para que sea legal en el combo
            self.cb_asignado["values"] = [texto_final] 
            self.cb_asignado.set(texto_final)
            
            # 4. Lo bloqueamos
            self.cb_asignado.config(state="readonly")

        # ---------------------------
        # Estado ANULADA
        # ---------------------------
        self.actividad_anulada = bool(anulada)
        if self.actividad_anulada:

            self.lbl_estado_anulada.config(
                text="*** ACTIVIDAD ANULADA ***",
                fg="white",
                bg="red",
                font=("Arial", 12, "bold")
            )
        else:
            self.lbl_estado_anulada.config(
                text="",
                bg=self.actividades_frame.cget("bg")
            )

        # --- 1. SETEAR VARIABLES CRÍTICAS DE CLASE ---
        # Guardamos todo antes de llamar a las funciones de refresco
        self.id_actividad_actual = id_
        self.var_id_actividad.set(str(id_))
        self.actividad_legajo = str(legajo).strip()
        self.actividad_asignado = str(asignado).split("-")[0].strip()
        
        # Guardamos firmas para que _puede_modificar_actividad las vea
        self.firma_bombero_fecha = firma_bombero_fecha
        self.firma_supervisor_fecha = firma_supervisor_fecha
        self.actividad_anulada = bool(anulada)
        
        # --- 2. REFRESCAR BOTONES Y UI ---
        self._estado_ui = "cargado"
        self._refrescar_actividades() 

        # --- 3. BLOQUEO FINAL DE COMBOS (Seguridad total) ---
        if hasattr(self, "cb_legajo_sel"):
            self.cb_legajo_sel.config(state="disabled")
        
        if hasattr(self, "cb_asignado"):
            # Obtenemos el nombre del dict para que el combo no quede solo con el número
            usuarios = getattr(self, "usuarios_dict", {})
            u_info = usuarios.get(self.actividad_asignado)
            texto_full = f"{self.actividad_asignado} - {u_info.get('nombre', '')}" if u_info else self.actividad_asignado
            
            self.cb_asignado.config(state="normal")
            self.cb_asignado.set(texto_full)
            self.cb_asignado.config(state="disabled")
            
        print(f">>> CARGA EXITOSA: ID={id_} | Autor={self.actividad_legajo} | Asignado={self.actividad_asignado}")
 
    def imprimir_actividad(self):
        import tempfile
        import os

        id_act = self.var_id_actividad.get().strip()
        if not id_act:
            messagebox.showwarning("Atención", "Debe seleccionar una actividad.")
            return

        # 📄 archivo temporal
        tmp_dir = tempfile.gettempdir()
        tmp_path = os.path.join(tmp_dir, f"Actividad_{id_act}.pdf")

        # reutiliza EXACTAMENTE el mismo armado
        self._crear_pdf_actividad(tmp_path)

        # abrir para imprimir
        try:
            os.startfile(tmp_path)
        except Exception as e:
#            from tkinter.messagebox import showwarning
            messagebox.showwarning(
                "Aviso",
                f"El PDF se creó pero no se pudo abrir automáticamente.\n{e}"
            )

    def _crear_pdf_actividad(self, path):
        import sqlite3
        from reportlab.platypus import Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors

        id_act = self.var_id_actividad.get().strip()
        if not id_act:
            raise Exception("ID de actividad no válido")

        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        c = conn.cursor()

        c.execute("""
            SELECT id, legajo, asignado, actividad, area,
                fecha_inicio, fecha_fin, hora_inicio, hora_fin,
                descripcion, horas,
                firma_bombero_usuario, firma_bombero_fecha,
                firma_supervisor_usuario, firma_supervisor_fecha,
                anulada, motivo_anulacion, usuario_anula, fecha_anulacion
            FROM actividades
            WHERE id = ?
        """, (id_act,))
        row = c.fetchone()

        if not row:
            conn.close()
            raise Exception("Actividad no encontrada")

        (
            id_act, legajo, asignado, actividad, area,
            fecha_ini, fecha_fin, hora_ini, hora_fin,
            descripcion, horas,
            fb_user, fb_fecha,
            fs_user, fs_fecha,
            anulada, motivo_anulacion, usuario_anula, fecha_anulacion
        ) = row

        # --- Normalizar formato de horas ---
        def formatear_hora(h):
            if not h:
                return ""
            h = str(h)
            if len(h) >= 5:
                return h[:5]   # HH:MM
            return h

        hora_ini = formatear_hora(hora_ini)
        hora_fin = formatear_hora(hora_fin)

        # ✅ Determinar si está anulada
        es_anulada = (anulada == 1)

        # === Datos bombero ===
        c.execute("SELECT apellido, nombre, dni FROM legajos WHERE legajo=?", (legajo,))
        r_leg = c.fetchone()
        leg_ap, leg_nom, leg_dni = r_leg if r_leg else ("", "", "")

        # === Datos supervisor ===
        asig_ap = asig_nom = asig_dni = ""
        asig_str = ""
        if asignado:
            c.execute("SELECT apellido, nombre, dni FROM legajos WHERE legajo=?", (asignado,))
            r_asig = c.fetchone()
            if r_asig:
                asig_ap, asig_nom, asig_dni = r_asig
                asig_str = f"{asignado} - {asig_ap} {asig_nom}"

        conn.close()

        # === Estado histórico (NO debe cambiar por anulación) ===
        if fs_fecha:
            estado = "FIRMADA POR SUPERVISOR"
            estado_color = colors.green
        elif fb_fecha:
            estado = "FIRMADA POR BOMBERO"
            estado_color = colors.orange
        else:
            estado = "PENDIENTE DE FIRMA"
            estado_color = colors.red
        
        # === Verificar si tiene historial (actividad modificada real) ===
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        c = conn.cursor()

        c.execute("""
            SELECT COUNT(*)
            FROM actividades_historial
            WHERE actividad_id = ?
        """, (id_act,))

        tiene_historial = c.fetchone()[0] > 0
        conn.close()

        if tiene_historial and not es_anulada:
            estado += " (MODIFICADA)"
            estado_color = colors.orange

        # === Texto firmas ===
        firma_bombero_txt = (
            f"Firmado digitalmente<br/>{fb_fecha}"
            if fb_fecha else "No firmada digitalmente"
        )

        firma_supervisor_txt = (
            f"Firmado digitalmente<br/>{fs_fecha}"
            if fs_fecha else "No firmada digitalmente"
        )

        styles = getSampleStyleSheet()
        elems = []

        estilo_anulada = ParagraphStyle(
            "AnuladaStyle",
            parent=styles["Heading2"],
            fontSize=14,
            textColor=colors.red,
            alignment=1,
            spaceAfter=6
        )

        estilo_motivo = ParagraphStyle(
            "MotivoStyle",
            parent=styles["Normal"],
            textColor=colors.red,
            spaceAfter=5
        )

        # === TÍTULO PRINCIPAL CON Nº ===
        estilo_titulo_principal = ParagraphStyle(
            "TituloActividad",
            parent=styles["Title"],
            fontSize=16,      # antes ~18-20
            spaceAfter=6
        )

        elems.append(
            Paragraph(f"REPORTE DE ACTIVIDAD Nº {id_act}", estilo_titulo_principal)
        )
        elems.append(Spacer(1, 10))

        # === ESTADO DE LA ACTIVIDAD ===
        estilo_estado = ParagraphStyle(
            "EstadoActividad",
            parent=styles["Normal"],
            fontSize=13,
            textColor=estado_color,
            alignment=1,
            spaceAfter=8
        )

        elems.append(
            Paragraph(f"ESTADO: {estado}", estilo_estado)
        )
        elems.append(Spacer(1, 8))

        if es_anulada:
            # Línea superior
            linea = Table([[""]], colWidths=[470])
            linea.setStyle(TableStyle([
                ("LINEABOVE", (0, 0), (-1, -1), 1.5, colors.red),
            ]))
            elems.append(linea)
            elems.append(Spacer(1, 6))

            # Título central
            elems.append(
                Paragraph("ACTIVIDAD ANULADA", estilo_anulada)
            )
            elems.append(Spacer(1, 6))

            # Motivo
            motivo_txt = motivo_anulacion if motivo_anulacion else "Sin motivo registrado"
            elems.append(
                Paragraph(f"<b>Motivo:</b> {motivo_txt}", estilo_motivo)
            )

            # Usuario y fecha
            if usuario_anula:
                elems.append(
                    Paragraph(f"<b>Anulada por:</b> {usuario_anula}", estilo_motivo)
                )

            if fecha_anulacion:
                elems.append(
                    Paragraph(f"<b>Fecha de anulación:</b> {fecha_anulacion}", estilo_motivo)
                )

            elems.append(Spacer(1, 6))

            # Línea inferior
            linea2 = Table([[""]], colWidths=[470])
            linea2.setStyle(TableStyle([
                ("LINEBELOW", (0, 0), (-1, -1), 1.5, colors.red),
            ]))
            elems.append(linea2)

            elems.append(Spacer(1, 15))

        # === BLOQUE BOMBERO (SEPARADO Y RESPIRADO) ===
        elems.append(
            Paragraph(f"<b>LEGAJO:</b> {legajo}", styles["Normal"])
        )
        elems.append(Spacer(1, 6))
        elems.append(
            Paragraph(f"<b>BOMBERO:</b> {leg_ap} {leg_nom}", styles["Normal"])
        )
        elems.append(Spacer(1, 16))

        # === TABLA DE DATOS ===
        labels = [
            "Asignado por",
            "Actividad", "Área",
            "Fecha Inicio", "Fecha Fin",
            "Hora Inicio", "Hora Fin",
            "Total Hs", "Descripción"
        ]

        valores = [
            asig_str,
            actividad,
            area,
            fecha_ini,
            fecha_fin,
            hora_ini,
            hora_fin,
            f"{horas:.2f}" if horas else "",
            descripcion
        ]

        data = []
        for lab, val in zip(labels, valores):
            data.append([
                Paragraph(f"<b>{lab}:</b>", styles["Normal"]),
                Paragraph(str(val or ""), styles["Normal"])
            ])

        tbl = Table(data, colWidths=[120, 350])
        tbl.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LINEBELOW", (0, 0), (-1, -1), 0.25, colors.grey),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        elems.append(tbl)

        # === FIRMAS ===
        elems.append(Spacer(1, 40))

        firma_tbl = Table([
            [
                Paragraph(
                    "<para alignment='left'>"
                    "__________________________<br/>"
                    "<b>Firma del Bombero</b><br/>"
                    f"{leg_ap} {leg_nom}<br/>"
                    f"DNI: {leg_dni}<br/><br/>"
                    f"<i>{firma_bombero_txt}</i>"
                    "</para>",
                    styles["Normal"]
                ),
                Paragraph(
                    "<para alignment='right'>"
                    "__________________________<br/>"
                    "<b>Firma del Responsable</b><br/>"
                    f"{asig_ap} {asig_nom}<br/>"
                    f"DNI: {asig_dni}<br/><br/>"
                    f"<i>{firma_supervisor_txt}</i>"
                    "</para>",
                    styles["Normal"]
                )
            ]
        ], colWidths=[250, 250])

        firma_tbl.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
            ("TOPPADDING", (0, 0), (-1, -1), 25),
        ]))

        elems.append(firma_tbl)

        # === PDF FINAL ===
        estado_real = self._determinar_estado_actividad(dict(row))

        self._crear_pdf_unificado(
            path,
            elems,
            titulo=None   # ← importante
        )

    def imprimir_listado_actividades(self):
        """Genera el listado de actividades con formato unificado (horizontal)."""
        periodo = self._pedir_periodo_mes_anio()
        if not periodo:
            return

        mes, anio = periodo
        periodo_sql = f"{anio:04d}-{mes:02d}"

        from datetime import date

        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        c = conn.cursor()

        periodo_sql = f"{anio:04d}-{mes:02d}"

        c.execute("""
            SELECT id, legajo, asignado, actividad, area,
                fecha_inicio, fecha_fin, hora_inicio, hora_fin,
                descripcion, horas,
                firma_bombero_fecha,
                firma_supervisor_fecha
            FROM actividades
            WHERE strftime('%Y-%m',
                substr(fecha_inicio, 7, 4) || '-' ||
                substr(fecha_inicio, 4, 2) || '-' ||
                substr(fecha_inicio, 1, 2)
            ) = ?
            ORDER BY fecha_inicio DESC, id DESC
        """, (periodo_sql,))

        rows = c.fetchall()
        conn.close()

        if not rows:
            self.ui.show_error("Sin datos", f"No hay actividades registradas en {mes:02d}/{anio}.")
            return

        sugerido = self._default_filename("Listado_Actividades")
        path = self.ui.ask_save_file(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            initialfile=sugerido
        )
        if not path:
            return

        styles = getSampleStyleSheet()
        elems = []

        # HEADERS CON ID - SIN ÁREA, FECHA FIN Y HORA FIN
        headers = ["ID", "Legajo", "Asignado", "Actividad", "F.Inicio", "H.Inicio", "Descripción", "Total Hs", "EST"]
        
        # FUNCIÓN SIMPLIFICADA - NO DIVIDE PALABRAS
        def dividir_texto(texto, max_caracteres):
            if not texto or len(texto) <= max_caracteres:
                return texto
            
            palabras = texto.split()
            lineas = []
            linea_actual = ""
            
            for palabra in palabras:
                # Si la palabra individual es demasiado larga, dejarla completa
                # en su propia línea en lugar de dividirla
                if len(palabra) > max_caracteres:
                    if linea_actual:
                        lineas.append(linea_actual)
                    lineas.append(palabra)  # Palabra completa en línea separada
                    linea_actual = ""
                else:
                    # Verificar si cabe en la línea actual
                    prueba_linea = linea_actual + " " + palabra if linea_actual else palabra
                    if len(prueba_linea) <= max_caracteres:
                        linea_actual = prueba_linea
                    else:
                        if linea_actual:
                            lineas.append(linea_actual)
                        linea_actual = palabra
            
            if linea_actual:
                lineas.append(linea_actual)
                
            return '\n'.join(lineas)
        
        def estado_simbolo(fs_fecha, fb_fecha):
            if fs_fecha:
                return "✔"   # Firmada por supervisor
            elif fb_fecha:
                return "✍"   # Firmada por bombero
            else:
                return "!"   # Pendiente

        # PREPARAR DATOS CON TEXTO FORMATEADO
        data = [headers]

        # ---- CONTADORES DE ESTADO ----
        cont_aprobadas = 0
        cont_firmadas = 0
        cont_pendientes = 0

        for r in rows:
            # Formatear texto - límites aumentados para 3 letras más
            actividad = dividir_texto(str(r[3] or ""), 38)  # 35 + 3 = 38 caracteres
            descripcion = dividir_texto(str(r[9] or ""), 43)  # 40 + 3 = 43 caracteres
            
            # Obtener estado
            fb_fecha = r[11]
            fs_fecha = r[12]
            estado = estado_simbolo(fs_fecha, fb_fecha)

            # Contadores
            if fs_fecha:
                cont_aprobadas += 1
            elif fb_fecha:
                cont_firmadas += 1
            else:
                cont_pendientes += 1

            data.append([
                r[0],  # ID
                r[1],  # Legajo
                r[2],  # Asignado
                actividad,
                r[5],  # F.Inicio
                r[7],  # H.Inicio
                descripcion,
                f"{r[10]:.2f}" if r[10] else "",
                estado
            ])

        # ANCHOS DE COLUMNA MÁS AMPLIOS (aproximadamente 3 letras más)
        col_widths = [
            30,   # ID
            40,   # Legajo
            60,   # Asignado
            165,  # Actividad (ligeramente reducido)
            55,   # F.Inicio
            45,   # H.Inicio
            205,  # Descripción (ligeramente reducido)
            45,   # Total Hs
            30    # EST
        ]

        tbl = Table(data, repeatRows=1, colWidths=col_widths)
        tbl.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.4, colors.black),
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("ALIGN", (3, 1), (3, -1), "LEFT"),
            ("ALIGN", (6, 1), (6, -1), "LEFT"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        elems.append(tbl)
        # ---- LEYENDA DE ESTADOS ----
        elems.append(Spacer(1, 10))

        leyenda_style = styles["Normal"]
        leyenda_style.fontSize = 8

        elems.append(Paragraph(
            "<b>Leyenda:</b> "
            "✔ Firmada por supervisor &nbsp;&nbsp;&nbsp; "
            "✍ Firmada por bombero &nbsp;&nbsp;&nbsp; "
            "⚠ Pendiente",
            leyenda_style
        ))

        # ---- RESUMEN DE ESTADOS ----
        total = cont_aprobadas + cont_firmadas + cont_pendientes

        elems.append(Spacer(1, 6))

        resumen_style = styles["Normal"]
        resumen_style.fontSize = 9

        texto_resumen = (
            f"<b>Resumen:</b> "
            f"Aprobadas: {cont_aprobadas} &nbsp;&nbsp;&nbsp; "
            f"Firmadas: {cont_firmadas} &nbsp;&nbsp;&nbsp; "
            f"Pendientes: {cont_pendientes} &nbsp;&nbsp;&nbsp; "
            f"Total: {total}"
        )

        elems.append(Paragraph(texto_resumen, resumen_style))

        try:
            # Nombre del mes en texto
            meses = [
                "", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
            ]
            titulo_pdf = f"Listado de Actividades – {meses[mes]} {anio}"

            self._crear_pdf_unificado(path, elems, titulo_pdf, landscape_mode=True)
            self.ui.show_info("OK", f"Listado PDF generado correctamente:\n{path}")

        except Exception as e:
            self.ui.show_error("Error", f"No se pudo generar el PDF:\n{e}")

    def _pedir_periodo_mes_anio(self):
        top = tk.Toplevel(self.master)
        top.title("Seleccionar período")
        top.geometry("320x180")
        top.resizable(False, False)
        top.grab_set()

        hoy = datetime.now()

        frame = tk.Frame(top, padx=20, pady=15)
        frame.pack(fill="both", expand=True)

        # ----- MES -----
        tk.Label(frame, text="Mes:", font=("Arial", 11, "bold")).grid(row=0, column=0, sticky="w", pady=5)

        meses = [
            "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
            "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
        ]

        combo_mes = ttk.Combobox(frame, values=meses, state="readonly", width=15)
        combo_mes.current(hoy.month - 1)
        combo_mes.grid(row=0, column=1, pady=5)

        # ----- AÑO -----
        tk.Label(frame, text="Año:", font=("Arial", 11, "bold")).grid(row=1, column=0, sticky="w", pady=5)

        combo_anio = ttk.Combobox(
            frame,
            values=[hoy.year - 1, hoy.year, hoy.year + 1],
            state="readonly",
            width=10
        )
        combo_anio.set(hoy.year)
        combo_anio.grid(row=1, column=1, pady=5)

        resultado = {"valor": None}

        def aceptar():
            mes = combo_mes.current() + 1
            anio = int(combo_anio.get())
            resultado["valor"] = (mes, anio)
            top.destroy()

        def cancelar():
            top.destroy()

        # ----- BOTONES -----
        frame_botones = tk.Frame(frame, pady=15)
        frame_botones.grid(row=2, column=0, columnspan=2, sticky="ew")

        tk.Button(
            frame_botones,
            text="Aceptar",
            width=12,
            bg="#2e7d32",
            fg="white",
            font=("Arial", 10, "bold"),
            command=aceptar
        ).pack(side="left", padx=10)

        tk.Button(
            frame_botones,
            text="Cancelar",
            width=12,
            bg="gray",
            fg="white",
            font=("Arial", 10, "bold"),
            command=cancelar
        ).pack(side="right", padx=10)

        combo_mes.focus()
        self.master.wait_window(top)
        return resultado["valor"]

    # ===================== pestaña CONCEPTOS =====================****************************************************************
    def imprimir_listado_conceptos(self):
        """Genera el listado de conceptos en formato HORIZONTAL con 3 columnas."""
        # --- 1. IMPORTACIONES NECESARIAS (Agregá estas líneas aquí) ---
        from reportlab.platypus import Table, TableStyle, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from datetime import datetime
        import os

        # --- Verificar si hay datos ---
        items = self.tree_conceptos.get_children()
        if not items:
            self.ui.show_info("Imprimir", "No hay conceptos para imprimir.")
            return

        # --- Nombre sugerido del archivo ---
        fecha_archivo = datetime.now().strftime("%Y-%m-%d")
        nombre_sugerido = f"Listado_Conceptos_{fecha_archivo}.pdf"
        file_path = self.ui.ask_save_file(
            defaultextension=".pdf",
            initialfile=nombre_sugerido,
            filetypes=[("PDF files", "*.pdf")],
            title="Guardar listado de conceptos como PDF"
        )
        if not file_path:
            return

        # --- 2. Definir Estilos ---
        styles = getSampleStyleSheet()
        style_header = styles["Heading5"]
        style_header.fontSize = 9
        
        style_normal = styles["Normal"]
        style_normal.fontSize = 8
        style_normal.leading = 10

        # --- 3. Armar datos (Solo 3 columnas) ---
        headers = ["Código", "Concepto / Actividad", "Puntos"]
        data = [[Paragraph(h, style_header) for h in headers]]

        for item in items:
            # [:3] evita la cuarta columna fantasma
            valores = list(self.tree_conceptos.item(item, "values"))[:3]
            fila = [Paragraph(str(v), style_normal) for v in valores]
            data.append(fila)

        # --- 4. Configuración de Tabla ---
        ancho_pag, _ = landscape(A4)
        ancho_util = ancho_pag - 100 

        colWidths = [
            60,                 # Código
            ancho_util - 120,   # Concepto / Actividad
            60                  # Puntos
        ]

        tbl = Table(data, repeatRows=1, colWidths=colWidths)
        tbl.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("BACKGROUND", (0, 0), (-1, 0), colors.gray),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (0, 0), (0, -1), "CENTER"), 
            ("ALIGN", (2, 0), (2, -1), "CENTER"), 
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))

        elems = [tbl]

        # --- 5. Crear PDF ---
        try:
            self._crear_pdf_unificado(
                file_path, 
                elems, 
                "Listado de Conceptos / Actividades",
                landscape_mode=True
            )
            
            self.ui.show_info("Imprimir", f"Listado guardado en:\n{file_path}")
            try:
                os.startfile(file_path)
            except Exception:
                pass
        except Exception as e:
            self.ui.show_error("Error", f"No se pudo generar el PDF:\n{e}")

    def init_conceptos(self):
        Label(self.conceptos_frame, text="CONCEPTOS", font=("Arial", 18, "bold"),
            fg="white", bg="red").place(x=400, y=6)

        y0, dy = 46, 28

        # ID
        Label(self.conceptos_frame, text="Código:", fg="white", bg="red").place(x=20, y=y0)
        self.var_id_concepto = StringVar()
        self.e_id_concepto = Entry(self.conceptos_frame, textvariable=self.var_id_concepto,
                                font=("Arial", 11), state="readonly", fg="black", bg="white")
        self.e_id_concepto.place(x=180, y=y0, width=160)

        # Concepto/Actividad
        Label(self.conceptos_frame, text="Concepto/Actividad:", fg="white", bg="red").place(x=20, y=y0+dy)
        self.var_concepto = StringVar()
        self.enforce_uppercase_var(self.var_concepto)
        self.e_concepto = Entry(self.conceptos_frame, textvariable=self.var_concepto,
                                font=("Arial", 11), state="disabled")
        self.e_concepto.place(x=180, y=y0+dy, width=360)

        # Puntos (solo números hasta 100)
        def validar_puntos(P):
            return (P.isdigit() and (int(P) <= 100)) or P == ""
        vcmd_puntos = (self.master.register(validar_puntos), "%P")
        Label(self.conceptos_frame, text="Puntos:", fg="white", bg="red").place(x=20, y=y0+2*dy)
        self.var_puntos = StringVar()
        self.e_puntos = Entry(self.conceptos_frame, textvariable=self.var_puntos,
                            font=("Arial", 11), validate="key", validatecommand=vcmd_puntos,
                            state="disabled")
        self.e_puntos.place(x=180, y=y0+2*dy, width=160)

        # Detalle (como Text)
        Label(self.conceptos_frame, text="Detalle:", fg="white", bg="red").place(x=20, y=y0+3*dy)
        self.txt_detalle = Text(self.conceptos_frame, font=("Arial", 11), height=3, width=40, state="disabled")
        self.txt_detalle.place(x=180, y=y0+3*dy, width=360, height=60)
        self.txt_detalle.bind("<KeyRelease>", self._detalle_uppercase)

        # --- BOTONES DIVIDIDOS EN DOS GRUPOS ARRIBA DE LA GRILLA ---
        botones_grupo1 = [
            ("Nuevo", self.nuevo_concepto),
            ("Guardar", self.guardar_concepto),
            ("Modificar", self._habilitar_modificar_concepto),
            ("Dar de baja", self.toggle_estado_concepto)
        ]
        
        botones_grupo2 = [
            ("Limpiar/Cancelar", self.limpiar_concepto),
            ("Buscar", self.buscar_concepto),
            ("Imprimir listado", self.imprimir_listado_conceptos)
        ]

        self.conc_btns = {}
        
        # Grupo 1 - Lado izquierdo
        x_grupo1 = 600
        y_grupo1 = y0 + 10  # Un poco más abajo que los campos
        
        for i, (txt, cmd) in enumerate(botones_grupo1):
            b = Button(self.conceptos_frame, text=txt, width=15, command=cmd)
            b.place(x=x_grupo1, y=y_grupo1 + (i * 35))
            self.conc_btns[txt] = b

        # Grupo 2 - Lado derecho
        x_grupo2 = x_grupo1 + 150  # Separación entre grupos
        
        for i, (txt, cmd) in enumerate(botones_grupo2):
            b = Button(self.conceptos_frame, text=txt, width=15, command=cmd)
            b.place(x=x_grupo2, y=y_grupo1 + (i * 35))
            self.conc_btns[txt] = b

        # 🔹 Depuración opcional
        self._estado_botones_concepto("inicial")
        # Forzar que el botón Buscar tenga el comando (seguridad para algunas versiones de Tk)
        self.conc_btns["Buscar"].config(command=self.buscar_concepto)
        # Bind de click (por si acaso alguna versión/overlay impide el command)
        self.conc_btns["Buscar"].bind("<Button-1>", lambda e: self.buscar_concepto())

        if self.usuario_actual["rol"] != "ADMIN":
            self.leg_btns["Eliminar"].config(state="disabled")

        # --- GRILLA MÁS ANCHA Y CON SCROLL HORIZONTAL FUNCIONAL ---
        cols = ("Código", "Concepto/Actividad", "Puntos", "Detalle")
        
        # Frame contenedor para Treeview y scrollbars - MÁS ANCHO
        tree_frame = Frame(self.conceptos_frame, bg="white")
        tree_frame.place(x=20, y=y_grupo1 + 165, width=1000, height=160)  # Ancho aumentado a 1050
        
        # Scrollbars
        scroll_y = ttk.Scrollbar(tree_frame, orient="vertical")
        scroll_x = ttk.Scrollbar(tree_frame, orient="horizontal")
        
        # Treeview configurado para scroll horizontal - CLAVE: establecer ancho total mayor que la suma de columnas
        self.tree_conceptos = ttk.Treeview(
            tree_frame, 
            columns=cols, 
            show="headings",
            yscrollcommand=scroll_y.set,
            xscrollcommand=scroll_x.set,
            height=6
        )
        
        # Configurar scrollbars
        scroll_y.config(command=self.tree_conceptos.yview)
        scroll_x.config(command=self.tree_conceptos.xview)
        
        # Posicionar elementos en el frame usando grid con weights
        self.tree_conceptos.grid(row=0, column=0, sticky="nsew")
        scroll_y.grid(row=0, column=1, sticky="ns")
        scroll_x.grid(row=1, column=0, sticky="ew")
        
        # Configurar el grid para que se expanda correctamente
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        tree_frame.grid_rowconfigure(1, weight=0)  # Fila para scroll horizontal

        # Configurar columnas con anchos específicos - SUMA MAYOR QUE EL ANCHO DEL TREEVIEW
        # Ancho total de columnas: 100 + 350 + 100 + 600 = 1150 (mayor que el ancho del treeview)
        self.tree_conceptos.column("Código", width=100, anchor="center", stretch=False)
        self.tree_conceptos.column("Concepto/Actividad", width=350, anchor="w", stretch=False)
        self.tree_conceptos.column("Puntos", width=100, anchor="center", stretch=False)
        self.tree_conceptos.column("Detalle", width=600, anchor="w", stretch=False)  # Ancho aumentado

        # Configurar headings
        self.tree_conceptos.heading("Código", text="Código",
            command=lambda: self._sort_conceptos("Código"))
        self.tree_conceptos.heading("Concepto/Actividad", text="Concepto/Actividad",
            command=lambda: self._sort_conceptos("Concepto/Actividad"))
        self.tree_conceptos.heading("Puntos", text="Puntos",
            command=lambda: self._sort_conceptos("Puntos"))
        self.tree_conceptos.heading("Detalle", text="Detalle",
            command=lambda: self._sort_conceptos("Detalle"))

        # 🔹 CONFIGURACIÓN CRÍTICA: Habilitar el scroll horizontal
        # Establecer el ancho mínimo total para forzar el scroll horizontal
        self.tree_conceptos.configure(style="Treeview")
        
        def update_scrollregion(event):
            """Actualizar la región de scroll cuando cambie el contenido"""
            self.tree_conceptos.update_idletasks()
        
        self.tree_conceptos.bind("<Configure>", update_scrollregion)

        self.tree_conceptos.bind("<Double-1>", self.on_concepto_double_click)

        # mover foco con Enter
        self.e_concepto.bind("<Return>", lambda e: self.e_puntos.focus_set())
        self.e_puntos.bind("<Return>", lambda e: self.txt_detalle.focus_set())
        self.txt_detalle.bind("<Return>", self._on_detalle_enter)
        self.conc_btns["Guardar"].bind("<Return>", lambda e: self.guardar_concepto())
        self.e_concepto.bind("<KeyRelease-Return>", lambda e: self.buscar_concepto())
        self.cargar_grilla_conceptos()
    
        # ===========================================
        #  Aplicar permisos si ya hay usuario cargado
        # ===========================================
        try:
            if hasattr(self, "usuario_actual"):
                rol = self.usuario_actual.get("rol", "")
                if rol == "BOMBERO":
                    # Deshabilitar todos
                    for k, b in self.conc_btns.items():
                        b.config(state="disabled")
                    # Habilitar solo Buscar e Imprimir listado
                    if "Buscar" in self.conc_btns:
                        self.conc_btns["Buscar"].config(state="normal")
                    if "Imprimir listado" in self.conc_btns:
                        self.conc_btns["Imprimir listado"].config(state="normal")
        except Exception as e:
            pass
        # ==========================================================
        #  Reforzar permisos visuales después de construir pestaña
        # ==========================================================
        def _forzar_color_botones_conceptos():
            try:
                if hasattr(self, "usuario_actual") and self.usuario_actual.get("rol") == "BOMBERO":
                    for k, b in self.conc_btns.items():
                        if "buscar" in k.lower() or "imprimir" in k.lower():
                            b.config(state="normal", bg="SystemButtonFace", fg="black", relief="raised")
                        else:
                            b.config(state="disabled")
            except Exception as e:
                pass

        # Ejecutar luego de 200 ms para asegurar que la UI esté lista
        self.master.after(200, _forzar_color_botones_conceptos)
    
    # helper: siguiente id para conceptos (muestra el número, no condiciona INSERT)

    def toggle_estado_concepto(self):

        concepto_id = self.var_id_concepto.get()
        id_concepto = self.var_id_concepto.get()
        
        if not concepto_id:
            self.ui.show_error("Error", "Debe seleccionar un concepto")
            return

        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()

        c.execute("SELECT activo FROM conceptos WHERE id=?", (concepto_id,))
        row = c.fetchone()

        if not row:
            conn.close()
            return

        activo = row[0]

        if activo == 1:
            nuevo_estado = 0
            mensaje = "Concepto dado de baja"
        else:
            nuevo_estado = 1
            mensaje = "Concepto reactivado"

        try:
            c.execute(
                "UPDATE conceptos SET activo=? WHERE id=?",
                (nuevo_estado, concepto_id)
            )
            conn.commit()

            self.ui.show_info("OK", mensaje)

        except Exception as e:
            self.ui.show_error("Error", f"No se pudo actualizar: {e}")

        finally:
            conn.close()

        self.cargar_grilla_conceptos()

        # volver a seleccionar el concepto en la grilla
        for item in self.tree_conceptos.get_children():
            if str(self.tree_conceptos.item(item)["values"][0]) == str(id_concepto):
                self.tree_conceptos.selection_set(item)
                self.tree_conceptos.focus(item)
                self.on_concepto_double_click(None)
                break

    def actualizar_boton_estado(self, activo):

        if activo == 1:
            self.conc_btns["Dar de baja"].config(
                text="Dar de baja",
                bg="#ffcccc"
            )
        else:
            self.conc_btns["Dar de baja"].config(
                text="Reactivar",
                bg="#ccffcc"
            )

    def _next_concepto_id(self):
        try:
            conn = sqlite3.connect(DB_PATH)
            c = conn.cursor()
            c.execute("SELECT MAX(id) FROM conceptos")
            row = c.fetchone()
            conn.close()
            return (row[0] + 1) if row and row[0] else 1
        except Exception:
            return ""

    def _formatear_horas_decimal_a_hhmm(self, horas_decimal):
        """Convierte horas decimales a formato HH:MM (versión sin dependencia de pandas)."""
        try:
            # Casos base
            if horas_decimal is None:
                return "0:00"
            
            # Verificar si es NaN sin usar pandas
            if isinstance(horas_decimal, float):
                import math
                if math.isnan(horas_decimal):
                    return "0:00"
            
            # Verificar si es string vacío
            if horas_decimal == "":
                return "0:00"
            
            # Si ya está en formato HH:MM
            if isinstance(horas_decimal, str):
                horas_str = str(horas_decimal).strip()
                if ':' in horas_str:
                    # Validar formato HH:MM
                    parts = horas_str.split(':')
                    if len(parts) == 2:
                        try:
                            h = int(parts[0])
                            m = int(parts[1])
                            if 0 <= h < 24 and 0 <= m < 60:
                                return f"{h:02d}:{m:02d}"
                        except:
                            pass
                # String vacío
                if not horas_str or horas_str.lower() in ['nan', 'none', 'null']:
                    return "0:00"
            
            # Convertir a string y limpiar
            str_val = str(horas_decimal).strip().replace(',', '.')
            
            # Intentar como float
            try:
                horas_float = float(str_val)
                
                # Si es muy pequeño o negativo
                if horas_float <= 0:
                    return "0:00"
                
                # Convertir a HH:MM
                horas_enteras = int(horas_float)
                minutos_decimal = (horas_float - horas_enteras) * 60
                minutos = int(round(minutos_decimal))
                
                # Ajustar redondeo
                if minutos >= 60:
                    horas_enteras += 1
                    minutos = 0
                
                # Asegurar que minutos estén entre 0-59
                if minutos < 0:
                    minutos = 0
                elif minutos > 59:
                    minutos = 59
                
                return f"{horas_enteras:02d}:{minutos:02d}"
                
            except ValueError:
                # Si no se puede convertir a número
                return str(horas_decimal)
                
        except Exception as e:
            print(f"ERROR _formatear_horas_decimal_a_hhmm: {e}, valor: {horas_decimal}")
            return "0:00"

    def _normalizar_entry_hora(self, entry):
        try:
            valor = entry.get().strip()

            if not valor:
                entry.delete(0, "end")
                entry.insert(0, "0:00")
                return

            valor = valor.replace(',', '.')

            if ':' in valor:
                partes = valor.split(':')
                horas = float(partes[0]) if partes[0] else 0
                minutos = float(partes[1]) if len(partes) > 1 else 0
                horas_decimal = horas + (minutos / 60.0)
            else:
                horas_decimal = float(valor)

            hhmm = self._formatear_horas_decimal_a_hhmm(horas_decimal)

            if not hhmm:
                hhmm = "0:00"

            entry.delete(0, "end")
            entry.insert(0, hhmm)

        except:
            entry.delete(0, "end")
            entry.insert(0, "0:00")

    def nuevo_concepto(self):
        """Prepara la pantalla para cargar un nuevo concepto:
        - muestra próximo ID
        - habilita campos y pone foco en Concepto
        """
        try:
            next_id = self._next_concepto_id()
            self.var_id_concepto.set(str(next_id) if next_id else "")
        except Exception:
            self.var_id_concepto.set("")

        # limpiar y habilitar campos
        self.var_concepto.set("")
        self.var_puntos.set("")
        self.txt_detalle.config(state="normal")
        self.txt_detalle.delete("1.0", "end")

        self.e_concepto.config(state="normal")
        self.e_puntos.config(state="normal")

        # estado y foco
        self.modo_concepto = "nuevo"
        self._estado_botones_concepto("nuevo")
        try:
            self.e_concepto.focus_set()
        except Exception:
            pass

    def guardar_concepto(self):
        """Inserta o actualiza según self.modo_concepto (nuevo/modificar)."""
        concepto = self.var_concepto.get().strip().upper()
        puntos = self.var_puntos.get().strip()
        detalle = self.txt_detalle.get("1.0", "end").strip()

        if not concepto:
            self.ui.show_error("Error", "Debe ingresar un concepto")
            return

        try:
            puntos = int(puntos) if puntos else 0
        except Exception:
            self.ui.show_error("Error", "Los puntos deben ser un número")
            return

        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        try:
            if getattr(self, "modo_concepto", "") == "modificar" and self.var_id_concepto.get():
                # UPDATE
                c.execute(
                    "UPDATE conceptos SET concepto=?, puntos=?, detalle=? WHERE id=?",
                    (concepto, puntos, detalle, int(self.var_id_concepto.get()))
                )
            else:
                # INSERT (autoincrement)
                c.execute(
                    "INSERT INTO conceptos (concepto, puntos, detalle) VALUES (?, ?, ?)",
                    (concepto, puntos, detalle)
                )
                # obtener id real asignado por SQLite y mostrarlo
                new_id = c.lastrowid
                try:
                    self.var_id_concepto.set(str(new_id))
                except Exception:
                    pass

            conn.commit()
            self.ui.show_info("OK", "Concepto guardado correctamente")
        except Exception as e:
            self.ui.show_error("Error", f"No se pudo guardar: {e}")
        finally:
            conn.close()

        self.cargar_grilla_conceptos()
        # después de guardar, volver a estado inicial
        self.limpiar_concepto()

    def _habilitar_modificar_concepto(self):
        """Habilita edición del concepto previamente cargado (desde doble clic)."""
        if not self.var_id_concepto.get():
            self.ui.show_error("Error", "Seleccione un concepto para modificar")
            return
        self.modo_concepto = "modificar"
        self.e_concepto.config(state="normal")
        self.e_puntos.config(state="normal")
        self.txt_detalle.config(state="normal")
        self._estado_botones_concepto("nuevo")
        try:
            self.e_concepto.focus_set()
        except Exception:
            pass
        
    def cargar_grilla_conceptos(self):
        """Llena la grilla de conceptos desde la base de datos."""

        # limpiar grilla
        for item in self.tree_conceptos.get_children():
            self.tree_conceptos.delete(item)

        # traer datos
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()

        try:
            c.execute("""
                SELECT id, concepto, puntos, detalle, activo
                FROM conceptos
                ORDER BY id ASC
            """)
            rows = c.fetchall()

        except Exception as e:
            self.ui.show_error("Error", f"No se pudo cargar la grilla de conceptos: {e}")
            rows = []

        finally:
            conn.close()

        # configurar estilos
        self.tree_conceptos.tag_configure("activo", foreground="black")
        self.tree_conceptos.tag_configure("baja", foreground="gray", background="#f0f0f0")

        # insertar en grilla
        for r in rows:

            if r[4] == 0:  # concepto dado de baja
                self.tree_conceptos.insert(
                    "",
                    "end",
                    values=r,
                    tags=("baja",)
                )

            else:  # concepto activo
                self.tree_conceptos.insert(
                    "",
                    "end",
                    values=r,
                    tags=("activo",)
                )

    def _estado_botones_concepto(self, estado):
        rol = self.usuario_actual.get("rol", "").upper()

        if estado == "inicio":
            self._toggle_concepto_editable(False)
            # Estado inicial: sólo permitir Buscar
            for k in self.conc_btns:
                self.conc_btns[k].config(state="disabled")
            if "Buscar" in self.conc_btns:
                self.conc_btns["Buscar"].config(state="normal")

        elif estado == "nuevo":
            self._toggle_concepto_editable(True)
            for k in self.conc_btns:
                self.conc_btns[k].config(state="disabled")
            for k in ["Guardar", "Limpiar/Cancelar"]:
                if k in self.conc_btns:
                    self.conc_btns[k].config(state="normal")

        elif estado == "cargado":
            self._toggle_concepto_editable(False)

            # Activar por defecto todos
            for k in ["Modificar", "Eliminar", "Limpiar/Cancelar"]:
                if k in self.conc_btns:
                    self.conc_btns[k].config(state="normal")
            if "Guardar" in self.conc_btns:
                self.conc_btns["Guardar"].config(state="disabled")

            # 🔒 Restricciones por rol
            if rol == "BOMBERO":
                # Deshabilitar todo excepto Buscar, Limpiar/Cancelar, Imprimir Concepto e Imprimir listado
                for k in self.conc_btns:
                    self.conc_btns[k].config(state="disabled")
                for k in ["Buscar", "Limpiar/Cancelar", "Imprimir Concepto", "Imprimir listado"]:
                    if k in self.conc_btns:
                        self.conc_btns[k].config(state="normal")
            elif rol == "SUPERVISOR":
                # SUPERVISOR puede modificar pero no eliminar
                if "Eliminar" in self.conc_btns:
                    self.conc_btns["Eliminar"].config(state="disabled")

    def buscar_concepto(self):
        """Busca conceptos por el texto del Entry Concepto/Actividad y los muestra en la grilla."""
        # debug en consola: confirmar que el método fue llamado
        try:
            print("DEBUG: buscar_concepto llamado. var_concepto:", repr(self.var_concepto.get()))
        except Exception:
            pass

        texto = (self.var_concepto.get() or "").strip().upper()

        # Si el Entry está disabled y no hay texto, preguntar al usuario (para poder buscar sin habilitar Edit)
        try:
            estado_entry = str(self.e_concepto.cget("state"))
        except Exception:
            estado_entry = ""

        if not texto and estado_entry == "disabled":
            # pedir texto (vacío = listar todo)
            from tkinter.simpledialog import askstring
            q = askstring("Buscar Concepto", "Ingrese texto a buscar (dejar vacío para listar todo):")
            if q is None:
                return
            texto = q.strip().upper()

        # limpiar grilla antes de cargar resultados
        for item in self.tree_conceptos.get_children():
            self.tree_conceptos.delete(item)

        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        rows = []
        try:
            if texto:
                # Buscar tanto en concepto como en detalle (insensible a mayúsculas)
                c.execute("""
                    SELECT id, concepto, puntos, detalle
                    FROM conceptos
                    WHERE UPPER(concepto) LIKE ? OR UPPER(detalle) LIKE ?
                    ORDER BY id DESC
                """, ('%' + texto + '%', '%' + texto + '%'))
            else:
                # si está vacío, trae todos
                c.execute("SELECT id, concepto, puntos, detalle FROM conceptos ORDER BY id DESC")
            rows = c.fetchall()
        except Exception as e:
            # mostrar error explícito
            self.ui.show_error("Error", f"No se pudo buscar: {e}")
            rows = []
        finally:
            conn.close()

    def eliminar_concepto(self):

        concepto_id = self.var_id_concepto.get()

        if not concepto_id:
            self.ui.show_error("Error", "Debe seleccionar un concepto")
            return

        if not messagebox.askyesno("Confirmar", "¿Dar de baja este concepto?"):
            return

        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()

        try:
            c.execute(
                "UPDATE conceptos SET activo = 0 WHERE id=?",
                (int(concepto_id),)
            )

            conn.commit()

            self.ui.show_info(
                "OK",
                "Concepto dado de baja correctamente"
            )

        except Exception as e:
            self.ui.show_error(
                "Error",
                f"No se pudo dar de baja: {e}"
            )

        finally:
            conn.close()

        self.cargar_grilla_conceptos()
        self.limpiar_concepto()

    def limpiar_concepto(self):
        """Limpia los campos de concepto y refresca la grilla."""
        # Limpiar entradas
        try:
            self.var_id_concepto.set("")
            self.var_concepto.set("")
            self.var_puntos.set("")
            self.txt_detalle.delete("1.0", "end")
        except Exception:
            pass

        # Refrescar la grilla de conceptos
        if hasattr(self, "cargar_grilla_conceptos"):
            self.cargar_grilla_conceptos()

        # --- Restaurar estado de botones según rol ---
        rol = self.usuario_actual["rol"].upper()
        if hasattr(self, "conc_btns"):
            # Primero deshabilita todos
            for b in self.conc_btns.values():
                b.config(state="disabled")

            # ADMIN: sólo habilitar Buscar, Nuevo e Imprimir
            if rol == "ADMIN":
                for nombre, boton in self.conc_btns.items():
                    texto = nombre.lower()
                    if (
                        "buscar" in texto
                        or "nuevo" in texto
                        or "imprimir" in texto
                    ):
                        boton.config(state="normal")

            # SUPERVISOR: puede todo menos eliminar
            elif rol == "SUPERVISOR":
                for nombre, boton in self.conc_btns.items():
                    boton.config(state="normal")
                if "Eliminar" in self.conc_btns:
                    self.conc_btns["Eliminar"].config(state="disabled")

            # BOMBERO: sólo Buscar e Imprimir
            elif rol == "BOMBERO":
                for nombre, boton in self.conc_btns.items():
                    texto = nombre.lower()
                    if "buscar" in texto or "imprimir" in texto:
                        boton.config(state="normal")

    def on_concepto_double_click(self, event):

        item = self.tree_conceptos.selection()
        if not item:
            return

        valores = self.tree_conceptos.item(item)["values"]

        concepto_id = valores[0]
        concepto = valores[1]
        puntos = valores[2]
        detalle = valores[3]
        estado = valores[4]

        self.var_id_concepto.set(concepto_id)
        self.var_concepto.set(concepto)
        self.var_puntos.set(puntos)

        self.txt_detalle.config(state="normal")
        self.txt_detalle.delete("1.0", "end")
        self.txt_detalle.insert("1.0", detalle)
        self.txt_detalle.config(state="disabled")

        boton_baja = self.conc_btns["Dar de baja"]

        if estado == 0:
            boton_baja.config(text="Reactivar", bg="#ccffcc")
        else:
            boton_baja.config(text="Dar de baja", bg="#ffcccc")

    def obtener_estado_concepto(self, concepto_id):
        cur = self.conn.cursor()
        cur.execute("SELECT estado FROM conceptos WHERE id=?", (concepto_id,))
        row = cur.fetchone()
        return row[0] if row else 1

    def _detalle_uppercase(self, event=None):
        """Forzar a mayúsculas el contenido del Text Detalle."""
        try:
            txt = self.txt_detalle.get("1.0", "end-1c")
            upper = txt.upper()
            if txt != upper:
                # Guardar posición del cursor
                pos = self.txt_detalle.index("insert")
                self.txt_detalle.delete("1.0", "end")
                self.txt_detalle.insert("1.0", upper)
                self.txt_detalle.mark_set("insert", pos)
        except Exception:
            pass
        return None

    def _on_detalle_enter(self, event=None):
        """Al presionar Enter en Detalle → ir al botón Guardar."""
        try:
            self.conc_btns["Guardar"].focus_set()
        except Exception:
            pass
        return "break"  # evita insertar salto de línea
    def _sort_conceptos(self, col, reverse=False):
        """Ordena la grilla de conceptos al clickear en el encabezado."""
        # nombre interno de columna
        col_map = {
            "Código": 0,
            "Concepto/Actividad": 1,
            "Puntos": 2,
            "Detalle": 3
        }
        idx = col_map[col]
        # tomar valores actuales
        data = [(self.tree_conceptos.set(k, col), k) for k in self.tree_conceptos.get_children('')]
        # convertir Puntos y Código a int si se puede
        if col in ("Puntos", "Código"):
            def safe_int(v): 
                try: return int(v)
                except: return 0
            data.sort(key=lambda t: safe_int(t[0]), reverse=reverse)
        else:
            data.sort(key=lambda t: t[0], reverse=reverse)

        for index, (val, k) in enumerate(data):
            self.tree_conceptos.move(k, '', index)
        self.tree_conceptos.heading(col,
            command=lambda: self._sort_conceptos(col, not reverse))

    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from datetime import datetime
    import os

    # ===================== pestaña INFORMES =====================
    def init_informes(self):
        Label(self.informes_frame, text="INFORMES Y ESTADÍSTICAS",
            font=("Arial", 20, "bold"), fg="white", bg="red").place(x=20, y=6)

        # --- Coordenadas ---
        y0, dy = 46, 28
        x_label, x_entry = 20, 160
        x_label2, x_entry2 = 320, 400
        x_label3, x_entry3 = 640, 700

        # ===== FILA 1 =====
        Label(self.informes_frame, text="N° Actividad:", fg="white", bg="red").place(x=x_label, y=y0)
        self.inf_id = Entry(self.informes_frame, font=("Arial", 11))
        self.inf_id.place(x=x_entry, y=y0, width=150)

        Label(self.informes_frame, text="Actividad:", fg="white", bg="red").place(x=x_label2, y=y0)
        
        import sqlite3
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("SELECT id, concepto FROM conceptos ORDER BY concepto ASC")
        rows = c.fetchall()
        conn.close()

        lista_conceptos = [" "] + [f"{r[0]} - {r[1]}" for r in rows]

        self.inf_actividad = ttk.Combobox(
            self.informes_frame,
            state="readonly",
            values=lista_conceptos
        )
        self.inf_actividad.place(x=x_entry2, y=y0, width=200)
        self.inf_actividad.current(0)
        self.inf_actividad.bind("<<ComboboxSelected>>", lambda e: self.informe_por_actividad())

        Label(self.informes_frame, text="Desde:", fg="white", bg="red").place(x=x_label3, y=y0)
        self.inf_desde = DateEntry(self.informes_frame, date_pattern='dd/mm/yyyy')
        self.inf_desde.place(x=x_entry3, y=y0, width=110)

        # ===== FILA 2 =====
        Label(self.informes_frame, text="Apellido/Nombre:", fg="white", bg="red").place(x=x_label, y=y0 + dy)
        self.inf_apynom = Entry(self.informes_frame, font=("Arial", 11))
        self.inf_apynom.place(x=x_entry, y=y0 + dy, width=150)

        Label(self.informes_frame, text="Leg:", fg="white", bg="red").place(x=x_label2, y=y0 + dy)
        self.inf_legajo_cb = ttk.Combobox(self.informes_frame, state="readonly")
        self.inf_legajo_cb.place(x=x_entry2, y=y0 + dy, width=200)

        # 🔥 ENTER → foco en botón Buscar
        def focus_buscar(event=None):
            self.btn_buscar.focus_set()

        self.inf_id.bind("<Return>", focus_buscar)
        self.inf_apynom.bind("<Return>", focus_buscar)

        Label(self.informes_frame, text="Hasta:", fg="white", bg="red").place(x=x_label3, y=y0 + dy)
        self.inf_hasta = DateEntry(self.informes_frame, date_pattern='dd/mm/yyyy')
        self.inf_hasta.place(x=x_entry3, y=y0 + dy, width=110)

        # ===== BOTONES DE CONTROL =====
        btn_y = y0 + 2 * dy + 7
        x0 = 205

        self.btn_buscar = Button(
            self.informes_frame,
            text="Buscar",
            command=self.buscar_informes,
            width=15
        )
        self.btn_buscar.place(x=x0, y=btn_y)
        Button(self.informes_frame, text="Limpiar", command=self.limpiar_informes, width=15).place(x=x0 + 130, y=btn_y)
        Button(self.informes_frame, text="Exportar Excel", command=self.exportar_excel, width=15).place(x=x0 + 260, y=btn_y)
        Button(self.informes_frame, text="Exportar PDF", command=self.exportar_pdf, width=15).place(x=x0 + 390, y=btn_y)
        self.inf_id.bind("<Return>", lambda e: self.buscar_informes())
        self.inf_apynom.bind("<Return>", lambda e: self.buscar_informes())

        # 🔥 Limpiar campos excluyentes

        def limpiar_si_escribe_en_id(event=None):
            if self.inf_id.get().strip():
                self.inf_apynom.delete(0, END)

        def limpiar_si_escribe_en_apynom(event=None):
            if self.inf_apynom.get().strip():
                self.inf_id.delete(0, END)

        self.inf_id.bind("<KeyRelease>", limpiar_si_escribe_en_id)
        self.inf_apynom.bind("<KeyRelease>", limpiar_si_escribe_en_apynom)

        # ===== BOTONES DE INFORMES GENERALES =====
        btn_y2 = btn_y + 45
        
        # Tres botones de informes generales - UNO AL LADO DEL OTRO
        Button(self.informes_frame, text="Informe por Bombero (todos)",
            command=self.informe_todos_bomberos,
            bg="white", fg="blue", font=("Arial", 10, "bold"),
            width=25).place(x=20, y=btn_y2)
        
        Button(self.informes_frame, text="Informe por Actividad (todas)",
            command=self.informe_todas_actividades,
            bg="white", fg="green", font=("Arial", 10, "bold"),
            width=25).place(x=290, y=btn_y2)  # 20 + 270 de separación
        
        Button(self.informes_frame, text="Informe Horas por Período",
            command=self.informe_horas_por_periodo,
            bg="white", fg="purple", font=("Arial", 10, "bold"),
            width=25).place(x=560, y=btn_y2)  # 290 + 270 de separación

        # ===== GRILLA DE RESULTADOS =====
        cols = (
            "ID", "Leg", "Actividad", "Área",
            "Fecha I.", "Fecha F.",
            "H.I.", "H.F.",
            "Horas",           # 👈 AGREGADA
            "Descripción",
            "Apellido",
            "Nombre"
        )

        frame_tabla_informes = Frame(self.informes_frame, bg="white")
        frame_tabla_informes.place(x=20, y=btn_y2 + 50, width=1000, height=190)
        
        scroll_y = Scrollbar(frame_tabla_informes, orient=VERTICAL)
        scroll_x = Scrollbar(frame_tabla_informes, orient=HORIZONTAL)

        self.inf_tree = ttk.Treeview(
            frame_tabla_informes,
            columns=cols,
            show="headings",
            yscrollcommand=scroll_y.set,
            xscrollcommand=scroll_x.set,
            height=10,
            style="Custom.Treeview"
        )

        # 🔧 Hacer que las columnas se adapten al espacio disponible
        self.inf_tree.pack_propagate(False)

        for col in cols:
            self.inf_tree.column(col, anchor="w", stretch=True)

        scroll_y.config(command=self.inf_tree.yview)
        scroll_x.config(command=self.inf_tree.xview)

        scroll_y.pack(side=RIGHT, fill=Y)
        scroll_x.pack(side=BOTTOM, fill=X)
        self.inf_tree.pack(side=LEFT, fill=BOTH, expand=True)

        # === Eventos para informes específicos ===
        # Estos permiten que al seleccionar una opción en las combos y hacer clic en "Buscar"
        # se generen los informes específicos automáticamente
        self.inf_legajo_cb.bind("<<ComboboxSelected>>", lambda e: self.informe_por_bombero())
        self.inf_actividad.bind("<<ComboboxSelected>>", lambda e: self.informe_por_actividad())

        # También puedes agregar atajos de teclado si quieres:
        self.inf_tree.bind("<Double-1>", self._on_informe_double_click)
        
        # Función para hacer que Enter en las combos también active la búsqueda
        self.inf_legajo_cb.bind("<Return>", lambda e: self.buscar_informes())
        self.inf_actividad.bind("<Return>", lambda e: self.buscar_informes())

        # --- Label de totales ---
        self.lbl_totales = Label(self.informes_frame, text="", font=("Arial", 10, "bold"),
                                bg="white", anchor="w", justify="left")
        self.lbl_totales.place(x=30, y=btn_y2 + 250)

        for c in cols:
            self.inf_tree.heading(c, text=c)
            self.inf_tree.column(c, width=100, anchor="center")

    # ---------------------------------------------------------
    def _pedir_rango_fechas(self):

        f1 = simpledialog.askstring("Informe", "Desde (dd/mm/yyyy):")
        f2 = simpledialog.askstring("Informe", "Hasta (dd/mm/yyyy):")
        if not f1 or not f2:
            return None, None
        try:
            d1 = datetime.strptime(f1, "%d/%m/%Y").date()
            d2 = datetime.strptime(f2, "%d/%m/%Y").date()
            return d1, d2
        except Exception:
            self.ui.show_error("Error", "Formato de fecha inválido (usa dd/mm/yyyy)")
            return None, None

# ==== INICIO BLOQUE INFORMES CORREGIDOS V2 ====
    import os

    # ---------- Compatibilidad / aliases (evitan errores por nombres distintos) ----------
    def _fmt_horas(self, valor):
        """Convierte horas decimales a formato HH:MM de manera CONSISTENTE."""
        try:
            if valor is None or pd.isna(valor):
                return "0:00"
            
            # Asegurar que sea float
            if isinstance(valor, str):
                # Si ya está en formato HH:MM, devolverlo tal cual
                if ':' in valor:
                    partes = valor.split(':')
                    if len(partes) >= 2:
                        horas = int(partes[0]) if partes[0] else 0
                        minutos = int(partes[1]) if partes[1] else 0
                        # Validar que sean números válidos
                        if 0 <= horas <= 23 and 0 <= minutos <= 59:
                            return f"{horas}:{minutos:02d}"
                # Si es string decimal
                try:
                    valor = float(valor)
                except:
                    return "0:00"
            
            valor = float(valor)
            
            # Redondear a 2 decimales para evitar errores de precisión
            valor = round(valor, 2)
            
            # Obtener horas enteras y minutos
            horas_enteras = int(valor)
            minutos = int(round((valor - horas_enteras) * 60))
            
            # Ajustar si minutos son 60
            if minutos >= 60:
                horas_enteras += 1
                minutos = 0
            
            # Ajustar si horas_enteras es negativo
            if horas_enteras < 0:
                return "0:00"
                
            return f"{horas_enteras}:{minutos:02d}"
        except Exception:
            return "0:00"
        
    def _horas_a_decimal(self, v):
        """Alias compatible para convertir horas formato 'HH:MM' a decimal."""
        return self._safe_horas_a_decimal(v) if hasattr(self, "_safe_horas_a_decimal") else 0.0

    # ----------------------- Helpers locales para el bloque -----------------------
    def _safe_horas_a_decimal(self, v):
        """Intenta convertir varios formatos ('HH:MM', float con coma/punto, etc.) a decimal horas."""
        try:
            if hasattr(self, "_horas_a_decimal") and self._horas_a_decimal != self._safe_horas_a_decimal:
                # si existe versión nativa la usamos para mantener compatibilidad
                return float(getattr(self, "_horas_a_decimal")(v))
        except Exception:
            pass
        try:
            s = str(v).strip()
            if not s:
                return 0.0
            if ":" in s:
                parts = s.split(":")
                h = int(parts[0]) if parts[0] else 0
                m = int(parts[1]) if len(parts) > 1 and parts[1] else 0
                return h + m / 60.0
            s = s.replace(",", ".")
            return float(s)
        except Exception:
            return 0.0

    def _safe_fmt_horas(self, v):
        """Formatea decimal horas a 'H:MM' (sin ceros excesivos) si existe _fmt_horas, sino formatea manualmente."""
        try:
            # si hay método original lo usamos
            if hasattr(self, "_fmt_horas") and self._fmt_horas != self._safe_fmt_horas:
                return getattr(self, "_fmt_horas")(v)
        except Exception:
            pass
        try:
            dec = float(v)
            h = int(dec)
            m = int(round((dec - h) * 60))
            # normalizar minutos a 0-59
            if m == 60:
                h += 1
                m = 0
            return f"{h}:{m:02d}"
        except Exception:
            # si era ya 'HH:MM' devolver tal cual
            try:
                s = str(v).strip()
                if s:
                    return s
            except:
                pass
            return "0:00"

    # -------------------- Ventana: tabla + gráfico + totales + Export PDF/Excel --------------------
    def _mostrar_tabla_y_grafico(self, titulo, df_graf, df_tabla=None, x_col=None, y_col=None, color="red"):
        if df_graf is None or df_graf.empty:
            self.ui.show_info("Sin datos", "No hay datos para mostrar.")
            return

        ventana = tk.Toplevel()
        ventana.title(titulo)
        ventana.geometry("1150x750")
        ventana.resizable(True, True)
        try:
            ventana.grab_set()
        except Exception:
            pass
        # === ICONO DE LA VENTANA (evita la pluma en Windows) ===
        try:
            ventana.iconbitmap(resource_path("bomberos.ico"))
        except Exception as e:
            pass

        # --- Layout principal ---
        main = tk.Frame(ventana, bg="#f0f0f0")
        main.pack(fill="both", expand=True, padx=10, pady=10)
        main.columnconfigure(0, weight=3)
        main.columnconfigure(1, weight=1)
        main.rowconfigure(0, weight=1)

        # --- IZQUIERDA: gráfico + tabla ---
        left = tk.Frame(main, bg="white")
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        left.rowconfigure(0, weight=2)
        left.rowconfigure(1, weight=3)

        # --- DERECHA: botones + resumen ---
        right = tk.Frame(main, bg="#f8f8f8")
        right.grid(row=0, column=1, sticky="ns")

        # === GRÁFICO ===
        frame_grafico = tk.Frame(left, bg="white")
        frame_grafico.grid(row=0, column=0, sticky="nsew", pady=(0, 6))

        filas = len(df_graf)
        alto_figura = max(4, filas * 0.38)
        fig, ax = plt.subplots(figsize=(8, alto_figura))

        if x_col not in df_graf.columns and len(df_graf.columns) >= 1:
            x_col = df_graf.columns[0]
        if y_col not in df_graf.columns and len(df_graf.columns) >= 2:
            y_col = df_graf.columns[1]

        df_graf_plot = df_graf.copy()
        try:
            df_graf_plot[y_col] = df_graf_plot[y_col].apply(lambda v: self._horas_a_decimal(v) if isinstance(v, str) else float(v))
        except Exception:
            df_graf_plot[y_col] = pd.to_numeric(df_graf_plot[y_col], errors="coerce").fillna(0.0)

        df_graf_plot = df_graf_plot.sort_values(by=y_col, ascending=True)
        # Detectar si el gráfico corresponde a actividades o a horas
        titulo_lower = titulo.lower()
        if "actividad" in titulo_lower or "actividades" in titulo_lower:
            eje_x_label = "Cantidad de actividades"
        else:
            eje_x_label = "Horas trabajadas"

        barras = ax.barh(df_graf_plot[x_col].astype(str), df_graf_plot[y_col], color=color)

        ax.set_xlabel(eje_x_label, fontsize=10)
        ax.set_ylabel("")                 # 👈 sin texto en eje Y
        ax.set_title(titulo, fontsize=12, fontweight="bold", pad=10)
        ax.tick_params(axis="y", labelsize=8)

        ax.bar_label(barras, fmt="%.1f", fontsize=8, padding=3)
        plt.tight_layout(pad=2)

        total_horas = float(df_graf_plot[y_col].sum())

        canvas_fig = FigureCanvasTkAgg(fig, master=frame_grafico)
        canvas_fig.draw()
        canvas_fig.get_tk_widget().pack(fill="both", expand=True)

        # === TABLA (Treeview) ===
        frame_tabla = tk.Frame(left)
        frame_tabla.grid(row=1, column=0, sticky="nsew")

        canvas_tabla = tk.Canvas(frame_tabla)
        vscroll = ttk.Scrollbar(frame_tabla, orient="vertical", command=canvas_tabla.yview)
        hscroll = ttk.Scrollbar(frame_tabla, orient="horizontal", command=canvas_tabla.xview)
        inner = tk.Frame(canvas_tabla)

        canvas_tabla.configure(yscrollcommand=vscroll.set, xscrollcommand=hscroll.set)
        vscroll.pack(side="right", fill="y")
        hscroll.pack(side="bottom", fill="x")
        canvas_tabla.pack(side="left", fill="both", expand=True)
        canvas_tabla.create_window((0, 0), window=inner, anchor="nw")

        def _on_config(e):
            canvas_tabla.configure(scrollregion=canvas_tabla.bbox("all"))
        inner.bind("<Configure>", _on_config)

        if df_tabla is not None and not df_tabla.empty:
            cols = list(df_tabla.columns)
            tree = ttk.Treeview(inner, columns=cols, show="headings", height=12)
            for col in cols:
                maxlen = max(len(str(col)), *(len(str(x)) for x in df_tabla[col].astype(str).values))
                ancho = max(140, min(800, maxlen * 8))  # más ancho horizontal
                anchor = "w" if df_tabla[col].dtype == object else "center"
                tree.heading(col, text=col)
                tree.column(col, anchor=anchor, width=ancho, stretch=True)
            for _, fila in df_tabla.iterrows():
                vals = [("" if pd.isna(x) else x) for x in fila.tolist()]
                tree.insert("", "end", values=vals)
            tree.pack(fill="both", expand=True)

        # === CÁLCULO DE TOTALES ===
        cant_meses = 0

        def _contar_meses(df):
            """Cuenta la cantidad de meses distintos en un DataFrame, buscando columnas con fecha o mes."""
            if df is None or df.empty:
                return 0
            for c in df.columns:
                colname = str(c).lower()
                serie = df[c].astype(str)
                # Si contiene fechas (dd/mm/yyyy o yyyy-mm)
                if serie.str.contains(r"\d{1,2}/\d{1,2}/\d{2,4}|\d{4}-\d{1,2}", regex=True).any():
                    try:
                        fechas = pd.to_datetime(serie, errors="coerce", dayfirst=True)
                        return fechas.dt.to_period("M").nunique()
                    except Exception:
                        continue
                # Si son nombres de meses (Enero, Febrero, etc.)
                meses = [
                    "enero", "febrero", "marzo", "abril", "mayo", "junio",
                    "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"
                ]
                if any(m in colname for m in ["mes", "periodo"]) or serie.str.lower().isin(meses).any():
                    return serie.nunique()
            return 0

        # Intentar primero con df_tabla, luego con df_graf
        cant_meses = _contar_meses(df_tabla)
        if cant_meses == 0:
            cant_meses = _contar_meses(df_graf)

        # === BOTONES ===
        ttk.Button(right, text="Exportar a Excel", width=20, command=lambda: self._exportar_excel_con_fechas(titulo, df_tabla, df_graf)).pack(pady=8)
        ttk.Button(right, text="Exportar a PDF", width=20, command=lambda: self._exportar_pdf_con_fechas(titulo, df_tabla, df_graf)).pack(pady=8)
        ttk.Button(right, text="Cerrar", width=20, command=ventana.destroy).pack(pady=8)

        # === RESUMEN DINÁMICO ===
        frame_resumen = tk.Frame(right, bg="#d9d9d9", relief="groove", bd=2)
        frame_resumen.pack(fill="x", pady=(20, 10))

        total_registros = 0
        if df_tabla is not None and not df_tabla.empty:
            total_registros = len(df_tabla)
        elif df_graf is not None and not df_graf.empty:
            total_registros = len(df_graf)

        titulo_lower = titulo.lower()
        if "actividad" in titulo_lower:
            etiqueta = "Actividades"
        elif "bombero" in titulo_lower:
            etiqueta = "Bomberos"
        else:
            etiqueta = "Meses"

        if "actividad" in titulo_lower:
            lbl_text = f"Registro de actividades: {total_registros}     |     Total actividades: {int(total_horas)}"
        elif etiqueta == "Meses":
            lbl_text = f"{etiqueta}: {cant_meses}     |     Total de horas: {total_horas:.1f}"
        else:
            lbl_text = f"{etiqueta}: {total_registros}     |     Total de horas: {total_horas:.1f}"

        lbl_resumen = tk.Label(
            frame_resumen, text=lbl_text,
            font=("Arial", 10, "bold"), fg="black", bg="#d9d9d9",
            anchor="center"
        )
        for linea in lbl_text.split("     |     "):
            tk.Label(frame_resumen, text=linea, font=("Arial", 10, "bold"),
                    fg="black", bg="#d9d9d9", anchor="center").pack(fill="x", padx=5, pady=2)

    def _on_close(self):

        self.cerrando = True  # 🔴 clave para frenar callbacks

        # 🔴 cancelar after si usás alguno
        try:
            if hasattr(self, "_after_id"):
                self.master.after_cancel(self._after_id)
        except:
            pass

        # limpiar campos
        for nombre in ("inf_actividad", "inf_legajo", "inf_nombre"):
            if hasattr(self, nombre):
                try:
                    getattr(self, nombre).set("")
                except:
                    pass

        # cerrar matplotlib si existe
        try:
            import matplotlib.pyplot as plt
            plt.close('all')
        except:
            pass

        # 🔴 destruir root al final
        try:
            self.master.destroy()
        except:
            pass

    # -------------------- CREAR PDF resumen (tabla + línea de totales debajo) --------------------
    # ==========================================================
    # Exportar a EXCEL con fechas en el nombre sugerido
    # ==========================================================
    def _exportar_excel_con_fechas(self, titulo, df_tabla, df_graf):
        # --- Detectar fechas en el título ---
        match = re.search(r"(\d{2}/\d{2}/\d{4}).*(\d{2}/\d{2}/\d{4})", titulo)
        fecha_txt = ""
        if match:
            desde, hasta = match.groups()
            fecha_txt = f"{desde.replace('/', '-')}_a_{hasta.replace('/', '-')}"
        else:
            fecha_txt = f"{datetime.now():%Y-%m-%d}"

        # --- Detectar tipo de informe ---
        titulo_lower = titulo.lower()
        if "bombero" in titulo_lower:
            base_name = "Informe_por_Bombero"
        elif "actividad" in titulo_lower:
            base_name = "Informe_por_Actividades"
        else:
            base_name = "Horas_por_Periodo"

        sugerido = f"{base_name}_{fecha_txt}.xlsx"

        ruta = self.ui.ask_save_file(
            defaultextension=".xlsx",
            initialfile=sugerido,
            filetypes=[("Excel", "*.xlsx")],
            title="Guardar como Excel"
        )
        if not ruta:
            return
        try:
            to_save = df_tabla if df_tabla is not None and not df_tabla.empty else df_graf
            to_save.to_excel(ruta, index=False)
            self.ui.show_info("Éxito", f"Archivo Excel guardado en:\n{ruta}")
        except Exception as e:
            self.ui.show_error("Error", f"No se pudo exportar a Excel:\n{e}")

    # ==========================================================
    # Exportar a PDF con fechas en el nombre sugerido
    # ==========================================================
    def _exportar_pdf_con_fechas(self, titulo, df_tabla, df_graf):

        # --- Detectar fechas en el título ---
        match = re.search(r"(\d{2}/\d{2}/\d{4}).*(\d{2}/\d{2}/\d{4})", titulo)
        fecha_txt = ""
        if match:
            desde, hasta = match.groups()
            fecha_txt = f"{desde.replace('/', '-')}_a_{hasta.replace('/', '-')}"
        else:
            fecha_txt = f"{datetime.now():%Y-%m-%d}"

        # --- Detectar tipo de informe ---
        titulo_lower = titulo.lower()
        if "bombero" in titulo_lower:
            base_name = "Informe_por_Bombero"
        elif "actividad" in titulo_lower:
            base_name = "Informe_por_Actividades"
        else:
            base_name = "Horas_por_Periodo"

        sugerido = f"{base_name}_{fecha_txt}.pdf"

        ruta = self.ui.ask_save_file(
            defaultextension=".pdf",
            initialfile=sugerido,
            filetypes=[("PDF", "*.pdf")],
            title="Guardar como PDF"
        )
        if not ruta:
            return

        try:
            df_res = (df_tabla if df_tabla is not None and not df_tabla.empty else df_graf.copy())
            df_det = None
            if hasattr(self, "_crear_pdf_resumen_informe"):
                self._crear_pdf_resumen_informe(ruta, titulo, df_resumen=df_res, df_detalle=df_det, landscape_mode=False)
            else:
                from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
                from reportlab.lib import colors
                from reportlab.lib.pagesizes import A4
                from reportlab.lib.styles import getSampleStyleSheet
                styles = getSampleStyleSheet()
                doc = SimpleDocTemplate(ruta, pagesize=A4)
                story = [Paragraph(titulo, styles["Title"]), Spacer(1, 12)]
                data = [df_res.columns.tolist()] + df_res.fillna("").astype(str).values.tolist()
                tabla = Table(data)
                tabla.setStyle(TableStyle([
                    ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey)
                ]))
                story.append(tabla)
                doc.build(story)
            self.ui.show_info("Éxito", f"Archivo PDF guardado en:\n{ruta}")
        except Exception as e:
            self.ui.show_error("Error", f"No se pudo exportar a PDF:\n{e}")

    # ---------------------- Mostrar tabla y grafico mejorado ----------------------------
    def _mostrar_form_informe_horas_periodo(self, titulo, df_graf, df_tabla, x_col="Mes", y_col="Horas", color="#4682B4"):
        """
        Muestra el formulario del informe de horas por período - VERSIÓN FINAL COMPLETA.
        Maximizado, botones funcionales, títulos uniformes, gráfico bajo.
        """
        # Cerrar ventanas anteriores del mismo tipo
        if hasattr(self, '_ventana_horas_periodo'):
            try:
                self._ventana_horas_periodo.destroy()
            except:
                pass
        
        # Crear ventana
        self._ventana_horas_periodo = tk.Toplevel(self.master)
        self._ventana_horas_periodo.title(titulo)
        self._ventana_horas_periodo.geometry("1150x700")
        self._ventana_horas_periodo.configure(bg='white')
        
        # 🔥 MAXIMIZAR LA VENTANA (NUEVO)
        try:
            # Intenta con state('zoomed') - funciona en Windows
            self._ventana_horas_periodo.state('zoomed')

        except Exception as e1:

            try:
                # Alternativa: attributes('-fullscreen', True)
                self._ventana_horas_periodo.attributes('-fullscreen', True)

            except Exception as e2:

                try:
                    # Segunda alternativa: usar tamaño de pantalla completo
                    ancho = self._ventana_horas_periodo.winfo_screenwidth()
                    alto = self._ventana_horas_periodo.winfo_screenheight()
                    self._ventana_horas_periodo.geometry(f"{ancho}x{alto}+0+0")

                except Exception as e3:
                    pass
        
        # Bloquear interacción con ventana principal
        self._ventana_horas_periodo.grab_set()
        self._ventana_horas_periodo.focus_force()
        
        # Ícono
        try:
            logo = tk.PhotoImage(file="C:/Actividades Bomberos/bomberos.png")
            self._ventana_horas_periodo.iconphoto(False, logo)
            self._ventana_horas_periodo._logo_ref = logo
        except Exception as e:
            pass
        
        # --- CONTENEDOR PRINCIPAL ---
        container = tk.Frame(self._ventana_horas_periodo, bg="#f0f0f0")
        container.pack(fill="both", expand=True)

        # --- PANEL IZQUIERDO (SCROLL) ---
        left_panel = tk.Frame(container, bg="#f0f0f0")
        left_panel.pack(side="left", fill="both", expand=True)

        canvas = tk.Canvas(left_panel, bg="#f0f0f0", highlightthickness=0)
        scrollbar = ttk.Scrollbar(left_panel, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # --- PANEL DERECHO DE ACCIONES (FIJO, FUERA DEL SCROLL) ---
        right_buttons = tk.Frame(
            container,
            bg="#f5f5f5",
            relief="groove",
            bd=2,
            width=220
        )
        right_buttons.pack(side="right", fill="y")
        right_buttons.pack_propagate(False)
        
        # Frame scrollable dentro del canvas
        scrollable_frame = tk.Frame(canvas, bg="white")
        
        # Configurar scroll
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        # Crear ventana en el canvas
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        
        # Posicionar elementos
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Permitir scroll con rueda del mouse
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        # --- TÍTULO PRINCIPAL (ESTILO UNIFICADO) ---
        tk.Label(
            scrollable_frame,
            text=titulo,
            font=("Arial", 14, "bold"),
            bg="white", 
            fg="red",
            pady=10
        ).pack()
        
        # --- SECCIÓN 1: GRÁFICO Y TABLA LADO A LADO ---
        section1_frame = tk.Frame(scrollable_frame, bg="white")
        section1_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        # --- COLUMNA IZQUIERDA: GRÁFICO DE LÍNEAS ---
        graph_frame = tk.Frame(section1_frame, bg="white", relief="groove", bd=1)
        graph_frame.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        # Título del gráfico
        tk.Label(
            graph_frame,
            text="GRÁFICO DE TENDENCIA",
            font=("Arial", 11, "bold"),
            bg="white", 
            fg="#2E5AAC",
            pady=5
        ).pack()
        
        # Crear figura para gráfico
        fig, ax = plt.subplots(figsize=(7, 3.5))
        
        # Preparar datos para gráfico
        meses = df_graf[x_col].tolist()
        horas = df_graf[y_col].tolist()
        
        # Gráfico de líneas con marcadores
        if meses and horas:
            ax.plot(meses, horas, marker='o', linewidth=2, markersize=6, 
                    color=color, label='Horas Totales')
            
            # Rellenar área bajo la curva
            ax.fill_between(meses, horas, alpha=0.2, color=color)
            
            # Personalizar gráfico
            ax.set_xlabel('Mes', fontsize=10, fontweight='bold')
            ax.set_ylabel('Horas Totales', fontsize=10, fontweight='bold')
            ax.grid(True, alpha=0.3, linestyle='--')
            ax.legend(loc='upper left', fontsize=9)
            
            # Rotar etiquetas del eje X si hay muchos meses
            if len(meses) > 6:
                plt.xticks(rotation=45, ha='right', fontsize=8)
            else:
                plt.xticks(fontsize=9)
                
            plt.yticks(fontsize=9)
        else:
            ax.text(0.5, 0.5, 'Sin datos para graficar', 
                    ha='center', va='center', transform=ax.transAxes,
                    fontsize=11, color='gray')
            ax.set_axis_off()
        
        plt.tight_layout()
        
        # Integrar gráfico en Tkinter
        canvas_graph = FigureCanvasTkAgg(fig, master=graph_frame)
        canvas_graph.draw()
        canvas_graph.get_tk_widget().pack(fill="both", expand=True, padx=10, pady=5)
        
        # --- COLUMNA DERECHA: TABLA RESUMEN ---
        table_frame = tk.Frame(section1_frame, bg="white", relief="groove", bd=1)
        table_frame.pack(side="right", fill="both", expand=True)
        
        # Título de la tabla
        tk.Label(
            table_frame,
            text="DETALLE POR MES",
            font=("Arial", 11, "bold"),
            bg="white", 
            fg="#2E5AAC",
            pady=5
        ).pack()
        
        # Crear Treeview para tabla
        table_container = tk.Frame(table_frame, bg="white")
        table_container.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Crear Treeview
        columns = ["Mes", "Horas", "Bomberos", "Registros", "Prom/HB"]
        tree = ttk.Treeview(table_container, columns=columns, show="headings", height=8)
        
        # Configurar estilo
        style = ttk.Style()
        style.theme_use("default")
        style.configure("Treeview", font=("Arial", 8), rowheight=22)
        style.configure("Treeview.Heading", font=("Arial", 9, "bold"), anchor="center")
        
        # Configurar columnas
        column_widths = [100, 80, 70, 70, 80]
        for col, width in zip(columns, column_widths):
            tree.heading(col, text=col)
            tree.column(col, width=width, anchor="center")
        
        # Insertar datos
        if not df_tabla.empty:
            for idx, row in df_tabla.iterrows():
                horas_formateadas = self._formatear_horas_decimal_a_hhmm(row['Horas_Decimal'])
                promedio = row.get('Promedio_Horas_Bombero', 0)
                promedio_formateado = self._formatear_horas_decimal_a_hhmm(promedio) if promedio > 0 else "0:00"
                
                tree.insert("", "end", values=(
                    row['Mes_Display'],
                    horas_formateadas,
                    int(row['Total_Bomberos']),
                    int(row['Total_Registros']),
                    promedio_formateado
                ))
        
        # Scrollbar para tabla
        table_scrollbar = ttk.Scrollbar(table_container, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=table_scrollbar.set)
        
        tree.pack(side="left", fill="both", expand=True)
        table_scrollbar.pack(side="right", fill="y")
        
        # --- SECCIÓN 2: ESTADÍSTICAS ---
        section2_frame = tk.Frame(scrollable_frame, bg="white", relief="groove", bd=2)
        section2_frame.pack(fill="x", padx=10, pady=15)
        
        # Título estadísticas
        tk.Label(
            section2_frame,
            text="📊 ESTADÍSTICAS DEL PERÍODO",
            font=("Arial", 11, "bold"),
            bg="white", 
            fg="red",
            pady=5
        ).pack()
        
        # Obtener totales
        totales = getattr(self, '_pdf_totales_horas_periodo', {})
        
        # Frame para estadísticas
        stats_container = tk.Frame(section2_frame, bg="white")
        stats_container.pack(fill="x", padx=15, pady=5)
        
        # Columna 1: Estadísticas generales
        col1 = tk.Frame(stats_container, bg="white")
        col1.pack(side="left", fill="both", expand=True, padx=10)
        
        tk.Label(
            col1,
            text="📈 GENERAL",
            font=("Arial", 9, "bold"),
            bg="white", fg="#2E5AAC"
        ).pack(anchor="w", pady=(0, 5))
        
        tk.Label(
            col1,
            text=f"• Meses analizados: {totales.get('total_meses', 0)}",
            font=("Arial", 9), bg="white", justify="left"
        ).pack(anchor="w")
        
        tk.Label(
            col1,
            text=f"• Horas totales: {self._formatear_horas_decimal_a_hhmm(totales.get('total_horas', 0))}",
            font=("Arial", 9), bg="white", justify="left"
        ).pack(anchor="w")
        
        tk.Label(
            col1,
            text=f"• Registros totales: {totales.get('total_registros', 0)}",
            font=("Arial", 9), bg="white", justify="left"
        ).pack(anchor="w")
        
        # Columna 2: Promedios
        col2 = tk.Frame(stats_container, bg="white")
        col2.pack(side="left", fill="both", expand=True, padx=10)
        
        tk.Label(
            col2,
            text="📊 PROMEDIOS",
            font=("Arial", 9, "bold"),
            bg="white", fg="#2E5AAC"
        ).pack(anchor="w", pady=(0, 5))
        
        tk.Label(
            col2,
            text=f"• Promedio mensual: {self._formatear_horas_decimal_a_hhmm(totales.get('promedio_horas_mensual', 0))}",
            font=("Arial", 9), bg="white", justify="left"
        ).pack(anchor="w")
        
        tk.Label(
            col2,
            text=f"• Promedio bombero/mes: {self._formatear_horas_decimal_a_hhmm(totales.get('promedio_horas_mensual', 0) / max(totales.get('total_meses', 1), 1))}",
            font=("Arial", 9), bg="white", justify="left"
        ).pack(anchor="w")
        
        # Columna 3: Análisis comparativo
        col3 = tk.Frame(stats_container, bg="white")
        col3.pack(side="left", fill="both", expand=True, padx=10)
        
        tk.Label(
            col3,
            text="📅 COMPARATIVO",
            font=("Arial", 9, "bold"),
            bg="white", fg="#2E5AAC"
        ).pack(anchor="w", pady=(0, 5))
        
        tk.Label(
            col3,
            text=f"• Mes pico: {totales.get('mes_pico', 'N/A')}",
            font=("Arial", 9), bg="white", justify="left"
        ).pack(anchor="w")
        
        tk.Label(
            col3,
            text=f"  Horas: {self._formatear_horas_decimal_a_hhmm(totales.get('horas_pico', 0))}",
            font=("Arial", 8), bg="white", justify="left"
        ).pack(anchor="w", padx=(10, 0))
        
        tk.Label(
            col3,
            text=f"• Mes valle: {totales.get('mes_valle', 'N/A')}",
            font=("Arial", 9), bg="white", justify="left"
        ).pack(anchor="w")
        
        tk.Label(
            col3,
            text=f"  Horas: {self._formatear_horas_decimal_a_hhmm(totales.get('horas_valle', 0))}",
            font=("Arial", 8), bg="white", justify="left"
        ).pack(anchor="w", padx=(10, 0))
        
        # Configurar cierre seguro
        def on_closing():
            plt.close('all')
            self._ventana_horas_periodo.destroy()

        # --- BOTONERA DERECHA ---
        ttk.Button(
            right_buttons,
            text="📄 Exportar PDF",
            command=self.exportar_pdf_horas_periodo
        ).pack(fill="x", padx=15, pady=5)

        ttk.Button(
            right_buttons,
            text="📊 Exportar Excel",
            command=self.exportar_excel_horas_periodo
        ).pack(fill="x", padx=15, pady=5)

        ttk.Separator(right_buttons, orient="horizontal").pack(fill="x", padx=10, pady=10)

        ttk.Button(
            right_buttons,
            text="❌ Cerrar",
            command=self._ventana_horas_periodo.destroy
        ).pack(fill="x", padx=15, pady=5)

        self._ventana_horas_periodo.protocol("WM_DELETE_WINDOW", on_closing)
        
    def _mostrar_form_informe_bombero_individual(
        self,
        titulo,
        df_graf,
        df_tabla,
        nombre_bombero,
        x_col="Actividad",
        y_col="Horas",
        color="steelblue"
    ):

        if df_graf is None or df_graf.empty:
            self.ui.show_info("Sin datos", "No hay datos para mostrar.")
            return
        # =========================
        # TÍTULO CORRECTO PARA BOMBERO INDIVIDUAL
        # =========================
        f1 = self.inf_desde.get_date()
        f2 = self.inf_hasta.get_date()

        titulo = (
            f"Informe de Actividades – {nombre_bombero} "
            f"({f1:%d/%m/%Y} a {f2:%d/%m/%Y})"
        )

        def cerrar_form():
            # Limpiar selección de la combo de legajos
            if hasattr(self, "inf_legajo_cb"):
                self.inf_legajo_cb.set("")

            # Opcional: limpiar variables asociadas
            if hasattr(self, "legajo_seleccionado"):
                self.legajo_seleccionado = None

            ventana.destroy()

        # =========================
        # VENTANA
        # =========================
        ventana = tk.Toplevel(self.master)

        if hasattr(self.master, "_icono_global") and self.master._icono_global:
            ventana.iconphoto(True, self.master._icono_global)

        ventana.title(titulo)
        ventana.geometry("1150x750")   # fallback
        try:
            ventana.state("zoomed")    # maximizado (Windows)
        except:
            pass
        ventana.configure(bg="white")
        ventana.grab_set()
        ventana.protocol("WM_DELETE_WINDOW", cerrar_form)

        # =========================
        # CONTENEDOR CON SCROLL GENERAL
        # =========================
        container = ttk.Frame(ventana)
        container.pack(fill="both", expand=True)

        main = ttk.Frame(container)
        main.pack(fill="both", expand=True)

        main.columnconfigure(0, weight=1)
        main.columnconfigure(1, weight=0)
        main.columnconfigure(2, weight=0)
        main.rowconfigure(0, weight=1)

        canvas = tk.Canvas(main, highlightthickness=0)
        scrollbar = ttk.Scrollbar(main, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")

        right = ttk.Frame(main, padding=10)
        right.grid(row=0, column=2, sticky="ns")

        frame = ttk.Frame(canvas, padding=10)
        canvas_window = canvas.create_window((0, 0), window=frame, anchor="nw")

        canvas.bind("<Configure>", lambda e: canvas.itemconfig(canvas_window, width=e.width))
        frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        # =========================
        # TÍTULO
        # =========================
        ttk.Label(
            frame,
            text=titulo,
            font=("Segoe UI", 13, "bold")
        ).pack(anchor="w", pady=(0, 8))

        # ======================================================
        # TOTALES – MISMA LÓGICA QUE OTROS INFORMES
        # ======================================================
        total_registros = len(df_tabla)

        # Limpiar nombre de actividad (solo para cálculo)
        df_tabla["_actividad_limpia"] = (
            df_tabla["actividad"]
            .astype(str)
            .str.replace(r"^\d+\s*-\s*", "", regex=True)
            .str.strip()
        )

        # Conversión de horas usando la lógica estándar del sistema
        df_tabla["_horas_decimal"] = df_tabla["horas"].apply(
            lambda x: self._obtener_horas_decimal(x)
        )

        resumen = (
            df_tabla
            .groupby("_actividad_limpia")["_horas_decimal"]
            .sum()
        )

        resumen = resumen[resumen > 0]

        total_actividades = resumen.shape[0]
        total_horas = resumen.sum()

        promedio_horas_actividad = (
            total_horas / total_actividades
            if total_actividades > 0 else 0
        )

        ttk.Label(
            frame,
            text=(
                f"Registros: {total_registros}   |   "
                f"Actividades: {total_actividades}   |   "
                f"Horas totales: {self._formatear_horas_decimal_a_hhmm(total_horas)}   |   "
                f"Prom. horas/actividad: {self._formatear_horas_decimal_a_hhmm(promedio_horas_actividad)}"
            ),
            font=("Segoe UI", 10, "bold")
        ).pack(anchor="w", pady=(0, 6))

        # ======================================================
        # DF PARA GRÁFICO (FUENTE ÚNICA: resumen ya calculado)
        # ======================================================
        df_graf = (
            resumen
            .reset_index()
            .rename(columns={
                "_actividad_limpia": x_col,
                "_horas_decimal": y_col
            })
        )

        # =========================
        # GRÁFICO (sin cambios de lógica)
        # =========================
        frame_grafico = ttk.Frame(frame)
        frame_grafico.pack(fill="x", pady=(0, 6))

        filas = len(df_graf)
        alto_figura = max(4, filas * 0.35)

        fig, ax = plt.subplots(figsize=(8.8, alto_figura))

        df_graf[x_col] = (
            df_graf[x_col]
            .astype(str)
            .str.replace(r"^\d+\s*-\s*", "", regex=True)
        )

        MAX_LEN = 32
        df_graf["_actividad_graf"] = df_graf[x_col].apply(
            lambda x: x if len(x) <= MAX_LEN else x[:MAX_LEN - 3] + "..."
        )

        df_plot = df_graf.sort_values(by=y_col, ascending=True)
        ax.barh(df_plot["_actividad_graf"], df_plot[y_col], color=color)

        ax.tick_params(axis="y", labelsize=7)
        ax.set_xlabel("Horas", fontsize=9)
        plt.subplots_adjust(left=0.30, right=0.96)
        ax.invert_yaxis()

        # === VALORES SOBRE LAS BARRAS ===
        max_val = df_plot[y_col].max() if not df_plot.empty else 0

        for i, val in enumerate(df_plot[y_col]):
            if val > 0:
                txt = self._formatear_horas_decimal_a_hhmm(val)
                dentro = val > max_val * 0.35
                offset = max_val * 0.04

                ax.text(
                    val - offset if dentro else val + offset * 0.25,
                    i,
                    txt,
                    va="center",
                    ha="right" if dentro else "left",
                    fontsize=8,
                    fontweight="bold",
                    color="white" if dentro else "black",
                    clip_on=False
                )

        canvas_fig = FigureCanvasTkAgg(fig, master=frame_grafico)
        canvas_fig.draw()
        canvas_fig.get_tk_widget().pack(fill="x")

        # =========================
        # TABLA
        # =========================
        frame_tabla = ttk.Frame(frame)
        frame_tabla.pack(fill="both", expand=True, pady=(0, 10))

        # 🔹 Estilo encabezados en negrita
        style = ttk.Style()
        style.configure(
            "Treeview.Heading",
            font=("Segoe UI", 9, "bold")
        )
        style.configure(
            "Treeview",
            font=("Segoe UI", 8)   # 👈 datos más chicos
        )

        encabezados = {
            "id": "ID",
            "legajo": "Legajo",
            "actividad": "Actividad",
            "area": "Área",
            "fecha_inicio": "Fecha inicio",
            "fecha_fin": "Fecha fin",
            "hora_inicio": "Hora inicio",
            "hora_fin": "Hora fin",
            "horas": "Horas",
            "descripcion": "Descripción",
            "asignado": "Asignado"
        }

        # ❌ excluir columnas técnicas
        cols = [c for c in df_tabla.columns if not c.startswith("_")]

        tree = ttk.Treeview(
            frame_tabla,
            columns=cols,
            show="headings",
            height=9   # 👈 9 filas visibles
        )

        def calcular_ancho(col):
            min_widths = {
                "fecha_inicio": 110,
                "hora_inicio": 100,
                "fecha_fin": 110,
                "hora_fin": 90
            }

            max_len = len(encabezados.get(col, col))
            for v in df_tabla[col].astype(str):
                max_len = max(max_len, len(v))

            ancho = max(max_len * 7, min_widths.get(col, 90))
            return min(ancho, 360)

        for col in cols:
            tree.heading(col, text=encabezados.get(col, col), anchor="center")
            tree.column(col, anchor="center", width=calcular_ancho(col), stretch=True)

        for _, row in df_tabla[cols].iterrows():
            fila = list(row)

            # Limpiar numeración en actividad SOLO para mostrar
            if "actividad" in cols:
                idx = cols.index("actividad")
                fila[idx] = (
                    str(fila[idx])
                    .replace("\n", " ")
                )
                fila[idx] = __import__("re").sub(r"^\d+\s*-\s*", "", fila[idx]).strip()

            tree.insert("", "end", values=fila)

        vsb = ttk.Scrollbar(frame_tabla, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(frame_tabla, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        frame_tabla.columnconfigure(0, weight=1)
        frame_tabla.rowconfigure(0, weight=1)

        # =========================
        # CONTEXTO DEL INFORME (IGUAL A OTROS FORMS)
        # =========================
        self.df_tabla_actual = df_tabla
        self.df_graf_actual = df_graf
        self.titulo_informe_actual = titulo
        self.tipo_informe_actual = "bombero_individual"
        self.nombre_bombero_actual = nombre_bombero

        self._df_resumen_bombero_individual = resumen.reset_index().rename(
            columns={
                "_actividad_limpia": "Actividad",
                "_horas_decimal": "Horas"
            }
        )

        self._pdf_totales_bombero_individual = {
            "total_actividades": total_actividades,
            "total_registros": total_registros,
            "total_horas": total_horas
        }

        # =========================
        # BOTONES
        # =========================
        frame_btn = ttk.Frame(right)
        frame_btn.pack(pady=10, fill="x")

        ttk.Button(
            frame_btn,
            text="Exportar PDF",
            command=self.exportar_pdf_bombero_individual
        ).pack(fill="x", pady=4)

        ttk.Button(
            frame_btn,
            text="Exportar Excel",
            command=lambda: self._exportar_informe_excel(
                titulo,
                self._df_resumen_bombero_individual
            )
        ).pack(fill="x", pady=4)

        ttk.Button(
            frame_btn,
            text="Cerrar",
            command=cerrar_form
        ).pack(fill="x", pady=12)

    def _mostrar_form_informe_actividad_individual(
        self,
        titulo,
        df_graf,
        df_tabla,
        nombre_actividad,
        x_col="Bombero",
        y_col="Horas",
        color="steelblue"
    ):

        if df_graf is None or df_graf.empty:
            self.ui.show_info("Sin datos", "No hay datos para mostrar.")
            return

        # Limpiar nombre de actividad (quitar "id - ")
        if " - " in nombre_actividad:
            nombre_actividad_limpio = nombre_actividad.split(" - ", 1)[1]
        else:
            nombre_actividad_limpio = nombre_actividad

        # =========================
        # TÍTULO
        # =========================
        f1 = self.inf_desde.get_date()
        f2 = self.inf_hasta.get_date()

        titulo = (
            f"Informe por Actividad – {nombre_actividad_limpio} "
            f"({f1:%d/%m/%Y} a {f2:%d/%m/%Y})"
        )

        # =========================
        # VENTANA
        # =========================
        def cerrar_form():
            if hasattr(self, "inf_actividad"):
                self.inf_actividad.set("")
                try:
                    self.inf_actividad.current(-1)
                except:
                    pass

            self.nombre_actividad_actual = None
            ventana.destroy()

        ventana = tk.Toplevel(self.master)
        ventana.title(titulo)
        ventana.geometry("1150x750")
        try:
            ventana.state("zoomed")
        except:
            pass
        ventana.configure(bg="white")
        ventana.grab_set()
        ventana.protocol("WM_DELETE_WINDOW", cerrar_form)

        # =========================
        # CONTENEDOR CON SCROLL
        # =========================
        container = ttk.Frame(ventana)
        container.pack(fill="both", expand=True)

        main = ttk.Frame(container)
        main.pack(fill="both", expand=True)

        main.columnconfigure(0, weight=1)
        main.columnconfigure(1, weight=0)
        main.columnconfigure(2, weight=0)
        main.rowconfigure(0, weight=1)

        canvas = tk.Canvas(main, highlightthickness=0)
        scrollbar = ttk.Scrollbar(main, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")

        right = ttk.Frame(main, padding=10)
        right.grid(row=0, column=2, sticky="ns")

        frame = ttk.Frame(canvas, padding=10)
        canvas_window = canvas.create_window((0, 0), window=frame, anchor="nw")

        canvas.bind("<Configure>", lambda e: canvas.itemconfig(canvas_window, width=e.width))
        frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        ttk.Label(frame, text=titulo, font=("Segoe UI", 13, "bold")).pack(anchor="w", pady=(0, 8))

        # =========================
        # TOTALES
        # =========================
        total_registros = len(df_tabla)

        df_tabla["_horas_decimal"] = df_tabla["horas"].apply(lambda x: self._obtener_horas_decimal(x))

        resumen = df_tabla.groupby("Bombero")["_horas_decimal"].sum()
        resumen = resumen[resumen > 0]

        total_bomberos = resumen.shape[0]
        total_horas = resumen.sum()

        # =========================
        # DF PARA GRÁFICO (fuente única: resumen)
        # =========================
        x_col = "Bombero"
        y_col = "Horas"

        df_graf = (
            resumen
            .reset_index()
            .rename(columns={
                "Bombero": x_col,
                "_horas_decimal": y_col
            })
        )

        promedio_horas_bombero = total_horas / total_bomberos if total_bomberos > 0 else 0

        ttk.Label(
            frame,
            text=(
                f"Registros: {total_registros}   |   "
                f"Bomberos: {total_bomberos}   |   "
                f"Horas totales: {self._formatear_horas_decimal_a_hhmm(total_horas)}   |   "
                f"Prom. horas/bombero: {self._formatear_horas_decimal_a_hhmm(promedio_horas_bombero)}"
            ),
            font=("Segoe UI", 10, "bold")
        ).pack(anchor="w", pady=(0, 6))

        # =========================
        # GRÁFICO
        # =========================
        frame_grafico = ttk.Frame(frame)
        frame_grafico.pack(fill="x", pady=(0, 6))

        filas = len(df_graf)
        alto_figura = max(4, filas * 0.35)

        fig, ax = plt.subplots(figsize=(8.8, alto_figura))

        MAX_LEN = 28
        df_graf["_bombero_graf"] = df_graf[x_col].astype(str).apply(
            lambda x: x if len(x) <= MAX_LEN else x[:MAX_LEN - 3] + "..."
        )

        df_plot = df_graf.sort_values(by=y_col, ascending=True)
        ax.barh(df_plot["_bombero_graf"], df_plot[y_col], color=color)

        ax.tick_params(axis="y", labelsize=7)
        ax.set_xlabel("Horas", fontsize=8)
        plt.subplots_adjust(left=0.30, right=0.96)
        ax.invert_yaxis()

        max_val = df_plot[y_col].max() if not df_plot.empty else 0

        for i, val in enumerate(df_plot[y_col]):
            if val > 0:
                txt = self._formatear_horas_decimal_a_hhmm(val)
                dentro = val > max_val * 0.35
                offset = max_val * 0.04
                ax.text(
                    val - offset if dentro else val + offset * 0.25,
                    i,
                    txt,
                    va="center",
                    ha="right" if dentro else "left",
                    fontsize=8,
                    fontweight="bold",
                    color="white" if dentro else "black",
                    clip_on=False
                )

        canvas_fig = FigureCanvasTkAgg(fig, master=frame_grafico)
        canvas_fig.draw()
        canvas_fig.get_tk_widget().pack(fill="x")

        # =========================
        # TABLA
        # =========================
        frame_tabla = ttk.Frame(frame, height=240)
        frame_tabla.pack(fill="x", expand=False, pady=(0, 10))
        frame_tabla.pack_propagate(False)  # 👈 evita que crezca por el contenido

        style = ttk.Style()
        style.configure("Treeview.Heading", font=("Segoe UI", 9, "bold"))
        style.configure("Treeview", font=("Segoe UI", 8))

        encabezados = {
            "id": "ID",
            "legajo": "Legajo",
            "bombero": "Bombero",
            "actividad": "Actividad", 
            "area": "Área",
            "fecha_inicio": "Fecha inicio",
            "fecha_fin": "Fecha fin",
            "hora_inicio": "Hora inicio",
            "hora_fin": "Hora fin",
            "horas": "Horas",
            "descripcion": "Descripción",
            "asignado": "Asignado"
        }

        # Limpiar columna actividad (quitar "id - ")
        if "actividad" in df_tabla.columns:
            df_tabla["actividad"] = df_tabla["actividad"].astype(str).apply(
                lambda x: x.split(" - ", 1)[1] if " - " in x else x
            )

        cols = [c for c in df_tabla.columns if not c.startswith("_")]

        tree = ttk.Treeview(frame_tabla, columns=cols, show="headings", height=9)

        def calcular_ancho(col):
            min_widths = {
                "fecha_inicio": 110,
                "hora_inicio": 100,
                "fecha_fin": 110,
                "hora_fin": 90
            }
            max_len = len(encabezados.get(col, col))
            for v in df_tabla[col].astype(str):
                max_len = max(max_len, len(v))
            ancho = max(max_len * 7, min_widths.get(col, 90))
            return min(ancho, 360)

        for col in cols:
            tree.heading(col, text=encabezados.get(col, col), anchor="center")
            tree.column(col, anchor="center", width=calcular_ancho(col), stretch=True)

        for _, row in df_tabla[cols].iterrows():
            tree.insert("", "end", values=list(row))

        vsb = ttk.Scrollbar(frame_tabla, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(frame_tabla, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        frame_tabla.columnconfigure(0, weight=1)
        frame_tabla.rowconfigure(0, weight=0)

        # =========================
        # CONTEXTO
        # =========================
        self.df_tabla_actual = df_tabla
        self.df_graf_actual = df_graf
        self.titulo_informe_actual = titulo
        self.tipo_informe_actual = "actividad_individual"
        self.nombre_actividad_actual = nombre_actividad

        self._df_resumen_actividad_individual = resumen.reset_index().rename(
            columns={"bombero": "Bombero", "_horas_decimal": "Horas"}
        )

        self._pdf_totales_actividad_individual = {
            "total_bomberos": total_bomberos,
            "total_registros": total_registros,
            "total_horas": total_horas
        }

        # =========================
        # BOTONES
        # =========================
        frame_btn = ttk.Frame(right)
        frame_btn.pack(pady=10, fill="x")

        ttk.Button(frame_btn, text="Exportar PDF",
                command=self.exportar_pdf_actividad_individual).pack(fill="x", pady=4)

        ttk.Button(frame_btn, text="Exportar Excel",
                command=lambda: self._exportar_informe_excel(
                    titulo,
                    self._df_resumen_actividad_individual
                )).pack(fill="x", pady=4)

        ttk.Button(frame_btn, text="Cerrar", command=cerrar_form).pack(fill="x", pady=12)

    def _exportar_pdf_actividad(self):
        id_act = self.var_id_actividad.get().strip()
        if not id_act:
            self.ui.show_error("Error", "Debe cargar primero una actividad.")
            return

        sugerido = self._default_informe_filename(f"Actividad_{id_act}")
        path = self.ui.ask_save_file(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            initialfile=sugerido
        )
        if not path:
            return

        try:
            self._crear_pdf_actividad(path)
            self.ui.show_info("OK", f"Actividad exportada correctamente:\n{path}")
        except Exception as e:
            self.ui.show_error("Error", f"No se pudo exportar el PDF:\n{e}")

    def exportar_pdf_actividad_individual(self):

        if self.tipo_informe_actual != "actividad_individual":
            messagebox.showwarning("Atención", "No hay un informe de actividad individual activo.")
            return

        df_resumen = getattr(self, "_df_resumen_actividad_individual", None)
        totales = getattr(self, "_pdf_totales_actividad_individual", None)

        if df_resumen is None or totales is None:
            messagebox.showwarning("Atención", "No hay datos para exportar.")
            return

        nombre_sugerido = re.sub(r'[\\/:*?"<>|]', '-', self.titulo_informe_actual)

        file_path = self.ui.ask_save_file(
            defaultextension=".pdf",
            filetypes=[("Archivos PDF", "*.pdf")],
            title="Guardar informe por actividad",
            initialfile=f"{nombre_sugerido}.pdf"
        )

        if not file_path:
            return

        try:
            self._crear_pdf_resumen_actividad_individual(
                file_path=file_path,
                titulo_pdf=self.titulo_informe_actual,
                df_resumen=df_resumen,
                totales=self._pdf_totales_actividad_individual
            )
            self.ui.show_info("Éxito", "Informe por actividad exportado correctamente.")
            # 👉 Abrir automáticamente el PDF
            try:
                if sys.platform.startswith("win"):
                    os.startfile(file_path)
                elif sys.platform.startswith("darwin"):
                    os.system(f"open '{file_path}'")
                else:
                    os.system(f"xdg-open '{file_path}'")
            except Exception as e:
                print("No se pudo abrir el PDF automáticamente:", e)
        except Exception as e:
            self.ui.show_error("Error", f"No se pudo exportar el PDF:\n{e}")

    def _crear_pdf_resumen_actividad_individual(
        self,
        file_path,
        titulo_pdf,
        df_resumen,
        totales
    ):
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle,
            Paragraph, Spacer
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm

        styles = getSampleStyleSheet()
        elems = []

        elems.append(Spacer(1, 20))

        # ==========================
        # TABLA RESUMEN POR BOMBERO
        # ==========================
        data = [["BOMBERO", "HORAS"]]

        for _, r in df_resumen.iterrows():
            data.append([
                r["Bombero"],
                self._formatear_horas_decimal_a_hhmm(r["Horas"])
            ])

        ancho = 595 - (30 * mm)
        tbl = Table(data, colWidths=[ancho * 0.75, ancho * 0.25])

        tbl.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.25, colors.gray),

            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E5AAC")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, 0), "MIDDLE"),
            ("ALIGN", (1, 1), (-1, -1), "CENTER"),
        ]))

        elems.append(tbl)
        elems.append(Spacer(1, 15))

        # ==========================
        # TOTALES
        # ==========================
        estilo = ParagraphStyle(
            "resumen",
            parent=styles["Normal"],
            fontSize=9,
            leftIndent=10,
            spaceAfter=4
        )

        total_bomberos = totales.get("total_bomberos", 0)
        total_registros = totales.get("total_registros", 0)
        total_horas = totales.get("total_horas", 0)

        promedio = (
            total_horas / total_bomberos
            if total_bomberos else 0
        )

        elems.append(Paragraph(
            f"<b>• Total de registros:</b> {total_registros}", estilo
        ))
        elems.append(Paragraph(
            f"<b>• Bomberos distintos:</b> {total_bomberos}", estilo
        ))
        elems.append(Paragraph(
            f"<b>• Total de horas:</b> "
            f"{self._formatear_horas_decimal_a_hhmm(total_horas)}", estilo
        ))
        elems.append(Paragraph(
            f"<b>• Promedio de horas por bombero:</b> "
            f"{self._formatear_horas_decimal_a_hhmm(promedio)}", estilo
        ))

        doc = SimpleDocTemplate(
            file_path,
            pagesize=A4,
            leftMargin=15 * mm,
            rightMargin=15 * mm,
            topMargin=40 * mm,
            bottomMargin=15 * mm
        )

        doc.build(
            elems,
            onFirstPage=lambda c, d:
                self._formato_pdf(c, d, titulo_pdf, False),
            onLaterPages=lambda c, d:
                self._formato_pdf(c, d, titulo_pdf, False)
        )

    def _sugerir_nombre_pdf(self, base, fecha_desde, fecha_hasta):
        """
        Genera nombre sugerido para PDFs con período
        """
        try:
            desde = fecha_desde.strftime("%Y-%m-%d")
            hasta = fecha_hasta.strftime("%Y-%m-%d")
        except Exception:
            desde = str(fecha_desde)
            hasta = str(fecha_hasta)

        return f"{base}_{desde}_a_{hasta}.pdf"

    def exportar_pdf_actividades_total(self):

        # 🔴 ESTE IF ES CLAVE
        if not hasattr(self, "_df_resumen_actividades_total") or self._df_resumen_actividades_total.empty:
            self.ui.show_error("Error", "Primero debe generar el informe.")
            return

        # 🔹 Verificar que el informe fue generado
        if not hasattr(self, "_df_resumen_actividades_total") \
        or self._df_resumen_actividades_total is None \
        or self._df_resumen_actividades_total.empty:
            self.ui.show_error("Error", "Primero debe generar el informe.")
            return

        try:
            f1 = self.inf_desde.get_date()
            f2 = self.inf_hasta.get_date()

            file_path = self._sugerir_nombre_pdf(
                "Informe_Actividades_Todas", f1, f2
            )

            if not file_path:
                return

            self._crear_pdf_resumen_actividades_total(
                file_path=file_path,
                titulo=f"Informe Actividades - Del {f1:%d/%m/%Y} al {f2:%d/%m/%Y}",
                df_resumen=self._df_resumen_actividades_total,
                totales=self._pdf_totales_actividades
            )

            self.ui.show_info("Éxito", f"PDF generado correctamente:\n{file_path}")
            os.startfile(os.path.abspath(file_path))

        except Exception as e:
            self.ui.show_error("Error", f"No se pudo generar el PDF:\n{e}")

#--------------------------------INFORME POR TODAS LAS ACTIVIDADES-----------------------------
    def informe_todas_actividades(self):
        """NUEVA: Informe para TODAS las actividades (sin selección)."""
        
        print("=== NUEVO >> informe_todas_actividades (GENÉRICO) ===")
        
        try:
            f1 = self.inf_desde.get_date()
            f2 = self.inf_hasta.get_date()
            s1, s2 = f1.strftime("%Y-%m-%d"), f2.strftime("%Y-%m-%d")
        except Exception as e:
            self.ui.show_error("Error", "Debe seleccionar fechas válidas.")
            return
        
        conn = sqlite3.connect(DB_PATH)
        try:
            # CONSULTA SIMPLE para actividades generales
            q = """
                SELECT 
                    COALESCE(actividad, 'SIN ACTIVIDAD') AS Actividad,
                    COUNT(DISTINCT legajo) AS Total_Bomberos,
                    COUNT(id) AS Total_Registros,
                    ROUND(SUM(
                        CASE 
                            WHEN typeof(horas) = 'text' AND horas LIKE '%:%' THEN
                                CAST(SUBSTR(horas, 1, INSTR(horas, ':')-1) AS REAL) + 
                                CAST(SUBSTR(horas, INSTR(horas, ':')+1) AS REAL) / 60.0
                            ELSE CAST(COALESCE(horas, 0) AS REAL)
                        END
                    ), 2) AS Horas_Decimal
                FROM actividades
                WHERE (CASE WHEN instr(fecha_inicio, '/')>0
                            THEN date(substr(fecha_inicio,7,4)||'-'||substr(fecha_inicio,4,2)||'-'||substr(fecha_inicio,1,2))
                            ELSE date(fecha_inicio) END)
                    BETWEEN date(?) AND date(?)
                GROUP BY actividad
                ORDER BY Horas_Decimal DESC
            """
            
            df_res = pd.read_sql_query(q, conn, params=(s1, s2))
            # 🔥 Limpiar prefijo "N - " en Actividad
            df_res["Actividad"] = (
                df_res["Actividad"]
                .astype(str)
                .str.replace(r"^\d+\s*-\s*", "", regex=True)
                .str.strip()
            )

            # 🔥 Volver a agrupar ya limpio (esto es CLAVE)
            df_res = (
                df_res
                .groupby("Actividad", as_index=False)
                .agg({
                    "Total_Bomberos": "sum",
                    "Total_Registros": "sum",
                    "Horas_Decimal": "sum"
                })
                .sort_values("Horas_Decimal", ascending=False)
            )

            if df_res.empty:
                self.ui.show_info("Sin datos", "No hay actividades en el período seleccionado.")
                return
            
            # Preparar datos para visualización simple
            titulo = f"Todas las Actividades ({f1:%d/%m/%Y} - {f2:%d/%m/%Y})"
            
            # Usar la función existente pero con parámetros simplificados
            df_graf = df_res[['Actividad', 'Horas_Decimal']].rename(columns={'Horas_Decimal': 'Horas'})

            df_tabla = df_res.rename(columns={
                'Total_Bomberos': 'Bomberos',
                'Total_Registros': 'Registros',
                'Horas_Decimal': 'Horas'
            })
            # =========================================================
            # 🔐 DATOS PARA EXPORTAR PDF (ACTIVIDADES TODAS)
            # =========================================================

            # PDF SOLO necesita Actividad + Horas
            self._df_resumen_actividades_total = df_tabla[["Actividad", "Horas"]].copy()

            # 🔹 DataFrame SOLO para PDF (Actividad + Horas)
            df_pdf = df_tabla[["Actividad", "Horas"]].copy()

            # 🔥 CALCULAR TOTALES PARA EL PANEL DERECHO
            total_actividades_distintas = len(df_res)
            total_bomberos = df_res["Total_Bomberos"].sum()
            total_registros = df_res["Total_Registros"].sum()
            total_horas = df_res["Horas_Decimal"].sum()

            # =========================================================
            # 🔐 GUARDAR DATOS PARA EXPORTACIÓN (PDF / EXCEL)
            # =========================================================

            # 🔹 ESTE DataFrame es el que usa el PDF
            # 🔹 DataFrame exclusivo para PDF (SIN Bomberos ni Registros)
            self._df_resumen_actividades_total = df_tabla[["Actividad", "Horas"]].copy()
            # 🔹 ESTE DataFrame lo usa el EXCEL (Hoja 1 y Hoja 2)
            self._df_tabla_actual = df_tabla.copy()
            self._df_detalle_actual = df_tabla.copy()

            # 🔹 Totales que lee el PDF (MISMO ESQUEMA que Bomberos)
            self._pdf_totales_actividades = {
                "total_actividades": total_actividades_distintas,
                "total_bomberos": total_bomberos,
                "total_registros": total_registros,
                "total_horas": total_horas
            }

            # 🔹 Marcar tipo de informe activo
            self.tipo_informe_actual = "actividades_total"

            promedio_horas_por_actividad = total_horas / total_actividades_distintas if total_actividades_distintas > 0 else 0
            promedio_registros_por_actividad = total_registros / total_actividades_distintas if total_actividades_distintas > 0 else 0

            self._mostrar_form_informe_actividades_todas(
                titulo,
                df_graf,
                df_tabla,
                x_col="Actividad",
                y_col="Horas",
                color="green",
                total_actividades_distintas=total_actividades_distintas,
                total_bomberos=total_bomberos,
                total_registros=total_registros,
                total_horas=total_horas,
                promedio_horas_por_registro=0,
                promedio_registros_por_actividad=promedio_registros_por_actividad,
                promedio_horas_por_bombero=0,
                promedio_actividades_por_bombero=0,
                promedio_horas_por_mes=0,
                report_type="Informe_Actividades"
            )
            
        except Exception as e:
            self.ui.show_error("Error", f"No se pudo generar el informe:\n{e}")
            import traceback
            traceback.print_exc()
        finally:
            conn.close()

#-------------------------------INFORME TODOS BOMBEROS -------------------------------------
    def informe_todos_bomberos(self):
        """Informe para TODOS los bomberos (sin selección)."""
        
        print("=== informe_todos_bomberos (GENERAL) - VERSIÓN CORREGIDA ===")
        
        try:
            f1 = self.inf_desde.get_date()
            f2 = self.inf_hasta.get_date()
            s1, s2 = f1.strftime("%Y-%m-%d"), f2.strftime("%Y-%m-%d")
        except Exception:
            self.ui.show_error("Error", "Debe seleccionar fechas válidas.")
            return
        
        conn = sqlite3.connect(DB_PATH)
        try:
            # 🔥 CONSULTA CORREGIDA - Separada en dos pasos para evitar errores
            # Paso 1: Obtener todos los legajos únicos en el período
            q_legajos = """
                SELECT DISTINCT a.legajo
                FROM actividades a
                WHERE a.legajo IS NOT NULL AND TRIM(a.legajo) != ''
                AND (CASE WHEN instr(a.fecha_inicio, '/')>0
                            THEN date(substr(a.fecha_inicio,7,4)||'-'||substr(a.fecha_inicio,4,2)||'-'||substr(a.fecha_inicio,1,2))
                            ELSE date(a.fecha_inicio) END)
                    BETWEEN date(?) AND date(?)
            """
            
            legajos_df = pd.read_sql_query(q_legajos, conn, params=(s1, s2))
            
            if legajos_df.empty:
                self.ui.show_info("Sin datos", "No hay actividades en el período seleccionado.")
                return
            
            legajos = legajos_df['legajo'].tolist()
            resultados = []
            
            # Paso 2: Calcular horas para CADA legajo por separado
            for legajo in legajos:
                q_horas = """
                    SELECT 
                        ? as legajo,
                        COALESCE(l.apellido, '') as apellido,
                        COALESCE(l.nombre, '') as nombre,
                        COUNT(a.id) as Total_Registros,
                        COUNT(DISTINCT a.actividad) as Actividades_Distintas,
                        -- 🔥 SUMA EXACTA para este legajo específico
                        ROUND(SUM(
                            CASE 
                                WHEN typeof(a.horas) = 'text' AND a.horas LIKE '%:%' THEN
                                    CAST(SUBSTR(a.horas, 1, INSTR(a.horas, ':')-1) AS REAL) + 
                                    CAST(SUBSTR(a.horas, INSTR(a.horas, ':')+1) AS REAL) / 60.0
                                WHEN typeof(a.horas) IN ('real', 'integer') THEN
                                    CAST(a.horas AS REAL)
                                ELSE 0.0
                            END
                        ), 2) as Horas_Totales
                    FROM actividades a
                    LEFT JOIN legajos l ON a.legajo = l.legajo
                    WHERE a.legajo = ?
                    AND (CASE WHEN instr(a.fecha_inicio, '/')>0
                                THEN date(substr(a.fecha_inicio,7,4)||'-'||substr(a.fecha_inicio,4,2)||'-'||substr(a.fecha_inicio,1,2))
                                ELSE date(a.fecha_inicio) END)
                        BETWEEN date(?) AND date(?)
                """
                
                c = conn.cursor()
                c.execute(q_horas, (legajo, legajo, s1, s2))
                row = c.fetchone()
                
                if row and row[5] > 0:  # Horas_Totales > 0
                    resultados.append({
                        'legajo': row[0],
                        'apellido': row[1],
                        'nombre': row[2],
                        'Total_Registros': row[3],
                        'Actividades_Distintas': row[4],
                        'Horas_Totales': row[5]
                    })
                    
                    # Verificación para Castro
                    if str(legajo) == '191':
                        print(f"  Legajo: {row[0]}")
                        print(f"  Apellido: {row[1]}")
                        print(f"  Nombre: {row[2]}")
                        print(f"  Total Registros: {row[3]}")
                        print(f"  Horas Totales: {row[5]}")
            
            if not resultados:
                self.ui.show_info("Sin datos", "No hay actividades en el período seleccionado.")
                return
            
            # 🔥 CREAR DATAFRAME RESUMEN (OBLIGATORIO)
            df_res = pd.DataFrame(resultados)

            # Ordenar por horas descendente
            df_res = df_res.sort_values('Horas_Totales', ascending=False)

            # 🔥 CREAR DATAFRAME DETALLE
            self._df_detalle_bomberos_total = None
            # 🔥 GUARDAR DATA PARA EXPORTAR PDF (DEFINITIVO)
            self._df_resumen_bomberos_total = df_res.copy()

         
            # Crear columna de visualización
            df_res['Bombero_Display'] = df_res.apply(
                lambda x: f"{x['legajo']} - {x['apellido']} {x['nombre']}" 
                if len(f"{x['apellido']} {x['nombre']}".strip()) <= 25
                else f"{x['legajo']} - {x['apellido']} {x['nombre']}"[:25] + "...",
                axis=1
            )
            
            print(f"\n=== RESUMEN CORREGIDO ===")
            for idx, row in df_res.iterrows():
                print(f"  {row['Bombero_Display']}: {row['Horas_Totales']:.2f} horas")
            
            # 🔥 CALCULAR TOTALES REALES
            total_horas = df_res['Horas_Totales'].sum()
           
            # Preparar datos para gráfico
            df_graf = df_res[['Bombero_Display', 'Horas_Totales']].copy()
            df_graf.columns = ['Bombero', 'Horas']
            df_graf = df_graf.sort_values('Horas', ascending=False)  # Ordenar por horas
            
            # Crear tabla formateada
            df_tabla = pd.DataFrame({
                'Legajo': df_res['legajo'],
                'Bombero': df_res.apply(lambda x: f"{x['apellido']} {x['nombre']}".strip(), axis=1),
                'Registros': df_res['Total_Registros'],
                'Actividades': df_res['Actividades_Distintas'],
                'Horas': df_res['Horas_Totales'].apply(self._formatear_horas_decimal_a_hhmm)
            })
            
            titulo = f"Todos los Bomberos ({f1:%d/%m/%Y} - {f2:%d/%m/%Y})"
            
            # Calcular estadísticas
            total_bomberos = len(df_res)
            total_actividades_distintas = df_res['Actividades_Distintas'].sum()
            total_registros = df_res['Total_Registros'].sum()
            promedio_horas_por_bombero = total_horas / total_bomberos if total_bomberos > 0 else 0
            promedio_actividades_por_bombero = total_registros / total_bomberos if total_bomberos > 0 else 0
            
            print(f"\n=== ESTADÍSTICAS FINALES CORREGIDAS ===")

            # 🔥 GUARDAR SIEMPRE PARA EXPORTAR PDF
            self._df_resumen_bomberos_total = df_res.copy()
            self._df_detalle_bomberos_total = None

            # 🔥 GUARDAR TOTALES PARA EL PDF (OBLIGATORIO)
            self._pdf_totales_bomberos = {
                "total_bomberos": total_bomberos,
                "total_registros": total_registros,
                "actividades_distintas": total_actividades_distintas,
                "total_horas": total_horas,
                "promedio_horas_bombero": promedio_horas_por_bombero,
                "cantidad_meses": cantidad_meses if 'cantidad_meses' in locals() else 0
            }

            # Mostrar el informe
            self._mostrar_form_informe_bomberos_todos(
                titulo,
                df_graf,
                df_tabla,
                x_col="Bombero",
                y_col="Horas",
                color="blue",
                total_actividades_distintas=total_actividades_distintas,
                total_registros=total_registros,
                total_bomberos=total_bomberos,
                total_horas=total_horas,
                promedio_horas_por_bombero=promedio_horas_por_bombero,
                promedio_actividades_por_bombero=promedio_actividades_por_bombero
            )
            
        except Exception as e:
            self.ui.show_error("Error", f"No se pudo generar el informe:\n{e}")
            import traceback
            traceback.print_exc()
        finally:
            conn.close()

    # Agrega esta función temporal para diagnóstico
    def probar_suma_horas(self):
        """Función de diagnóstico para probar la suma de horas."""
        import sqlite3
        
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        # Probemos con un bombero específico
        legajo = '191'  # CASTRO
        
        print(f"\n=== DIAGNÓSTICO SUMA HORAS PARA {legajo} ===")
        
        # 1. Ver todas las actividades de Castro
        c.execute("""
            SELECT id, horas, actividad, fecha_inicio 
            FROM actividades 
            WHERE legajo = ? 
            ORDER BY fecha_inicio
        """, (legajo,))
        
        actividades = c.fetchall()
        print(f"Total actividades de {legajo}: {len(actividades)}")
        
        total_horas = 0
        for act in actividades[:10]:  # Mostrar primeras 10
            id_act, horas_str, actividad, fecha = act
            print(f"  ID {id_act}: {fecha} - {actividad} - Horas: '{horas_str}'")
            
            # Convertir a decimal
            if horas_str and ':' in str(horas_str):
                try:
                    h, m = map(float, str(horas_str).split(':'))
                    horas_dec = h + (m / 60.0)
                except:
                    horas_dec = 0
            elif horas_str:
                try:
                    horas_dec = float(horas_str)
                except:
                    horas_dec = 0
            else:
                horas_dec = 0
                
            total_horas += horas_dec
        
        print(f"\nSuma manual de horas para {legajo}: {total_horas:.2f}")
        
        # 2. Probar la consulta de suma
        c.execute("""
            SELECT 
                legajo,
                SUM(
                    CASE 
                        WHEN typeof(horas) = 'text' AND horas LIKE '%:%' THEN
                            CAST(SUBSTR(horas, 1, INSTR(horas, ':')-1) AS REAL) + 
                            CAST(SUBSTR(horas, INSTR(horas, ':')+1) AS REAL) / 60.0
                        WHEN typeof(horas) IN ('real', 'integer') THEN
                            CAST(horas AS REAL)
                        ELSE 0.0
                    END
                ) as total_horas
            FROM actividades 
            WHERE legajo = ?
            GROUP BY legajo
        """, (legajo,))
        
        resultado = c.fetchone()
        if resultado:
            print(f"Suma SQL para {legajo}: {resultado[1]:.2f} horas")
        
        conn.close()

# ===================== INFORME POR ACTIVIDADES =====================
    def generar_informe_por_actividad_click(self):
        """Función principal que decide qué informe de actividad mostrar."""
        
        try:
            f1 = self.inf_desde.get_date()
            f2 = self.inf_hasta.get_date()
        except Exception as e:
            self.ui.show_error("Error", "Debe seleccionar fechas válidas.")
            return
        
        actividad_sel = (self.inf_actividad.get() or "").strip()
        
        if actividad_sel:
            # Tiene selección → Detalle
            self.informe_detalle_por_actividad()
        else:
            # Sin selección → Resumen
            self.informe_resumen_por_actividad()
        
    def generar_informe_horas_click(self):
        """Versión mejorada del informe de horas por período"""
        import pandas as pd
        try:
            f1 = self.inf_desde.get_date()
            f2 = self.inf_hasta.get_date()
        except Exception:
            self.ui.show_error("Error", "Debe seleccionar fechas válidas.")
            return

        s1, s2 = f1.strftime("%Y-%m-%d"), f2.strftime("%Y-%m-%d")

        conn = sqlite3.connect(DB_PATH)
        try:
            # Consulta mejorada que agrupa por mes y actividad
            q = """
                SELECT 
                    strftime('%Y-%m', 
                        CASE WHEN instr(a.fecha_inicio, '/')>0
                            THEN date(substr(a.fecha_inicio,7,4)||'-'||substr(a.fecha_inicio,4,2)||'-'||substr(a.fecha_inicio,1,2))
                            ELSE date(a.fecha_inicio) END
                    ) AS Mes,
                    a.actividad AS Actividad,
                    COUNT(a.id) AS Cantidad_Actividades,
                    ROUND(SUM(COALESCE(a.horas, 0)), 2) AS Total_Horas,
                    COUNT(DISTINCT a.legajo) AS Bomberos_Involucrados
                FROM actividades a
                WHERE (CASE WHEN instr(a.fecha_inicio, '/')>0
                            THEN date(substr(a.fecha_inicio,7,4)||'-'||substr(a.fecha_inicio,4,2)||'-'||substr(a.fecha_inicio,1,2))
                            ELSE date(a.fecha_inicio) END)
                    BETWEEN date(?) AND date(?)
                GROUP BY Mes, a.actividad
                ORDER BY Mes, Total_Horas DESC
            """
            df_resumen = pd.read_sql_query(q, conn, params=(s1, s2))
            
            if df_resumen.empty:
                self.ui.show_info("Sin datos", "No hay actividades en el período seleccionado.")
                return

            # Preparar datos para gráfico (agrupar por mes)
            df_graf_mes = df_resumen.groupby('Mes', as_index=False)['Total_Horas'].sum()
            df_graf_mes.columns = ['Mes', 'Horas']
            
            # Formatear nombres de meses
            meses_esp = {
                '01': 'Enero', '02': 'Febrero', '03': 'Marzo', '04': 'Abril',
                '05': 'Mayo', '06': 'Junio', '07': 'Julio', '08': 'Agosto',
                '09': 'Septiembre', '10': 'Octubre', '11': 'Noviembre', '12': 'Diciembre'
            }
            
            def formatear_mes(mes_str):
                try:
                    anio, mes = mes_str.split('-')
                    return f"{meses_esp.get(mes, mes)} {anio}"
                except:
                    return mes_str
            
            df_graf_mes['Mes'] = df_graf_mes['Mes'].apply(formatear_mes)

            titulo = f"Horas por Período ({f1:%d/%m/%Y} - {f2:%d/%m/%Y})"
            
            # Calcular totales
            total_actividades_distintas = df_resumen['Actividad'].nunique()
            total_registros = df_resumen['Cantidad_Actividades'].sum()
            total_bomberos = df_resumen['Bomberos_Involucrados'].sum()
            total_horas = df_resumen['Total_Horas'].sum()
            cantidad_meses = df_graf_mes['Mes'].nunique()
            promedio_horas_por_mes = total_horas / cantidad_meses if cantidad_meses > 0 else 0

            self._mostrar_form_informe_actividades_todas(
                titulo, df_graf_mes, df_resumen,
                x_col="Mes", y_col="Horas",
                color="green",
                total_actividades_distintas=total_actividades_distintas,
                total_registros=total_registros,
                total_bomberos=total_bomberos,
                total_horas=total_horas,
                cantidad_meses=cantidad_meses,
                promedio_horas_por_mes=promedio_horas_por_mes
            )

        except Exception as e:
            self.ui.show_error("Error", f"No se pudo generar el informe:\n{e}")
        finally:
            conn.close()

    def _exportar_informe_excel(self, titulo, df):
        def limpiar_nombre_archivo(texto):
            if not texto:
                return "Sin_Nombre"
            return re.sub(r'[\\/:*?"<>|]+', '-', str(texto)).strip()

        if df is None or df.empty:
            self.ui.show_info("Exportar", "No hay datos para exportar.")
            return

        # =============================
        # NOMBRE SUGERIDO
        # =============================
        fecha_actual = datetime.now().strftime("%Y-%m-%d")

        titulo_lower = titulo.lower()

        if self.tipo_informe_actual == "bombero_individual":
            base_name = f"Informe_Bombero_{self.nombre_bombero_actual}"

        elif self.tipo_informe_actual == "actividad_individual":
            nombre_limpio = self.nombre_actividad_actual
            if " - " in nombre_limpio:
                nombre_limpio = nombre_limpio.split(" - ", 1)[1]
            base_name = f"Informe_Actividad_{nombre_limpio}"

        elif self.tipo_informe_actual == "horas_periodo":
            base_name = "Informe_Horas_por_Periodo"

        elif "bombero" in titulo_lower:
            base_name = "Informe_Todos_Bomberos"

        else:
            base_name = "Informe_Todas_Actividades"

        # 👉 LIMPIEZA DEL NOMBRE
        base_name = limpiar_nombre_archivo(base_name)

        fecha_match = re.search(r"(\d{2}/\d{2}/\d{4}).*?(\d{2}/\d{2}/\d{4})", titulo)
        if fecha_match:
            desde, hasta = fecha_match.groups()
            fecha_str = f"{desde.replace('/', '-')}_a_{hasta.replace('/', '-')}"
        else:
            fecha_str = fecha_actual

        nombre_archivo = f"{base_name}_{fecha_str}.xlsx"

        ruta = self.ui.ask_save_file(
            defaultextension=".xlsx",
            initialfile=nombre_archivo,
            filetypes=[("Excel files", "*.xlsx")],
            title="Guardar informe como Excel"
        )

        if not ruta:
            return

        try:
            # =============================
            # LIMPIEZA DE COLUMNAS
            # =============================
            df_export = df[[c for c in df.columns if not c.startswith("_")]].copy()
            # 👉 Quitar columna actividad del Excel
            if "actividad" in df_export.columns:
                df_export.drop(columns=["actividad"], inplace=True)

            if "actividad" in df_export.columns:
                df_export["actividad"] = (
                    df_export["actividad"]
                    .astype(str)
                    .str.replace(r"^\d+\s*-\s*", "", regex=True)
                    .str.strip()
                )

            # 🔹 AGREGAR COLUMNA PROMEDIO (si existe total/registro)
            if hasattr(self, "_pdf_totales_bombero_individual") and "total_horas" in self._pdf_totales_bombero_individual:
                t = self._pdf_totales_bombero_individual
                if t.get("total_actividades", 0) > 0:
                    df_export["Promedio"] = df_export["Horas"].apply(
                        lambda x: self._formatear_horas_decimal_a_hhmm(x / t.get("total_actividades", 1))
                    )
            # ===== FORMATEAR HORAS PARA EXCEL (AL FINAL) =====
            for col in df_export.columns:
                if "hora" in col.lower():
                    df_export[col] = df_export[col].apply(self._formatear_horas_decimal_a_hhmm)

            # =============================
            # ESCRITURA EXCEL (CLAVE)
            # =============================
            with pd.ExcelWriter(ruta, engine="openpyxl") as writer:

                # 🔵 HOJA 1: DATOS (SIEMPRE)
                df_export.to_excel(writer, sheet_name="Datos", index=False)

                # 🟣 HOJA 2: TOTALES
                totales = []

                if self.tipo_informe_actual == "bombero_individual":
                    t = getattr(self, "_pdf_totales_bombero_individual", {})

                    totales.append(["Bombero", self.nombre_bombero_actual])
                    totales.append(["Total registros", t.get("total_registros", 0)])
                    totales.append(["Actividades distintas", t.get("total_actividades", 0)])

                    total_horas = t.get("total_horas", 0)
                    totales.append([
                        "Total horas",
                        self._formatear_horas_decimal_a_hhmm(total_horas)
                    ])

                    promedio = (
                        total_horas / t.get("total_actividades", 1)
                        if t.get("total_actividades", 0) else 0
                    )
                    totales.append([
                        "Promedio horas / actividad",
                        self._formatear_horas_decimal_a_hhmm(promedio)
                    ])

                elif self.tipo_informe_actual == "actividad_individual":
                    t = getattr(self, "_pdf_totales_actividad_individual", {})

                    totales.append(["Actividad", self.nombre_actividad_actual])
                    totales.append(["Total registros", t.get("total_registros", 0)])
                    totales.append(["Bomberos distintos", t.get("total_bomberos", 0)])

                    total_horas = t.get("total_horas", 0)
                    totales.append([
                        "Total horas",
                        self._formatear_horas_decimal_a_hhmm(total_horas)
                    ])

                    promedio = (
                        total_horas / t.get("total_bomberos", 1)
                        if t.get("total_bomberos", 0) else 0
                    )
                    totales.append([
                        "Promedio horas / bombero",
                        self._formatear_horas_decimal_a_hhmm(promedio)
                    ])

                elif hasattr(self, "_pdf_totales_bomberos"):
                    t = self._pdf_totales_bomberos

                    totales.append(["Bomberos participantes", t.get("total_bomberos", 0)])
                    totales.append(["Total registros", t.get("total_registros", 0)])
                    totales.append(["Actividades distintas", t.get("actividades_distintas", 0)])
                    totales.append([
                        "Total horas",
                        self._formatear_horas_decimal_a_hhmm(t.get("total_horas", 0))
                    ])
                    totales.append([
                        "Promedio horas/bombero",
                        self._formatear_horas_decimal_a_hhmm(t.get("promedio_horas_bombero", 0))
                    ])
                elif self.tipo_informe_actual == "horas_periodo":
                    t = getattr(self, "_pdf_totales_horas_periodo", {})

                    totales.append(["Total meses", t.get("total_meses", 0)])
                    totales.append(["Total registros", t.get("total_registros", 0)])

                    totales.append([
                        "Total horas",
                        self._formatear_horas_decimal_a_hhmm(t.get("total_horas", 0))
                    ])

                    totales.append([
                        "Promedio horas mensual",
                        self._formatear_horas_decimal_a_hhmm(t.get("promedio_horas_mensual", 0))
                    ])

                    totales.append([
                        "Mes pico",
                        f"{t.get('mes_pico', '')} ({self._formatear_horas_decimal_a_hhmm(t.get('horas_pico', 0))})"
                    ])

                    totales.append([
                        "Mes valle",
                        f"{t.get('mes_valle', '')} ({self._formatear_horas_decimal_a_hhmm(t.get('horas_valle', 0))})"
                    ])

                    totales.append([
                        "Variación tendencia (%)",
                        t.get("variacion_tendencia", 0)
                    ])

                elif self.tipo_informe_actual == "actividades_total":
                    # 🔹 Totales para todas las actividades (USAR DECIMALES)
                    t = getattr(self, "_pdf_totales_actividades", {})

                    total_actividades = t.get("total_actividades", 0)
                    total_registros = t.get("total_registros", 0)
                    total_horas = t.get("total_horas", 0)  # ← FLOAT, NO STRING

                    totales.append(["Total actividades", total_actividades])
                    totales.append(["Total registros", total_registros])
                    totales.append([
                        "Total horas",
                        self._formatear_horas_decimal_a_hhmm(total_horas)
                    ])

                    promedio = total_horas / total_actividades if total_actividades > 0 else 0
                    totales.append([
                        "Promedio horas / actividad",
                        self._formatear_horas_decimal_a_hhmm(promedio)
                    ])

                    # 🔹 Agregar columna "Promedio" a la hoja Datos
                    if "Horas" in df_export.columns:
                        df_export["Promedio"] = df_export["Horas"].apply(
                            lambda x: self._formatear_horas_decimal_a_hhmm(promedio)
                        )

                if not totales:
                    totales.append(["Información", "No disponible"])

                df_totales = pd.DataFrame(totales, columns=["Concepto", "Valor"])
                # ===== ELIMINAR SOLO FILA 'Concepto' =====
                df_totales = df_totales[df_totales["Concepto"].str.lower() != "concepto"]
                # Escribir sin cabecera (header=False)
                df_totales.to_excel(writer, sheet_name="Totales", index=False, header=True)

            # =============================
            # FINAL
            # =============================
            self.ui.show_info("Éxito", f"Informe exportado correctamente:\n{ruta}")
            try:
                os.startfile(ruta)
            except Exception:
                pass

        except Exception as e:
            self.ui.show_error("Error", f"No se pudo exportar a Excel:\n{e}")

    def _exportar_informe_pdf(self, titulo, df):
        """Exporta el informe actual a PDF con nombre sugerido"""

        if df is None or df.empty:
            self.ui.show_info("Exportar", "No hay datos para exportar.")
            return
        
        # Generar nombre de archivo (misma lógica que Excel)
        fecha_actual = datetime.now().strftime("%Y-%m-%d")
        base_name = "Informe"
        
        titulo_lower = titulo.lower()
        if "bombero" in titulo_lower:
            base_name = "Informe_Bomberos"
        elif "actividad" in titulo_lower:
            base_name = "Informe_Actividades"
        elif "horas" in titulo_lower or "período" in titulo_lower:
            base_name = "Informe_Horas"
        
        fecha_match = re.search(r"(\d{2}/\d{2}/\d{4}).*?(\d{2}/\d{2}/\d{4})", titulo)
        if fecha_match:
            desde, hasta = fecha_match.groups()
            fecha_str = f"{desde.replace('/', '-')}_a_{hasta.replace('/', '-')}"
        else:
            fecha_str = fecha_actual
        
        nombre_archivo = f"{base_name}_{fecha_str}.pdf"
        
        ruta = self.ui.ask_save_file(
            defaultextension=".pdf", 
            initialfile=nombre_archivo,
            filetypes=[("PDF files", "*.pdf")],
            title="Guardar informe como PDF"
        )
        
        if not ruta:
            return
        
        try:
            # 👉 USAR SIEMPRE tu PDF con formato completo
            self._crear_pdf_resumen_informe(
                file_path=ruta,
                titulo=titulo,
                df_resumen=df
            )

            self.ui.show_info("Éxito", f"Informe PDF exportado correctamente:\n{ruta}")

            # 🔥 Abrir el PDF automáticamente
            import os
            try:
                os.startfile(ruta)
            except Exception as e:
                pass

        except Exception as e:
            self.ui.show_error("Error", f"No se pudo exportar a PDF:\n{e}")

    # ==========================================================
    #   INFORMES FILTRADOS (al seleccionar opción en combo)
    #   -> usa el mismo diseño visual unificado (rojo bombero)
    # ==========================================================
    def _mostrar_tabla_y_grafico_filtrado(
        self, titulo, df_graf, df_detalle,
        x_col="Actividad", y_col="Horas", color="blue",
        report_type=None, principal_name=None
    ):
        print(f"🚨 📊 _mostrar_tabla_y_grafico_filtrado INICIADA - df_graf: {type(df_graf)}")
        df_resumen = df_graf

        if df_graf is None or len(df_graf) == 0:
            self.ui.show_info("Sin datos", "No hay información para graficar.")
            return

        """Ventana con gráfico + tabla detallada (detección corregida de tipo de informe)."""
        def is_not_na(val):
            """Versión segura de pd.notna() sin depender de pandas."""
            if val is None:
                return False
            val_str = str(val).lower()
            if val_str in ['nan', 'none', 'null', '']:
                return False
            return True

        # Formatear horas en df_detalle si existen
        if df_detalle is not None:
            for col in df_detalle.columns:
                col_lower = str(col).lower()
                if any(keyword in col_lower for keyword in ['hora', 'total', 'duracion']):
                    # Verificar si necesita formateo
                    try:
                        # Obtener muestra sin usar pandas.dropna()
                        sample_vals = []
                        for val in df_detalle[col].head(3):
                            if val is not None and str(val).lower() not in ['nan', 'none', 'null', '']:
                                sample_vals.append(val)
                        
                        if sample_vals:
                            # Si alguna muestra tiene formato decimal (contiene punto)
                            has_decimal = any('.' in str(val) for val in sample_vals)
                            
                            if has_decimal:
                                # Aplicar formateo seguro
                                def formatear_horas_seguro(x):
                                    if x is None:
                                        return "0:00"
                                    x_str = str(x).lower()
                                    if x_str in ['nan', 'none', 'null', '']:
                                        return "0:00"
                                    return self._formatear_horas_decimal_a_hhmm(x)
                                
                                df_detalle[col] = df_detalle[col].apply(formatear_horas_seguro)
                                
                    except Exception as e:
                        continue
            
        # --- Ventana secundaria ---
        ventana = tk.Toplevel(self.master)
        if hasattr(self.master, "_icono_global") and self.master._icono_global:
            ventana.iconphoto(True, self.master._icono_global)
        win.title(titulo)
        win.configure(bg="white")
        win.geometry("1150x720")
        try:
            icon_img = tk.PhotoImage(file="Bomberos.png")
            win.iconphoto(True, icon_img)
        except Exception as e:
            print("No se pudo cargar icono PNG:", e)

        win.focus_set()
        win.grab_set()

        # --- Normalizar ---
        df_graf = df_graf.copy()

        # 🔹 Si existe una columna con nombre completo en el detalle, unirla al graf
        if df_detalle is not None and "Bombero" in df_detalle.columns and "Bombero_ID" in df_graf.columns:
            # Extraer diccionario Legajo → Nombre completo
            mapa = {}
            for v in df_detalle["Bombero"].astype(str):
                leg = v.split("-")[0].strip()
                mapa[leg] = v.replace("\n", " ").replace("  ", " ").strip()

            # Crear la columna completa para el resumen PDF
            df_graf["BomberoCompleto"] = df_graf["Bombero_ID"].astype(str).apply(
                lambda x: mapa.get(x.split("-")[0].strip(), x)
            )

        # --- Normalizar nombres de columnas y helper case-insensitive ---
        cols_lower = {c.lower(): c for c in df_graf.columns}

        def col_exist_ci(name):
            """Devuelve el nombre real de columna en df_graf que coincide case-insensitive con `name`,
            o None si no existe."""
            try:
                if name is None:
                    return None
                return cols_lower.get(str(name).lower())
            except Exception:
                return None

        # Priorizar lo que pidió el caller (x_col, y_col) si existen
        x_final = col_exist_ci(x_col) if 'x_col' in locals() or 'x_col' in globals() else None
        y_final = col_exist_ci(y_col) if 'y_col' in locals() or 'y_col' in globals() else None

        # Si no vino/No existe, elegir fallback ordenado
        if not x_final:
            for cand in ("Actividad", "Bombero", "Legajo", "Área", "Mes"):
                x_final = col_exist_ci(cand)
                if x_final:
                    break

        if not y_final:
            for cand in ("Horas", "horas", "Total_Horas", "Total Horas", "Cantidad"):
                y_final = col_exist_ci(cand)
                if y_final:
                    break

        if not x_final or not y_final:
            # Mensaje claro en vez de NameError: raise controlado
            raise Exception(f"No se encontraron columnas válidas para graficar. x_final={x_final}, y_final={y_final}")

        self.df_graf = df_graf   # <<< FIX
        df_graf.columns = [str(c).strip() for c in df_graf.columns]
        df_detalle.columns = [str(c).strip() for c in df_detalle.columns]

        # 🔥 DETECCIÓN ESPECÍFICA PARA INFORME POR BOMBERO
        es_informe_por_bombero = (
            report_type == "Informe_Bomberos" and 
            principal_name is not None and
            "Detalle de" not in titulo  # Para distinguir entre detalle y resumen
        )

        # ==============================
        #  SELECCIÓN SIMPLE DE COLUMNAS (ÚNICA VERSIÓN)
        # ==============================

        # Normalizar nombres de columnas (lowercase para comparación)
        cols_lower = {c.lower(): c for c in df_graf.columns}

        def buscar_columna(nombre_buscado, alternativas):
            """Busca una columna por nombre o alternativas."""
            if nombre_buscado:
                # Intentar con el nombre exacto que pidieron
                col = cols_lower.get(str(nombre_buscado).lower())
                if col:
                    return col
            
            # Buscar en alternativas
            for alt in alternativas:
                col = cols_lower.get(str(alt).lower())
                if col:
                    return col
            
            # Si no encuentra, usar primera columna disponible
            if df_graf.columns.any():
                return df_graf.columns[0]
            return None

        # Determinar columnas para eje X
        x_alternativas = ["Actividad", "Bombero", "Legajo", "Área", "Mes", "Bombero_ID"]
        x_final = buscar_columna(x_col, x_alternativas)

        # Determinar columnas para eje Y (horas)
        y_alternativas = ["Horas", "horas", "Total_Horas", "Total Horas", "Cantidad"]
        y_final = buscar_columna(y_col, y_alternativas)

        # Validación final
        if not x_final or not y_final:
            self.ui.show_error("Error", "No se pudo determinar qué columnas usar para el gráfico.")
            return

        # Si aún no encontró nada válido → error explicativo
        if not x_final or not y_final:
            raise Exception(f"No se encontraron columnas válidas: x={x_final}, y={y_final}")

        # --- Corrección si columnas numéricas están mal detectadas ---
        if y_final and y_final in df_graf.columns and not pd.api.types.is_numeric_dtype(df_graf[y_final]):
            try:
                df_graf[y_final] = pd.to_numeric(df_graf[y_final], errors="coerce").fillna(0)
            except Exception:
                pass

        if not x_final or not y_final or df_graf.empty:
            self.ui.show_error("Error", "No hay datos válidos para graficar.")
            win.destroy()
            return

        # --- Apellido + inicial si X = Bombero ---
        if x_final.lower() == "bombero":
            df_graf[x_final] = df_graf[x_final].apply(
                lambda n: f"{str(n).split()[0]} {str(n).split()[1][0]}."
                if isinstance(n, str) and len(str(n).split()) > 1
                else str(n).split()[0] if isinstance(n, str) else n
            )

        # --- Asegurar numéricos ---
        df_graf[y_final] = pd.to_numeric(df_graf[y_final], errors="coerce").fillna(0)

        # 🔥 LIMPIAR NOMBRES DE ACTIVIDAD ANTES DE AGRUPAR (FIX TOTAL)
        import re

        if x_final and x_final.lower() == "actividad":
            df_graf[x_final] = df_graf[x_final].astype(str).str.strip()
            df_graf[x_final] = df_graf[x_final].apply(
                lambda x: re.sub(r'^\d+\s*-\s*', '', x)
            )

        # --- Agrupar para no duplicar barras ---
        import textwrap

        def wrap_words(text, width=12):
            """Envuelve texto sin cortar palabras."""
            return textwrap.fill(
                str(text),
                width=width,
                break_long_words=False,
                break_on_hyphens=False
            )

        try:
            plot_df = df_graf.groupby(x_final, as_index=False)[y_final].sum()

            # 🔥 Mantener nombres completos para PDF/Excel si existen
            if "ActividadCompleta" in df_graf.columns:
                # Tomar el primer valor completo de cada grupo
                df_aux = (
                    df_graf.groupby(x_final, as_index=False)["ActividadCompleta"]
                    .first()
                )
                # Unir a plot_df
                plot_df = plot_df.merge(df_aux, on=x_final, how="left")

            # 🔥 Nombre REAL para tabla y PDF (NO tocar)
            df_graf[x_final] = df_graf[x_final].astype(str)
            df_detalle[x_final] = df_detalle[x_final].astype(str)

            # Copia SOLO PARA EJE X DEL GRÁFICO
            import re
            def limpiar_nombre_actividad(t):
                t = str(t).strip()
                # Quitar "N - " al inicio
                t = re.sub(r'^\d+\s*-\s*', '', t)
                return t

            plot_df[x_final] = plot_df[x_final].apply(limpiar_nombre_actividad)
            plot_df[x_final] = plot_df[x_final].apply(lambda t: wrap_words(t, 12))

            # 🔥 Copia SOLO PARA EJE X DEL GRÁFICO
            plot_df["x_label_original"] = plot_df[x_final].astype(str)

            plot_df[x_final] = plot_df[x_final].apply(lambda t: wrap_words(t, 12))
        except Exception:
            plot_df = df_graf[[x_final, y_final]].copy()
            plot_df[x_final] = plot_df[x_final].apply(lambda t: wrap_words(t, 12))

        # ------------------------------------------
        # ETIQUETA FINAL (promedio + totales)
        # ------------------------------------------
        # Corrección: nombre visible del eje X
        mapa_etiquetas = {
            "bombero_id": "Bombero",
            "bombero": "Bombero",
            "actividad_id": "Actividad",
            "actividad": "Actividad",
            "mes": "Mes",
        }

        # --- DECISIÓN DE COLOR: si es detalle por ACTIVIDAD y existe columna BOMBERO, forzamos VERDE ---
        usar_color_verde = False
        if ("Actividad" in df_detalle.columns) and ("Bombero" in df_detalle.columns):
            usar_color_verde = True

        bar_color = "green" if usar_color_verde else color  # <-- forzamos verde en detalle por actividad

        # --- Gráfico ---
        fig, ax = plt.subplots(figsize=(10, 4.2))  # Ancho aumentado

        # FUNCIÓN PARA ACORTAR TEXTOS LARGOS EN EL EJE X
        def acortar_texto(texto, max_longitud=20):
            texto_str = str(texto)
            if len(texto_str) <= max_longitud:
                return texto_str
            # Acortar y agregar "..."
            return texto_str[:max_longitud-3] + "..."

        # Aplicar acortamiento a los textos del eje X (ya vienen limpios)
        plot_df[x_final] = plot_df[x_final].apply(acortar_texto)

        bars = ax.bar(plot_df[x_final], plot_df[y_final], color=bar_color)

        # CONFIGURAR FUENTE MÁS PEQUEÑA EN EL EJE X
        plt.xticks(rotation=45, ha="right", fontsize=5)

        max_val = plot_df[y_final].max() if not plot_df.empty else 0
        threshold = max_val * 0.1 if max_val > 0 else 0

        import matplotlib.ticker as ticker

        def decimal_a_hhmm_formatter(x, pos):
            """Formatter para el eje Y: convierte decimal a HH:MM"""
            try:
                horas = int(x)
                minutos = int(round((x - horas) * 60))
                if minutos >= 60:
                    horas += 1
                    minutos -= 60
                return f"{horas}:{minutos:02d}"
            except:
                return f"{x:.1f}"

        # Aplicar el formatter al eje Y
        ax.yaxis.set_major_formatter(ticker.FuncFormatter(decimal_a_hhmm_formatter))

        for bar in bars:
            h = bar.get_height()
            # CONVERTIR a HH:MM para la etiqueta
            h_display = self._formatear_horas_decimal_a_hhmm(h)
            
            if h > threshold:
                ax.text(bar.get_x() + bar.get_width() / 2, h / 2, f"{h_display}",
                        ha="center", va="center", color="white", fontsize=8)
            else:
                ax.text(bar.get_x() + bar.get_width() / 2, h + (max_val * 0.01),
                        f"{h_display}", ha="center", va="bottom", color="black", fontsize=8)

        ax.set_title(titulo, fontsize=11, pad=12)
        ax.set_ylabel(y_final)
        # 🔹 Quitar etiqueta del eje X si es "Actividad" o "Bombero" (innecesario)
        if x_final.lower() in ("actividad", "bombero"):
            ax.set_xlabel("")
        else:
            etiqueta_visible = mapa_etiquetas.get(x_final.lower(), x_final.capitalize())
            ax.set_xlabel(etiqueta_visible, labelpad=10)

        if len(plot_df[x_final]) > 10:
            plt.xticks(rotation=60, ha="right", fontsize=6)
            fig.subplots_adjust(bottom=0.28)
            ax.tick_params(axis='x', labelsize=7)

        plt.tight_layout(rect=[0, 0, 1, 0.92])

        plt.xticks(rotation=60, ha="right", fontsize=6)

        # --- Rotar etiquetas si se enciman ---
        plt.xticks(rotation=45, ha="right")

        # --- Ajustar márgenes automáticos ---
        plt.tight_layout()

        canvas = FigureCanvasTkAgg(fig, master=win)
        canvas.draw()
        canvas.get_tk_widget().place(x=20, y=20, width=700, height=360)

        # --- TABLA ---
        frame_tabla = tk.Frame(win, bg="white")
        frame_tabla.place(x=20, y=400, width=1090, height=280)

        cols = list(df_detalle.columns)

        # ===== ESTILO (IGUAL A OTROS INFORMES) =====
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except:
            pass

        style.configure(
            "Informe.Treeview",
            font=("Arial", 8),
            rowheight=26,
            background="white",
            fieldbackground="white",
            foreground="black"
        )

        style.configure(
            "Informe.Treeview.Heading",
            font=("Arial", 9, "bold"),
            background="#e6e6e6",
            foreground="black"
        )

        style.map(
            "Informe.Treeview",
            background=[("selected", "#cce6ff")],
            foreground=[("selected", "black")]
        )

        # ===== SCROLLBARS (PRIMERO) =====
        scroll_y = ttk.Scrollbar(frame_tabla, orient="vertical")
        scroll_y.pack(side="right", fill="y")

        scroll_x = ttk.Scrollbar(frame_tabla, orient="horizontal")
        scroll_x.pack(side="bottom", fill="x")

        # ===== TREEVIEW (DESPUÉS) =====
        tree = ttk.Treeview(
            frame_tabla,
            columns=cols,
            show="headings",
            height=12,
            style="Informe.Treeview",
            yscrollcommand=scroll_y.set,
            xscrollcommand=scroll_x.set
        )

        tree.pack(fill="both", expand=True)

        scroll_y.config(command=tree.yview)
        scroll_x.config(command=tree.xview)

        # ===== HEADERS Y ANCHOS =====
        for c in cols:
            col_lower = c.lower()

            if col_lower == "actividad":
                tree.heading(c, text="Actividad")
                ancho = 300

            elif col_lower in ("total_bomberos", "bomberos"):
                tree.heading(c, text="Bomberos")
                ancho = 120

            elif col_lower in ("total_registros", "registros"):
                tree.heading(c, text="Registros")
                ancho = 120

            elif col_lower in ("horas_decimal", "horas", "total_horas"):
                tree.heading(c, text="Total horas")
                ancho = 120

            elif report_type == "Informe_Bomberos" and "actividad" in col_lower:
                tree.heading(c, text="Actividades")
                ancho = 400

            elif "bombero" in col_lower:
                tree.heading(c, text=c)
                ancho = 250

            else:
                tree.heading(c, text=c)
                max_len = max(len(str(c)), *(len(str(x)) for x in df_detalle[c].astype(str).values))
                ancho = min(300, max(100, max_len * 8))

            tree.column(c, width=ancho, anchor="center", stretch=True)

        # ===== FILAS =====
        for _, row in df_detalle.iterrows():
            valores = list(row)
            for i, col in enumerate(df_detalle.columns):
                if "hora" in col.lower():
                    try:
                        valores[i] = self._formatear_horas_decimal_a_hhmm(valores[i])
                    except:
                        pass
            tree.insert("", "end", values=valores)

        # ==========================================
        #   TOTales INTELIGENTES (FIX DEFINITIVO)
        # ETIQUETA FINAL (promedio + totales) - FORMATEADO
        # ------------------------------------------
        if not plot_df.empty and y_final in plot_df.columns:
            total_y_decimal = plot_df[y_final].sum()
        else:
            total_y_decimal = 0
            
        total_registros = len(df_detalle) if df_detalle is not None else 0

        if not plot_df.empty and y_final in plot_df.columns:
            promedio_decimal = plot_df[y_final].mean()
        else:
            promedio_decimal = 0

        # 🔥 CONVERTIR a HH:MM para mostrar
        total_y_display = self._formatear_horas_decimal_a_hhmm(total_y_decimal)
        promedio_display = self._formatear_horas_decimal_a_hhmm(promedio_decimal)

        # Normalizar nombres de columnas para chequeos seguros (case-insensitive)
        cols_l = {c.lower(): c for c in df_detalle.columns} if df_detalle is not None else {}

        has_act = "actividad" in cols_l
        has_bom = "bombero" in cols_l
        has_area = "área" in cols_l or "area" in cols_l

        col_act = cols_l.get("actividad")
        col_bom = cols_l.get("bombero")
        col_area = cols_l.get("área") or cols_l.get("area")

        # ANTES de calcular el resumen, verifica si es informe por un solo bombero
        es_detalle_bombero = (
            report_type == "Informe_Bomberos" and 
            principal_name is not None and
            has_bom and 
            df_detalle[col_bom].nunique() == 1
        )

        # ==========================================
        #   TOTALES Y RESUMEN (FIX DEFINITIVO)
        # ==========================================

        # 🔥 Total REAL de actividades distintas (desde df_detalle limpio)
        if has_act and col_act:
            total_actividades = df_detalle[col_act].astype(str)\
                .str.replace(r'^\d+\s*-\s*', '', regex=True)\
                .str.strip()\
                .nunique()
        else:
            total_actividades = len(plot_df)

        if es_detalle_bombero:
            # INFORME POR UN SOLO BOMBERO
            resumen = (
                f"Total de registros: {total_registros}\n"
                f"Total de actividades: {total_actividades}\n"
                f"Total de horas: {total_y_display}\n"
                f"Promedio de horas por actividad: {promedio_display}"
            )

        else:
            # INFORME GENERAL

            if has_act:
                total_extra_texto = f"Total actividades distintas: {total_actividades}"
            elif has_bom:
                total_extra_texto = f"Total bomberos distintos: {df_detalle[col_bom].nunique()}"
            else:
                total_extra_texto = "Totales no disponibles"

            resumen = (
                f"Total registros: {total_registros}\n"
                f"Total horas: {total_y_display}\n"
                f"Promedio por {x_final}: {promedio_decimal:.2f}\n"
                f"{total_extra_texto}"
            )

        # ==========================================
        # MOSTRAR ETIQUETAS EN LA INTERFAZ
        # ==========================================
        tk.Label(
            win, text=resumen, bg="white", fg="blue",
            font=("Arial", 10, "bold"), justify="left"
        ).place(x=740, y=40)

        # --- NUEVO: etiqueta dedicada y visible para 'Cantidad total bomberos' cuando corresponda ---
        try:
            if not es_detalle_bombero and has_act and has_bom and df_detalle is not None:
                cant_bom = int(df_detalle[col_bom].nunique())
                tk.Label(
                    win,
                    text=f"Cantidad total bomberos: {cant_bom}",
                    bg="white",
                    fg="green",
                    font=("Arial", 12, "bold"),
                    justify="left"
                ).place(x=740, y=100)
        except Exception as e:
            pass

        # --- Helper periodo limpio ---
        def _periodo_from_title_or_fields():
            try:
                periodo = titulo.split("(")[-1].split(")")[0]
            except Exception:
                periodo = ""
            if not periodo:
                try:
                    d1 = getattr(self, "inf_desde", None)
                    d2 = getattr(self, "inf_hasta", None)
                    if d1 and d2:
                        v1 = d1.get() if callable(getattr(d1, "get", None)) else str(d1)
                        v2 = d2.get() if callable(getattr(d2, "get", None)) else str(d2)
                        periodo = f"{v1}_{v2}"
                except Exception:
                    periodo = datetime.now().strftime("%Y%m%d")
            periodo = re.sub(r'[\/:*?"<>|()]+', "", periodo).replace(" ", "_")
            return periodo or datetime.now().strftime("%Y%m%d")

        periodo = _periodo_from_title_or_fields()
        tipo_informe = x_final.capitalize() if x_final else "Informe"

        # --- Exportar ---

        def exportar_excel_desde_grafico():
            """Exporta DataFrames desde la ventana de gráfico a Excel."""
           
            try:
                # 🔹 1. Crear una copia limpia SIN columnas no deseadas
                df_detalle_limpio = df_detalle.copy()
                
                # 🔹 2. Eliminar columnas problemáticas (incluyendo horas_formateadas)
                columnas_a_eliminar = []
                for col in df_detalle_limpio.columns:
                    col_lower = str(col).lower()
                    # Buscar columnas que NO queremos en el Excel
                    if any(keyword in col_lower for keyword in ['formateada', '_fmt', 'temporal', 'completa']):
                        columnas_a_eliminar.append(col)
                
                if columnas_a_eliminar:
                    df_detalle_limpio = df_detalle_limpio.drop(columns=columnas_a_eliminar, errors='ignore')
                
                # 🔹 3. Función para convertir decimal a HH:MM
                def decimal_a_hhmm(valor):
                    """Convierte horas decimales a formato HH:MM."""
                    try:
                        # Si ya está en formato HH:MM, dejarlo igual
                        if isinstance(valor, str) and ':' in valor:
                            return valor
                        
                        # Si es None o NaN
                        if pd.isna(valor):
                            return "0:00"
                        
                        # Convertir a float
                        horas_float = float(valor)
                        
                        # Obtener horas enteras y minutos
                        horas_enteras = int(horas_float)
                        minutos = int(round((horas_float - horas_enteras) * 60))
                        
                        # Ajustar si minutos son 60
                        if minutos >= 60:
                            horas_enteras += 1
                            minutos -= 60
                        
                        return f"{horas_enteras}:{minutos:02d}"
                    except:
                        return str(valor) if not pd.isna(valor) else "0:00"
                
                # 🔹 4. Convertir columnas de horas principales a HH:MM
                columnas_horas = []
                for col in df_detalle_limpio.columns:
                    col_lower = str(col).lower()
                    # Buscar solo columnas principales de horas
                    if col_lower in ['horas', 'total_horas', 'horas_trabajadas', 'duracion']:
                        columnas_horas.append(col)
                
                for col_horas in columnas_horas:
                    if col_horas in df_detalle_limpio.columns:
                        df_detalle_limpio[col_horas] = df_detalle_limpio[col_horas].apply(decimal_a_hhmm)
                
                # 🔹 5. Cargar nombres reales de legajos para columna "asignado"
                conn = sqlite3.connect(DB_PATH)
                c = conn.cursor()
                c.execute("SELECT legajo, apellido, nombre FROM legajos")
                rows = c.fetchall()
                conn.close()
                
                legajos_a_nombres = {
                    str(r[0]).strip(): f"{r[1]} {r[2]}".strip()
                    for r in rows
                }
                
                # 🔹 6. Procesar columna "asignado" si existe
                if "asignado" in df_detalle_limpio.columns:
                    nueva_columna = []
                    for val in df_detalle_limpio["asignado"].fillna(""):
                        s = str(val).strip()
                        
                        if s == "" or s.lower() == "nan":
                            nueva_columna.append("")
                            continue
                        
                        # Extraer legajo del formato "245 - ADARO"
                        try:
                            leg = int(s.split("-")[0].strip())
                            nombre_real = legajos_a_nombres.get(str(leg), "")
                            
                            if nombre_real:
                                # Formato: "245 - APELLIDO NOMBRE"
                                nuevo_valor = f"{leg} - {nombre_real}"
                            else:
                                nuevo_valor = s
                        except:
                            nuevo_valor = s
                        
                        nueva_columna.append(nuevo_valor)
                    
                    df_detalle_limpio["asignado"] = nueva_columna
                
                # 🔹 7. Preparar nombre del archivo
                periodo = _periodo_from_title_or_fields()
                
                # Determinar tipo de informe
                tipo = "Informe"
                if report_type:
                    rt = str(report_type).lower()
                    if "bombero" in rt:
                        tipo = "Informe_Bomberos"
                    elif "area" in rt or "área" in rt:
                        tipo = "Informe_Areas"
                    elif "actividad" in rt:
                        tipo = "Informe_Actividades"
                
                # Obtener nombre principal
                nombre_principal = ""
                if principal_name:
                    nombre_principal = str(principal_name).strip()
                else:
                    try:
                        if "detalle de" in titulo.lower():
                            nombre_principal = titulo.split("Detalle de", 1)[1].split("(")[0].strip()
                    except:
                        pass
                
                # Limpiar nombre principal
                nombre_limpio = ""
                if nombre_principal:
                    tmp = unicodedata.normalize("NFKD", nombre_principal)
                    tmp = "".join(c for c in tmp if not unicodedata.combining(c))
                    tmp = re.sub(r"[^A-Za-z0-9_ -]", "", tmp).replace(" ", "_")[:40]
                    nombre_limpio = tmp
                
                # Construir nombre final
                fecha_actual = datetime.now().strftime("%Y%m%d_%H%M")
                partes = [tipo]
                
                if nombre_limpio:
                    partes.append(nombre_limpio)
                elif nombre_principal:
                    nombre_corto = nombre_principal[:20].replace(" ", "_")
                    partes.append(nombre_corto)
                
                if periodo:
                    partes.append(periodo)
                
                partes.append(fecha_actual)
                nombre_final = "_".join(p for p in partes if p) + ".xlsx"
                
                # 🔹 8. Pedir ruta de guardado
                path = self.ui.ask_save_file(
                    defaultextension=".xlsx",
                    initialfile=nombre_final,
                    filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                    title="Guardar informe Excel"
                )
                
                if not path:
                    return
                
                # 🔹 9. Exportar a Excel
                with pd.ExcelWriter(path, engine='openpyxl') as writer:
                    # Hoja 1: Detalle (con datos limpios)
                    df_detalle_limpio.to_excel(writer, index=False, sheet_name="Detalle")
                    
                    # Hoja 2: Resumen del gráfico (si existe)
                    if plot_df is not None and not plot_df.empty:
                        plot_df_limpio = plot_df.copy()
                        
                        # También limpiar plot_df si es necesario
                        for col in plot_df_limpio.columns:
                            col_lower = str(col).lower()
                            if 'hora' in col_lower and col in plot_df_limpio.columns:
                                plot_df_limpio[col] = plot_df_limpio[col].apply(decimal_a_hhmm)
                        
                        plot_df_limpio.to_excel(writer, index=False, sheet_name="Resumen")
                
                self.ui.show_info("Éxito", f"Archivo Excel exportado correctamente:\n{path}")
                
                # 🔹 10. Intentar abrir automáticamente
                try:
                    import os
                    os.startfile(path)
                except:
                    pass
                    
            except Exception as e:
                self.ui.show_error("Error", f"No se pudo exportar a Excel:\n{str(e)}")
                print(f"ERROR en exportar_excel_desde_grafico: {e}")

    # ----------------------------------------------------------
    def on_detalle(self, actividad, fecha_desde, fecha_hasta):
        """
        Muestra el detalle de una actividad o bombero seleccionado desde combo.
        Reutiliza la nueva ventana unificada (rojo bombero).
        """
        conn = sqlite3.connect(DB_PATH)
        try:
            query = """
                SELECT a.id, a.legajo, 
                       COALESCE(l.apellido || ', ' || l.nombre, '(Sin nombre)') AS Bombero,
                       a.actividad, a.area, a.fecha_inicio, a.fecha_fin,
                       COALESCE(a.horas, 0) AS Horas,
                       COALESCE(a.descripcion, '') AS Descripción
                FROM actividades a
                LEFT JOIN legajos l ON a.legajo = l.legajo
                WHERE a.actividad = ? 
                  AND (CASE WHEN instr(a.fecha_inicio, '/')>0
                            THEN date(substr(a.fecha_inicio,7,4)||'-'||substr(a.fecha_inicio,4,2)||'-'||substr(a.fecha_inicio,1,2))
                            ELSE date(a.fecha_inicio) END)
                      BETWEEN date(?) AND date(?)
                ORDER BY a.fecha_inicio DESC
            """
            df = pd.read_sql_query(query, conn, params=(actividad, fecha_desde, fecha_hasta))
            if df.empty:
                self.ui.show_info("Sin datos", f"No hay registros para {actividad}")
                return

            # Convertir formato horas
            if "Horas" in df.columns:
                try:
                    df["Horas"] = df["Horas"].apply(self._fmt_horas)
                except:
                    pass

            # Gráfico de horas por bombero
            df_graf = df.copy()
            if "Horas" in df_graf.columns:
                df_graf["HorasNum"] = df_graf["Horas"].apply(lambda h: self._horas_a_decimal(str(h)))
                df_graf = df_graf.groupby("Bombero", as_index=False)["HorasNum"].sum().rename(columns={"HorasNum": "Horas"})
            else:
                df_graf = pd.DataFrame(columns=["Bombero", "Horas"])

            titulo = f"Detalle de {actividad} ({fecha_desde:%d/%m/%Y} - {fecha_hasta:%d/%m/%Y})"
            self._mostrar_tabla_y_grafico_filtrado(
                titulo,
                df_graf,
                df,
                x_col="Mes",
                y_col="Horas",
                color="green",
                report_type="horas",
                principal_name="Horas_por_Mes"
            )

        finally:
            conn.close()

    def _exportar_pdf_con_logo(self, df, titulo, path=None):
        """
        Exporta df a PDF con el formato similar a los 3 informes:
        - Cabecera roja con logo a la izquierda y título en blanco
        - Tabla y totales al final
        Si se recibe 'path', lo usa; si no, abre dialog (pero en _mostrar... ya le pasamos path).
        """
        try:
            if df.empty:
                self.ui.show_info("Sin datos", "No hay datos para exportar.")
                return

            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.units import mm
            import os

            if not path:
                path = self.ui.ask_save_file(defaultextension=".pdf",
                                                    filetypes=[("PDF", "*.pdf")])
                if not path:
                    return

            doc = SimpleDocTemplate(path, pagesize=A4)
            styles = getSampleStyleSheet()
            elems = []

            # Cabecera: tabla de 2 columnas: logo | título (con fondo rojo)
            header_data = []
            logo_exists = os.path.exists(LOGO_PATH)
            if logo_exists:
                # logo en la izquierda (usamos imagen reducida)
                header_data.append([RLImage(LOGO_PATH, width=25*mm, height=25*mm),
                                    Paragraph(f"<b>{titulo}</b>", ParagraphStyle(name="h", fontSize=14, alignment=1, textColor=colors.white))])
            else:
                header_data.append([Paragraph("", styles["Normal"]),
                                    Paragraph(f"<b>{titulo}</b>", ParagraphStyle(name="h", fontSize=14, alignment=1, textColor=colors.white))])

            header = Table(header_data, colWidths=[30*mm, 150*mm])
            header.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,-1), colors.red),
                ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
                ("ALIGN", (1,0), (1,0), "CENTER"),
            ]))
            elems.append(header)
            elems.append(Spacer(1, 10))

             # --- Corrección encabezados ---
            cols = list(df.columns)
            if len(cols) >= 2:
                cols[1] = "Bombero"   # segunda columna muestra nombres

            # --- Datos tabla ---
            data = [cols] + df.values.tolist()

            # --- Anchos de columna (modificables) ---
            colWidths = [40, 200, 80, 70, 60, 160]

            table = Table(data, repeatRows=1, colWidths=colWidths)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
                ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
                ("ALIGN", (0,0), (-1,-1), "CENTER"),
                ("FONT", (0,0), (-1,0), "Helvetica-Bold"),
                ("WORDWRAP", (0,0), (-1,-1), None),    # evita cortar nombres
                ("VALIGN", (0,0), (-1,-1), "MIDDLE"),  # centra vertical
            ]))

            elems.append(table)
            elems.append(Spacer(1, 8))

            # Totales
            # detectar columna horas robustamente
            hours_cols = [c for c in df.columns if c.lower().startswith("hora")]
            total_horas = 0.0
            if hours_cols:
                # sumar la primera columna que parezca horas
                try:
                    total_horas = df[hours_cols[0]].astype(float).sum()
                except Exception:
                    # intentar convertir reemplazando comas
                    total_horas = df[hours_cols[0]].apply(lambda x: float(str(x).replace(",",".")) if str(x).strip()!="" else 0).sum()
            elems.append(Paragraph(f"Total registros: {len(df)}", styles["Normal"]))
            elems.append(Paragraph(f"Total horas: {total_horas:.2f}", styles["Normal"]))
            # actividades distintas (si aplica)
            if "actividad" in [c.lower() for c in df.columns]:
                col_act = [c for c in df.columns if c.lower()=="actividad"][0]
                elems.append(Paragraph(f"Actividades distintas: {df[col_act].nunique()}", styles["Normal"]))

            doc.build(elems)
            # messagebox mostrado por caller
        except Exception as e:
            self.ui.show_error("Error", f"No se pudo exportar a PDF:\n{e}")

    # ---------------------------------------------------------
    def _obtener_datos_por_actividad(self, actividad, fecha_desde=None, fecha_hasta=None):
        import pandas as pd, sqlite3, unicodedata
        from datetime import datetime

        conn = sqlite3.connect(DB_PATH)
        try:
            query = """
                SELECT 
                    a.id AS id_actividad,
                    a.actividad,
                    b.legajo,
                    b.apellido || ', ' || b.nombre AS Bombero,
                    a.area,
                    a.fecha_inicio AS Fecha,
                    COALESCE(a.horas, 0) AS horas,
                    COALESCE(a.descripcion, '') AS Descripcion
                FROM actividades a
                LEFT JOIN legajos b ON a.legajo = b.legajo
                WHERE 1=1
            """
            params = []

            if fecha_desde and fecha_hasta:
                try:
                    f1 = datetime.strptime(fecha_desde, "%d/%m/%Y").strftime("%Y-%m-%d")
                    f2 = datetime.strptime(fecha_hasta, "%d/%m/%Y").strftime("%Y-%m-%d")
                    query += """
                        AND (
                            CASE WHEN instr(a.fecha_inicio,'/')>0
                            THEN date(substr(a.fecha_inicio,7,4)||'-'||substr(a.fecha_inicio,4,2)||'-'||substr(a.fecha_inicio,1,2))
                            ELSE date(a.fecha_inicio) END
                        ) BETWEEN date(?) AND date(?)
                    """
                    params.extend([f1, f2])
                except Exception:
                    pass

            query += " ORDER BY a.fecha_inicio ASC"
            df = pd.read_sql_query(query, conn, params=params)

        except Exception as e:
            df = pd.DataFrame()
        finally:
            conn.close()

        if df.empty:
            return pd.DataFrame(), pd.DataFrame()

        # --- Filtro en memoria robusto por actividad ---
        def _norm(s):
            if s is None:
                return ""
            s = str(s).strip().upper()
            s = unicodedata.normalize("NFKD", s)
            s = "".join(ch for ch in s if not unicodedata.combining(ch))
            s = " ".join(s.split())
            return s

        actividad_norm = _norm(actividad)
        df["actividad_norm"] = df["actividad"].apply(_norm)

        before = len(df)
        df = df[df["actividad_norm"] == actividad_norm].copy()
        after = len(df)

        if df.empty:
            return pd.DataFrame(), pd.DataFrame()

        # --- Procesamiento de fechas y horas ---
        try:
            df["Fecha_dt"] = pd.to_datetime(df["Fecha"], errors="coerce", dayfirst=True)
            df["Mes"] = df["Fecha_dt"].dt.strftime("%b %Y").fillna("(sin fecha)")
        except Exception as e:
            df["Mes"] = "(sin fecha)"

        df["HorasNum"] = pd.to_numeric(df["horas"], errors="coerce").fillna(0)
        df_graf = df.groupby("Mes", as_index=False)["HorasNum"].sum().rename(columns={"HorasNum": "Horas"})

        return df_graf, df

    def _sql_actividades_base(self):
        """
        Consulta base de la tabla actividades.
        Siempre devuelve las columnas en el orden correcto
        para _cargar_actividad_row.
        """
        return """
            SELECT
                a.id,
                a.legajo,
                l.apellido,
                l.nombre,
                a.actividad,
                a.area,
                a.fecha_inicio,
                a.fecha_fin,
                a.hora_inicio,
                a.hora_fin,
                a.horas,
                a.descripcion,
                a.asignado,
                a.concepto_id,
                a.firma_bombero_usuario,
                a.firma_bombero_fecha,
                a.firma_supervisor_usuario,
                a.firma_supervisor_fecha,
                a.anulada
            FROM actividades a
            LEFT JOIN legajos l ON l.legajo = a.legajo
        """

    def _obtener_datos_horas_periodo(self, fecha_desde, fecha_hasta):
        import pandas as pd, sqlite3
        from datetime import datetime

        try:
            fd = datetime.strptime(fecha_desde, "%d/%m/%Y").strftime("%Y-%m-%d")
            fh = datetime.strptime(fecha_hasta, "%d/%m/%Y").strftime("%Y-%m-%d")
        except:
            return pd.DataFrame(), pd.DataFrame()

        conn = sqlite3.connect(DB_PATH)
        try:
            sql_det = """
                SELECT a.id,
                    a.legajo || ' - ' || IFNULL(b1.apellido, '') AS legajo,
                    a.actividad,
                    a.area,
                    a.fecha_inicio,
                    a.fecha_fin,
                    a.hora_inicio,
                    a.hora_fin,
                    a.horas,
                    a.descripcion,
                    a.asignado || ' - ' || IFNULL(b2.apellido, '') AS asignado
                FROM actividades a
                LEFT JOIN legajos b1 ON b1.legajo = a.legajo OR b1.legajo = CAST(a.legajo AS TEXT)
                LEFT JOIN legajos b2 ON b2.legajo = a.asignado OR b2.legajo = CAST(a.asignado AS TEXT)
                WHERE (CASE WHEN instr(a.fecha_inicio,'/')>0
                            THEN date(substr(a.fecha_inicio,7,4)||'-'||substr(a.fecha_inicio,4,2)||'-'||substr(a.fecha_inicio,1,2))
                            ELSE date(a.fecha_inicio) END)
                    BETWEEN date(?) AND date(?)
                ORDER BY a.fecha_inicio DESC
            """
            df_tabla = pd.read_sql_query(sql_det, conn, params=[fd, fh])

            sql_graf = """
                SELECT strftime('%Y-%m',
                                CASE WHEN instr(a.fecha_inicio,'/')>0
                                    THEN substr(a.fecha_inicio,7,4)||'-'||substr(a.fecha_inicio,4,2)||'-'||substr(a.fecha_inicio,1,2)
                                    ELSE a.fecha_inicio END
                            ) AS Mes,
                    ROUND(SUM(COALESCE(a.horas,0)),2) AS Total_Horas
                FROM actividades a
                WHERE (CASE WHEN instr(a.fecha_inicio,'/')>0
                            THEN date(substr(a.fecha_inicio,7,4)||'-'||substr(a.fecha_inicio,4,2)||'-'||substr(a.fecha_inicio,1,2))
                            ELSE date(a.fecha_inicio) END)
                    BETWEEN date(?) AND date(?)
                GROUP BY Mes
                ORDER BY Mes
            """
            df_graf = pd.read_sql_query(sql_graf, conn, params=[fd, fh])
            df_graf.columns = ["Mes", "Horas Trabajadas"]

            # Agregar totales al final de df_tabla
            try:
                total_row = {}
                for col in df_tabla.columns:
                    if pd.api.types.is_numeric_dtype(df_tabla[col]):
                        total_row[col] = df_tabla[col].sum()
                    else:
                        total_row[col] = ""
                if any(total_row.values()):
                    total_row[df_tabla.columns[0]] = "TOTAL GENERAL"
                    df_tabla = pd.concat([df_tabla, pd.DataFrame([total_row])], ignore_index=True)
            except Exception as e:
                print("Error al calcular totales (horas periodo):", e)

            return df_graf, df_tabla
        finally:
            try:
                if "horas" in df_tabla.columns:
                    df_tabla["horas"] = df_tabla["horas"].apply(self._fmt_horas)
                if "Total_Horas" in df_graf.columns:
                    df_graf["Total_Horas"] = df_graf["Total_Horas"].apply(self._fmt_horas)
            except Exception:
                pass
            conn.close()

    # ============================================================
    # 🔹 Informe automático al seleccionar un LEGAJO
    # ============================================================
    def on_seleccionar_legajo_informe(self, event=None):
        """Genera informe detallado del legajo seleccionado en combo."""
        print(">>> EVENTO: on_seleccionar_legajo_informe ejecutado")
        print(">> Legajo seleccionado:", self.inf_legajo_cb.get())
        leg_sel = (self.inf_legajo_cb.get() or "").strip()
        act_sel = (self.inf_actividad.get() or "").strip()

        # ⚠️ Evitar conflicto con combo de actividad
        if act_sel and act_sel != "(TODAS)":
            messagebox.showwarning("Atención", "Debe deseleccionar la actividad antes de elegir un legajo.")
            self.inf_legajo_cb.set("")  # limpia selección de legajo
            return

        if not leg_sel or leg_sel == "()":
            return

        try:
            f1 = self.inf_desde.get()
            f2 = self.inf_hasta.get()
            # Convertir a formato ISO solo para la consulta SQL
            f1_sql = datetime.strptime(f1, "%d/%m/%Y").strftime("%Y-%m-%d")
            f2_sql = datetime.strptime(f2, "%d/%m/%Y").strftime("%Y-%m-%d")
        except Exception:
            self.ui.show_error("Error", "Debe seleccionar fechas válidas.")
            return

        conn = sqlite3.connect(DB_PATH)
        try:
            legajo = leg_sel.split(" - ")[0].strip()
            q = """
                SELECT a.fecha_inicio AS Fecha,
                    a.actividad AS Actividad,
                    a.area AS Área,
                    COALESCE(a.horas, 0) AS horas,
                    COALESCE(a.descripcion, '') AS Descripción
                FROM actividades a
                WHERE a.legajo = ?
                AND (
                    CASE 
                    WHEN instr(a.fecha_inicio,'/')>0
                    THEN date(substr(a.fecha_inicio,7,4)||'-'||substr(a.fecha_inicio,4,2)||'-'||substr(a.fecha_inicio,1,2))
                    ELSE date(a.fecha_inicio)
                    END
                ) BETWEEN date(?) AND date(?)
                ORDER BY a.fecha_inicio DESC
            """
            df = pd.read_sql_query(q, conn, params=(legajo, f1_sql, f2_sql))
            if df.empty:
                self.ui.show_info("Sin datos", "No hay registros para el legajo seleccionado en ese período.")
                return

            titulo = f"Actividades de {leg_sel} ({f1} - {f2})"
            # Convertimos horas a número para gráfico
            # después de df = pd.read_sql_query(...)
            # Crear df_graf seguro para el gráfico mensual (Mes/Horas)
            df_graf = df.copy()
            # intentar detectar columna de horas con varios nombres
            horas_col = None
            for c in ("horas", "Horas", "HorasNum"):
                if c in df_graf.columns:
                    horas_col = c
                    break

            # Si hay Fecha -> construir columna Mes (YYYY-MM) y agregar horas por mes
            if "Fecha" in df_graf.columns:
                try:
                    df_graf["Fecha_dt"] = pd.to_datetime(df_graf["Fecha"], dayfirst=True, errors="coerce")
                    df_graf["Mes"] = df_graf["Fecha_dt"].dt.strftime("%Y-%m")
                except Exception:
                    df_graf["Mes"] = ""
            else:
                df_graf["Mes"] = ""

            # Normalizar columna de horas numéricas
            if horas_col:
                df_graf["Horas_num"] = pd.to_numeric(df_graf[horas_col], errors="coerce").fillna(0)
            else:
                df_graf["Horas_num"] = 0

            # Agrupar por Mes para el gráfico (si corresponde)
            if "Mes" in df_graf.columns and df_graf["Mes"].notna().any():
                df_graf_plot = df_graf.groupby("Mes", as_index=False)["Horas_num"].sum().rename(columns={"Mes": "Mes", "Horas_num": "Horas"})
            else:
                # fallback: si no hay Mes, usar Actividad/Fecha u otra columna
                df_graf_plot = df_graf.rename(columns={horas_col: "Horas"})[["Mes", "Horas"]] if horas_col else pd.DataFrame(columns=["Mes","Horas"])

            # pasar df_graf_plot como df_graf al método
            self.df_graf = df_graf_plot
            self.informe_por_bombero()    
            print(">>> EVENTO FINALIZÓ sin errores")
        except Exception as e:
            self.ui.show_error("Error", f"No se pudo generar el informe:\n{e}")
        finally:
            conn.close()

    # ============================================================
    # 🔹 Informe automático al seleccionar una ACTIVIDAD
    # ============================================================
    #---------- DETALLE POR ACTIVIDAD -------------------------------
    def informe_detalle_por_actividad(self):
        """Informe detallado de la actividad seleccionada."""
        act_sel = self.inf_actividad.get().strip()

        if not act_sel:
            return

        # --- Validación combo legajo ---
        leg_sel = (self.inf_legajo_cb.get() or "").strip()
        if leg_sel:
            messagebox.showwarning(
                "Atención",
                "Debe deseleccionar el legajo antes de elegir una actividad."
            )
            self.inf_actividad.set("")
            return

        # --- Fechas ---
        try:
            f1 = self.inf_desde.get()
            f2 = self.inf_hasta.get()
            f1_sql = datetime.strptime(f1, "%d/%m/%Y").strftime("%Y-%m-%d")
            f2_sql = datetime.strptime(f2, "%d/%m/%Y").strftime("%Y-%m-%d")
        except Exception:
            self.ui.show_error("Error", "Debe seleccionar fechas válidas.")
            return

        # --- ID del concepto ---
        try:
            concepto_id = int(act_sel.split(" - ", 1)[0])
        except Exception:
            self.ui.show_error("Error", "No se pudo obtener el ID del concepto.")
            return

        conn = sqlite3.connect(DB_PATH)
        try:
            q = """
                SELECT a.legajo AS Legajo,
                    COALESCE(l.apellido || ' ' || l.nombre, '') AS Bombero,
                    a.area AS Área,
                    a.fecha_inicio AS Fecha,
                    COALESCE(a.horas, 0) AS horas,
                    COALESCE(a.descripcion, '') AS Descripción
                FROM actividades a
                LEFT JOIN legajos l ON a.legajo = l.legajo
                WHERE a.concepto_id = ?
                AND (
                    CASE 
                    WHEN instr(a.fecha_inicio,'/')>0
                    THEN date(substr(a.fecha_inicio,7,4)||'-'||substr(a.fecha_inicio,4,2)||'-'||substr(a.fecha_inicio,1,2))
                    ELSE date(a.fecha_inicio)
                    END
                ) BETWEEN date(?) AND date(?)
                ORDER BY a.fecha_inicio DESC
            """

            df = pd.read_sql_query(q, conn, params=(concepto_id, f1_sql, f2_sql))

            if df.empty:
                self.ui.show_info("Sin datos", "No hay registros para esa actividad en el período seleccionado.")
                self.inf_actividad.set("")
                return

            # --- df para gráfico ---
            df_graf = df.copy()
            df_graf.rename(columns={"horas": "Horas"}, inplace=True)
            df_graf["Horas"] = pd.to_numeric(df_graf["Horas"], errors="coerce").fillna(0)

            df_graf["Bombero_ID"] = df_graf["Legajo"].astype(str) + " - " + df_graf["Bombero"]

            df_graf = df_graf.groupby(["Bombero", "Bombero_ID"], as_index=False)["Horas"].sum()
            df_graf = df_graf.sort_values("Bombero")
            df_graf = df_graf[["Bombero_ID", "Horas"]]

            self.df_graf = df_graf

            titulo = f"Detalle de {act_sel} ({f1} - {f2})"

            df_norm = df.rename(columns={
                "Área": "area",
                "Fecha": "fecha_inicio",
                "Descripción": "descripcion"
            }).copy()

            df_norm["actividad"] = act_sel

            self._mostrar_form_informe_actividad_individual(
                titulo,
                df_graf,
                df_norm,
                act_sel
            )

        except Exception as e:
            import traceback
            traceback.print_exc()
            self.ui.show_error("Error", f"No se pudo generar el informe:\n{e}")


    #---------- INFORME POR ACTIVIDAD (CONTROLADOR) -----------------
    def informe_por_actividad(self):
        """Decide si mostrar resumen o detalle según la combo de actividad."""
        act_sel = (self.inf_actividad.get() or "").strip()

        if act_sel:
            self.informe_detalle_por_actividad()
        else:
            self.informe_resumen_por_actividad()

    def _llenar_treeview_desde_df(self, tree, df):
        """
        Vacía el treeview y lo llena con el df. Mantiene el orden de columnas del df.
        Espera que 'tree' ya tenga tantas columnas como df.columns (o las ajusta).
        """
        import pandas as pd
        # limpiar
        for it in tree.get_children():
            tree.delete(it)

        cols = list(df.columns)
        # si el tree no tiene las columnas configuradas, establecerlas
        try:
            tree["columns"] = cols
            for c in cols:
                tree.heading(c, text=c)
                tree.column(c, width=100, stretch=True)
        except Exception:
            pass

        # insertar filas (convertir NaN)
        for _, row in df.fillna("").iterrows():
            vals = [("" if pd.isna(row[c]) else row[c]) for c in cols]
            tree.insert("", "end", values=vals)

     # ------------------ Exportaciones estadísticas (añadir dentro de App) ------------------
    def exportar_excel_estadistico(self, df, titulo, columnas=None):
        if df is None or df.empty:
            messagebox.showwarning("Exportar Excel", "No hay datos para exportar.")
            return
        try:
            # Limpiar caracteres no válidos
            titulo_limpio = re.sub(r"[\\/:*?\"<>|]", "-", titulo or "")
            base = self._default_filename(titulo_limpio.replace(" ", "_")).replace(".pdf", ".xlsx")

            path = self.ui.ask_save_file(
                defaultextension=".xlsx",
                initialfile=base,
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            if not path:
                return

            if columnas:
                cols = [c for c in columnas if c in df.columns]
                df_to_save = df.loc[:, cols]
            else:
                df_to_save = df.copy()

            df_to_save.to_excel(path, index=False)
            self.ui.show_info("Exportar Excel", f"Exportado correctamente:\n{path}")
        except Exception as e:
            import traceback
            self.ui.show_error("Error exportar Excel", f"{e}\n\n{traceback.format_exc()}")

    def exportar_pdf_estadistico(self, df, titulo, columnas=None):
        """
        Exporta 'df' a PDF usando el formato institucional unificado (_crear_pdf_unificado).
        """
        if df is None or df.empty:
            messagebox.showwarning("Exportar PDF", "No hay datos para exportar.")
            return

        try:
            # --- limpiar título y preparar nombre ---
            titulo_limpio = re.sub(r"\(Legajo\s*\d+\)", "", titulo or "").strip()
            titulo_limpio = re.sub(r"[\\/:*?\"<>|]", "-", titulo_limpio)

            # columnas a exportar
            if columnas:
                cols = [c for c in columnas if c in df.columns]
            else:
                cols = list(df.columns)

            if not cols:
                self.ui.show_error("Exportar PDF", "No hay columnas para exportar.")
                return

            # decidir orientación
            texto_largo = any(df[c].astype(str).str.len().mean() > 25 for c in cols)
            landscape_mode = len(cols) > 6 or texto_largo

            # sugerir nombre de archivo
            sugerido = self._default_filename(titulo_limpio.replace(" ", "_")).replace(".pdf", ".pdf")
            path = self.ui.ask_save_file(
                defaultextension=".pdf",
                initialfile=sugerido,
                filetypes=[("PDF files", "*.pdf")],
                title="Guardar informe PDF"
            )
            if not path:
                return

            styles = getSampleStyleSheet()
            style_cell = ParagraphStyle(
                name="CellStyle",
                parent=styles["Normal"],
                fontSize=8,
                leading=10,
                alignment=1,
                wordWrap="CJK"
            )

            elems = []
            elems.append(Paragraph(f"<b>{titulo_limpio}</b>", styles["Title"]))
            elems.append(Spacer(1, 8))

            # construir tabla de valores (representamos TOTAL GENERAL con salto de línea si existe)
            data = [cols]
            for _, row in df[cols].iterrows():
                fila = []
                for j, val in enumerate(row):
                    texto = str(val) if val is not None else ""
                    if j == 0 and "TOTAL GENERAL" in texto.upper():
                        texto = "TOTAL<br/>GENERAL"
                    fila.append(Paragraph(texto, style_cell))
                data.append(fila)

            tbl = Table(data, repeatRows=1)
            tbl.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#a50000")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.4, colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('LEFTPADDING', (0, 0), (-1, -1), 2),
                ('RIGHTPADDING', (0, 0), (-1, -1), 2),
            ]))
            elems.append(tbl)

            # --- crear con envoltorio unificado (encabezado/pie) ---
            self._crear_pdf_unificado(path, elems, titulo_limpio, landscape_mode=landscape_mode)

            self.ui.show_info("Exportar PDF", f"Informe exportado correctamente.\n\n{path}")

        except Exception as e:
            self.ui.show_error("Error exportar PDF", f"{e}\n\n{traceback.format_exc()}")

    # Wrappers seguros para usar en los botones (evitan excepciones sin capturar)
    def _cmd_export_excel_safe(self, df, titulo, columnas=None):
        try:
            self.exportar_excel_estadistico(df, titulo, columnas)
        except Exception as e:
            self.ui.show_error("Error", f"Error al exportar Excel: {e}")

    def _cmd_export_pdf_safe(self, df, titulo, columnas=None):
        try:
            self.exportar_pdf_estadistico(df, titulo, columnas)
        except Exception as e:
            self.ui.show_error("Error", f"Error al exportar PDF: {e}")

    # ---------------------------------------------------------
    def informe_por_periodo(self):
        """Muestra total de horas trabajadas por bombero en el rango de fechas elegido."""
        try:
            f1 = self.inf_desde.get_date()
            f2 = self.inf_hasta.get_date()
        except Exception:
            self.ui.show_error("Error", "Debe seleccionar fechas válidas.")
            return

        s1 = f1.strftime("%Y-%m-%d")
        s2 = f2.strftime("%Y-%m-%d")

        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        try:
            # 🔹 Agrupar por bombero y sumar horas (no actividades)
            c.execute("""
                SELECT a.legajo,
                    COALESCE(l.apellido || ' ' || l.nombre, '(Sin nombre)') AS Bombero,
                    ROUND(SUM(COALESCE(a.horas, 0)), 2) AS Total_Horas
                FROM actividades a
                LEFT JOIN legajos l ON a.legajo = l.legajo
                WHERE (CASE WHEN instr(a.fecha_inicio, '/')>0
                            THEN date(substr(a.fecha_inicio,7,4)||'-'||substr(a.fecha_inicio,4,2)||'-'||substr(a.fecha_inicio,1,2))
                            ELSE date(a.fecha_inicio) END)
                    BETWEEN date(?) AND date(?)
                GROUP BY a.legajo, Bombero
                ORDER BY Bombero
            """, (s1, s2))
            rows = c.fetchall()
        except Exception as e:
            conn.close()
            self.ui.show_error("Error", f"No se pudo consultar la base de datos:\n{e}")
            return
        conn.close()

        if not rows:
            self.ui.show_info("Sin datos", "No hay registros en el período seleccionado.")
            return

        df = pd.DataFrame(rows, columns=["Legajo", "Bombero", "Total Horas"])
        titulo = f"Horas por Período ({f1:%d/%m/%Y} - {f2:%d/%m/%Y})"
        self._mostrar_tabla_y_grafico(titulo, df, x_col="Bombero", y_col="Total Horas", color="green")

    def generar_informe_estadistico(self, fecha_desde=None, fecha_hasta=None):
        """Genera informe estadístico: actividades y horas por bombero y por tipo de actividad."""
        # === Pedir fechas si no se pasan como argumento ===
        if not fecha_desde:
            fecha_desde = simpledialog.askstring("Informe", "Desde (dd/mm/yyyy):")
        if not fecha_hasta:
            fecha_hasta = simpledialog.askstring("Informe", "Hasta (dd/mm/yyyy):")
        try:
            f1 = datetime.strptime(fecha_desde, "%d/%m/%Y")
            f2 = datetime.strptime(fecha_hasta, "%d/%m/%Y")
        except Exception:
            self.ui.show_error("Error", "Fechas inválidas.")
            return

        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("""
            SELECT a.legajo, l.apellido, l.nombre,
                   a.actividad, a.horas, a.fecha_inicio
            FROM actividades a
            LEFT JOIN legajos l ON a.legajo = l.legajo
            WHERE date(a.fecha_inicio) BETWEEN date(?) AND date(?)
        """, (f1.strftime("%Y-%m-%d"), f2.strftime("%Y-%m-%d")))
        rows = c.fetchall()
        conn.close()

        if not rows:
            self.ui.show_info("Sin datos", "No hay actividades en el período seleccionado.")
            return

        df = pd.DataFrame(rows, columns=["Legajo", "Apellido", "Nombre", "Actividad", "Horas", "Fecha"])

        # --- Resumenes ---
        resumen_bombero = df.groupby(["Legajo", "Apellido", "Nombre"]).agg(
            Actividades=("Actividad", "count"),
            Total_Horas=("Horas", "sum")
        ).reset_index()

        resumen_actividad = df.groupby(["Actividad"]).agg(
            Cantidad=("Actividad", "count"),
            Total_Horas=("Horas", "sum")
        ).reset_index()

        # --- Mostrar resumenes ---
        win = Toplevel(self.master)
        win.title(f"Informe de Actividades ({fecha_desde} a {fecha_hasta})")
        win.geometry("900x500")

        ttk.Label(win, text=f"Período: {fecha_desde} a {fecha_hasta}",
                  font=("Arial", 10, "italic")).pack(pady=5)

        # --- Grilla 1: por bombero ---
        ttk.Label(win, text="Resumen por Bombero", font=("Arial", 12, "bold")).pack()
        tree1 = ttk.Treeview(win, columns=list(resumen_bombero.columns), show="headings", height=8)
        for col in resumen_bombero.columns:
            tree1.heading(col, text=col)
            tree1.column(col, width=120)
        for _, row in resumen_bombero.iterrows():
            tree1.insert("", "end", values=list(row))
        tree1.pack(pady=5, fill="x")

        # --- Grilla 2: por tipo de actividad ---
        ttk.Label(win, text="Resumen por Actividad", font=("Arial", 12, "bold")).pack(pady=(15, 0))
        tree2 = ttk.Treeview(win, columns=list(resumen_actividad.columns), show="headings", height=8)
        for col in resumen_actividad.columns:
            tree2.heading(col, text=col)
            tree2.column(col, width=150)
        for _, row in resumen_actividad.iterrows():
            tree2.insert("", "end", values=list(row))
        tree2.pack(pady=5, fill="x")

        self.ui.show_info("OK", "Informe generado con éxito.")

    def mostrar_informe_filtrado(self, titulo, df_graf, df_detalle, x_col, y_col, color="red"):
        """
        Muestra una ventana con el mismo formato visual que los informes principales
        (rojo bombero, logo, totales, botones exportar PDF/Excel/Cerrar)
        para el caso filtrado (cuando se selecciona opción en combo).
        """

        # --- Crear ventana ---
        win = Toplevel(self.master)
        win.title(titulo)
        win.geometry("1000x600")
        win.configure(bg="white")

        # === LOGO y TÍTULO ===
        if os.path.exists(LOGO_PATH):
            try:
                from PIL import Image, ImageTk
                img = Image.open(LOGO_PATH).resize((45, 45))
                logo = ImageTk.PhotoImage(img)
                Label(win, image=logo, bg="white").place(x=15, y=10)
                win.logo_ref = logo
            except:
                pass

        Label(
            win,
            text="SOCIEDAD BOMBEROS VOLUNTARIOS DE ALMAFUERTE",
            font=("Arial", 16, "bold"),
            fg="red",
            bg="white"
        ).place(x=70, y=15)

        Label(
            win,
            text=titulo,
            font=("Arial", 11, "bold"),
            fg="black",
            bg="white"
        ).place(x=70, y=45)

        # === GRÁFICO ===
        frame_graf = tk.Frame(win, bg="white")
        frame_graf.place(x=30, y=90, width=520, height=300)

        fig, ax = plt.subplots(figsize=(6, 3))
        if not df_graf.empty:
            ax.bar(range(len(df_graf[x_col])), df_graf[y_col], color=color)
            ax.set_xlabel(x_col)
            ax.set_ylabel(y_col)
            ax.set_xticks(range(len(df_graf[x_col])))
            ax.set_xticklabels(df_graf[x_col], rotation=45, ha="right")
            ax.set_title("")
        canvas = FigureCanvasTkAgg(fig, master=frame_graf)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)

        # === TOTALES ===
        total_actividades = len(df_detalle)
        total_horas = 0.0
        col_horas = None

        # Buscar la columna que represente horas (en cualquier formato)
        for posible in ["Horas", "horas", "Total_Horas", "Total Horas"]:
            if df_detalle is not None and posible in df_detalle.columns:
                col_horas = posible
                break

        # Si encontramos una, la sumamos
        if col_horas:
            for h in df_detalle[col_horas]:
                try:
                    total_horas += self._horas_a_decimal(str(h))
                except Exception:
                    pass

        Label(win, text=f"Total registros: {total_actividades}", font=("Arial", 11, "bold"), bg="white", fg="red").place(x=600, y=120)
        Label(win, text=f"Total horas: {self._fmt_horas(total_horas)}", font=("Arial", 11, "bold"), bg="white", fg="red").place(x=600, y=150)

        # === TREEVIEW ===
        frame_tabla = tk.Frame(win, bg="white")
        frame_tabla.place(x=30, y=400, width=930, height=160)

        if not df_detalle.empty:
            cols = list(df_detalle.columns)
            tree = ttk.Treeview(frame_tabla, columns=cols, show="headings")
            for c in cols:
                tree.heading(c, text=c)
                tree.column(c, width=100, anchor="center")
            for _, row in df_detalle.iterrows():
                tree.insert("", "end", values=list(row))
            tree.pack(fill="both", expand=True)
            win.tree = tree

        # === BOTONES ===
        if df_detalle.empty:
            self.ui.show_info(
                "Sin datos",
                f"No hay actividades para {leg_sel} en el período seleccionado."
            )
            return None   # <<< importante

            fpath = self.ui.ask_save_file(defaultextension=".xlsx",
                                                filetypes=[("Excel", "*.xlsx")],
                                                title="Guardar Excel filtrado")
            if not fpath:
                return
            df_detalle.to_excel(fpath, index=False)
            self.ui.show_info("Éxito", f"Archivo guardado:\n{fpath}")

        def export_pdf():
            if df_detalle.empty:
                self.ui.show_info("Sin datos", "No hay datos para exportar.")
                return
            fpath = self.ui.ask_save_file(defaultextension=".pdf",
                                                filetypes=[("PDF", "*.pdf")],
                                                title="Guardar PDF filtrado")
            if not fpath:
                return
            self._crear_pdf_resumen_informe(fpath, titulo, df_detalle, df_detalle)
            self.ui.show_info("Éxito", f"PDF generado:\n{fpath}")

        Button(win, text="Exportar Excel", bg="white", fg="red", font=("Arial", 10, "bold"),
            command=export_excel, width=15).place(x=600, y=200)
        Button(win, text="Exportar PDF", bg="white", fg="red", font=("Arial", 10, "bold"),
            command=export_pdf, width=15).place(x=600, y=240)
        Button(win, text="Cerrar", bg="white", fg="red", font=("Arial", 10, "bold"),
            command=win.destroy, width=15).place(x=600, y=280)

        return win

    def buscar_informes(self):
        try:
            conn = sqlite3.connect(DB_PATH)
            c = conn.cursor()

            query = """
                SELECT a.id, a.legajo, a.actividad, a.area,
                    a.fecha_inicio, a.fecha_fin, a.hora_inicio, a.hora_fin,
                    COALESCE(a.horas, 0),
                    a.descripcion, l.apellido, l.nombre
                FROM actividades a
                LEFT JOIN legajos l ON a.legajo = l.legajo
                WHERE 1=1
            """
            params = []

            # --- Filtros ---
            if self.inf_id.get().strip():
                query += " AND a.id=?"
                params.append(self.inf_id.get().strip())

            if self.inf_actividad.get().strip():
                query += " AND a.actividad=?"
                params.append(self.inf_actividad.get().strip())

            if self.inf_legajo_cb.get().strip() and self.inf_legajo_cb.get().strip() != "()":
                legajo = self.inf_legajo_cb.get().split(" - ")[0]
                query += " AND a.legajo=?"
                params.append(legajo)

            if self.inf_apynom.get().strip():
                query += " AND (UPPER(l.apellido) LIKE ? OR UPPER(l.nombre) LIKE ?)"
                like = f"%{self.inf_apynom.get().strip().upper()}%"
                params.extend([like, like])

            desde = parse_ddmmyyyy(self.inf_desde.get())
            hasta = parse_ddmmyyyy(self.inf_hasta.get())
            if desde and hasta:
                query += """
                    AND (date(substr(a.fecha_inicio,7,4)||'-'||substr(a.fecha_inicio,4,2)||'-'||substr(a.fecha_inicio,1,2))
                        BETWEEN ? AND ?)
                """
                params.extend([desde.strftime("%Y-%m-%d"), hasta.strftime("%Y-%m-%d")])

            query += " ORDER BY a.id DESC"

            # --- Ejecutar y cerrar conexión ---
            c.execute(query, params)
            rows = c.fetchall()
            conn.close()

            # --- Limpiar Treeview ---
            for item in self.inf_tree.get_children():
                self.inf_tree.delete(item)

            # --- Insertar filas ---
            total_horas = 0
            for i, r in enumerate(rows):
                tag = 'even' if i % 2 == 0 else 'odd'
                self.inf_tree.insert("", "end", values=r, tags=(tag,))
                try:
                    total_horas += float(r[8]) if r[8] else 0
                except:
                    pass

            # --- Ajustar columnas ---
            font = tkFont.nametofont("TkDefaultFont")

            for col in self.inf_tree["columns"]:
                # ancho base = título
                max_width = font.measure(col)

                for item in self.inf_tree.get_children():
                    val = str(self.inf_tree.set(item, col))
                    ancho = font.measure(val)
                    if ancho > max_width:
                        max_width = ancho

                # 🔥 padding extra
                max_width += 20

                # 🔥 regla especial para Descripción
                if col == "Descripción":
                    max_width = min(max_width, 600)
                else:
                    max_width = min(max_width, 200)

                self.inf_tree.column(col, width=max_width)

            # --- Mostrar totales debajo de la tabla ---
            total_registros = len(rows)
            texto_totales = f"Total de registros: {total_registros} | Total de horas: {total_horas:.2f}"
            if hasattr(self, "lbl_totales"):
                self.lbl_totales.config(text=texto_totales)
            else:
                print(texto_totales)  # fallback si no está creada la etiqueta

            self.ui.show_info("Resultados", f"Se encontraron {total_registros} registros.")

        except Exception as e:
            self.ui.show_error("Error", f"No se pudo buscar: {e}")

    def limpiar_informes(self):
        self.inf_id.delete(0, END); self.inf_actividad.set("")
        self.inf_legajo_cb.set(""); self.inf_apynom.delete(0, END)
        self.inf_desde.set_date(date.today()); self.inf_hasta.set_date(date.today())
        for i in self.inf_tree.get_children(): self.inf_tree.delete(i)

    def _datos_grilla_informes(self):
        return [self.inf_tree.item(i, "values") for i in self.inf_tree.get_children()]

    def exportar_excel(self):
        import pandas as pd
        from datetime import datetime

        columnas = self.inf_tree["columns"]
        datos = []

        for item in self.inf_tree.get_children():
            valores = self.inf_tree.item(item, "values")
            datos.append(valores)

        if not datos:
            self.ui.show_info("Sin datos", "No hay datos para exportar.")
            return

        # 🔧 NORMALIZAR DATOS
        datos_limpios = []
        for fila in datos:
            fila = list(fila)

            if len(fila) > len(columnas):
                fila = fila[:len(columnas)]
            elif len(fila) < len(columnas):
                fila += [""] * (len(columnas) - len(fila))

            datos_limpios.append(fila)

        df = pd.DataFrame(datos_limpios, columns=columnas)

        fecha_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
        sugerido = f"informe_{fecha_str}.xlsx"

        ruta = self.ui.ask_save_file(
            defaultextension=".xlsx",
            initialfile=sugerido,
            filetypes=[("Excel", "*.xlsx")],
            title="Guardar Informe Excel"
        )

        if ruta:
            df.to_excel(ruta, index=False)
            self.ui.show_info("Exportar a Excel", f"Archivo guardado en:\n{ruta}")

    def _on_informe_double_click(self, event=None):
        """Imprime la actividad seleccionada o el listado completo si no hay selección."""
        try:
            selected = self.inf_tree.selection()
            if selected:
                item = selected[0]
                valores = self.inf_tree.item(item, "values")
                if valores:
                    id_act = valores[0]  # primera columna = ID
                    self.var_id_actividad.set(str(id_act))
                    self.imprimir_actividad()
                    return
            # Si no hay selección, imprime todo el listado
            self.imprimir_listado_actividades()
        except Exception as e:
            self.ui.show_error("Error", f"No se pudo imprimir: {e}")

    def exportar_pdf(self):
        rows = self._datos_grilla_informes()

        if not rows:
            messagebox.showwarning("Atención", "No hay datos para exportar.")
            return

        rows = self._datos_grilla_informes()

        # 🔥 Detectar si hay filtro por persona
        filtro_apynom = self.inf_apynom.get().strip()
        filtro_id = self.inf_id.get().strip()

        if rows and (filtro_apynom or filtro_id):
            primer = rows[0]

            try:
                cols = list(self.inf_tree["columns"])
                idx_apellido = cols.index("Apellido")
                idx_nombre = cols.index("Nombre")

                apellido = str(primer[idx_apellido]).strip()
                nombre = str(primer[idx_nombre]).strip()

                nombre_base = f"{apellido}_{nombre}"
            except:
                nombre_base = "informe"
        else:
            # 🔥 SIN FILTRO → nombre genérico
            nombre_base = "informe_general"

        nombre_base = nombre_base.replace(" ", "_").lower()

        fecha_actual = datetime.now().strftime("%Y-%m-%d")
        sugerido = f"{nombre_base}_{fecha_actual}.pdf"

        file_path = self.ui.ask_save_file(
            defaultextension=".pdf",
            initialfile=sugerido,
            filetypes=[("PDF files", "*.pdf")],
            title="Guardar Informe PDF"
        )
        if not file_path:
            return

        styles = getSampleStyleSheet()
        elems = []

        elems.append(Paragraph("<b>INFORME DE ACTIVIDADES</b>", styles["Title"]))
        elems.append(Spacer(1, 10))
        elems.append(Spacer(1, 8))

        cols = self.inf_tree["columns"]

        # 🔧 NORMALIZAR FILAS
        data_rows = []
        for r in rows:
            fila = list(r)

            if len(fila) > len(cols):
                fila = fila[:len(cols)]
            elif len(fila) < len(cols):
                fila += [""] * (len(cols) - len(fila))

            data_rows.append(fila)

        data = [cols] + data_rows

        tabla = Table(data, repeatRows=1)
        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#a50000")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
        ]))

        elems.append(tabla)

        # ===== TOTALES SEGUROS =====
        total = len(rows)

        elems.append(Spacer(1, 15))
        elems.append(Paragraph(f"<b>Total registros:</b> {total}", styles["Normal"]))

        try:
            self._crear_pdf_unificado(
                buffer_or_path=file_path,
                elems=elems,
                titulo="",
                landscape_mode=True
            )
            self.ui.show_info("Éxito", f"Informe guardado en:\n{file_path}")
        except Exception as e:
            self.ui.show_error("Error", f"No se pudo exportar el PDF:\n{e}")

    def _default_informe_filename(self, tipo):
        """Genera un nombre de archivo para informes con fecha."""
        from datetime import datetime
        fecha = datetime.now().strftime("%Y-%m-%d")
        tipo = tipo.replace(" ", "_")
        return f"Informe_{tipo}_{fecha}.pdf"

    # -------------------- utils: navegación ENTER y text handling --------------------
    def _focus_next_widget(self, widget):
        """
        Mover foco al siguiente widget en el orden natural.
        Para widgets en actividades usamos self._actividades_order cuando corresponda.
        """
        try:
            # si está en actividades y pertenece al orden definido, usar ese orden
            if hasattr(self, '_actividades_order') and widget in self._actividades_order:
                lst = self._actividades_order
                try:
                    idx = lst.index(widget)
                    next_widget = lst[idx + 1]
                    next_widget.focus_set()
                except (ValueError, IndexError):
                    # si no hay siguiente, poner foco en primer botón Guardar
                    self.act_btns.get("Guardar", None).focus_set()
            else:
                widget.tk_focusNext().focus_set()
        except Exception:
            pass

    def _descripcion_enter(self, event):
        try:
            self.act_btns["Guardar"].focus_set()
        except Exception:
            pass
        return "break"


    def _descripcion_shift_enter(self, event):
        # permite salto de línea normal
        return None
    
    def salir(self):
        """Cierra la aplicación limpiamente evitando bloqueos."""
        import os
        try:
            # Evitar que siga ejecutándose _aplicar_permisos
            self._cerrando = True

            if hasattr(self, "_after_permisos_id"):
                try:
                    self.master.after_cancel(self._after_permisos_id)
                except Exception:
                    pass

            # Destruir ventana principal
            self.master.destroy()
        except Exception:
            # Si por alguna razón tkinter ya está destruido
            os._exit(0)

def set_icon(window):
    from PIL import Image, ImageTk
    import os, sys
    
    try:
        # Usar la lógica de base_dir que ya tenías
        if hasattr(sys, "_MEIPASS"):
            base_dir = sys._MEIPASS
        else:
            base_dir = os.path.dirname(os.path.abspath(__file__))
            
        ruta = os.path.join(base_dir, "Bomberos.png")
        
        if os.path.exists(ruta):
            # Forzamos con PIL
            img = Image.open(ruta).resize((32, 32), Image.Resampling.LANCZOS)
            render = ImageTk.PhotoImage(img)
            # ANCLA: Sin esto, verás el rectángulo blanco
            window._icon_ref = render 
            window.iconphoto(True, render)
            print(f">>> Icono cargado correctamente desde: {ruta}")
        else:
            print(f"⚠ ERROR: No se encuentra el archivo en {ruta}")
    except Exception as e:
        print(f"⚠ ERROR CRITICO ICONO: {e}")

# -------------------- MAIN --------------------
if __name__ == "__main__":
    import tkinter as tk   # 🔴 SIEMPRE ARRIBA
    from ui_helpers import UIHelpers # Asegurate de tener el import

    # 1. Inicializar base de datos MySQL
    init_db()

    # 2. Crear el root principal
    root = tk.Tk()
    root.title("Agenda de Bomberos")

    # 🔴 ICONO ROOT
    root._icono_global = cargar_icono_global()
    if root._icono_global:
        root.iconphoto(True, root._icono_global)

    # 3. OCULTAR root
    root.withdraw() 

    # --- PASO NUEVO: Crear el ayudante de interfaz ---
    ui = UIHelpers(root) 

    print(">>> Mostrando login")

    # 4. Login (Ahora pasamos 'root' y 'ui')
    login = LoginWindow(root, ui) 
    root.wait_window(login.top)

    # 5. Verificar login
    if not hasattr(login, 'logged_user') or not login.logged_user:
        print(">>> Login cancelado o fallido, cerrando")
        root.destroy()
        sys.exit()

    print(f">>> Usuario autenticado: {login.logged_user}")

    # 6. App principal (Pasamos también el ui para que App lo use)
    app = App(root, login.logged_user, ui) 
    
    # 7. Mostrar ventana
    root.deiconify() 
    root.after(200, lambda: root.focus_force())
    
    # 8. Loop
    root.mainloop()